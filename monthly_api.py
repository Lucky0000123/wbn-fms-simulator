"""Monthly plan roll-up + comparison (owner 2026-08-13).

"when we make a plan, and it is for one day ... we want to use continuous
plan for the whole month ... we have an Excel file in which we made a plan
for the whole month ... you have to compare these plans ... make an Excel
file out of this, which showing our prediction model plans and the normal
plans, and also the charts on that one. So better make a new page for that."

Design:
  • A month is built from ONE saved daily plan (data/saved_plans/DATE.json):
    the same holding plan runs every day of the month. Day = 2 × 12 h shifts,
    so the daily figure is the saved per-shift prediction × 2 (the plan page's
    meta.predict totals are per shift). Nothing is invented beyond that.
  • The "manual" plan (the owner's Excel, made from historical trips with no
    variables) is uploaded as .xlsx/.csv or pasted; per-day rows are matched
    by date. Its numbers are stored verbatim.
    • Comparison + export: Key sheet (target, predicted plan
    predicted plan) plus one sheet per month. Default Excel hides achievable.
    ⬇ with achievable (achv=1) adds old / optimized simulate next to predicted,
    matching Plan Allocate (Your plan vs after Allocate). Never averaged.
    Charts sit under their tables. White report except coverage badges.

State lives in data/monthly_plans/YYYY-MM.json, one file per month, same
local-disk pattern as saved_plans.
"""

import calendar
import io
import json
import math
import os
import re
import threading
from datetime import date, datetime

from flask import Blueprint, jsonify, request, send_file

bp = Blueprint("monthly_api", __name__)

_ROOT = os.path.dirname(os.path.abspath(__file__))
_SAVED_DIR = os.path.join(_ROOT, "data", "saved_plans")
_MONTH_DIR = os.path.join(_ROOT, "data", "monthly_plans")


def _month_path(month):
    # \d{2} alone let "2026-13" through to calendar.monthrange -> 500
    # (found in the failure-mode battery). Month must be 01-12.
    if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", month or ""):
        return None
    return os.path.join(_MONTH_DIR, month + ".json")


def _load_state(month):
    p = _month_path(month)
    if not p or not os.path.isfile(p):
        return None
    try:
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError):
        return None


def _save_state(month, state):
    p = _month_path(month)
    os.makedirs(_MONTH_DIR, exist_ok=True)
    tmp = p + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=1)
        f.write("\n")
    os.replace(tmp, p)


def _days_in(month):
    y, m = int(month[:4]), int(month[5:7])
    return [date(y, m, d).isoformat() for d in range(1, calendar.monthrange(y, m)[1] + 1)]


@bp.route("/api/monthly/state")
def api_monthly_state():
    month = (request.args.get("month") or "").strip()
    if not _month_path(month):
        return jsonify({"ok": False, "error": "supply month=YYYY-MM"}), 400
    st = _load_state(month)
    n = len(_days_in(month))
    alloc, src = _resolve_allocation(month, st)
    view = _alloc_view(alloc, n, src, include_detail=True)
    return jsonify({
        "ok": True, "month": month, "state": st, "exists": st is not None,
        "alloc": view,
    })


@bp.route("/api/monthly/months")
def api_monthly_months():
    out = []
    if os.path.isdir(_MONTH_DIR):
        for f in sorted(os.listdir(_MONTH_DIR), reverse=True):
            if re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])\.json", f):
                out.append(f[:-5])
    return jsonify({"ok": True, "months": out})


def _simulate_holding(plan):
    """Same engine as Plan → Check capacity. Returns simulate payload or None."""
    raw = plan.get("paths") or {}
    items = raw.values() if isinstance(raw, dict) else raw
    plans = []
    for p in items:
        if not isinstance(p, dict) or p.get("foreign"):
            continue
        key = (p.get("key") or "").strip()
        if ">" not in key:
            continue
        src, dst = key.split(">", 1)
        dt = float(p.get("dt") or 0)
        if dt <= 0:
            continue
        plans.append({"route": key, "source": p.get("source") or src,
                      "destination": p.get("dest") or dst,
                      "n_trucks": int(round(dt)),
                      "contractor": p.get("contractor") or ""})
    if not plans:
        return None
    import plan_simulator
    return plan_simulator.simulate({"plans": plans})


def _achv_day_from_pred(p):
    """Plan-tab achievable (shift) × 2, or the stored per-day field."""
    if not p:
        return None
    if p.get("per_day_achv_wmt") is not None:
        return p["per_day_achv_wmt"]
    sh = p.get("sim_achievable_shift")
    if sh is None:
        return None
    try:
        return round(float(sh) * 2)
    except (TypeError, ValueError):
        return None


def _fill_achv_from_plan(p):
    """Run Plan Check capacity on stored paths; write achievable onto the prediction."""
    if not p or _achv_day_from_pred(p) is not None:
        return _achv_day_from_pred(p)
    paths = p.get("paths") or []
    if not paths:
        return None
    sim = _simulate_holding({"paths": {str(i): row for i, row in enumerate(paths)}})
    if not sim or sim.get("error"):
        return None
    sh = (sim.get("summary") or {}).get("achievable_production_t")
    if sh is None:
        return None
    try:
        day = round(float(sh) * 2)
    except (TypeError, ValueError):
        return None
    p["per_day_achv_wmt"] = day
    p["sim_achievable_shift"] = sh
    results = sim.get("results") or []
    for i, row in enumerate(paths):
        if not isinstance(row, dict):
            continue
        if i < len(results) and row.get("achv_wmt_day") is None:
            row["achv_wmt_day"] = round(float(results[i].get("achievable_production_t") or 0) * 2)
    for d in p.get("days") or []:
        if d.get("wmt_achv") is None:
            d["wmt_achv"] = day
    return day


@bp.route("/api/monthly/build", methods=["POST"])
def api_monthly_build():
    """Roll one saved daily plan across a month (prediction side)."""
    body = request.get_json(silent=True) or {}
    month = (body.get("month") or "").strip()
    src_date = (body.get("source_date") or "").strip()
    if not _month_path(month):
        return jsonify({"ok": False, "error": "supply month=YYYY-MM"}), 400
    sp = os.path.join(_SAVED_DIR, src_date + ".json")
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", src_date) or not os.path.isfile(sp):
        return jsonify({"ok": False, "error": "no saved plan for %s" % src_date}), 404
    with open(sp, encoding="utf-8") as f:
        plan = json.load(f)
    meta = plan.get("meta") or {}
    pred = meta.get("predict") or {}
    per_shift_wmt = float(pred.get("wmt") or 0)
    per_shift_trips = float(pred.get("trips") or 0)
    achv_shift = meta.get("sim_achievable")
    if achv_shift is None:
        sim = _simulate_holding(plan)
        if sim and not sim.get("error"):
            achv_shift = (sim.get("summary") or {}).get("achievable_production_t")
    if not per_shift_wmt:
        return jsonify({"ok": False, "error": "saved plan for %s carries no prediction totals - open it in the Plan tab and re-save" % src_date}), 400
    # Day = 2 x 12 h shifts. Same plan every day; rain is whatever the plan was
    # saved with (labelled, not hidden).
    day_wmt = per_shift_wmt * 2
    day_trips = per_shift_trips * 2
    day_achv = round(float(achv_shift) * 2) if achv_shift is not None else None
    days = [{"date": d, "wmt": round(day_wmt), "trips": round(day_trips),
             "wmt_achv": day_achv} for d in _days_in(month)]
    st = _load_state(month) or {"month": month}
    st["prediction"] = {
        "source_date": src_date,
        "per_shift_wmt": round(per_shift_wmt),
        "per_day_wmt": round(day_wmt),
        "per_day_achv_wmt": day_achv,
        "dt": pred.get("dt"),
        "rain_mm": plan.get("rain_mm"),
        "sim_achievable_shift": achv_shift,
        "paths": [
            {"key": p.get("key"), "contractor": p.get("contractor"),
             "dt": p.get("dt"), "material": p.get("material"),
             "otype": p.get("otype"), "targetWmt": p.get("targetWmt"),
             "dt_before": ((p.get("_preAlloc") or {}) if isinstance(p.get("_preAlloc"), dict) else {}).get("dt"),
             "wbSel": p.get("wbSel")}
            for p in ((plan.get("paths") or {}).values()
                      if isinstance(plan.get("paths"), dict)
                      else (plan.get("paths") or []))
            if isinstance(p, dict)
        ],
        "days": days,
        "built_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "note": ("Same holding plan every day; day = 2 x 12 h shifts x saved "
                 "per-shift prediction. Target is the matrix. "
                 "Rain fixed at the saved plan's value."),
    }
    day_alloc = plan.get("allocation")
    if isinstance(day_alloc, dict) and day_alloc.get("frozen"):
        st["prediction"]["allocation"] = day_alloc
        st["saved_day_allocation"] = day_alloc
    _save_state(month, st)
    return jsonify({"ok": True, "month": month, "state": st})


def _parse_manual_rows(rows, month):
    """[[cell,...],...] -> [{date, wmt, trips?}] for days in the month.

    Header detection is forgiving: a date-ish column plus a tonnage-ish column
    (wmt/tonnage/tons/production). Day numbers (1..31) also accepted."""
    days = _days_in(month)
    out, header = {}, None
    for r in rows:
        cells = ["" if c is None else str(c).strip() for c in r]
        if not any(cells):
            continue
        low = [c.lower() for c in cells]
        if header is None and any(re.search(r"date|day|tanggal", c) for c in low) \
           and any(re.search(r"wmt|ton|prod", c) for c in low):
            di = next(i for i, c in enumerate(low) if re.search(r"date|day|tanggal", c))
            wi = next(i for i, c in enumerate(low) if re.search(r"wmt|ton|prod", c))
            ti = next((i for i, c in enumerate(low) if re.search(r"trip", c)), None)
            header = (di, wi, ti)
            continue
        di, wi, ti = header if header else (0, 1, None)
        if len(cells) <= max(di, wi):
            continue
        dcell, wcell = cells[di], cells[wi]
        iso = None
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", dcell)
        if m:
            iso = m.group(0)
        elif re.fullmatch(r"\d{1,2}(\.0)?", dcell):
            n = int(float(dcell))
            if 1 <= n <= len(days):
                iso = days[n - 1]
        else:
            m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", dcell)
            if m:  # d/m/yyyy
                try:
                    iso = date(int(m.group(3)), int(m.group(2)), int(m.group(1))).isoformat()
                except ValueError:
                    pass
        if not iso or not iso.startswith(month):
            continue
        try:
            wmt = float(re.sub(r"[^\d.\-]", "", wcell) or 0)
        except ValueError:
            continue
        rec = {"date": iso, "wmt": round(wmt)}
        if ti is not None and len(cells) > ti and cells[ti]:
            try:
                rec["trips"] = round(float(re.sub(r"[^\d.\-]", "", cells[ti])))
            except ValueError:
                pass
        out[iso] = rec
    return [out[d] for d in days if d in out]


@bp.route("/api/monthly/manual", methods=["POST"])
def api_monthly_manual():
    """Manual (Excel) plan: file upload (xlsx/csv) or pasted rows."""
    month = (request.form.get("month") or (request.get_json(silent=True) or {}).get("month") or "").strip()
    if not _month_path(month):
        return jsonify({"ok": False, "error": "supply month=YYYY-MM"}), 400
    rows, src_name = [], None
    f = request.files.get("file")
    if f and f.filename:
        src_name = f.filename
        data = f.read()
        if not data:
            return jsonify({"ok": False, "error": "the file is empty"}), 400
        if f.filename.lower().endswith((".xlsx", ".xlsm")):
            from openpyxl import load_workbook
            try:
                wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            except Exception:  # noqa: BLE001 — corrupt/renamed file, not an xlsx
                return jsonify({"ok": False, "error": "could not read that file as .xlsx - is it really an Excel file?"}), 400
            ws = wb.active
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
        else:  # csv / tsv
            text = data.decode("utf-8", "replace")
            lines = text.splitlines()
            if not lines:
                return jsonify({"ok": False, "error": "the file is empty"}), 400
            sep = "\t" if "\t" in lines[0] else ","
            rows = [line.split(sep) for line in lines]
    else:
        body = request.get_json(silent=True) or {}
        text = body.get("pasted") or ""
        src_name = "pasted"
        sep = "\t" if "\t" in text else ","
        rows = [line.split(sep) for line in text.splitlines()]
    days = _parse_manual_rows(rows, month)
    if not days:
        return jsonify({"ok": False, "error": "could not find any %s rows - need a date (or day number 1-31) column and a WMT/tonnage column" % month}), 400
    st = _load_state(month) or {"month": month}
    st["manual"] = {"source": src_name, "days": days,
                    "loaded_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")}
    _save_state(month, st)
    return jsonify({"ok": True, "month": month, "state": st, "parsedDays": len(days)})


@bp.route("/api/monthly/clear", methods=["POST"])
def api_monthly_clear():
    body = request.get_json(silent=True) or {}
    month = (body.get("month") or "").strip()
    which = body.get("which")  # 'prediction' | 'manual' | 'all'
    p = _month_path(month)
    if not p:
        return jsonify({"ok": False, "error": "supply month=YYYY-MM"}), 400
    st = _load_state(month)
    if st:
        if which == "all" and os.path.isfile(p):
            os.remove(p)
            st = None
        elif which in ("prediction", "manual"):
            st.pop(which, None)
            _save_state(month, st)
    return jsonify({"ok": True, "month": month, "state": st})


# Line colours for charts: gold target, slate old plan, navy optimized.
# Table numbers are always ink (black) — never red.
_XLSX_PRED, _XLSX_TGT = "1F2937", "EAB308"
# Last workbook sheet — same two Congestion-tab corridor curves, with the
# sweep table so a reader can see how tonnes are priced. J72 treats this as
# a non-month sheet (with Year / Paths / Road crowding).
_XLSX_SAT_SHEET = "Saturation"
_XLSX_NAVY, _XLSX_MUTED, _XLSX_INK = "1F4E79", "64748B", "1F2937"
_XLSX_ACHV = "1F2937"
# Target, predicted — (hex, dotted). No red.
_XLSX_CLOCKS = (
    ("EAB308", False),
    ("64748B", True),
    ("1F4E79", False),
)
# Same three, then old achievable (dotted green) and optimized achievable.
_XLSX_CLOCKS_ACHV3 = (
    _XLSX_CLOCKS[0],       # target
    _XLSX_CLOCKS[2],       # optimized predicted
    ("059669", False),     # optimized achievable
)
_XLSX_CLOCKS_ACHV = _XLSX_CLOCKS + (
    ("86EFAC", True),
    ("059669", False),
)


def _xlsx_send(wb, name):
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    # Excel "file format is not valid" is what you get when a timeout or
    # JSON error is saved with an .xlsx name. Refuse to ship a non-zip.
    if raw[:2] != b"PK":
        return jsonify({
            "ok": False,
            "error": "workbook did not serialise as Excel (got %r)" % (raw[:40],),
        }), 500
    if not str(name).lower().endswith(".xlsx"):
        name = "%s.xlsx" % name
    buf.seek(0)
    rv = send_file(
        buf, as_attachment=True, download_name=name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # Quoted filename so Safari/Excel do not treat the download as a nameless
    # blob; nosniff so a proxy cannot relabel this as text/html.
    rv.headers["Content-Disposition"] = 'attachment; filename="%s"' % name
    rv.headers["X-Content-Type-Options"] = "nosniff"
    rv.headers["Content-Length"] = str(len(raw))
    return rv


def _xlsx_font(bold=False, size=11, color=None):
    from openpyxl.styles import Font
    return Font(name="Calibri", bold=bold, size=size, color=color or _XLSX_INK)


def _xlsx_card_border(color):
    """Thin box with a coloured top bar — the Year / month scorecard look."""
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color="D0D5DD")
    bar = Side(style="medium", color=color or _XLSX_NAVY)
    return Border(left=thin, right=thin, top=bar, bottom=thin)


def _xlsx_sides():
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color="D0D5DD")
    med = Side(style="medium", color=_XLSX_NAVY)
    return Border(left=thin, right=thin, top=thin, bottom=thin), Border(
        left=thin, right=thin, top=thin, bottom=med)


def _xlsx_sheet_setup(ws):
    from openpyxl.worksheet.page import PageMargins
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.5,
                                  header=0.25, footer=0.25)
    ws.print_options.horizontalCentered = True
    ws.oddFooter.left.text = "WBN FMS · monthly plan"
    ws.oddFooter.right.text = "Page &P of &N"


def _xlsx_headers(ws, row, headers, start=1, center=False):
    from openpyxl.styles import Alignment
    _box, head = _xlsx_sides()
    mid = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for i, h in enumerate(headers, start=start):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _xlsx_font(True, 10, _XLSX_NAVY)
        if center:
            c.alignment = mid
        else:
            c.alignment = Alignment(wrap_text=True, vertical="center",
                                    horizontal="right" if i > start else "left")
        c.border = head
    ws.row_dimensions[row].height = 26


def _xlsx_mid():
    from openpyxl.styles import Alignment
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _xlsx_num(cell, value, bold=False, center=False):
    cell.value = value
    cell.font = _xlsx_font(bold, 11)
    cell.number_format = "#,##0"
    cell.border = _xlsx_sides()[0]
    if center:
        cell.alignment = _xlsx_mid()


def _xlsx_text(cell, value, bold=False, color=None, size=11, center=False):
    cell.value = value
    cell.font = _xlsx_font(bold, size, color)
    cell.border = _xlsx_sides()[0]
    if center:
        cell.alignment = _xlsx_mid()


def _xlsx_widths(ws, widths, start=1):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=start):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = max(
            ws.column_dimensions[letter].width or 0, w)


def _xlsx_paint_lines(chart, specs):
    """specs: hex strings or (hex, dotted) tuples. Dotted = old plan only.
    Target (gold, solid) is a hard line — thicker, no dash."""
    for s, spec in zip(chart.series, specs):
        if isinstance(spec, (tuple, list)):
            col = spec[0]
            dotted = bool(spec[1]) if len(spec) > 1 else False
        else:
            col, dotted = spec, False
        hard = (not dotted) and str(col).upper() == _XLSX_TGT
        s.graphicalProperties.line.solidFill = col
        s.graphicalProperties.line.width = 20000 if dotted else (35000 if hard else 28000)
        if dotted:
            s.graphicalProperties.line.dashStyle = "sysDash"
        s.marker.symbol = "diamond" if hard else "circle"
        s.marker.size = 8 if hard else (5 if dotted else 7)
        s.marker.graphicalProperties.solidFill = col
        s.marker.graphicalProperties.line.solidFill = col


def _xlsx_rate(cell, value, bold=False, center=True):
    cell.value = value
    cell.font = _xlsx_font(bold, 10)
    cell.number_format = "0.00"
    cell.border = _xlsx_sides()[0]
    if center:
        cell.alignment = _xlsx_mid()


def _xlsx_total_border(cell):
    from openpyxl.styles import Border, Side
    med = Side(style="medium", color=_XLSX_NAVY)
    thin = Side(style="thin", color="D0D5DD")
    cell.border = Border(left=thin, right=thin, top=med, bottom=thin)


def _xlsx_line_chart(ws, title, y_title, min_col, max_col, header_row, last_row, anchor,
                     height=7.5, width=15, cat_col=1, colors=None):
    from openpyxl.chart import LineChart, Reference
    lc = LineChart()
    lc.title = title
    lc.y_axis.title = y_title
    lc.y_axis.scaling.min = 0
    lc.y_axis.numFmt = "#,##0"
    lc.height, lc.width = height, width
    lc.legend.position = "t"
    lc.style = None
    data = Reference(ws, min_col=min_col, max_col=max_col,
                     min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=cat_col, min_row=header_row + 1, max_row=last_row)
    lc.add_data(data, titles_from_data=True)
    lc.set_categories(cats)
    n = max_col - min_col + 1
    palette = colors or ((_XLSX_PRED, _XLSX_TGT) + _XLSX_CLOCKS)
    _xlsx_paint_lines(lc, palette[:n])
    ws.add_chart(lc, anchor)
    return lc


def _xlsx_five_clock_block(ws, row, title, sub, points, start=1, chart_col="I", achv=False,
                           total_target=None):
    """Month (or day) × clocks + % of target. Chart sits UNDER the table.
    achv=True adds old / optimized achievable (simulate), same as Plan Allocate."""
    from openpyxl.utils import get_column_letter
    r = _xlsx_section(ws, row, title, sub)
    # Owner 2026-08-26: the workbook shows ONE prediction — the optimized
    # plan, labelled plainly. The pre-Allocate "old predicted" clock is gone
    # from every sheet (it confused readers into comparing two internal
    # stages instead of plan vs target).
    if achv:
        heads = ["Month", "Target", "Predicted plan", "Achievable"]
        keys = ["target", "new_pred", "new_achv"]
    else:
        heads = ["Month", "Target", "Predicted plan"]
        keys = ["target", "new_pred"]
    heads.append("Predicted %")
    if achv:
        heads.append("Achievable %")
    if points and points[0].get("label") == "Date":
        heads[0] = "Date"
    # Owner 2026-08-27: ONE clock. When the caller supplies the sales-line
    # total, the month rows' targets are the sales target DISTRIBUTED over
    # the months in the plan's own shape (scaled so they sum exactly), so
    # a row can never show green while the TOTAL shows short. The plan's
    # internal monthly amounts are the shape, not the judgement.
    if total_target:
        _plan_sum = sum((p.get("target") or 0) for p in points)
        if _plan_sum > 0:
            _f = float(total_target) / _plan_sum
            points = [dict(p) for p in points]
            _run = 0
            for _i, p in enumerate(points):
                if p.get("target") is not None:
                    if _i == len(points) - 1:
                        p["target"] = int(round(total_target - _run))
                    else:
                        p["target"] = int(round(p["target"] * _f))
                        _run += p["target"]
    _xlsx_headers(ws, r, heads, start=start, center=True)
    header_row = r
    rr = r
    n_clocks = len(keys)
    pct_col = start + 1 + n_clocks
    achv_pct_col = pct_col + 1
    for p in points:
        rr += 1
        _xlsx_text(ws.cell(row=rr, column=start), p.get("name"), center=True)
        for i, key in enumerate(keys):
            cell = ws.cell(row=rr, column=start + 1 + i)
            _xlsx_num(cell, p.get(key), center=True)
        tgt, np_ = p.get("target"), p.get("new_pred")
        _xlsx_paint_cov(ws.cell(row=rr, column=pct_col), _cov_pct(np_, tgt))
        if achv:
            _xlsx_paint_cov(ws.cell(row=rr, column=achv_pct_col),
                            _cov_pct(p.get("new_achv"), tgt))
    data_last = rr
    tot = {k: 0 for k in keys}
    n_have = {k: 0 for k in keys}
    for p in points:
        for k in keys:
            v = p.get(k)
            if v is not None:
                tot[k] += v
                n_have[k] += 1
    if data_last > header_row:
        rr += 1
        lab = ws.cell(row=rr, column=start, value="TOTAL")
        lab.font = _xlsx_font(True, 11, _XLSX_NAVY)
        lab.alignment = _xlsx_mid()
        _xlsx_total_border(lab)
        # Owner 2026-08-27: the TOTAL row's target is the SALES line, not
        # the sum of the internal monthly plan amounts (which no longer
        # exists as a requirement). Callers pass total_target; the monthly
        # rows above keep their own per-month figures.
        if total_target is not None:
            tot = dict(tot)
            tot["target"] = total_target
            n_have["target"] = 1
        for i, k in enumerate(keys):
            cell = ws.cell(row=rr, column=start + 1 + i, value=tot[k] if n_have[k] else None)
            cell.font = _xlsx_font(True, 11, _XLSX_INK)
            cell.number_format = "#,##0"
            cell.alignment = _xlsx_mid()
            _xlsx_total_border(cell)
        tgt, np_ = tot["target"], tot["new_pred"]
        c5 = ws.cell(row=rr, column=pct_col)
        _xlsx_paint_cov(c5, _cov_pct(np_, tgt) if n_have["target"] else None)
        _xlsx_total_border(c5)
        if achv:
            c6 = ws.cell(row=rr, column=achv_pct_col)
            _xlsx_paint_cov(c6, _cov_pct(tot.get("new_achv"), tgt) if n_have["target"] else None)
            _xlsx_total_border(c6)
        chart_row = rr + 2
        anchor = "%s%d" % (get_column_letter(start), chart_row)
        _xlsx_line_chart(
            ws, title, "tonnes", start + 1, start + n_clocks, header_row, data_last,
            anchor, height=8, width=18 if achv else 16, cat_col=start,
            colors=_XLSX_CLOCKS_ACHV3 if achv else _XLSX_CLOCKS)
        return chart_row + 16
    return rr + 2


def _xlsx_board_header(ws, heading, sub, start=1):
    """Year-board chrome: title, then a one-line reading note."""
    from openpyxl.utils import get_column_letter
    _xlsx_sheet_setup(ws)
    ws.cell(row=1, column=start, value=heading).font = _xlsx_font(True, 18, _XLSX_NAVY)
    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + 5)
    ws.cell(row=2, column=start, value=sub).font = _xlsx_font(False, 10, _XLSX_MUTED)
    ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=start + 5)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18
    for i, w in enumerate([24, 24, 28, 22, 16, 16], start=start):
        ws.column_dimensions[get_column_letter(i)].width = w
    return 4


def _xlsx_kpi_strip(ws, row, kpis, start=1):
    """Scorecard: colour bar on the label, large figure below.
    Each item is (label, value, color) or (label, value, color, unit).
    Pass value as a 0–100 coverage figure with unit='pct' to format as percent."""
    from openpyxl.styles import Alignment
    mid = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, item in enumerate(kpis):
        label, value, _color = item[0], item[1], item[2]
        unit = item[3] if len(item) > 3 else None
        col = start + i
        if unit == "pct":
            lab_txt = label
        elif unit:
            lab_txt = "%s\n%s" % (label, unit)
        else:
            lab_txt = label
        lab = ws.cell(row=row, column=col, value=lab_txt)
        lab.font = _xlsx_font(True, 8, _XLSX_NAVY)
        lab.alignment = mid
        lab.border = _xlsx_card_border(_XLSX_NAVY)
        cell = ws.cell(row=row + 1, column=col)
        cell.alignment = mid
        cell.border = _xlsx_sides()[0]
        if unit == "pct":
            _xlsx_paint_cov(cell, value, size=20)
            cell.border = _xlsx_sides()[0]
            cell.alignment = mid
        else:
            cell.value = value
            cell.font = _xlsx_font(True, 20, _XLSX_INK)
            cell.number_format = "#,##0"
    ws.row_dimensions[row].height = 28
    ws.row_dimensions[row + 1].height = 32
    return row + 3


def _xlsx_month_cards(ws, row, cards, year, start=2):
    """Month matrix: names across, Predicted / Target down column A."""
    box = _xlsx_sides()[0]
    mid = _xlsx_mid()
    ws.cell(row=row, column=start, value="Months").font = _xlsx_font(True, 13, _XLSX_NAVY)
    ws.row_dimensions[row].height = 19
    names_row = row + 1
    meta_row = row + 2
    for i, c in enumerate(cards):
        col = start + i
        name = ("%s %s" % (c.get("name") or "", year)).strip()
        meta = "%s days" % (c.get("n_days") or "—")
        if c.get("dt") is not None:
            meta += " · %s DT" % format(int(c["dt"]), ",")
        if not c.get("built"):
            meta += " · not built yet"
        cell_n = ws.cell(row=names_row, column=col, value=name)
        cell_n.font = _xlsx_font(True, 12, _XLSX_INK)
        cell_n.alignment = mid
        cell_n.border = box
        cell_m = ws.cell(row=meta_row, column=col, value=meta)
        cell_m.font = _xlsx_font(False, 9, _XLSX_MUTED)
        cell_m.alignment = mid
        cell_m.border = box
    metric_rows = (
        (row + 3, "Predicted plan", "pred_month", _XLSX_INK),
        (row + 4, "Target", "target_month", _XLSX_INK),
    )
    for mrow, label, key, color in metric_rows:
        lab = ws.cell(row=mrow, column=1, value=label)
        lab.font = _xlsx_font(False, 9, _XLSX_MUTED)
        lab.alignment = mid
        lab.border = box
        ws.row_dimensions[mrow].height = 20
        for i, c in enumerate(cards):
            cell = ws.cell(row=mrow, column=start + i, value=c.get(key))
            cell.font = _xlsx_font(True, 14, color)
            cell.number_format = "#,##0"
            cell.alignment = mid
            cell.border = box
    return row + 6


def _daily_triples(st):
    """[(date, pred, achv, target), ...] — Achievable from Plan simulate, never averaged."""
    p = (st or {}).get("prediction") or {}
    m = (st or {}).get("manual") or {}
    pred = {d["date"]: d for d in (p.get("days") or [])}
    man = {d["date"]: d for d in (m.get("days") or [])}
    month = (st or {}).get("month")
    if not month:
        return [], p, m
    if _achv_day_from_pred(p) is None:
        _fill_achv_from_plan(p)
    achv_flat = _achv_day_from_pred(p)
    rows = []
    for d in _days_in(month):
        pd_ = pred.get(d, {})
        aw = pd_.get("wmt_achv")
        if aw is None:
            aw = achv_flat
        rows.append((d, pd_.get("wmt"), aw, man.get(d, {}).get("wmt")))
    return rows, p, m


def _simulate_for_paths(paths):
    """Re-run Check capacity (cycle clock). Never send path-model trips."""
    plans = []
    for p in paths or []:
        key = (p.get("key") or "").strip()
        if ">" not in key:
            continue
        src, dst = key.split(">", 1)
        plans.append({"route": key, "source": src, "destination": dst,
                      "n_trucks": int(round(p.get("dt") or 0)),
                      "contractor": p.get("contractor")})
    if not plans:
        return None, []
    try:
        import plan_simulator
        sim = plan_simulator.simulate({"plans": plans})
    except Exception:
        return None, plans
    if sim.get("error"):
        return None, plans
    return sim, plans


def _xlsx_section(ws, row, title, sub=None):
    ws.cell(row=row, column=1, value=title).font = _xlsx_font(True, 13, _XLSX_NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    if sub:
        ws.cell(row=row + 1, column=1, value=sub).font = _xlsx_font(False, 9, _XLSX_MUTED)
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=8)
        return row + 2
    return row + 1


def _xlsx_saved_day_allocation(ws, r, alloc):
    """Plan-tab Allocate snapshot carried from the saved daily plan."""
    box = _xlsx_sides()[0]
    mid = _xlsx_mid()
    old = alloc.get("old") or {}
    new = alloc.get("new") or {}
    fleet = alloc.get("fleet") or {}
    r = _xlsx_section(
        ws, r, "Saved daily allocation (Plan tab)",
        "Your plan (old) stays as checked. New Allocation Plan is the optimized fleet. "
        "Predicted = path model. Target = matrix.")
    goals = alloc.get("goals") or {}
    buckets = alloc.get("buckets") or {}
    labels = [("sap", "SAP · must-move"), ("tos", "LIM-TOS"), ("ld", "Other LIM · LD")]
    r = _xlsx_section(
        ws, r, "Plan by material",
        "GP targets vs the allocated plan (after Allocate DT).")
    _xlsx_headers(ws, r, ["", "SAP", "LIM-TOS", "Other LIM"], center=True)
    metric_rows = [
        ("Target t/day", "target", _XLSX_INK, False),
        ("Predicted plan", "pred_after", _XLSX_INK, True),
        ("DT", "dt_after", _XLSX_INK, True),
    ]
    for lab, key, color, bold in metric_rows:
        r += 1
        lab_c = ws.cell(row=r, column=1, value=lab)
        lab_c.font = _xlsx_font(False, 10, _XLSX_MUTED)
        lab_c.border = box
        lab_c.alignment = mid
        for i, (bkey, _) in enumerate(labels):
            b = buckets.get(bkey) or {}
            val = b.get(key) if key != "target" else (b.get("target") or goals.get(bkey))
            cell = ws.cell(row=r, column=2 + i, value=val)
            cell.border = box
            cell.alignment = mid
            cell.font = _xlsx_font(bold, 14 if bold else 12, color)
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0"
    r += 2
    old = alloc.get("old") or {}
    new = alloc.get("new") or {}
    fleet = alloc.get("fleet") or {}
    note = (
        "Fleet %s DT (same trucks, after allocate). "
        "Total predicted plan %s."
        % (fleet.get("after") or new.get("dt") or "—",
           new.get("pred") if new.get("pred") is not None else "—")
    )
    ws.cell(row=r, column=1, value=note).font = _xlsx_font(False, 9, _XLSX_MUTED)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2
    rows = alloc.get("rows") or []
    if rows:
        r = _xlsx_path_alloc_table(
            ws, r, rows, "New Allocation Plan table",
            "Allocated plan (after Allocate DT). "
            "WMT/DT is tonnes per truck-trip (payload).")
    return r


def _xlsx_pct_cell(cell, value):
    cell.value = None if value is None else value / 100.0
    cell.font = _xlsx_font(True, 11)
    cell.number_format = "0.0%"
    cell.border = _xlsx_sides()[0]
    cell.alignment = _xlsx_mid()


def _xlsx_cov_tone(pct):
    """Coverage % is ink on a clear cell — no red/gold/green pills in Excel."""
    if pct is None:
        return None, None
    return _XLSX_INK, None


def _xlsx_paint_cov(cell, pct, size=11):
    """Percent of target: GREEN when the target is met (>=100%), plain black
    below. Owner 2026-08-27: "if we pass 100% of our targets make it green,
    no matter which — total or anything." This is deliberately NOT the old
    three-colour traffic light (removed by the 2026-08-23 ruling): under
    100% stays a plain figure, no amber/red judgement — only success is
    coloured."""
    from openpyxl.styles import PatternFill
    _xlsx_pct_cell(cell, pct)
    if pct is not None and pct >= 100:
        cell.font = _xlsx_font(True, size, "1B7A41")
        cell.fill = PatternFill("solid", fgColor="D9F2E2")
    else:
        cell.font = _xlsx_font(True, size, _XLSX_INK)
    return cell


def _pick_achv(d, old=True, grain="month"):
    """Raw simulate achievable; fall back to the capped companion if raw is missing."""
    d = d or {}
    if old:
        v = d.get("old_achv_raw_" + grain)
        return v if v is not None else d.get("old_achv_" + grain)
    v = d.get("new_achv_raw_" + grain)
    return v if v is not None else d.get("new_achv_" + grain)


def _pick_mat_achv(m, before=True, grain="month"):
    m = m or {}
    if before:
        v = m.get("achv_before_raw_" + grain)
        return v if v is not None else m.get("achv_before_" + grain)
    v = m.get("achv_after_raw_" + grain)
    return v if v is not None else m.get("achv_after_" + grain)


def _ten_key(row):
    """Key into _tenant_trips_per_dt.

    Identity of a ROW, not of a road: id() of the row dict, because the same
    (path, contractor) legitimately appears more than once in one month —
    TF>HUAFEI/RIM is both a P2 LIM-TOS row and a P3 LIM-LD row, at different
    rates. Keying on (path, contractor) let the second row's answer overwrite
    the first, and the P2 row then printed the P3 row's 2.15 over its own
    2.11: a +1.9% GAIN from added traffic.

    The ROAD key (path, contractor) is still what the engine ratio is cached
    under inside _tenant_trips_per_dt — that one really is a road property.
    ONE home for each, so the lookup and the build cannot drift apart.
    """
    return id(row)


def _tenant_cap_note():
    """Live official-capacity sentence so the Excel caption cannot go stale."""
    try:
        from congestion.segments import SEGMENTS
        from congestion.speed_limits import FOLLOWING_DISTANCE_M
        s1 = next((s["cap_hr"] for s in SEGMENTS if s["id"] == "S1"), None)
        s4 = next((s["cap_hr"] for s in SEGMENTS if s["id"] == "S4"), None)
        if s1 and s4:
            return ("capacity (S1–S3 %.0f/hr, S4 %.0f/hr) is derived from posted "
                    "speed limits and a %.0f m following distance"
                    % (s1, s4, FOLLOWING_DISTANCE_M))
    except ImportError:
        pass
    return "capacity is derived from posted speed limits and following distance"


def _tenant_trips_per_dt(rows):
    """{(path, contractor) -> Trips/DT with the OTHER TENANTS on the road}.

    Owner, 2026-08-24: several tenants (MHM, POSITION, PMA, HSM, KR>RSF,
    HUAFEI>RSF — 1,340 DT) run our haul road and give us no tonnage. Every
    Trips/DT this workbook has ever shown was therefore the clear-road answer.
    This adds one column with the same routes re-priced under that traffic.

    It does NOT move tonnage, targets, DT or the allocation. Tenant trucks are
    road load only, so the honest report is "here is the rate you would
    actually turn", beside the plan, not a silently re-planned month.

    ## Why this is a RATIO applied to the sheet's own number

    The first version printed the engine's tenant-priced rate directly and it
    came out HIGHER than the clear-road column beside it (BLB>FENI KM0 4.61
    against 4.58) — adding traffic appearing to speed trucks up. The two
    numbers had two causes, not one: the sheet's Trips/DT comes from the
    FROZEN allocation saved on the plan date, which predates the current
    hybrid calibration, while the new column was a fresh engine call. The
    tenant effect (~-1%) was smaller than the drift between the two vintages,
    so the drift showed and the tenants did not.

    So the engine is asked BOTH questions — same route, same fleet, same
    contractor, same segment map, tenants off then on — and only their RATIO
    is used, applied to whatever the sheet already shows. The column is then
    exactly "your own number, degraded by the tenant traffic and nothing
    else", and it can never disagree with its neighbour for a reason the
    caption does not name. Re-freezing the saved allocations under the hybrid
    model is a separate open thread (HANDOFF §14) and this must not pre-empt
    it silently.

    Returns {} when the congestion package or its calibration is unavailable
    (fixture mode, fresh clone) — the caller then omits the column rather than
    printing a blank one that looks like zero traffic.
    """
    try:
        from congestion.config import route_params
        from congestion.predictor import predict
        from congestion.segments import segment_trucks
    except ImportError:
        return {}
    combined = {}
    for row in rows or []:
        k = row.get("key")
        dt = _finite(row.get("dt_after"))
        if k and dt:
            combined[k] = combined.get(k, 0.0) + float(dt)
    if not combined:
        return {}
    # Every route the plan runs is on the road together, so the segment map is
    # built from the WHOLE plan and both calls share it. Only the tenant flag
    # differs between them.
    seg = segment_trucks(combined)
    out = {}
    # Cache the RATIO per (road, contractor), never the finished value: the
    # ratio is a property of the road and the fleet, but the value it is
    # applied to is the ROW's own rate. Caching the value made the second
    # TF>HUAFEI/SMA row (217 DT, 2.36) print the first row's answer (52 DT,
    # 2.37) and so appear to GAIN 0.4% from added traffic.
    ratios = {}
    for row in rows or []:
        k = row.get("key")
        shown = _path_rates(row).get("trips_per_dt_after")
        # The engine ratio is cached per (road, contractor): RIM and SMA run
        # the same road at different overheads, and caching on the path alone
        # gave both of them whichever was priced first (TF>HUAFEI SMA printed
        # RIM's 2.17 against its own 2.36).
        ck = (k, (row.get("contractor") or "").strip().upper())
        if not k or not shown:
            continue
        if ck not in ratios:
            ratios[ck] = None
            try:
                if route_params(k).get("calibrated"):
                    kw = dict(segment_fleet=seg, mode="road",
                              contractor=ck[1] or None)
                    clear = predict(k, combined[k], None, **kw)
                    withn = predict(k, combined[k], None,
                                    tenant_flow_hr=True, **kw)
                    base = clear.get("trips_per_DT_per_day")
                    tenv = withn.get("trips_per_DT_per_day")
                    # A route not on the shared mainline (the BLB spur) comes
                    # back with tenant_traffic False and its clear-road answer.
                    # Leave it blank rather than repeating the neighbouring
                    # column, so "no tenant traffic here" cannot be misread as
                    # "tenants made no difference".
                    if base and tenv and withn.get("tenant_traffic"):
                        ratios[ck] = float(tenv) / float(base)
            except (ValueError, ArithmeticError, KeyError, TypeError, OSError):
                ratios[ck] = None
        if ratios[ck]:
            # THREE decimals, not two. At official capacities the tenants cost
            # 0.1-1.5% of trips/DT, and 2 dp rounds that straight back onto the
            # neighbouring column: TF>FENI KM15 — the road carrying 800 of the
            # 1,340 tenant DT — printed an identical number to its own rate, so
            # the one row the reader checks first said the tenants did nothing.
            # A column whose answer is always invisible is not a column.
            out[_ten_key(row)] = round(float(shown) * ratios[ck], 3)
    return out


def _xlsx_path_alloc_table(ws, r, rows, title, sub, achv=False):
    """Allocated-plan path table: WMT, DT, trips, plus trips/DT and WMT/DT.
    The pre-Allocate (old) columns are omitted — month sheets show the
    optimized plan only, with no 'old' / 'new' suffix on the headers
    (owner, 2026-08-24). achv=True appends the engine's achievable (t/day)."""
    box = _xlsx_sides()[0]
    mid = _xlsx_mid()
    rows = _ensure_tenant_rows(rows)
    rows.sort(key=lambda x: (
        2 if _is_tenant_row(x) else (1 if x.get("foreign") else 0),
        x.get("prio") or 9,
        x.get("contractor") or "",
        x.get("key") or "",
    ))
    if not rows:
        return r
    # One extra column: the same paths priced with the other tenants' trucks
    # on the road (congestion/tenants.py). Empty dict -> no column at all.
    # Tenant rows themselves are the register, not something to re-price.
    ours = [x for x in rows if not _is_tenant_row(x)]
    ten = _tenant_trips_per_dt(ours)
    if ten:
        try:
            from congestion.tenants import TENANTS as _TEN_REG
            _n_dt = sum(t["dt"] for t in _TEN_REG)
            _names = ", ".join(t["name"] for t in _TEN_REG)
        except ImportError:
            _n_dt, _names = 0, ""
        # State the basis in the sheet. A lower Trips/DT with no explanation
        # beside it is how a reader concludes the model changed its mind.
        sub = (sub or "") + (
            " · 'Trips/DT w/ other tenants' re-prices the SAME fleet with the "
            "other tenants' %s DT on the shared road (%s). They carry no "
            "tonnage for us and take no trucks from us — only road. Tonnes, "
            "targets and DT in this table are unchanged. Both fleets are on ONE "
            "clock: a tenant truck occupies the loaded lane for one pass per "
            "road cycle, exactly as ours does, so they take the busiest section "
            "(POS 12–KM15) against that official capacity. Remaining caveat: that "
            "%s, NOT a counted "
            "traffic survey — if the real lane carries less, this cost grows "
            "sharply, because delay rises with the fourth power of v/c. Blank = "
            "that route is off the shared mainline (the BLB spur), not that the "
            "tenants made no difference. Other-tenant fleets are listed at the "
            "bottom of this table (material 'other tenant') — the same rows as "
            "Plan → New Allocation Plan. They take road, not WMT, and are not "
            "in TOTAL DT. IWIP POS-transit (material 'road') is its OWN "
            "fleet too: on the road, 0 WMT, counted on its own line "
            "under TOTAL, never out of the contractor pools."
            % (_n_dt, _names, _tenant_cap_note()))
    r = _xlsx_section(ws, r, title, sub)
    # Optimized plan only. Year-board Path sheet already uses these names
    # (no "new" prefix); month sheets now match.
    # Weighbridge column (owner, 2026-08-26): each plan row names the
    # bridge(s) its trips weigh over, from the SAME deterministic allocator
    # the Plan tab uses (simulator_api min-max water-fill over the owner
    # matrix). Dominant bridge first with its share; tenant rows are never
    # allocated a bridge (they are not ours to weigh).
    wb_by_id = {}
    try:
        import simulator_api as _sim
        _basis = _sim._wb_basis()
        if _basis:
            _req = []
            for x in rows:
                # Tenant rows are PASSED IN now (owner, 2026-08-31): they are
                # still never allocated a bridge, but their trucks queue on the
                # same bridges, so they must load them before our water-fill.
                _tenant = _is_tenant_row(x)
                _rates = _path_rates(x)
                _tr = _rates.get("trips_after") or 0
                if not _tr and _tenant:
                    _tr = x.get("trips") or 0
                if _tr and x.get("key"):
                    _req.append({"id": x.get("id") or x.get("key"),
                                 "route": x.get("key"), "trips": _tr,
                                 "foreign": bool(x.get("foreign")),
                                 "tenant": _tenant})
            if _req:
                out_rows, _pb, _fl, _uv = _sim.wb_assign_rows(_basis, _req,
                                                              hours=20.0)
                for a in out_rows:
                    names = a.get("assigned") or []
                    if names:
                        # "T15 - 34% - 2min" per bridge: bridge, this
                        # row's share of trips over it, M/M/1 wait.
                        # Compact on purpose (owner, 2026-08-27: "keep it
                        # simple, WB name - percentage - time, don't add
                        # too much text"). Util dropped from the cell.
                        parts = []
                        for b in names:
                            if round(100 * (b.get("share") or 0)) < 1:
                                continue   # 0% rows are noise in a compact cell
                            parts.append("%s - %d%% - %.0fmin%s" % (
                                b["bridge"].replace("WB_IWIP_", "").replace("WB_RIM_", "RIM "),
                                round(100 * (b.get("share") or 0)),
                                b.get("wait_min") or 0,
                                " SAT" if b.get("saturated") else ""))
                        wb_by_id[a["id"]] = " · ".join(parts)
    except Exception:
        wb_by_id = {}
    heads = [
        "P", "Path", "Contractor", "Material", "Target WMT/day",
        "DT", "Trips", "WMT", "WMT/DT", "Trips/DT",
    ]
    if achv:
        heads.append("Achievable")
    if ten:
        heads.append("Trips/DT w/ other tenants")
    # Weighbridge is the LAST column (owner, 2026-08-26: numbers first,
    # then the wide text).
    heads.append("Weighbridge")
    _xlsx_headers(ws, r, heads, center=True)
    tot = {
        "tgt": 0, "dt_b": 0, "dt_a": 0, "tr_b": 0, "tr_a": 0,
        "pr_b": 0, "pr_a": 0,
    }
    for row in rows:
        r += 1
        rates = _path_rates(row)
        av_new = _finite(row.get("achv_sim"))
        if av_new is None:
            av_new = _finite(row.get("achv_after"))
        vals = [
            _prio_cell(row), row.get("key"), row.get("contractor"),
            _path_mat_label(row), row.get("target"),
            row.get("dt_after"), rates["trips_after"], row.get("pred_after"),
            rates["wmt_per_trip_after"], rates["trips_per_dt_after"],
        ]
        if achv:
            vals.append(av_new)
        ten_col = len(vals) + 1 if ten else None
        if ten:
            vals.append(ten.get(_ten_key(row)))
        wb_col = len(vals) + 1
        vals.append("—" if _is_tenant_row(row)
                    else wb_by_id.get(row.get("id") or row.get("key")) or "")
        rate_cols = {9, 10}  # WMT/DT, Trips/DT
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = box
            cell.alignment = mid
            cell.font = _xlsx_font(col == 2, 9)
            if ten_col and col == ten_col:
                if isinstance(val, (int, float)):
                    # 3 dp: the tenant effect is sub-1% on most roads and a
                    # "0.00" format hides it behind the column it is meant to
                    # be compared against.
                    cell.number_format = "0.000"
                cell.font = _xlsx_font(True, 9, _XLSX_NAVY)
                continue
            if col in rate_cols and isinstance(val, (int, float)):
                cell.number_format = "0.00"
            elif col >= 5 and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
            if col == wb_col:
                cell.font = _xlsx_font(False, 8, _XLSX_MUTED)
                continue
            if col in (6, 7, 8) or (achv and col == 11):
                cell.font = _xlsx_font(True, 9, _XLSX_INK)
        if _is_tenant_row(row):
            continue
        if row.get("foreign"):
            # IWIP POS-transit is its OWN fleet and never comes out of the
            # contractor pools (rules §10.8; the Plan navbar has said
            # "581 DT + 27 IWIP" since 2026-08-21). Summing it into TOTAL DT
            # put Sep at 707 against the pool figure 581 printed elsewhere on
            # the same sheet. It also silently deflated Trips/DT, because these
            # rows carry DT but 0 trips/0 WMT/0 achievable — so our trips were
            # being divided by a fleet that included them (Sep 21.7% low).
            # Counted on its own line under TOTAL instead.
            tot["dt_iwip"] = tot.get("dt_iwip", 0) + (row.get("dt_after") or 0)
            tot["dt_iwip_b"] = tot.get("dt_iwip_b", 0) + (row.get("dt_before") or 0)
            tot["n_iwip"] = tot.get("n_iwip", 0) + 1
            continue
        tot["tgt"] += row.get("target") or 0
        tot["dt_b"] += row.get("dt_before") or 0
        tot["dt_a"] += row.get("dt_after") or 0
        tot["pr_b"] += row.get("pred_before") or 0
        tot["pr_a"] += row.get("pred_after") or 0
        tot["av_b"] = tot.get("av_b", 0) + (_finite(row.get("achv_before")) or 0)
        av_a = _finite(row.get("achv_sim"))
        if av_a is None:
            av_a = _finite(row.get("achv_after"))
        tot["av_a"] = tot.get("av_a", 0) + (av_a or 0)
        if rates["trips_before"] is not None:
            tot["tr_b"] += rates["trips_before"]
        elif rates["trips_per_dt_before"] is not None and row.get("dt_before"):
            tot["tr_b"] += rates["trips_per_dt_before"] * (row.get("dt_before") or 0)
        if rates["trips_after"] is not None:
            tot["tr_a"] += rates["trips_after"]
        elif rates["trips_per_dt_after"] is not None and row.get("dt_after"):
            tot["tr_a"] += rates["trips_per_dt_after"] * (row.get("dt_after") or 0)
    r += 1
    tpd_a = round(tot["tr_a"] / tot["dt_a"], 2) if tot["dt_a"] else None
    pay_a = round(tot["pr_a"] / tot["tr_a"], 2) if tot["tr_a"] else None
    n_ours = sum(1 for x in rows
                 if not _is_tenant_row(x) and not x.get("foreign"))
    tot_vals = [
        "TOTAL", "%s paths" % n_ours, "", "",
        tot["tgt"] or None, tot["dt_a"] or None,
        int(round(tot["tr_a"])) if tot["tr_a"] else None,
        tot["pr_a"] or None, pay_a, tpd_a,
    ]
    if achv:
        tot_vals.append(int(round(tot.get("av_a") or 0)) or None)
    if ten:
        # DT-weighted over EVERY row, exactly the denominator tpd_a uses.
        # Summing only the rows that have a tenant value put the two totals on
        # different fleets: the BLB spur and the IWIP rows dropped out and the
        # "with tenants" total read 2.92 against a clear-road 2.74 — 6.6% HIGH
        # for a column that can only ever be lower. A row the tenants do not
        # touch keeps its own rate, because that IS its rate under this traffic.
        t_tr = 0.0
        for x in rows:
            # Skip foreign/IWIP for the SAME reason tot["dt_a"] now does:
            # numerator and denominator must be one fleet. These rows carry DT
            # but no trips, so leaving them in the denominator alone would
            # under-state the with-tenants rate exactly as it under-stated
            # tpd_a.
            if _is_tenant_row(x) or x.get("foreign"):
                continue
            dt_x = _finite(x.get("dt_after")) or 0
            if not dt_x:
                continue
            rate = ten.get(_ten_key(x))
            if rate is None:
                rate = _path_rates(x).get("trips_per_dt_after")
            t_tr += (rate or 0) * dt_x
        # 3 dp for the same reason as the per-row value: at 2 dp the total sat
        # on top of the clear-road total while every one of its own rows had
        # moved down.
        tot_vals.append(round(t_tr / tot["dt_a"], 3) if tot["dt_a"] else None)
    for col, val in enumerate(tot_vals, start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = _xlsx_font(True, 9, _XLSX_NAVY)
        cell.alignment = mid
        _xlsx_total_border(cell)
        _rate_col = col in (9, 10)
        _ten_total = bool(ten) and col == len(tot_vals)
        if _ten_total:
            _rate_col = True
        if _rate_col and isinstance(val, (int, float)):
            cell.number_format = "0.000" if _ten_total else "0.00"
        elif col >= 5 and isinstance(val, (int, float)):
            cell.number_format = "#,##0"
    r += 1
    pct_lab = ws.cell(row=r, column=1, value="% of target")
    pct_lab.font = _xlsx_font(True, 9, _XLSX_MUTED)
    pct_lab.alignment = mid
    pct_lab.border = box
    last_col = (11 if achv else 10) + (1 if ten else 0)
    for col in range(2, last_col + 1):
        ws.cell(row=r, column=col).border = box
        ws.cell(row=r, column=col).alignment = mid
    pred_col = 8  # WMT (predicted)
    _xlsx_paint_cov(ws.cell(row=r, column=pred_col),
                    _cov_pct(tot["pr_a"], tot["tgt"]), size=9)
    if achv:
        _xlsx_paint_cov(ws.cell(row=r, column=11),
                        _cov_pct(tot.get("av_a"), tot["tgt"]), size=9)
    # Three fleets, three lines, each named: ours in TOTAL, IWIP POS-transit
    # and the other tenants beside it. Before this, IWIP was invisible because
    # it was folded INTO TOTAL DT, which is how Sep read 707 against a pool of
    # 581 on the same sheet.
    if tot.get("dt_iwip"):
        r += 1
        note = ws.cell(
            row=r, column=1,
            value=("IWIP POS-transit · %s DT on the road · %s paths · 0 WMT · "
                   "own fleet, not from the contractor pools, not in TOTAL"
                   % (int(round(tot["dt_iwip"])), tot.get("n_iwip", 0))))
        note.font = _xlsx_font(False, 8, _XLSX_MUTED)
        note.alignment = mid
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=last_col)
    ten_rows = [x for x in rows if _is_tenant_row(x)]
    if ten_rows:
        r += 1
        dt_ten = int(round(sum((_finite(x.get("dt_after")) or 0) for x in ten_rows)))
        note = ws.cell(
            row=r, column=1,
            value=("Other tenants · %s DT on the road · 0 WMT · not in TOTAL"
                   % dt_ten))
        note.font = _xlsx_font(False, 8, _XLSX_MUTED)
        note.alignment = mid
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=last_col)
    return r + 1


def _split_route_key(key):
    """'BLB>FENI KM13' → (BLB, FENI KM13)."""
    k = (key or "").strip()
    if ">" not in k:
        return k, ""
    src, dst = k.split(">", 1)
    return src.strip(), dst.strip()


# Pit → plant KPI (owner 2026-08-25): dest rows × pit columns, t/day.
# Lead dests stay visible even at 0 so the sheet layout does not jump.
_PIT_ORDER = ("TF", "BLB", "KR")
_PIT_ALIASES = {"TOFU": "TF", "KRENE": "KR"}
_DEST_LEAD = ("FENI KM0", "FENI KM15", "POS")
_MAT_LEAD = (("sap", "SAP"), ("tos", "LIM-TOS"), ("ld", "LIM-LD"))


def _pit_name(origin):
    o = (origin or "").strip().upper()
    o = _ORIGIN_MAP.get(o, o)
    return _PIT_ALIASES.get(o, o) or "?"


def _dest_plant(dest):
    """Group tips the planner actually asks about: FeNi plants + POS together."""
    d = _canon_dest(dest)
    if "KM15" in d:
        return "FENI KM15"
    if "FENI" in d:
        return "FENI KM0"
    if d.startswith("POS"):
        return "POS"
    if "HUAFEI" in d:
        return "HUAFEI"
    if d == "BSE" or d.startswith("BSE"):
        return "BSE"
    return d or "Other"


def _mat_bucket(row):
    """sap / tos / ld, or None for IWIP, tenants, and unlabelled road rows."""
    if row.get("foreign") or _is_tenant_row(row):
        return None
    mat = (row.get("material") or "").strip().upper()
    otype = (row.get("otype") or "").strip().upper()
    if " - " in mat and not otype:
        left, right = mat.split(" - ", 1)
        mat, otype = left.strip(), right.strip()
    if mat in ("ROAD", "OTHER TENANT") or not mat:
        return None
    if mat == "SAP":
        return "sap"
    if otype == "LD" or mat.endswith("LD"):
        return "ld"
    if mat.startswith("LIM") or otype == "TOS":
        return "tos"
    return None


def _dest_pit_flow(rows):
    """{(mat, dest, pit) -> t/day} from allocation / path rows. Production WMT only."""
    out = {}
    for row in rows or []:
        b = _mat_bucket(row)
        t = _finite(row.get("pred_after"))
        if t is None:
            t = _finite(row.get("pred_wmt_day"))
        if not b or not t:
            continue
        origin, dest = _split_route_key(row.get("key"))
        if not origin:
            origin = row.get("source") or row.get("origin") or ""
        if not dest:
            dest = row.get("dest") or row.get("destination") or ""
        pit = _pit_name(origin)
        plant = _dest_plant(dest)
        key = (b, plant, pit)
        out[key] = out.get(key, 0.0) + float(t)
    return out


def _xlsx_dest_from_pits(ws, r, rows, grain="t / day"):
    """KPI table: how much of each material each pit sends to each plant, per day."""
    flow = _dest_pit_flow(rows)
    if not flow:
        return r
    pits = {pit for (_, _, pit) in flow}
    dests = {dest for (_, dest, _) in flow}
    pit_cols = [p for p in _PIT_ORDER if p in pits] + sorted(pits - set(_PIT_ORDER))
    extra_dests = sorted(dests - set(_DEST_LEAD))
    dest_rows = list(_DEST_LEAD) + extra_dests
    r = _xlsx_section(
        ws, r, "Where material goes · %s" % grain,
        "One day of this month's plan (same every day). Rows are plants; "
        "columns are pits. POS is POS 6 / 10 / 12 / 14 / 15 / 16 together. "
        "Predicted tonnes — not target, not achievable. HUAFEI / BSE rows "
        "appear when this plan tips there, so the totals still add up.")
    box = _xlsx_sides()[0]
    mid = _xlsx_mid()
    heads = ["Material", "To plant"] + list(pit_cols) + ["Total"]
    last_col = len(heads)

    def _cell_t(mat, dest, pit):
        if mat is None:
            return sum(v for (m, d, p), v in flow.items()
                       if d == dest and p == pit)
        return flow.get((mat, dest, pit), 0.0)

    def _write_block(mat_key, mat_label, plants):
        nonlocal r
        _xlsx_headers(ws, r, heads, center=True)
        for dest in plants:
            r += 1
            tot = 0.0
            ws.cell(row=r, column=1, value=mat_label).border = box
            ws.cell(row=r, column=1).alignment = mid
            ws.cell(row=r, column=1).font = _xlsx_font(False, 10, _XLSX_MUTED)
            _xlsx_text(ws.cell(row=r, column=2), dest, center=True)
            for i, pit in enumerate(pit_cols):
                val = _cell_t(mat_key, dest, pit)
                tot += val
                cell = ws.cell(row=r, column=3 + i)
                _xlsx_num(cell, val if val else None, center=True)
            _xlsx_num(ws.cell(row=r, column=last_col), tot if tot else None,
                      True, center=True)
        r += 1
        _xlsx_text(ws.cell(row=r, column=1), mat_label, True, _XLSX_NAVY, center=True)
        _xlsx_text(ws.cell(row=r, column=2), "Total", True, _XLSX_NAVY, center=True)
        grand = 0.0
        for i, pit in enumerate(pit_cols):
            col_tot = sum(_cell_t(mat_key, dest, pit) for dest in plants)
            grand += col_tot
            _xlsx_num(ws.cell(row=r, column=3 + i),
                      col_tot if col_tot else None, True, center=True)
        _xlsx_num(ws.cell(row=r, column=last_col),
                  grand if grand else None, True, center=True)
        _xlsx_total_border(ws.cell(row=r, column=1))
        _xlsx_total_border(ws.cell(row=r, column=2))
        for col in range(3, last_col + 1):
            _xlsx_total_border(ws.cell(row=r, column=col))
        r += 2

    mats_present = {m for (m, _, _) in flow}
    for mk, title in _MAT_LEAD:
        if mk not in mats_present:
            continue
        used = {d for (m, d, _) in flow if m == mk}
        plants = [d for d in _DEST_LEAD] + sorted(used - set(_DEST_LEAD))
        _write_block(mk, title, plants)
    _write_block(None, "Together", dest_rows)
    return r


def _paths_as_flow_rows(paths):
    """Month-file path dicts → the shape _dest_pit_flow reads."""
    out = []
    for prow in paths or []:
        if not isinstance(prow, dict) or prow.get("foreign"):
            continue
        out.append({
            "key": prow.get("key"),
            "source": prow.get("source") or prow.get("origin"),
            "dest": prow.get("dest") or prow.get("destination"),
            "material": prow.get("material"),
            "otype": prow.get("otype") or "",
            "pred_after": prow.get("pred_wmt_day"),
            "foreign": False,
        })
    return out


def _path_mat_label(row):
    """Material cell: production ore, IWIP POS-transit as 'road', other tenants named."""
    if _is_tenant_row(row):
        return "other tenant"
    mat = (row.get("material") or "").strip()
    otype = (row.get("otype") or "").strip()
    if row.get("foreign") and not mat:
        return "road"
    if mat.lower() in ("other tenant", "road"):
        return mat
    if mat.upper() == "SAP":
        return "SAP"
    if mat and otype:
        return "%s - %s" % (mat, otype)
    return mat or otype


def _prio_cell(row):
    if _is_tenant_row(row):
        return "—"
    p = row.get("prio")
    return "P%s" % p if p not in (None, "") else ""


def _card_alloc_detail(c):
    """Saved Allocate snapshot for a year-board card, with path rows."""
    month = c.get("month")
    st = _load_state(month) or {"month": month}
    st["month"] = month
    if c.get("alloc_raw"):
        raw, src = c["alloc_raw"], c.get("alloc_source")
    else:
        raw, src = _resolve_allocation(month, st, day=c.get("_alloc_day"))
    n = c.get("n_days") or len(_days_in(month))
    return _alloc_view(raw, n, src, include_detail=True), n


def _collect_year_path_rows(cards):
    """One row per allocated path per month. WMT new is t/day; predicted/month = × NB days."""
    out = []
    for c in cards:
        alloc, n = _card_alloc_detail(c)
        if not alloc:
            continue
        for row in _ensure_tenant_rows(alloc.get("rows") or []):
            dt = _finite(row.get("dt_after"))
            wmt = _finite(row.get("pred_after"))
            if not dt and not wmt:
                continue
            is_ten = _is_tenant_row(row)
            rates = _path_rates(row)
            origin, dest = _split_route_key(row.get("key"))
            achv_new = _finite(row.get("achv_sim"))
            if achv_new is None:
                achv_new = _finite(row.get("achv_after"))
            out.append({
                "month": c.get("name") or c.get("month"),
                "month_key": c.get("month") or "",
                "prio": None if is_ten else (row.get("prio") or 9),
                "origin": origin,
                "dest": dest,
                "key": row.get("key"),
                "contractor": row.get("contractor"),
                "material": _path_mat_label(row),
                "target": _finite(row.get("target")),
                "dt": dt,
                "trips": rates["trips_after"],
                "wmt": wmt,
                "wmt_per_trip": rates["wmt_per_trip_after"],
                "trips_per_dt": rates["trips_per_dt_after"],
                "achv": achv_new,
                "n_days": n,
                "pred_month": int(round(wmt * n)) if wmt is not None and n else None,
                "achv_month": int(round(achv_new * n)) if achv_new is not None and n else None,
                "_tenant": is_ten,
                "foreign": bool(row.get("foreign")),
            })
    out.sort(key=lambda x: (
        x["month_key"],
        2 if x.get("_tenant") else (1 if x.get("foreign") else 0),
        x["prio"] if x["prio"] is not None else 9,
        x.get("contractor") or "",
        x.get("key") or "",
    ))
    return out


def _wb_assignments_for_month(rows_month):
    """Weighbridge assignment for one month's path rows via the SAME
    deterministic allocator the Plan tab and month sheets use
    (simulator_api.wb_assign_rows: owner-matrix eligibility, min-max
    water-fill). Returns {(key, contractor): text} listing EVERY bridge
    (no "+N more" truncation — owner, 2026-08-27) in the compact
    "name - share% - wait" form (owner, 2026-08-27: "keep it simple,
    WB name - percentage - time, don't add too much text")."""
    try:
        import simulator_api as _sim
        basis = _sim._wb_basis()
    except Exception:
        return {}
    if not basis:
        return {}
    req = []
    for x in rows_month:
        tr = x.get("trips") or 0
        if tr and x.get("key"):
            req.append({"id": "%s|%s" % (x.get("key"), x.get("contractor")),
                        "route": x.get("key"), "trips": tr,
                        "foreign": bool(x.get("foreign")),
                        "tenant": bool(x.get("_tenant"))})
    if not req:
        return {}
    try:
        out_rows, _pb, _fl, _uv = _sim.wb_assign_rows(basis, req, hours=20.0)
    except Exception:
        return {}
    matrix = basis.get("matrix") or {}
    res = {}
    for a in out_rows:
        names = a.get("assigned") or []
        if not names:
            continue
        route = a.get("route")
        allowed = set(matrix.get(route) or [])
        parts = []
        off_matrix = []
        for b in names:
            nm = b["bridge"]
            label = nm.replace("WB_IWIP_", "").replace("WB_RIM_", "RIM ")
            if round(100 * (b.get("share") or 0)) >= 1:
                parts.append("%s - %d%% - %.0fmin%s" % (
                    label, round(100 * (b.get("share") or 0)),
                    b.get("wait_min") or 0,
                    " SAT" if b.get("saturated") else ""))
            if allowed and nm not in allowed:
                off_matrix.append(label)
        txt = " · ".join(parts)
        if off_matrix:
            txt += "  ⚠ off-matrix: " + ", ".join(off_matrix)
        elif allowed:
            txt += "  ✓ matrix"
        res[a.get("id")] = txt
    return res


def _xlsx_all_paths_table(ws, r, cards, achv=False):
    """All months × allocated paths. NB Days then Predicted / month = WMT × days."""
    from openpyxl.styles import Alignment, PatternFill
    rows = _collect_year_path_rows(cards)
    if not rows:
        return r
    # weighbridge text per row, computed month by month (the water-fill
    # balances within a month's plan, so each month is its own allocation).
    wb_txt = {}
    for mk in sorted({x["month_key"] for x in rows}):
        month_rows = [x for x in rows if x["month_key"] == mk]
        for rid, txt in _wb_assignments_for_month(month_rows).items():
            wb_txt[(mk, rid)] = txt
    r = _xlsx_section(
        ws, r, "Paths — all months",
        "Optimized plan only. WMT is t/day. Predicted / month = WMT × NB Days. "
        "WMT/DT is tonnes per truck-trip (payload). Other-tenant fleets sit at "
        "the end of each month (material 'other tenant') — same rows as Plan → "
        "New Allocation Plan, 0 WMT, not in TOTAL DT. IWIP POS-transit "
        "('road') is its own fleet and is listed under TOTAL, not in it."
        + (" Achievable is /api/simulate, t/day." if achv else ""))
    heads = [
        "Month", "Priority", "Origin", "Destination", "Path", "Contractor", "Material",
        "Target WMT/day", "DT", "Trips", "WMT",
        "WMT/DT", "Trips/DT",
    ]
    if achv:
        heads.append("Achievable")
    heads += ["NB Days", "Predicted / month", "Weighbridge (name - share - wait)"]
    _xlsx_headers(ws, r, heads, center=True)
    navy = PatternFill("solid", fgColor=_XLSX_NAVY)
    for col in range(1, len(heads) + 1):
        cell = ws.cell(row=r, column=col)
        cell.fill = navy
        cell.font = _xlsx_font(True, 9, "FFFFFF")
    header_row = r
    box = _xlsx_sides()[0]
    mid = _xlsx_mid()
    tot = {"tgt": 0, "dt": 0, "tr": 0, "wmt": 0, "achv": 0, "pm": 0}
    pred_col = 15 if achv else 14
    n_cols = pred_col
    for row in rows:
        r += 1
        vals = [
            row["month"],
            "—" if row.get("_tenant") else "P%s" % row["prio"],
            row["origin"], row["dest"],
            row["key"], row["contractor"], row["material"],
            row["target"], row["dt"], row["trips"], row["wmt"],
            row["wmt_per_trip"], row["trips_per_dt"],
        ]
        if achv:
            vals.append(row["achv"])
        vals += [row["n_days"], row["pred_month"],
                 ("—" if row.get("_tenant")
                  else wb_txt.get((row["month_key"],
                                   "%s|%s" % (row.get("key"), row.get("contractor")))) or "")]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = box
            wb_col = len(heads)
            cell.alignment = (mid if col not in (5, wb_col) else Alignment(
                horizontal="left", vertical="center"))
            cell.font = (_xlsx_font(False, 8, _XLSX_MUTED) if col == wb_col
                         else _xlsx_font(col == 5, 9))
            if col in (12, 13) and isinstance(val, (int, float)):
                cell.number_format = "0.00"
            elif col >= 8 and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
            if 9 <= col <= 13:
                cell.font = _xlsx_font(True, 9, _XLSX_PRED)
            if achv and col == 14:
                cell.font = _xlsx_font(True, 9, _XLSX_ACHV)
            if col == pred_col:
                cell.font = _xlsx_font(True, 9, _XLSX_PRED)
        if row.get("_tenant"):
            continue
        if row.get("foreign"):
            # Same rule as the month sheets: IWIP POS-transit is its own fleet.
            # The Year Paths TOTAL read 5,261 DT = 4,073 ours + 1,188 IWIP.
            tot["dt_iwip"] = tot.get("dt_iwip", 0) + (row["dt"] or 0)
            tot["n_iwip"] = tot.get("n_iwip", 0) + 1
            continue
        tot["tgt"] += row["target"] or 0
        tot["dt"] += row["dt"] or 0
        tot["tr"] += row["trips"] or 0
        tot["wmt"] += row["wmt"] or 0
        tot["achv"] += row["achv"] or 0
        tot["pm"] += row["pred_month"] or 0
    r += 1
    pay = round(tot["wmt"] / tot["tr"], 2) if tot["tr"] else None
    tpd = round(tot["tr"] / tot["dt"], 2) if tot["dt"] else None
    n_ours = sum(1 for x in rows
                 if not x.get("_tenant") and not x.get("foreign"))
    tot_vals = [
        "TOTAL", "%s paths" % n_ours, "", "", "", "", "",
        tot["tgt"] or None, tot["dt"] or None,
        int(round(tot["tr"])) if tot["tr"] else None,
        tot["wmt"] or None, pay, tpd,
    ]
    if achv:
        tot_vals.append(int(round(tot["achv"])) if tot["achv"] else None)
    tot_vals += [None, tot["pm"] or None]
    for col, val in enumerate(tot_vals, start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = _xlsx_font(True, 9, _XLSX_NAVY)
        cell.alignment = mid
        _xlsx_total_border(cell)
        if col in (12, 13) and isinstance(val, (int, float)):
            cell.number_format = "0.00"
        elif col >= 8 and isinstance(val, (int, float)):
            cell.number_format = "#,##0"
        if col == pred_col:
            cell.font = _xlsx_font(True, 9, _XLSX_PRED)
        if achv and col == 14:
            cell.font = _xlsx_font(True, 9, _XLSX_ACHV)
    if tot.get("dt_iwip"):
        r += 1
        note = ws.cell(
            row=r, column=1,
            value=("IWIP POS-transit · %s DT-months on the road · %s rows · "
                   "0 WMT · own fleet, not in TOTAL"
                   % (int(round(tot["dt_iwip"])), tot.get("n_iwip", 0))))
        note.font = _xlsx_font(False, 8, _XLSX_MUTED)
        note.alignment = mid
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    from openpyxl.utils import get_column_letter
    ws.auto_filter.ref = "A%d:%s%d" % (
        header_row, get_column_letter(n_cols), r - 1)
    ws.freeze_panes = "A%d" % (header_row + 1)
    _xlsx_widths(ws, [
        10, 10, 10, 14, 22, 12, 14,
        14, 10, 12, 12, 12, 12,
        14 if achv else 10, 12, 16,
    ])
    return r + 2


def _xlsx_append_paths_sheet(wb, year, cards, used, prefix="", achv=False,
                             after_sheet=None):
    """Sheet after Year: every allocated path in every month."""
    if not _collect_year_path_rows(cards):
        return
    name = _xlsx_unique_sheet_name(prefix + "Paths", used)
    if after_sheet and after_sheet in wb.sheetnames:
        pos = wb.sheetnames.index(after_sheet) + 1
    else:
        pos = 1 if wb.sheetnames else 0
    ws = wb.create_sheet(name, pos)
    _xlsx_sheet_setup(ws)
    ws["A1"] = "Paths · %s" % year
    ws["A1"].font = _xlsx_font(True, 16, _XLSX_NAVY)
    ws.merge_cells("A1:P1")
    ws["A2"] = (
        "One row per route per month. Filter Month. "
        "Predicted / month = WMT (t/day) × NB Days. "
        "Other-tenant fleets (material 'other tenant') are listed, 0 WMT."
        + (" Achievable is /api/simulate." if achv else ""))
    ws["A2"].font = _xlsx_font(False, 10, _XLSX_MUTED)
    ws.merge_cells("A2:P2")
    _xlsx_all_paths_table(ws, 4, cards, achv=achv)
    from openpyxl.utils import get_column_letter
    # Weighbridge text is the last column — give it room.
    ws.column_dimensions[get_column_letter(17 if achv else 16)].width = 90
    ws.row_dimensions[1].height = 24


def _xlsx_month_park_routes(ws, r, park, month_name="", alloc=None,
                            n_days=30, col=1):
    """The NEW predicted plan for this month: park the surplus LIM-LD trucks.

    Owner, 2026-08-28 (final): "Option 1 is just that we park the trucks from
    LIM-LD to make it 100% - we are NOT redistributing them to FeNi KM0, that
    was option 2 and we do not want it. Show only option 1: this is the extra
    number of trucks we park, and LIM-LD becomes 100% for the ANNUAL, not for
    that month."

    So: trucks come off the LIM-LD rows and stop. Nothing else in the plan
    moves. SAP and LIM-TOS are untouched by construction, and the target the
    parking is sized against is the YEAR's LIM-LD line."""
    from openpyxl.styles import PatternFill, Alignment
    if not park or not park.get("dt") or not alloc:
        return r
    box = _xlsx_sides()[0]
    red = PatternFill("solid", fgColor="FBE9E9")
    rows = [x for x in (alloc.get("rows") or [])
            if not x.get("foreign") and not x.get("_tenant") and (x.get("dt_after") or 0) > 0]
    if not rows:
        return r
    cut = {}
    for c in park.get("cuts") or []:
        cut[(c["key"], c["con"])] = c["n"]
    yr = park.get("year") or {}
    _t = ws.cell(row=r, column=col,
                 value="NEW PREDICTED PLAN — park %d DT off LIM-LD" % park["dt"])
    _t.font = _xlsx_font(True, 13, _XLSX_NAVY)
    ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 7)
    r += 1
    _sub = ws.cell(row=r, column=col, value=(
        "LIM-LD is running past its YEAR line, so %d trucks come off it and are "
        "PARKED (red). Nothing is re-routed: every other path keeps exactly the "
        "trucks and tonnes it had, which is why SAP and LIM-TOS do not move. "
        "The parking is sized on the YEAR: across Sep-Dec LIM-LD goes %.1f%% -> "
        "%.1f%% of its line."
        % (park["dt"], yr.get("cov_before") or 0, yr.get("cov_after") or 0)))
    _sub.font = _xlsx_font(False, 9, _XLSX_MUTED)
    _sub.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=col, end_row=r + 2, end_column=col + 7)
    r += 3
    _xlsx_headers(ws, r, ["Path", "Contractor", "Material", "DT now", "DT NEW",
                          "Δ DT", "t/day now", "t/day NEW"], start=col, center=True)
    r += 1
    mats_before = {"SAP": 0.0, "LIM-TOS": 0.0, "LIM-LD": 0.0}
    mats_after = {"SAP": 0.0, "LIM-TOS": 0.0, "LIM-LD": 0.0}
    dt_before = dt_after = 0
    order = {"LIM-LD": 0, "LIM-TOS": 1, "SAP": 2}

    def _mk(x):
        mat = str(x.get("material") or "").upper()
        ot = str(x.get("otype") or "").upper()
        return "SAP" if mat == "SAP" else ("LIM-LD" if ot == "LD" else "LIM-TOS")

    for x in sorted(rows, key=lambda z: (order[_mk(z)], -(z.get("pred_after") or 0))):
        mk = _mk(x)
        key, con = x.get("key"), x.get("contractor")
        dt = x.get("dt_after") or 0
        t_day = x.get("pred_after") or 0
        rate = t_day / dt if dt else 0
        n_cut = cut.get((key, con), 0) if mk == "LIM-LD" else 0
        new_dt = dt - n_cut
        new_t = rate * new_dt if n_cut else t_day
        _xlsx_text(ws.cell(row=r, column=col + 0), key, bool(n_cut), center=True)
        _xlsx_text(ws.cell(row=r, column=col + 1), con, center=True)
        _xlsx_text(ws.cell(row=r, column=col + 2), mk, center=True)
        _xlsx_num(ws.cell(row=r, column=col + 3), dt, center=True)
        _xlsx_num(ws.cell(row=r, column=col + 4), new_dt, bool(n_cut), center=True)
        if n_cut:
            _xlsx_num(ws.cell(row=r, column=col + 5), -n_cut, True, center=True)
            ws.cell(row=r, column=col + 5).font = _xlsx_font(True, 11, "A52929")
        else:
            _xlsx_text(ws.cell(row=r, column=col + 5), "—", color=_XLSX_MUTED, center=True)
        _xlsx_num(ws.cell(row=r, column=col + 6), round(t_day), center=True)
        _xlsx_num(ws.cell(row=r, column=col + 7), round(new_t), bool(n_cut), center=True)
        if n_cut:
            for _c in range(col, col + 8):
                ws.cell(row=r, column=_c).fill = red
        for _c in range(col, col + 8):
            ws.cell(row=r, column=_c).border = box
        mats_before[mk] += t_day
        mats_after[mk] += new_t
        dt_before += dt
        dt_after += new_dt
        r += 1
    _xlsx_text(ws.cell(row=r, column=col + 0), "PARKED", True, "A52929", center=True)
    _xlsx_text(ws.cell(row=r, column=col + 1), "", center=True)
    _xlsx_text(ws.cell(row=r, column=col + 2), "", center=True)
    _xlsx_text(ws.cell(row=r, column=col + 3), "—", color=_XLSX_MUTED, center=True)
    _xlsx_num(ws.cell(row=r, column=col + 4), park["dt"], True, center=True)
    ws.cell(row=r, column=col + 4).font = _xlsx_font(True, 12, "A52929")
    _xlsx_text(ws.cell(row=r, column=col + 5), "", center=True)
    _xlsx_text(ws.cell(row=r, column=col + 6), "", center=True)
    _xlsx_text(ws.cell(row=r, column=col + 7), "these trucks do not run", size=9,
               color="A52929", center=True)
    for _c in range(col, col + 8):
        ws.cell(row=r, column=_c).fill = red
        ws.cell(row=r, column=_c).border = box
    r += 1
    _xlsx_text(ws.cell(row=r, column=col + 0), "TOTAL", True, _XLSX_NAVY, center=True)
    _xlsx_text(ws.cell(row=r, column=col + 1), "", center=True)
    _xlsx_text(ws.cell(row=r, column=col + 2), "", center=True)
    _xlsx_num(ws.cell(row=r, column=col + 3), dt_before, True, center=True)
    _xlsx_num(ws.cell(row=r, column=col + 4), dt_after, True, center=True)
    _xlsx_num(ws.cell(row=r, column=col + 5), dt_after - dt_before, True, center=True)
    ws.cell(row=r, column=col + 5).font = _xlsx_font(True, 11, "A52929")
    _xlsx_num(ws.cell(row=r, column=col + 6), round(sum(mats_before.values())), True, center=True)
    _xlsx_num(ws.cell(row=r, column=col + 7), round(sum(mats_after.values())), True, center=True)
    for _c in range(col, col + 8):
        _xlsx_total_border(ws.cell(row=r, column=_c))
    r += 2
    tgt = {}
    for k, api in (("SAP", "sap"), ("LIM-TOS", "tos"), ("LIM-LD", "ld")):
        tgt[k] = ((alloc.get("materials") or {}).get(api) or {}).get("target_day") or 0
    tot_t = sum(tgt.values())
    _xlsx_headers(ws, r, ["NEW % of target", "SAP", "LIM-TOS", "LIM-LD", "Together"],
                  start=col, center=True)
    r += 1
    _xlsx_text(ws.cell(row=r, column=col + 0), "Target t/day", True, center=True)
    for i, k in enumerate(("SAP", "LIM-TOS", "LIM-LD")):
        _xlsx_num(ws.cell(row=r, column=col + 1 + i), round(tgt[k]), center=True)
    _xlsx_num(ws.cell(row=r, column=col + 4), round(tot_t), True, center=True)
    for _c in range(col, col + 5):
        ws.cell(row=r, column=_c).border = box
    r += 1
    _xlsx_text(ws.cell(row=r, column=col + 0), "NEW predicted plan", True, center=True)
    for i, k in enumerate(("SAP", "LIM-TOS", "LIM-LD")):
        _xlsx_num(ws.cell(row=r, column=col + 1 + i), round(mats_after[k]), True, center=True)
    _xlsx_num(ws.cell(row=r, column=col + 4), round(sum(mats_after.values())), True, center=True)
    for _c in range(col, col + 5):
        ws.cell(row=r, column=_c).border = box
    r += 1
    _xlsx_text(ws.cell(row=r, column=col + 0), "NEW % of target", True, center=True)
    for i, k in enumerate(("SAP", "LIM-TOS", "LIM-LD")):
        _xlsx_paint_cov(ws.cell(row=r, column=col + 1 + i),
                        (100 * mats_after[k] / tgt[k]) if tgt[k] else None)
    _xlsx_paint_cov(ws.cell(row=r, column=col + 4),
                    (100 * sum(mats_after.values()) / tot_t) if tot_t else None)
    for _c in range(col, col + 5):
        ws.cell(row=r, column=_c).border = box
    r += 1
    _xlsx_text(ws.cell(row=r, column=col + 0), "was", True, _XLSX_MUTED, center=True)
    for i, k in enumerate(("SAP", "LIM-TOS", "LIM-LD")):
        _xlsx_pct_cell(ws.cell(row=r, column=col + 1 + i),
                       (100 * mats_before[k] / tgt[k]) if tgt[k] else None)
        ws.cell(row=r, column=col + 1 + i).font = _xlsx_font(False, 10, _XLSX_MUTED)
    _xlsx_pct_cell(ws.cell(row=r, column=col + 4),
                   (100 * sum(mats_before.values()) / tot_t) if tot_t else None)
    ws.cell(row=r, column=col + 4).font = _xlsx_font(False, 10, _XLSX_MUTED)
    for _c in range(col, col + 5):
        ws.cell(row=r, column=_c).border = box
    r += 2
    # THE YEAR RESULT — the number that is actually signed off (owner,
    # 2026-08-29: "your total yearly target should be reached 100%").
    _ym = yr.get("materials") or {}
    if _ym:
        _t2 = ws.cell(row=r, column=col, value=(
            "YEAR RESULT with this fleet (%s DT every month)"
            % format(int(yr.get("flat_level") or 0), ",")))
        _t2.font = _xlsx_font(True, 12, "1B7A41")
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 7)
        r += 1
        _xlsx_headers(ws, r, ["Sep–Dec total", "SAP", "LIM-TOS", "LIM-LD", "Together"],
                      start=col, center=True)
        r += 1
        _tg = yr.get("together") or {}
        _xlsx_text(ws.cell(row=r, column=col), "Sales line t", True, center=True)
        for i, k in enumerate(("SAP", "LIM-TOS", "LIM-LD")):
            _xlsx_num(ws.cell(row=r, column=col + 1 + i),
                      round((_ym.get(k) or {}).get("tgt") or 0), center=True)
        _xlsx_num(ws.cell(row=r, column=col + 4), round(_tg.get("tgt") or 0), True,
                  center=True)
        for _c in range(col, col + 5):
            ws.cell(row=r, column=_c).border = box
        r += 1
        _xlsx_text(ws.cell(row=r, column=col), "Year predicted", True, center=True)
        for i, k in enumerate(("SAP", "LIM-TOS", "LIM-LD")):
            _xlsx_num(ws.cell(row=r, column=col + 1 + i),
                      round((_ym.get(k) or {}).get("pred") or 0), True, center=True)
        _xlsx_num(ws.cell(row=r, column=col + 4), round(_tg.get("pred") or 0), True,
                  center=True)
        for _c in range(col, col + 5):
            ws.cell(row=r, column=_c).border = box
        r += 1
        _xlsx_text(ws.cell(row=r, column=col), "YEAR % of target", True, center=True)
        for i, k in enumerate(("SAP", "LIM-TOS", "LIM-LD")):
            _xlsx_paint_cov(ws.cell(row=r, column=col + 1 + i),
                            (_ym.get(k) or {}).get("cov"))
        _xlsx_paint_cov(ws.cell(row=r, column=col + 4), _tg.get("cov"))
        for _c in range(col, col + 5):
            ws.cell(row=r, column=_c).border = box
        r += 2
    if yr.get("cov_after") is not None:
        note = ws.cell(row=r, column=col, value=(
            "The judged line is the YEAR, not this month. Across Sep-Dec LIM-LD was "
            "%.1f%% of its year line and parking these trucks lands it on %.1f%%. The "
            "earlier months run under their own lines, so this month can still read "
            "above 100%% on its own while the year is exactly on target."
            % (yr.get("cov_before") or 0, yr.get("cov_after") or 0)))
        note.font = _xlsx_font(False, 10, _XLSX_MUTED)
        note.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=col, end_row=r + 2, end_column=col + 7)
        r += 2
        r += 2
    return r


def _xlsx_month_dt_box(ws, alloc, top_row=1, col=14, month=None, park=None):
    """Fleet-in-use box at the top right of a month sheet (owner,
    2026-08-27: "clearly mention how many DTs we are using for this
    month"). Counts the ALLOCATED rows: ours by contractor, IWIP
    POS-transit on its own line (it is not contractor fleet), tenants
    excluded entirely (not ours). Also names the matrix POOL and the
    idle rest (owner, 2026-08-27: "why is December showing 1190, not
    1280?" — 1281 is the pool, 1190 is fielded, 91 SMA DT sit idle
    because LD is already at its line and SMA cannot enter RIM-only
    pits)."""
    from openpyxl.utils import get_column_letter
    rows = (alloc.get("rows") or []) if alloc else []
    by_c = {}
    iwip = 0.0
    for r in rows:
        if r.get("_tenant"):
            continue
        dt = r.get("dt_after") or 0
        if not dt:
            continue
        if r.get("foreign"):
            iwip += dt
        else:
            c = r.get("contractor") or "?"
            by_c[c] = by_c.get(c, 0) + dt
    total = sum(by_c.values())
    if not total and not iwip:
        return
    # matrix pool for this month: what the pasted yearly matrix fields
    pool_c = {}
    try:
        _y = _load_yearly()
        _mo = str(int(str(month)[5:7])) if month else None
        for e in (_y or {}).get("entries") or []:
            v = (e.get("dt") or {}).get(_mo)
            if v:
                pool_c[e.get("contractor") or "?"] = \
                    pool_c.get(e.get("contractor") or "?", 0) + v
    except Exception:
        pool_c = {}
    pool = sum(pool_c.values())
    box = _xlsx_sides()[0]
    mid = _xlsx_mid()
    c0 = col
    label = "DT USED THIS MONTH: %s" % format(int(round(total)), ",")
    if pool and int(round(pool)) != int(round(total)):
        label += " of %s" % format(int(round(pool)), ",")
    head = ws.cell(row=top_row, column=c0, value=label)
    head.font = _xlsx_font(True, 14, "FFFFFF")
    head.alignment = mid
    from openpyxl.styles import PatternFill
    fill = PatternFill("solid", fgColor="0F4C81")
    ws.merge_cells(start_row=top_row, start_column=c0,
                   end_row=top_row, end_column=c0 + 2)
    for cc in range(c0, c0 + 3):
        cell = ws.cell(row=top_row, column=cc)
        cell.fill = fill
        cell.border = box
    parts = [" + ".join("%s %s" % (k, format(int(round(v)), ","))
                        for k, v in sorted(by_c.items()))]
    if iwip:
        parts.append("IWIP transit %s (own fleet)" % format(int(round(iwip)), ","))
    idle = {c: pool_c[c] - by_c.get(c, 0) for c in pool_c
            if pool_c[c] - by_c.get(c, 0) >= 1}
    if idle:
        parts.append("idle: " + ", ".join(
            "%s %d" % (c, round(v)) for c, v in sorted(idle.items())))
    det = ws.cell(row=top_row + 1, column=c0, value=" · ".join(parts))
    det.font = _xlsx_font(False, 10, _XLSX_MUTED)
    det.alignment = mid
    # Parking line for THIS month (owner, 2026-08-28: "show in each month
    # the number of DT you parked"). Filled by the year book, which is the
    # only caller that knows the year-level LIM-LD surplus.
    if park is not None:
        _p_dt = int(park.get("dt") or 0)
        _p_keep = int(park.get("absorbed") or 0)
        _p_park = int(park.get("park") or 0)
        if _p_dt:
            _txt = ("TO HIT THE YEAR TARGET: PARK %d DT off LIM-LD" % _p_dt)
            _tone = "A52929"
        else:
            _txt = "TO HIT THE YEAR TARGET: nothing to park this month"
            _tone = "1B7A41"
        _pc = ws.cell(row=top_row + 2, column=c0, value=_txt)
        _pc.font = _xlsx_font(True, 10, _tone)
        _pc.alignment = mid
        ws.merge_cells(start_row=top_row + 2, start_column=c0,
                       end_row=top_row + 2, end_column=c0 + 2)
        for _cc in range(c0, c0 + 3):
            ws.cell(row=top_row + 2, column=_cc).border = box
    ws.merge_cells(start_row=top_row + 1, start_column=c0,
                   end_row=top_row + 1, end_column=c0 + 2)
    for cc in range(c0, c0 + 3):
        ws.cell(row=top_row + 1, column=cc).border = box


def _xlsx_fill_month_alloc(ws, month, title, alloc, st=None, achv=False, park=None):
    """One month: predicted plan, materials, path table. No DT-move list.
    achv=True adds the engine's achievable everywhere predicted appears."""
    _xlsx_sheet_setup(ws)
    n_days = len(_days_in(month))
    src = alloc.get("source_date") or ""
    ws["A1"] = title
    ws["A1"].font = _xlsx_font(True, 16, _XLSX_NAVY)
    ws.merge_cells("A1:M1")
    _xlsx_month_dt_box(ws, alloc, top_row=1, col=14, month=month, park=park)
    if achv:
        ws["A2"] = (
            "Target = matrix. Predicted plan = after Allocate DT. "
            "Achievable = /api/simulate (effective cycle + loader clip). "
            "Not averaged with predicted."
            + ((" Saved %s." % src) if src else "")
            + " Month = day × %s days." % n_days)
    else:
        ws["A2"] = (
            "Target = matrix. Predicted plan = the allocated plan "
            "(after Allocate DT)."
            + ((" Saved %s." % src) if src else "")
            + " Month = day × %s days." % n_days)
    ws["A2"].font = _xlsx_font(False, 10, _XLSX_MUTED)
    ws.merge_cells("A2:M2")

    box = _xlsx_sides()[0]
    mid = _xlsx_mid()
    r = 4
    cov = alloc.get("cov_new_pred")
    if achv:
        month_kpis = [
            ("Target", alloc.get("target_month"), _XLSX_TGT, "Month tonnes"),
            ("Predicted plan", alloc.get("new_pred_month"), _XLSX_PRED, "Month tonnes"),
            ("Achievable", _pick_achv(alloc, False, "month"), _XLSX_ACHV, "Month tonnes"),
            ("Optimized vs target", cov, "059669" if (cov or 0) >= 100 else "D97706", "pct"),
        ]
        day_kpis = [
            ("Target", alloc.get("target_day"), _XLSX_TGT, "t / day"),
            ("Predicted plan", alloc.get("new_pred_day"), _XLSX_PRED, "t / day"),
            ("Achievable", _pick_achv(alloc, False, "day"), _XLSX_ACHV, "t / day"),
            ("Fleet after allocate", alloc.get("dt_after"), _XLSX_INK, "DT"),
        ]
    else:
        month_kpis = [
            ("Target", alloc.get("target_month"), _XLSX_TGT, "Month tonnes"),
            ("Predicted plan", alloc.get("new_pred_month"), _XLSX_PRED, "Month tonnes"),
            ("Predicted vs target", cov, "059669" if (cov or 0) >= 100 else "D97706", "pct"),
        ]
        day_kpis = [
            ("Target", alloc.get("target_day"), _XLSX_TGT, "t / day"),
            ("Predicted plan", alloc.get("new_pred_day"), _XLSX_PRED, "t / day"),
            ("Fleet after allocate", alloc.get("dt_after"), _XLSX_INK, "DT"),
        ]
    r = _xlsx_kpi_strip(ws, r, month_kpis, start=1)
    r = _xlsx_kpi_strip(ws, r, day_kpis, start=1)

    r = _xlsx_dest_from_pits(ws, r, alloc.get("rows") or [])

    mats = alloc.get("materials") or {}
    r = _xlsx_section(
        ws, r, "Materials — t / day",
        "Same clocks as the year sheet. Coverage is predicted plan ÷ target.")
    labels = [("sap", "SAP"), ("tos", "LIM-TOS"), ("ld", "LIM-LD")]
    _xlsx_headers(ws, r, ["", "SAP", "LIM-TOS", "LIM-LD", "Together"], center=True)
    if achv:
        metric = [
            ("Target t/day", lambda k: (mats.get(k) or {}).get("target_day"), alloc.get("target_day"), _XLSX_INK, False),
            ("Predicted plan", lambda k: (mats.get(k) or {}).get("pred_after_day"), alloc.get("new_pred_day"), _XLSX_INK, True),
            ("Achievable",
             lambda k: _pick_mat_achv(mats.get(k) or {}, False, "day"),
             _pick_achv(alloc, False, "day"), _XLSX_INK, True),
            ("New DT", lambda k: (mats.get(k) or {}).get("dt_after"), alloc.get("dt_after"), _XLSX_INK, True),
        ]
    else:
        metric = [
            ("Target t/day", lambda k: (mats.get(k) or {}).get("target_day"), alloc.get("target_day"), _XLSX_INK, False),
            ("Predicted plan", lambda k: (mats.get(k) or {}).get("pred_after_day"), alloc.get("new_pred_day"), _XLSX_INK, True),
            ("DT", lambda k: (mats.get(k) or {}).get("dt_after"), alloc.get("dt_after"), _XLSX_INK, True),
        ]
    for lab, fn, tot, color, bold in metric:
        r += 1
        lab_c = ws.cell(row=r, column=1, value=lab)
        lab_c.font = _xlsx_font(False, 10, _XLSX_MUTED)
        lab_c.border = box
        lab_c.alignment = mid
        for i, (k, _) in enumerate(labels):
            cell = ws.cell(row=r, column=2 + i, value=fn(k))
            cell.border = box
            cell.alignment = mid
            cell.font = _xlsx_font(bold, 13 if bold else 11, color)
            if isinstance(fn(k), (int, float)):
                cell.number_format = "#,##0"
        tot_c = ws.cell(row=r, column=5, value=tot)
        tot_c.border = box
        tot_c.alignment = mid
        tot_c.font = _xlsx_font(True, 13, color)
        if isinstance(tot, (int, float)):
            tot_c.number_format = "#,##0"
    r += 1
    lab_c = ws.cell(row=r, column=1, value="Predicted % of target")
    lab_c.font = _xlsx_font(True, 10, _XLSX_NAVY)
    lab_c.border = box
    lab_c.alignment = mid
    for i, (k, _) in enumerate(labels):
        _xlsx_paint_cov(ws.cell(row=r, column=2 + i), (mats.get(k) or {}).get("cov_pred"), size=13)
    _xlsx_paint_cov(ws.cell(row=r, column=5), alloc.get("cov_new_pred"), size=13)
    ws.row_dimensions[r].height = 22
    r += 2
    left_lab = ws.cell(row=r, column=1, value="Delta vs target · month t")
    left_lab.font = _xlsx_font(False, 10, _XLSX_MUTED)
    left_lab.border = box

    def _delta_cell(cell, v, bold=False):
        """Signed gap to target: red under, green over (owner 2026-08-31)."""
        _xlsx_num(cell, v, bold, center=True)
        if isinstance(v, (int, float)) and v:
            cell.number_format = "+#,##0;-#,##0"
            cell.font = _xlsx_font(True, 11, "1B7A41" if v > 0 else "A52929")
        return cell

    for i, (k, _) in enumerate(labels):
        _m = mats.get(k) or {}
        _tg = _m.get("target_month")
        _pv = _m.get("pred_after_month")
        _d = ((_pv - _tg) if (_tg is not None and _pv is not None)
              else _m.get("delta_pred_month"))
        if _d is None and _m.get("left_pred_month") is not None:
            _d = -_m["left_pred_month"]
        _delta_cell(ws.cell(row=r, column=2 + i), _d)
    _tgt_m = alloc.get("target_month")
    _prd_m = alloc.get("new_pred_month")
    _dt = ((_prd_m - _tgt_m) if (_tgt_m is not None and _prd_m is not None)
           else alloc.get("delta_new_pred_month"))
    if _dt is None and alloc.get("left_new_pred_month") is not None:
        _dt = -alloc["left_new_pred_month"]
    _delta_cell(ws.cell(row=r, column=5), _dt, True)

    rows = list(alloc.get("rows") or [])
    _paths_top = None
    if rows:
        r += 2
        _paths_top = r          # the NEW plan table is placed beside this one
        r = _xlsx_path_alloc_table(
            ws, r, rows, "Paths — predicted plan",
            "P1 SAP · P2 LIM-TOS · P3 LIM-LD. WMT is predicted tonnes. DT is trucks. "
            "Trips are predicted trips. WMT/DT is tonnes per truck-trip (payload)."
            + (" Achievable is /api/simulate (effective cycle + loader clip), t/day." if achv else ""),
            achv=achv)

    cap = alloc.get("capacity") or {}
    if cap:
        # Two labelled numbers, never merged: LIM-LD capacity is never clipped,
        # only the tonnage credited against the target stops at it.
        r += 2
        r = _xlsx_section(
            ws, r, "LIM-LD capacity vs credited production",
            "Capacity is what the free fleet could move and is never clipped. "
            "Credited is what counts against the target. The difference is "
            "unused/excess capacity — reported, never folded into production.")
        _xlsx_headers(ws, r, ["", "DT", "t/day", "t/month"], start=1, center=True)
        for lab, dt_k, day_k, mon_k in (
                ("Capacity (never clipped)", "ld_dt_capacity",
                 "ld_t_day_capacity", "ld_t_month_capacity"),
                ("Credited against target", "ld_dt_credited",
                 "ld_t_day_credited", "ld_t_month_credited"),
                ("Unused / excess capacity", "ld_dt_unused",
                 "ld_t_day_excess", "ld_t_month_excess")):
            r += 1
            _xlsx_text(ws.cell(row=r, column=1), lab)
            _xlsx_num(ws.cell(row=r, column=2), cap.get(dt_k), center=True)
            _xlsx_num(ws.cell(row=r, column=3), cap.get(day_k), center=True)
            _xlsx_num(ws.cell(row=r, column=4), cap.get(mon_k), center=True)
        r += 1
        if not cap.get("feasible", True):
            ws.cell(row=r, column=1, value=(
                "NOT FEASIBLE: targets this month need %s DT more than the pool "
                "holds. Credited tonnage is what the fielded trucks can move, "
                "not what was asked." % cap.get("infeasible_dt"))
            ).font = _xlsx_font(True, 10, _XLSX_INK)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            r += 1

    # Hour grid sits ABOVE the 30-day daily table so it is not buried under
    # a chart. Same table as Plan → C · Road crowding by hour.
    r = _xlsx_month_corridor_block(ws, r + 2, alloc=alloc)

    r += 2
    days = _days_in(month)
    points = [{
        "name": d,
        "label": "Date",
        "target": alloc.get("target_day"),
        "old_pred": alloc.get("old_pred_day"),
        "new_pred": alloc.get("new_pred_day"),
        "old_achv": _pick_achv(alloc, True, "day") if achv else None,
        "new_achv": _pick_achv(alloc, False, "day") if achv else None,
    } for d in days]
    r = _xlsx_five_clock_block(
        ws, r, "Daily WMT (flat — the same plan every day)",
        "Target vs predicted plan"
        + (" · achievable from /api/simulate." if achv else ". Same day every day."),
        points, start=1, chart_col="A", achv=achv)

    # Which trucks move to land the YEAR's LIM-LD on 100% (owner 2026-08-28).
    # The NEW plan sits SIDE BY SIDE with the old one (owner, 2026-08-28:
    # "move this table next to the old plan so both tables are together").
    # Column N clears the old table's A-L span and its weighbridge tail.
    _xlsx_month_park_routes(ws, _paths_top if _paths_top else r + 2, park,
                            month_name=month, alloc=alloc, n_days=n_days,
                            col=14)

    # Weighbridge text is the LAST column; give the tail columns room.
    _xlsx_widths(ws, [16, 22, 14, 14, 14, 11, 11, 11, 11, 12, 14, 58, 58])
    # The NEW plan table sits in N.. beside the old one; give it real widths
    # and a gutter so the two read as a pair (owner, 2026-08-28).
    if park and park.get("dt"):
        ws.column_dimensions["M"].width = 3
        _xlsx_widths(ws, [20, 13, 13, 11, 11, 9, 13, 13], start=14)
    ws.freeze_panes = "A4"
    ws.row_dimensions[1].height = 24
    return alloc.get("new_pred_month"), alloc.get("target_month")


def _xlsx_fill_month(ws, st, title, achv=False):
    """One month: KPIs, Production & capacity, SAP targets / required DT, daily chart.
    achv=True adds the engine's achievable (Plan Step 2 /api/simulate)."""
    from openpyxl.styles import Alignment
    _xlsx_sheet_setup(ws)
    rows, p, m = _daily_triples(st)
    paths = list(p.get("paths") or [])
    sim, _plans = _simulate_for_paths(paths)
    sim_rows = (sim or {}).get("results") or []
    sim_sum = (sim or {}).get("summary") or {}
    n_days = len(rows) or 1

    ws["A1"] = title
    ws["A1"].font = _xlsx_font(True, 16, _XLSX_NAVY)
    ws.merge_cells("A1:L1")
    ws["A2"] = ("Same day every day. Predicted plan = path model · Target = matrix."
                + (" Achievable = /api/simulate (effective cycle + loader clip), same as Plan Step 2."
                   if achv else ""))
    ws["A2"].font = _xlsx_font(False, 10, _XLSX_MUTED)
    ws.merge_cells("A2:L2")

    tot_p = sum((pw or 0) for _, pw, _, _ in rows)
    tot_a = sum((aw or 0) for _, _, aw, _ in rows)
    tot_t = sum((tw or 0) for _, _, _, tw in rows)
    pred_day = (p.get("per_day_wmt") if p.get("per_day_wmt") is not None
                else (tot_p / n_days if n_days else None))
    achv_day = (p.get("per_day_achv_wmt") if p.get("per_day_achv_wmt") is not None
                else (tot_a / n_days if n_days else None))
    tgt_day = (tot_t / n_days) if tot_t else None
    planned_day = None
    if sim_sum.get("planned_production_t") is not None:
        planned_day = float(sim_sum["planned_production_t"]) * 2
    shortfall_day = None
    if planned_day is not None and achv_day is not None:
        shortfall_day = max(0.0, planned_day - float(achv_day))

    r = 4
    month_kpis = [
        ("Target", tot_t or None, _XLSX_TGT, "Month tonnes"),
        ("Predicted plan", tot_p or None, _XLSX_PRED, "Month tonnes"),
    ]
    if achv:
        month_kpis.append(("Achievable", tot_a or None, _XLSX_ACHV, "Month tonnes"))
    month_kpis.append(("Trucks", p.get("dt"), _XLSX_INK, "DT"))
    r = _xlsx_kpi_strip(ws, r, month_kpis, start=1)
    day_kpis = [
        ("Target", tgt_day, _XLSX_TGT, "t / day"),
        ("Predicted plan", pred_day, _XLSX_PRED, "t / day"),
    ]
    if achv:
        day_kpis.append(("Achievable", achv_day, _XLSX_ACHV, "t / day"))
    r = _xlsx_kpi_strip(ws, r, day_kpis, start=1)
    r = _xlsx_dest_from_pits(ws, r, _paths_as_flow_rows(paths))
    box = _xlsx_sides()[0]
    mid = _xlsx_mid()
    r = _xlsx_section(ws, r, "Production",
                      "Day = 2 × 12 h shifts. Predicted plan is the path model. Target is the matrix.")

    cap_heads = ["Path", "Contractor", "Material", "DT", "Cycle min",
                 "Eff. cycle min", "Trips/DT", "Predicted plan t/day"]
    if achv:
        cap_heads.append("Achievable t/day")
    cap_heads += ["Target t/day", "Roster"]
    _xlsx_headers(ws, r, cap_heads, center=True)
    tot_dt = tot_pred = tot_tgt = tot_path_achv = 0
    pred_col = 8
    achv_col = 9 if achv else None
    tgt_col = 10 if achv else 9
    roster_col = tgt_col + 1
    for i, prow in enumerate(paths):
        r += 1
        sr = sim_rows[i] if i < len(sim_rows) else {}
        key = prow.get("key")
        dt = prow.get("dt") or 0
        pred = prow.get("pred_wmt_day")
        tgt = prow.get("manual_wmt_day")
        trips_dt = sr.get("trips_per_shift_per_truck")
        path_achv = prow.get("achv_wmt_day")
        if path_achv is None and sr.get("achievable_production_t") is not None:
            path_achv = float(sr["achievable_production_t"]) * 2
        vals = [
            key, prow.get("contractor"), prow.get("material"), dt,
            sr.get("predicted_cycle_time_min"), sr.get("effective_cycle_min"),
            trips_dt, pred,
        ]
        if achv:
            vals.append(path_achv)
        vals += [tgt, sr.get("trucks_to_roster")]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = box
            cell.font = _xlsx_font(col == 1, 10)
            cell.alignment = mid
            if col in (4, 5, 6, 7, pred_col, tgt_col, roster_col) + ((achv_col,) if achv_col else ()) and val is not None:
                cell.number_format = "#,##0.0" if col in (5, 6, 7) else "#,##0"
        tot_dt += dt or 0
        tot_pred += pred or 0
        tot_tgt += tgt or 0
        tot_path_achv += path_achv or 0
        ws.row_dimensions[r].height = 18
    if paths:
        r += 1
        _xlsx_text(ws.cell(row=r, column=1), "TOTAL", True, _XLSX_NAVY, center=True)
        for col in range(2, roster_col + 1):
            ws.cell(row=r, column=col).border = box
            ws.cell(row=r, column=col).alignment = mid
        _xlsx_num(ws.cell(row=r, column=4), tot_dt or None, True, center=True)
        _xlsx_num(ws.cell(row=r, column=pred_col), tot_pred or None, True, center=True)
        if achv:
            ac = ws.cell(row=r, column=achv_col)
            _xlsx_num(ac, tot_path_achv or None, True, center=True)
        _xlsx_num(ws.cell(row=r, column=tgt_col), tot_tgt or None, True, center=True)

    warns = sim_sum.get("capacity_warnings") or p.get("extrapolated") or []
    if warns:
        r += 2
        r = _xlsx_section(ws, r, "Capacity warnings")
        for w in warns:
            ws.cell(row=r, column=1, value=str(w)).font = _xlsx_font(False, 9, _XLSX_MUTED)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
            ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r].height = 28
            r += 1

    r += 2
    sap_rows = _sap_table_rows(st, paths, sim_rows)
    r = _xlsx_section(
        ws, r, "Priority targets — P1 SAP · P2 LIM from TOS",
        "Supplied LD tonnage is the P3 target. Predicted = path model.")
    sap_heads = ["P", "Path", "Mat · type", "Contractor", "Target t/day", "Predicted plan t/day",
                 "Allocated DT", "Required DT", "Status"]
    _xlsx_headers(ws, r, sap_heads, center=True)
    if not sap_rows:
        r += 1
        _xlsx_text(ws.cell(row=r, column=1),
                   "No P1/P2 rows in the matrix for this month.",
                   False, _XLSX_MUTED, 10)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    for sap in sap_rows:
        r += 1
        # c840db8 removed the status colour coding but deleted the `status`
        # binding with it, leaving `status` in vals — every workbook export
        # (year books, scenario zip, J72) died with NameError.
        status = sap.get("status") or ""
        vals = ["P%s" % sap.get("prio", 1), sap["path"],
                "%s · %s" % (sap.get("mat") or "", sap.get("otype") or "—"),
                sap["contractor"], sap["target"], sap["pred"],
                sap["alloc_dt"], sap["req_dt"], status]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = box
            cell.font = _xlsx_font(col == 2 or col == 9, 10, _XLSX_INK)
            cell.alignment = mid
            if col in (5, 6, 7, 8) and val is not None:
                cell.number_format = "#,##0"

    alloc = (st or {}).get("allocation") or {}
    if alloc.get("new") or alloc.get("old"):
        r += 2
        r = _xlsx_section(
            ws, r, "After priority allocation",
            "Original production above is the matrix fleet. "
            "These clocks are after moving DT to P1 then P2 inside each contractor.")
        r = _xlsx_kpi_strip(ws, r, [
            ("Target", (alloc.get("new") or alloc.get("old") or {}).get("target"),
             _XLSX_TGT, "t / day"),
            ("Predicted plan", (alloc.get("new") or {}).get("pred"),
             _XLSX_PRED, "t / day"),
        ] + ([
            ("Achievable",
             (alloc.get("new") or {}).get("achv_sim")
             if (alloc.get("new") or {}).get("achv_sim") is not None
             else (alloc.get("new") or {}).get("achv"),
             _XLSX_ACHV, "t / day"),
        ] if achv else []), start=1)
        if alloc.get("shortfalls"):
            r += 1
            r = _xlsx_section(ws, r, "Fleet shortfalls")
            for s in alloc["shortfalls"]:
                ws.cell(row=r, column=1, value=str(s)).font = _xlsx_font(False, 9, _XLSX_MUTED)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
                r += 1

    day_alloc = (st or {}).get("saved_day_allocation") or {}
    if day_alloc.get("frozen"):
        r += 2
        r = _xlsx_saved_day_allocation(ws, r, day_alloc)

    r += 3
    daily_heads = ["Date", "Predicted plan"]
    if achv:
        daily_heads.append("Achievable")
    daily_heads.append("Target")
    daily_head = _xlsx_section(
        ws, r, "Daily WMT",
        "Same day every day — the lines are flat on purpose."
        + (" Achievable is /api/simulate." if achv else ""))
    _xlsx_headers(ws, daily_head, daily_heads, center=True)
    dr = daily_head
    tgt_dcol = 4 if achv else 3
    for d, pw, aw, tw in rows:
        dr += 1
        _xlsx_text(ws.cell(row=dr, column=1), d, center=True)
        pred_c = ws.cell(row=dr, column=2)
        _xlsx_num(pred_c, pw, center=True)
        if achv:
            achv_c = ws.cell(row=dr, column=3)
            _xlsx_num(achv_c, aw, center=True)
        tgt_c = ws.cell(row=dr, column=tgt_dcol)
        _xlsx_num(tgt_c, tw, center=True)
    last_daily = dr
    dr += 1
    _xlsx_text(ws.cell(row=dr, column=1), "TOTAL", True, _XLSX_NAVY, center=True)
    tot_p_c = ws.cell(row=dr, column=2)
    _xlsx_num(tot_p_c, tot_p or None, True, center=True)
    if achv:
        tot_a_c = ws.cell(row=dr, column=3)
        _xlsx_num(tot_a_c, tot_a or None, True, center=True)
    tot_t_c = ws.cell(row=dr, column=tgt_dcol)
    _xlsx_num(tot_t_c, tot_t or None, True, center=True)
    if last_daily > daily_head:
        pal = (_XLSX_MUTED, _XLSX_ACHV, _XLSX_TGT) if achv else (_XLSX_MUTED, _XLSX_TGT)
        _xlsx_line_chart(ws, "Daily WMT", "t/day", 2, tgt_dcol, daily_head, last_daily,
                         "A%d" % (dr + 2), colors=pal)

    last_chart_row = dr + 18
    r = _xlsx_month_corridor_block(ws, last_chart_row, alloc=day_alloc if day_alloc.get("frozen") else None,
                                  paths=paths)

    _xlsx_widths(ws, [22, 14, 14, 13, 12, 14, 13, 16, 20, 16, 14, 10])
    ws.freeze_panes = "A4"
    ws.row_dimensions[1].height = 24
    return tot_p, tot_a, tot_t


def _xlsx_month_book(month, st):
    from openpyxl import Workbook
    wb = Workbook()
    n = len(_days_in(month))
    raw, src = _resolve_allocation(month, st)
    alloc = _alloc_view(raw, n, src, include_detail=True)
    label = "%s %s" % (calendar.month_name[int(month[5:7])], month[:4])
    key = wb.active
    if alloc:
        key.title = "Key"
        _xlsx_fill_month_alloc(key, month, "%s — old vs new" % label, alloc, st)
        return wb
    key.title = "Key"
    rows, _p, _m = _daily_triples(st)
    tot_p = sum((pw or 0) for _, pw, _, _ in rows)
    tot_a = sum((aw or 0) for _, _, aw, _ in rows)
    tot_t = sum((tw or 0) for _, _, _, tw in rows)
    r = _xlsx_board_header(
        key, label,
        "Same day every day. Predicted plan · Target.")
    kpis = [
        ("Target", tot_t or None, _XLSX_TGT, "Month tonnes"),
        ("Predicted plan", tot_p or None, _XLSX_PRED, "Month tonnes"),
    ]
    if tot_p and tot_t:
        kpis.append(("Prediction vs target", tot_p - tot_t,
                     "059669" if tot_p >= tot_t else "B91C1C", "tonnes"))
    r = _xlsx_kpi_strip(key, r, kpis, start=1)
    table_row = r
    key.cell(row=table_row, column=1, value="Daily totals").font = _xlsx_font(True, 13, _XLSX_NAVY)
    table_row += 1
    _xlsx_headers(key, table_row, ["Date", "Predicted plan", "Target"], start=1)
    rr = table_row
    for d, pw, aw, tw in rows:
        rr += 1
        _xlsx_text(key.cell(row=rr, column=1), d)
        pc = key.cell(row=rr, column=2)
        _xlsx_num(pc, pw)
        tc = key.cell(row=rr, column=3)
        _xlsx_num(tc, tw)
    last = rr
    rr += 1
    _xlsx_text(key.cell(row=rr, column=1), "TOTAL", True, _XLSX_NAVY)
    pc = key.cell(row=rr, column=2)
    _xlsx_num(pc, tot_p or None, True)
    tc = key.cell(row=rr, column=3)
    _xlsx_num(tc, tot_t or None, True)
    if last > table_row:
        _xlsx_line_chart(key, "Daily WMT (flat — the same plan every day)", "t/day",
                         2, 3, table_row, last, "A%d" % (rr + 2), cat_col=1,
                         colors=(_XLSX_MUTED, _XLSX_TGT))
    month_ws = wb.create_sheet(label[:31])
    _xlsx_fill_month(month_ws, st, "%s — production, capacity & SAP" % label)
    return wb


def _xlsx_unique_sheet_name(base, used):
    name = (base or "Sheet")[:31]
    n = 2
    while name in used:
        name = ("%s %d" % (base, n))[:31]
        n += 1
    used.add(name)
    return name


def _xlsx_plan_source_block(ws, r, cards):
    """WHICH PLAN THIS WORKBOOK IS. Printed on the Year sheet, above the totals.

    The Year TOTAL used to sum months that came from different scenarios with
    no disclosure anywhere on the sheet (month sheets said "Saved 2026-09-04"
    in row 2; the Year sheet said nothing). One workbook is now one scenario,
    and it says so here — scenario, day, and the exact saved file per month,
    plus any month that has no plan for this scenario rather than a silent
    hole where another day's plan used to be substituted."""
    first = cards[0] if cards else {}
    note = first.get("_source_note")
    scen = first.get("_scenario")
    if not note and not scen:
        return r
    r = _xlsx_section(ws, r, "Plan source — one workbook, one scenario", note)
    _xlsx_headers(ws, r, ["Month", "Saved plan read", "Target (t)"], start=1)
    for c in cards:
        r += 1
        a = c.get("alloc") or {}
        _xlsx_text(ws.cell(row=r, column=1), c.get("name"))
        _xlsx_text(ws.cell(row=r, column=2),
                   c.get("alloc_source_date") or a.get("source_date")
                   or c.get("alloc_source") or "—")
        _xlsx_num(ws.cell(row=r, column=3), a.get("target_month") or c.get("target_month"))
    r += 2
    missing = first.get("_missing_months") or []
    ws.cell(row=r, column=1, value=(
        "Months with no %s plan: %s — not substituted from another day."
        % (scen or "scenario", ", ".join(str(m) for m in missing))
        if missing else
        "Every month in this year has a %s plan; nothing was substituted."
        % (scen or "scenario"))).font = _xlsx_font(False, 9, _XLSX_MUTED)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    cap = (first.get("alloc") or {}).get("capacity") or {}
    if cap:
        ws.cell(row=r, column=1, value=(
            "LIM-LD: capacity is reported in full and never clipped; only the "
            "tonnage CREDITED against the target stops at it. See each month "
            "sheet for that month's capacity / credited / unused split."
        )).font = _xlsx_font(False, 9, _XLSX_MUTED)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
    return r + 1



def _is_tenant_row(r):
    """Is this saved allocation row one of the OTHER tenants' fleets?

    ONE home: congestion.tenants.is_tenant_plan. Recognised by name as well
    as by flag so a plan frozen before `_tenant` was stamped still drops
    them — sending them into shared_flow would charge the register twice.
    """
    from congestion.tenants import is_tenant_plan
    return is_tenant_plan(r)


_TENANT_SHEET_CACHE = None


def _tenant_sheet_rows():
    """Allocation-shaped rows for the other-tenant register (display only).

    Saved plans usually omit these — the Plan tab injects them live from
    congestion/tenants.py. Excel was listing only IWIP POS-transit (blank /
    'road' material) and dropping MHM / POSITION / PMA / HSM / KR>RSF /
    HUAFEI>RSF. Same fleets as New Allocation Plan; still never simulated.
    """
    global _TENANT_SHEET_CACHE
    if _TENANT_SHEET_CACHE is not None:
        return [dict(x) for x in _TENANT_SHEET_CACHE]
    from congestion.tenants import tenant_rows as _trows
    out = []
    for t in _trows():
        dt = int(round(float(t.get("dt") or 0)))
        if dt <= 0:
            continue
        tpd = t.get("trips_per_dt")
        trips_day = t.get("trips_per_day")
        if trips_day is None and tpd:
            trips_day = dt * float(tpd)
        trips = int(round(trips_day)) if trips_day else 0
        out.append({
            "id": "TENANT|%s|road" % t["name"],
            "key": t.get("route"),
            "contractor": t["name"],
            "material": "other tenant",
            "otype": "",
            "prio": None,
            "target": 0,
            "foreign": True,
            "_tenant": True,
            "dt_before": dt,
            "dt_after": dt,
            "pred_before": 0,
            "pred_after": 0,
            "achv_before": 0,
            "achv_after": 0,
            "achv_sim": 0,
            "trips": trips,
            "trips_before": trips,
        })
    _TENANT_SHEET_CACHE = out
    return [dict(x) for x in out]


def _ensure_tenant_rows(rows):
    """Plan UI shows other-tenant DT; Excel must list the same fleets.

    Does not send them into shared_flow — `_plans_from_alloc_rows` still
    drops them so the register is charged once as background flow.
    """
    out = []
    seen = set()
    for r in rows or []:
        if not isinstance(r, dict):
            continue
        rr = dict(r)
        if _is_tenant_row(rr):
            if not (rr.get("material") or "").strip():
                rr["material"] = "other tenant"
            seen.add(str(rr.get("contractor") or "").upper())
        elif rr.get("foreign") and not (rr.get("material") or "").strip():
            rr["material"] = "road"
        out.append(rr)
    for t in _tenant_sheet_rows():
        name = str(t.get("contractor") or "").upper()
        if name in seen:
            continue
        out.append(t)
        seen.add(name)
    return out


def _plans_from_alloc_rows(rows):
    """Allocation / holding rows → shared_flow plans. Includes IWIP on the plan."""
    plans = []
    for i, r in enumerate(rows or []):
        if not isinstance(r, dict):
            continue
        dt = r.get("dt_after")
        if dt is None:
            dt = r.get("n_trucks") if r.get("n_trucks") is not None else r.get("dt")
        if not (dt or 0) > 0:
            continue
        # Other tenants never come from a saved plan. They are a register that
        # the road model injects itself as background flow, so a saved copy
        # would be counted twice — and plans saved before the save-side guard
        # existed carry them WITHOUT the _tenant flag, so they must be
        # recognised by name here too. Left in, an unrecognised "KR>RSF" row
        # cannot be placed on the stick (RSF is not one of our nodes) and the
        # corridor invents a "<SOURCE> spur" for it: the phantom "KR spur" and
        # "HUAFEI spur" the owner reported on the Road crowding card.
        if _is_tenant_row(r):
            continue
        key = str(r.get("key") or "")
        if ">" in key:
            src, dst = key.split(">", 1)
        else:
            src = r.get("source") or r.get("origin")
            dst = r.get("destination") or r.get("dest")
        if not src or not dst:
            continue
        plans.append({
            "id": r.get("id") or ("r%d" % i),
            "source": src, "destination": dst,
            "n_trucks": int(round(dt)),
            "contractor": r.get("contractor"),
        })
    return plans


# Year Excel used to call shared_flow 12 times for 4 months (Year + Road
# crowding + each month sheet) — ~14 s locally, long enough for ngrok/Safari
# to cut the request and save an HTML error as .xlsx. Same trucks, same rain,
# one DES per plan fingerprint; a process cache makes a retry instant.
_CORRIDOR_CACHE = {}
_CORRIDOR_LOCK = threading.Lock()
_CORRIDOR_CACHE_MAX = 32


def _corridor_cache_key(plans):
    from congestion.speed_limits import FOLLOWING_DISTANCE_M
    items = tuple(sorted(
        (str(p.get("source") or ""),
         str(p.get("destination") or ""),
         int(round(p.get("n_trucks") or 0)),
         str(p.get("contractor") or ""))
        for p in (plans or [])
    ))
    return (FOLLOWING_DISTANCE_M, items)


def _corridor_run(plans):
    """Time the given trucks onto the stick. Advisory — never clips tonnes."""
    if not plans:
        return None
    key = _corridor_cache_key(plans)
    with _CORRIDOR_LOCK:
        hit = _CORRIDOR_CACHE.get(key)
    if hit is not None:
        return hit
    try:
        import plan_shared_flow as _sf
        # tenants=True: the workbook's road view must match the Plan tab's, and
        # the road carries the other tenants whether or not our plan mentions
        # them. The block states the DT and that they are not ours to move.
        res = _sf.shared_flow(plans, shift_hours=12, rain_mm=0, start_hour=7,
                              whole_day=True, tenants=True)
    except Exception:  # noqa: BLE001 — a report must not die on an advisory panel
        return None
    if not res.get("ok"):
        return None
    got = res, sum(p["n_trucks"] for p in plans)
    with _CORRIDOR_LOCK:
        if key not in _CORRIDOR_CACHE:
            if len(_CORRIDOR_CACHE) >= _CORRIDOR_CACHE_MAX:
                _CORRIDOR_CACHE.pop(next(iter(_CORRIDOR_CACHE)))
            _CORRIDOR_CACHE[key] = got
    return got


def _corridor_for_alloc(alloc):
    """Hourly corridor from the SAME rows the month sheet already printed."""
    if not isinstance(alloc, dict):
        return None
    plans = _plans_from_alloc_rows(alloc.get("rows") or [])
    got = _corridor_run(plans)
    if got:
        return got
    src = alloc.get("source_date") or alloc.get("alloc_source_date")
    if not src:
        return None
    return _corridor_for_month({"alloc_source_date": src})


def _corridor_for_month(card):
    """Per-section road occupancy for ONE finalised month, or None.

    The corridor sits DOWNSTREAM of the plan (owner, 2026-08-24): once a
    month's allocation is finalised, the corridor's only job is to distribute
    THOSE trucks across sections and hours. It does not re-derive or
    second-guess the plan's tonnage — that was settled upstream.

    Rain is forced to the normal-day basis. The owner plans normal days and
    applies rain deliberately as a scenario; the wet response is physically
    derived but has never been validated on site (zero wet days in every
    measurement window), so it must not silently shape a published number.

    IWIP / POS-transit rows that are already on the allocated plan ARE
    included — they share the road when this plan runs. Extra measured
    Other-trips (the Plan-tab checkbox, not saved on the allocation) are not.
    """
    # Prefer rows already on the card/view so the grid cannot drift from
    # the path table above it. Fall back to the named saved file.
    rows = (card.get("alloc") or {}).get("rows") or card.get("rows") or []
    plans = _plans_from_alloc_rows(rows)
    got = _corridor_run(plans)
    if got:
        return got
    src = card.get("alloc_source_date") or (card.get("alloc") or {}).get("source_date")
    if not src:
        return None
    path = os.path.join(_ROOT, "data", "saved_plans", "%s.json" % str(src)[:10])
    if not os.path.isfile(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as fh:
            saved = json.load(fh)
    except (OSError, ValueError):
        return None
    plans = _plans_from_alloc_rows((saved.get("allocation") or {}).get("rows") or [])
    return _corridor_run(plans)


def _xlsx_road_corridor_block(ws, r, cards):
    """Page-one road corridor: trucks per section, and WHEN each section peaks.

    Advisory, and labelled as such on the sheet — this never clips simulate
    tonnes (J53). What is measured and what is not is stated on the face of the
    block rather than in a footnote nobody reads: the release shape comes from
    273,222 weighbridge loads over 234 days; the section split from 463,060
    measured traversals; the capacity from the official speed-limit sheets.
    Presence trails the loading peak by each section's transit time, which is
    physics, not a fitted lag.
    """
    from openpyxl.styles import Font
    ws.cell(row=r, column=1, value="Road corridor — trucks on each section")
    ws.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=1, value=(
        "Hour-by-hour tables (same numbers and colours as Plan → Road crowding) "
        "are on the Road crowding sheet and each month tab."))
    ws.cell(row=r, column=1).font = Font(italic=True, size=9)
    r += 1
    # Peak here is the BIN-FREE instantaneous maximum; the month grid shows the
    # MEAN concurrent within each hour. Both are real and they are not the same
    # number — the peak legitimately exceeds every cell of the grid. Say so on
    # the face of the table: two panels quoting one concept is how this project
    # has been misread before.
    ws.cell(row=r, column=1, value=(
        "Peak = most trucks on the LOADED lane at any instant (empty uses the "
        "other carriageway). Colour and % are the same one-lane packing as Plan "
        "(50 m). The month tabs show the AVERAGE across each hour, so the peak "
        "sits above every cell there."))
    ws.cell(row=r, column=1).font = Font(italic=True, size=9)
    r += 1
    ws.cell(row=r, column=1, value=(
        "Advisory. Distribution of the FINALISED plan across the road — it never "
        "changes plan tonnage. Normal-day basis (0-1 mm rain). Includes the other "
        "tenants' 1,340 DT, which share the road and are NOT ours to move: only "
        "our own trucks can be replanned."))
    ws.cell(row=r, column=1).font = Font(italic=True, size=9)
    r += 2
    for c, h in enumerate(["Month", "Section", "Peak loaded-lane trucks (any instant)",
                           "Average loaded-lane trucks (per hour)",
                           "Busiest hour", "One-lane packing %"], start=1):
        ws.cell(row=r, column=c, value=h).font = Font(bold=True)
    r += 1
    from openpyxl.styles import Alignment, PatternFill
    fills = {
        "high": PatternFill("solid", fgColor=_RC_HIGH),
        "hot": PatternFill("solid", fgColor=_RC_HOT),
        "watch": PatternFill("solid", fgColor=_RC_WATCH),
        "open": PatternFill("solid", fgColor=_RC_OPEN),
        "idle": PatternFill("solid", fgColor=_RC_IDLE),
    }
    mid = Alignment(horizontal="center")
    any_row = False
    for card in cards:
        got = _corridor_for_month(card)
        if not got:
            continue
        res, dt = got
        first = True
        for s in (res.get("sections") or []):
            occ = s.get("occupancy") or []
            if not occ:
                continue
            peak = max(occ)
            hr = (int(res.get("start_hour") or 7) + occ.index(peak)) % 24
            pack = _xlsx_pack_cap(s)
            inst = round(s.get("peak_concurrent") or peak, 1)
            vc = (inst / pack) if pack > 0 else 0.0
            tone = _xlsx_crowd_tone(vc)
            ws.cell(row=r, column=1,
                    value=("%s (%d DT)" % (card.get("name") or "", dt)) if first else "")
            ws.cell(row=r, column=2, value=s.get("section"))
            peak_c = ws.cell(row=r, column=3, value=inst)
            peak_c.fill = fills.get(tone) or fills["idle"]
            peak_c.alignment = mid
            ws.cell(row=r, column=4, value=round(sum(occ) / len(occ), 1))
            ws.cell(row=r, column=5, value="%02d:00" % hr)
            pct_c = ws.cell(row=r, column=6, value=round(vc, 3) if pack else None)
            pct_c.number_format = "0%"
            pct_c.fill = fills.get(tone) or fills["idle"]
            first = False
            any_row = True
            r += 1
    if not any_row:
        ws.cell(row=r, column=1, value="No finalised allocation to distribute.")
        return r + 1
    return r + 1


# Presence colours match Plan → C · Road crowding: sitting ÷ ONE loaded
# lane at FOLLOWING_DISTANCE_M. GREEN <40% · YELLOW 40–70% · ORANGE 70–100%
# · RED at/over packing (614 on TF–KR is red against 576, not 1,152).
_RC_OPEN, _RC_WATCH, _RC_HOT, _RC_HIGH, _RC_IDLE = (
    "BBF7D0", "FDE68A", "FDBA74", "FCA5A5", "F8FAFC")
_RC_DAY_H, _RC_NIGHT_H = "E2E8F0", "DBEAFE"


def _xlsx_pack_cap(section):
    """One loaded lane at the live following distance — same as the Plan grid."""
    try:
        cap = float(section.get("cap_trucks_lane") or 0)
    except (TypeError, ValueError):
        cap = 0.0
    if cap > 0:
        return cap
    try:
        km = float(section.get("section_km") or 0)
    except (TypeError, ValueError):
        km = 0.0
    from congestion.speed_limits import FOLLOWING_DISTANCE_M
    if km > 0:
        return km * 1000.0 / FOLLOWING_DISTANCE_M
    try:
        both = float(section.get("cap_trucks_present")
                     or section.get("cap_trucks_bin") or 0)
    except (TypeError, ValueError):
        both = 0.0
    return both / 2.0 if both > 0 else 0.0


def _xlsx_crowd_tone(vc):
    if not vc or vc <= 0 or vc != vc:
        return "idle"
    if vc >= 1:
        return "high"
    if vc >= 0.7:
        return "hot"
    if vc >= 0.4:
        return "watch"
    return "open"


def _xlsx_road_corridor_hourly(ws, r, res, dt, source=None):
    """The Plan-tab hour grid: mean concurrent trucks per section × hour.

    Same engine as /api/plan/shared-flow (whole day, 2 × 12 h, rain 0).
    Colour = sitting ÷ one loaded lane at 50 m — same as Plan (J53: advisory).
    """
    from openpyxl.styles import Alignment, PatternFill, Font
    from congestion.speed_limits import FOLLOWING_DISTANCE_M

    n_hour_cols = 24
    title = "Road crowding by hour"
    if source:
        title = "Road crowding by hour · %s" % source
    r = _xlsx_section(ws, r, title)
    if not res or not res.get("ok"):
        ws.cell(row=r, column=1, value="No finalised allocation to time onto the road.")
        ws.cell(row=r, column=1).font = _xlsx_font(False, 9, _XLSX_MUTED)
        return r + 1

    follow = float((res.get("basis") or {}).get("following_distance_m")
                   or FOLLOWING_DISTANCE_M)
    # WHO IS ON THE ROAD (owner, 2026-08-27: "it's not just the 1,281 trucks
    # running there — the other tenants crowd the road too, show them"). The
    # engine has always ADDED the tenant register to the occupancy numbers
    # (tenants=True), but the header counted only our fleet, so a reader
    # could not tell whether the 1,340 tenant DT were in the colour or not.
    # State the whole road: ours + theirs = what the colour is priced on.
    _ten_dt = 0
    try:
        _ten_dt = int(round(float((res.get("tenants") or {}).get("total_dt") or 0)))
    except Exception:  # noqa: BLE001
        _ten_dt = 0
    if not _ten_dt:
        try:
            from congestion.tenants import TENANTS as _TREG
            _ten_dt = int(sum(t["dt"] for t in _TREG))
        except Exception:  # noqa: BLE001
            _ten_dt = 0
    if _ten_dt and dt:
        ws.cell(row=r, column=1, value=(
            "ON THIS ROAD: %s DT total = %s ours (plan + IWIP transit) + "
            "%s other tenants (POSITION, HUAFEI>RSF, PMA, MHM, HSM, KR>RSF — "
            "not ours to schedule, but they take lane space). Every number "
            "below already includes them."
            % (format(int(dt) + _ten_dt, ","), format(int(dt), ","),
               format(_ten_dt, ","))))
        ws.cell(row=r, column=1).font = _xlsx_font(True, 9, _XLSX_NAVY)
        r += 1
    ws.cell(row=r, column=1, value=(
        "Loaded-lane trucks sitting each hour (empty is the other carriageway). "
        "Colour vs one loaded lane at %.0f m "
        "(TF–KR 576): GREEN <40%% · YELLOW 40–70%% · ORANGE 70–100%% · RED over."
        % follow))
    ws.cell(row=r, column=1).font = _xlsx_font(False, 9, _XLSX_MUTED)
    r += 1

    secs = res.get("sections") or []
    start_h = int(res.get("start_hour") or 7) % 24
    bin_h = float(res.get("bin_hours") or 1) or 1
    n_bins = max((len(s.get("occupancy") or []) for s in secs), default=0)
    if n_bins <= 0:
        ws.cell(row=r, column=1, value="Corridor engine returned no hourly bins.")
        ws.cell(row=r, column=1).font = _xlsx_font(False, 9, _XLSX_MUTED)
        return r + 1
    n_bins = min(n_bins, n_hour_cols)
    hours = [((start_h + int(round(b * bin_h))) % 24) for b in range(n_bins)]
    shift_len = int(round(float(res.get("shift_hours") or 12) / bin_h))

    mid = _xlsx_mid()
    box = _xlsx_sides()[0]
    day_fill = PatternFill("solid", fgColor=_RC_DAY_H)
    night_fill = PatternFill("solid", fgColor=_RC_NIGHT_H)
    fills = {
        "high": PatternFill("solid", fgColor=_RC_HIGH),
        "hot": PatternFill("solid", fgColor=_RC_HOT),
        "watch": PatternFill("solid", fgColor=_RC_WATCH),
        "open": PatternFill("solid", fgColor=_RC_OPEN),
        "idle": PatternFill("solid", fgColor=_RC_IDLE),
    }

    # Band row: Day 07–18 / Night 19–06
    band = ws.cell(row=r, column=1, value="Corridor")
    band.font = _xlsx_font(True, 9, _XLSX_NAVY)
    band.border = box
    day_end = min(shift_len, n_bins)
    if day_end > 0:
        c = ws.cell(row=r, column=2, value="Day shift 07–18")
        c.font = _xlsx_font(True, 8, _XLSX_NAVY)
        c.fill = day_fill
        c.alignment = mid
        c.border = box
        if day_end > 1:
            ws.merge_cells(start_row=r, start_column=2, end_row=r,
                           end_column=1 + day_end)
            for col in range(3, 2 + day_end):
                ws.cell(row=r, column=col).fill = day_fill
                ws.cell(row=r, column=col).border = box
    if n_bins > shift_len:
        c = ws.cell(row=r, column=2 + shift_len, value="Night shift 19–06")
        c.font = _xlsx_font(True, 8, _XLSX_NAVY)
        c.fill = night_fill
        c.alignment = mid
        c.border = box
        night_n = n_bins - shift_len
        if night_n > 1:
            ws.merge_cells(start_row=r, start_column=2 + shift_len, end_row=r,
                           end_column=1 + n_bins)
            for col in range(3 + shift_len, 2 + n_bins):
                ws.cell(row=r, column=col).fill = night_fill
                ws.cell(row=r, column=col).border = box
    r += 1

    heads = ["Corridor"] + ["%02d" % h for h in hours]
    _xlsx_headers(ws, r, heads, start=1, center=True)
    for b, h in enumerate(hours):
        cell = ws.cell(row=r, column=2 + b)
        cell.fill = night_fill if b >= shift_len else day_fill
    r += 1

    for s in secs:
        occ = list(s.get("occupancy") or [])
        cap = _xlsx_pack_cap(s)
        name = s.get("section") or ""
        shared = s.get("shared")
        peak_vc = max(((float(x or 0) / cap) if cap else 0.0) for x in occ) if occ else 0.0
        tone = _xlsx_crowd_tone(peak_vc)
        tag = {"high": "RED · over", "hot": "ORANGE", "watch": "YELLOW",
               "open": "GREEN"}.get(tone, "")
        if tag and peak_vc < 1:
            tag = "%s · %d%% left" % (tag.split(" · ")[0], round(100 * max(0, 1 - peak_vc)))
        lab = name
        if tag:
            lab = "%s  %s" % (name, tag)
        if shared:
            lab += "  (shared)"
        lab_c = ws.cell(row=r, column=1, value=lab)
        lab_c.font = _xlsx_font(True, 9, _XLSX_INK)
        lab_c.border = box
        lab_c.alignment = Alignment(vertical="center")
        lab_c.fill = fills.get(tone) or fills["idle"]
        for b in range(n_bins):
            val = occ[b] if b < len(occ) else 0
            try:
                num = float(val or 0)
            except (TypeError, ValueError):
                num = 0.0
            cell = ws.cell(row=r, column=2 + b)
            cell.border = box
            cell.alignment = mid
            cell.font = _xlsx_font(False, 9)
            if num <= 0:
                cell.value = None
                cell.fill = fills["idle"]
                continue
            cell.value = int(round(num))
            cell.number_format = "0"
            ratio = (num / cap) if cap > 0 else 0.0
            cell.fill = fills.get(_xlsx_crowd_tone(ratio)) or fills["open"]
            if ratio >= 1:
                cell.font = Font(name="Calibri", bold=True, size=9, color="7F1D1D")
        r += 1

    # Name the CONSTANT in the cells, or every plan looks the same. The other
    # tenants (owner register) are ~50-80% of each mainline cell and identical
    # across every plan and scenario — the owner read two different plans'
    # grids as "the same numbers" because the only part a plan can move (our
    # own trucks) is the minority share. Cells stay TOTALS (capacity is
    # consumed by totals); this line says which part is yours to move.
    ten_bits, our_bits = [], []
    for s in secs:
        tp = float(s.get("tenant_trucks_present") or 0)
        op = float(s.get("our_peak_concurrent") or 0)
        if tp > 0.5:
            ten_bits.append("%s +%d" % (s.get("section"), round(tp)))
        if op > 0.5:
            our_bits.append("%s %d" % (s.get("section"), round(op)))
    if ten_bits:
        note = ws.cell(row=r, column=1, value=(
            "Other tenants are a CONSTANT background across every plan and "
            "scenario (owner register, measured road clock): %s trucks. Your "
            "own trucks peak at: %s — plan changes move only these."
            % (" · ".join(ten_bits), " · ".join(our_bits))))
        note.font = _xlsx_font(False, 8, _XLSX_MUTED)
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=1 + n_bins)
        r += 1

    _xlsx_widths(ws, [28] + [5.5] * n_bins, start=1)
    return r + 1


def _xlsx_month_corridor_block(ws, r, alloc=None, card=None, paths=None):
    """Hourly grid for one month sheet. alloc/card preferred; paths as fallback."""
    got = None
    source = None
    if alloc:
        got = _corridor_for_alloc(alloc)
        source = alloc.get("source_date")
    if not got and card:
        got = _corridor_for_month(card)
        source = card.get("alloc_source_date") or (card.get("alloc") or {}).get("source_date")
    if not got and paths:
        got = _corridor_run(_plans_from_alloc_rows(paths))
    if not got:
        return _xlsx_road_corridor_hourly(ws, r, None, 0)
    res, dt = got
    return _xlsx_road_corridor_hourly(ws, r, res, dt)


def _xlsx_append_crowding_sheet(wb, cards, used, prefix="", after_sheet=None):
    """Workbook sheet of Plan-tab hour grids, one block per month.

    Year keeps a peak summary; this sheet is the same 07..06 colour grid as
    Plan → C · Road crowding (one loaded lane at 50 m).
    """
    name = _xlsx_unique_sheet_name((prefix or "") + "Road crowding", used)
    idx = None
    if after_sheet and after_sheet in wb.sheetnames:
        idx = wb.sheetnames.index(after_sheet) + 1
    ws = wb.create_sheet(name, idx)
    _xlsx_sheet_setup(ws)
    r = _xlsx_board_header(
        ws, "Road crowding by hour",
        "Same numbers and colours as Plan → C · Road crowding. "
        "One loaded lane at 50 m. Advisory — never changes plan tonnes.",
        start=1)
    any_grid = False
    for card in cards:
        got = _corridor_for_month(card)
        if not got:
            continue
        res, dt = got
        any_grid = True
        # The title counts EVERY truck on the road, ours plus the tenant
        # register (owner, 2026-08-27). "Dec (1,310 DT)" read as if the road
        # held only our fleet when the colour was already priced on 2,650.
        try:
            from congestion.tenants import TENANTS as _TREG
            _ten = int(sum(t["dt"] for t in _TREG))
        except Exception:  # noqa: BLE001
            _ten = 0
        r = _xlsx_road_corridor_hourly(
            ws, r + 1, res, dt,
            source=("%s (%s DT on the road: %s ours + %s tenants)"
                    % (card.get("name") or card.get("month") or "",
                       format(int(dt) + _ten, ","), format(int(dt), ","),
                       format(_ten, ","))
                    if _ten else
                    "%s (%d DT)" % (card.get("name") or card.get("month") or "", dt)))
        r += 1
    if not any_grid:
        ws.cell(row=r, column=1, value="No finalised allocation to time onto the road.")
        ws.cell(row=r, column=1).font = _xlsx_font(False, 9, _XLSX_MUTED)
    ws.freeze_panes = "B4"
    return name


def _xlsx_plan_sat_payload():
    """Same payload as Congestion → Mainline corridor saturation charts."""
    try:
        from simulator_api import plan_saturation_payload
        d = plan_saturation_payload(max_dt=6000)
    except Exception:  # noqa: BLE001 — export must not die on a missing model
        return None
    if not d or not d.get("ok") or not (d.get("curve") or []):
        return None
    return d


def _xlsx_sat_tpd_ylim(payload):
    """Same y window as Congestion #cong-wmtdt-chart — NOT from 0.

    floor(min(data, measured−10)) … ceil(max(data, measured+5)) so 168 sits
    in the plot and the fall to ~118 is visible. Excel otherwise autoscales
    positive series from 0 and the curve looks like a flat line on the roof.
    """
    import math
    base = payload.get("baseline") or {}
    hist = base.get("wmt_per_dt_day")
    ys = [p.get("wmt_per_dt_day") for p in (payload.get("curve") or [])
          if isinstance(p.get("wmt_per_dt_day"), (int, float))]
    if not ys:
        return 120, 173
    lo, hi = min(ys), max(ys)
    if isinstance(hist, (int, float)):
        lo = min(lo, hist - 10)
        hi = max(hi, hist + 5)
    return int(math.floor(lo)), int(math.ceil(hi))


def _xlsx_sat_line_chart(ws, title, y_title, min_col, max_col, header_row,
                         last_row, anchor, cat_col, colors, y_min=None,
                         y_max=None, y_fmt="#,##0", height=8.5, width=14):
    """Line chart for the corridor saturation sweep. No markers — 60 points."""
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.marker import Marker
    lc = LineChart()
    lc.title = title
    lc.y_axis.title = y_title
    lc.y_axis.numFmt = y_fmt
    if y_min is not None:
        lc.y_axis.scaling.min = y_min
    if y_max is not None:
        lc.y_axis.scaling.max = y_max
    if y_min is not None:
        # Default crosses="autoZero" parks the x-axis at y=0 even when the
        # scale starts at 120, which is the "from zero" look we are killing.
        lc.y_axis.crosses = "min"
    lc.height, lc.width = height, width
    lc.legend.position = "t"
    lc.style = None
    lc.x_axis.title = "corridor fleet DT (BLB excluded)"
    data = Reference(ws, min_col=min_col, max_col=max_col,
                     min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=cat_col, min_row=header_row + 1, max_row=last_row)
    lc.add_data(data, titles_from_data=True)
    lc.set_categories(cats)
    _xlsx_paint_lines(lc, colors)
    for s in lc.series:
        s.marker = Marker(symbol="none")
        s.smooth = False
    ws.add_chart(lc, anchor)
    return lc


def _xlsx_write_sat_sweep(ws, row, col, payload, with_origin=True):
    """Write DT / t/DT / 168-held / t/day / 168×fleet. Returns (header, last)."""
    base = payload.get("baseline") or {}
    hist_tpd = base.get("wmt_per_dt_day")
    curve = list(payload.get("curve") or [])
    if with_origin and curve and curve[0].get("total_dt"):
        curve = [{"total_dt": 0, "wmt_day": 0,
                  "wmt_per_dt_day": curve[0].get("wmt_per_dt_day")} ] + curve
    heads = ["Corridor fleet DT", "Model t/DT", "Jan–Jun measured t/DT",
             "Model t/day", "168 t/DT × fleet"]
    _xlsx_headers(ws, row, heads, start=col, center=True)
    header = row
    rr = row
    for p in curve:
        rr += 1
        dt = p.get("total_dt")
        tpd = p.get("wmt_per_dt_day")
        wmt = p.get("wmt_day")
        _xlsx_num(ws.cell(row=rr, column=col), dt, center=True)
        cell_tpd = ws.cell(row=rr, column=col + 1)
        cell_h = ws.cell(row=rr, column=col + 2)
        if dt == 0:
            # Origin is only for the total-tonnes chart (0, 0). Leave t/DT
            # blank so that series starts at the first real fleet — same as
            # Congestion #cong-wmtdt-chart, which does not plot DT=0.
            cell_tpd.border = _xlsx_sides()[0]
            cell_h.border = _xlsx_sides()[0]
        else:
            _xlsx_rate(cell_tpd, tpd, center=True)
            _xlsx_rate(cell_h, hist_tpd, center=True)
        _xlsx_num(ws.cell(row=rr, column=col + 3), wmt, center=True)
        held_t = (round(hist_tpd * dt) if hist_tpd is not None and dt is not None
                  else None)
        _xlsx_num(ws.cell(row=rr, column=col + 4), held_t, center=True)
    return header, rr


def _xlsx_fill_saturation_block(ws, r, payload, table_col=1, chart_col=7,
                                point_to_last_sheet=False):
    """Both Congestion-tab corridor curves + the numbers they are drawn from.

    Year: table_col=16 (P) so the dashboard stays A–N; charts at A.
    Last sheet: table_col=1 (visible), charts at G beside the table.
    """
    if not payload:
        r = _xlsx_section(
            ws, r, "Corridor saturation — how tonnes are priced",
            "Sweep unavailable (no Jan–Jun mix or congestion model). "
            "The Congestion tab shows the same two charts when the model is live.")
        return r + 1
    from openpyxl.utils import get_column_letter
    base = payload.get("baseline") or {}
    hist_tpd = base.get("wmt_per_dt_day")
    hist_wmt = base.get("avg_wmt_day")
    hist_dt = base.get("avg_fleet_dt")
    n_mix = payload.get("mix_routes")
    w = base.get("window") or []
    win = " → ".join(str(x) for x in w) if isinstance(w, (list, tuple)) else str(w or "")
    last = (payload.get("curve") or [None])[-1] or {}
    held_last = (hist_tpd * last["total_dt"]) if hist_tpd and last.get("total_dt") else None
    r = _xlsx_section(
        ws, r, "Corridor saturation — how tonnes are priced",
        "Same two curves as Congestion → Mainline corridor (km 0–68). "
        "Jan–Jun mix of %s routes, shared-road pricing, loaders at calibrated faces. "
        "Dashed 168 t/DT is the measured average over %s at ~%s DT (109 kt/day). "
        "The model meets that point, then t/DT falls as extra trucks queue at the "
        "same faces and share the same road — so total tonnes bend below 168 × fleet."
        % (n_mix, win or "Jan–Jun",
           int(round(hist_dt)) if hist_dt else "650"))
    header, last_row = _xlsx_write_sat_sweep(ws, r, table_col, payload)
    stacked = chart_col == 1
    a1 = get_column_letter(chart_col) + str(r)
    a2 = get_column_letter(chart_col) + str(r + 18) if stacked \
        else get_column_letter(chart_col + 8) + str(r)
    tpd_lo, tpd_hi = _xlsx_sat_tpd_ylim(payload)
    _xlsx_sat_line_chart(
        ws, "WMT/day per DT vs corridor fleet", "WMT/day per DT",
        table_col + 1, table_col + 2, header, last_row, a1,
        cat_col=table_col, colors=("059669", (_XLSX_TGT, True)),
        y_min=tpd_lo, y_max=tpd_hi, y_fmt="0.0")
    _xlsx_sat_line_chart(
        ws, "Total WMT/day vs corridor fleet", "WMT/day, t",
        table_col + 3, table_col + 4, header, last_row, a2,
        cat_col=table_col, colors=("6366F1", (_XLSX_TGT, True)),
        y_min=0, y_fmt="#,##0")
    note_bits = []
    if hist_tpd and hist_dt and hist_wmt:
        note_bits.append(
            "At ~%s DT the model sits on the measured %s t/DT / %s kt/day."
            % (int(round(hist_dt)), hist_tpd, int(round(hist_wmt / 1000))))
    if last.get("total_dt") and last.get("wmt_day") and held_last:
        note_bits.append(
            "At %s DT the model is %s kt/day vs %s kt if %s t/DT had held."
            % (int(last["total_dt"]), int(round(last["wmt_day"] / 1000)),
               int(round(held_last / 1000)), hist_tpd))
    if point_to_last_sheet:
        note_bits.append(
            "The sweep numbers are on the Saturation sheet at the end of this workbook.")
    cap_row = r + 38 if stacked else last_row + 2
    ws.cell(row=cap_row, column=1, value=" ".join(note_bits)).font = _xlsx_font(
        False, 9, _XLSX_MUTED)
    ws.merge_cells(start_row=cap_row, start_column=1, end_row=cap_row, end_column=8)
    return max(cap_row, last_row) + 2


def _xlsx_append_saturation_sheet(wb, used, prefix=""):
    """Last sheet: both corridor saturation curves + the sweep they are drawn from."""
    name = _xlsx_unique_sheet_name((prefix or "") + _XLSX_SAT_SHEET, used)
    ws = wb.create_sheet(name)  # append = last
    _xlsx_sheet_setup(ws)
    payload = _xlsx_plan_sat_payload()
    r = _xlsx_board_header(
        ws, "Corridor saturation — WMT/day per DT and total WMT/day",
        "Same model as Congestion → Mainline corridor (km 0–68). "
        "This sheet is the numbers; the Year page shows the same two charts.",
        start=1)
    _xlsx_fill_saturation_block(ws, r, payload, table_col=1, chart_col=7)
    _xlsx_widths(ws, [18, 14, 22, 14, 18])
    ws.freeze_panes = "A4"
    return name


def _xlsx_scenario_constraints_block(ws, start_col, scenario_label=""):
    """Right-side panel on the Year sheet (owner, 2026-08-26): every rule the
    scenario runs under, written where the reader of the workbook can see it
    without opening the app. Numbers come from the live constants
    (scenario_api.SAP_ROUTING / LIM_LD_TARGET_T), so this panel cannot drift
    from the engine."""
    import scenario_api as _sa
    from openpyxl.utils import get_column_letter
    col = start_col
    lab = scenario_label or ""
    is_split = "3.0.2" in lab or "3.1.2" in lab or lab.startswith("S4") or lab.startswith("S6")
    is_31 = "3.1" in lab
    is_41 = "4.1" in lab or lab.startswith("S7")
    is_42 = "4.2" in lab or lab.startswith("S8") or lab.startswith("S9")
    is_421 = "4.2.1" in lab or lab.startswith("S9")
    rows = [
        ("H", "Scenario constraints%s" % ((" — " + lab) if lab else "")),
        ("S", "P1 — SAP routing (owner 2026-08-26)"),
        ("T", "Each pit sends ~2,000 t/day SAP to its POS buffer; the REST goes DIRECT to FeNi per the mine plan:"),
    ]
    for pit, rule in sorted(_sa.SAP_ROUTING.items()):
        fx = ", ".join("%s @ %s t/day" % (d, format(int(v), ",")) for d, v in rule["fixed"])
        rows.append(("T", "  %s: buffer %s -> rest DIRECT %s" % (pit, fx, rule["rest"])))
    if is_42:
        rows += [
            ("S", "P2 — LIM-TOS (plant split, client 2026-09-01)"),
            ("T", "Fresh LIM split by RECEIVING plant every month: 2/3 Huafei (3,054,091 t)"),
            ("T", "+ 1/3 BSE (1,527,046 t) = the %s t line. Fills only after P1 is met."
                  % format(_sa.LIM_TOS_SALES_42_T, ",")),
            ("S", "P3 — LIM-LD (yard + POS split)"),
            ("T", "LD line %s t: DIRECT to yard 2,000,000 (Huafei 1,500,000 + BSE 500,000)"
                  % format(_sa.LIM_LD_SALES_42_T, ",")),
            ("T", "+ via POS 6,000,000 (Huafei 4,000,000 + BSE 2,000,000), reclaimed on IWIP (input = output)."),
            ("T", ("POS balance 4.2.1 (owner): Sep-Oct all via-POS LD transits POS 12; from Nov 1 a "
                   "%d%% share stays on POS 12 and %d%% goes to POS 6 - tuned so the year lands 100%%"
                   " with the SAME fleet (no trucks added)."
                   % (round(_sa.POS12_SHARE_NOVDEC_421 * 100),
                      round((1 - _sa.POS12_SHARE_NOVDEC_421) * 100))) if is_421 else
             "POS switch (owner 2026-09-01): Sep-Oct LD transits TF -> POS 12; from Nov 1 LIM is"),
            ("T", "" if is_421 else
             "NOT stocked in POS 12 - both plants' LD transits POS 6. SAP buffers keep POS 12/14."),
            ("T", "capacity above target stays visible as excess, never folded into the credited number."),
        ]
    elif is_41:
        rows += [
            ("S", "P2 — LIM-TOS"),
            ("T", "Fresh LIM direct to HUAFEI/BSE. Fills only after P1 is met. Target %s t"
                  % format(_sa.LIM_TOS_SALES_41_T, ",")),
            ("S", "P3 — LIM-LD"),
            ("T", "ALL LD goes TF -> POS 12 (Huafei meeting 2026-08-31: POS 6 not ready),"),
            ("T", "then reclaims POS 12 -> HUAFEI on IWIP trucks (input = output)."),
            ("T", "Sales target %s t Sep-Dec (stockpile reclaim, 20260828 mine plan);"
                  % format(_sa.LIM_LD_SALES_41_T, ",")),
            ("T", "capacity above target stays visible as excess, never folded into the credited number."),
        ]
    else:
        rows += [
            ("S", "P2 — LIM-TOS"),
            ("T", ("All LIM-TOS to HUAFEI/BSE. Fills only after P1 is met. Target %s"
                   % ("4,640,201 t (the sales table — includes the ~1 Mt addition)"
                      if is_31 else "3,650,201 t (without the 3.1 addition)"))),
            ("S", "P3 — LIM-LD"),
            ("T", "Leftover trucks haul LD (Tofu dump -> HUAFEI). Sales target %s t Sep-Dec"
                  % format(_sa.LIM_LD_TARGET_T if is_31 else _sa.LIM_LD_TARGET_30_T, ",")),
            ("T", ("(3.1: the ~1 Mt addition sits in LIM-TOS)" if is_31 else
                   "(3.0: includes the ~1 Mt transferred from LIM-TOS — total stays 17.0 Mt);")),
            ("T", "capacity above target stays visible as excess, never folded into the credited number."),
        ]
    if is_41 or is_42:
        pass
    elif is_split:
        rows += [
            ("S", "Hauling variant .2 — POS 6 split"),
            ("T", "Half of the leftover LD trucks go TF -> POS 6 instead of HUAFEI/BSE,"),
            ("T", "STARTING OCTOBER (September allocates like the .1 variant)."),
        ]
    else:
        rows += [
            ("S", "Hauling variant .1"),
            ("T", "All leftover LD trucks go to HUAFEI/BSE (no POS 6 split)."),
        ]
    if is_31:
        rows += [
            ("S", "Mining plan 3.1"),
            ("T", "+330,000 t/month BLB LIM October-December on top of the 3.0 plan."),
        ]
    rows += [
        ("S", "Standing rules"),
        ("T", "BLB pit accepts RIM trucks only. POS is transit: inbound tonnes leave on IWIP"),
        ("T", "reclaim sized so input = output. IWIP trucks are not contractor fleet."),
        ("T", ("SAP target %s wmt (20260828 mine plan ROM table: 6,541,121 ROM x 88%% + 627,239 HGS stock)."
               % format(_sa.SAP_SALES_41_T, ",")) if (is_41 or is_42) else
              "SAP target 5,718,686 wmt (sales table). LIM-TOS: sales table 4,640,201 = the 3.1 scenario."),
    ]
    # Section titles merge A:H for chrome, which puts every row's merge
    # across column G. The titles' text lives in column A, so shrinking
    # those merges to end at F costs nothing visually and frees the panel's
    # room (owner: "right side started from column G").
    from openpyxl.worksheet.cell_range import CellRange
    shrink = []
    for m in list(ws.merged_cells.ranges):
        if m.max_col >= col and m.min_col < col:
            shrink.append((m.min_row, m.min_col, m.max_row))
            ws.unmerge_cells(str(m))
        elif m.min_col >= col:
            ws.unmerge_cells(str(m))
    for min_row, min_col, max_row in shrink:
        if col - 1 > min_col:
            ws.merge_cells(start_row=min_row, start_column=min_col,
                           end_row=max_row, end_column=col - 1)
    r = 4
    width_cols = 5
    ws.column_dimensions[get_column_letter(col)].width = 4
    for i in range(1, width_cols + 1):
        ws.column_dimensions[get_column_letter(col + i)].width = 18
    for kind, text in rows:
        if kind == "H":
            c = ws.cell(row=r, column=col, value=text)
            c.font = _xlsx_font(True, 14, _XLSX_NAVY)
            ws.merge_cells(start_row=r, start_column=col,
                           end_row=r, end_column=col + width_cols)
            r += 2
        elif kind == "S":
            c = ws.cell(row=r, column=col, value=text)
            c.font = _xlsx_font(True, 11, _XLSX_INK)
            ws.merge_cells(start_row=r, start_column=col,
                           end_row=r, end_column=col + width_cols)
            r += 1
        else:
            c = ws.cell(row=r, column=col, value=text)
            c.font = _xlsx_font(False, 10, _XLSX_MUTED)
            ws.merge_cells(start_row=r, start_column=col,
                           end_row=r, end_column=col + width_cols)
            r += 1
    return r


# ── Year-level LIM-LD adjustment (owner, 2026-08-28) ────────────────────
# "Together 101.7%, SAP 100%, LIM-TOS 100%, but LIM-LD 103% — reduce LD
# from 103 to 100 and tell me how many DTs we are NOT using. December can
# stay at 110%; see the bigger picture, keep the rules in the months."
#
# THE YEAR is the judged line, not each month: remove only the year's LD
# surplus, taken from the most-over month first, and never push a month
# BELOW its own line. The freed trucks are optional capacity — first
# re-absorbed by moving the SAME SAP tonnage onto a LONGER haul (KM15 ->
# KM0, POS 12 -> KM0 need 1.34x / 1.56x the trucks for the same tonnes),
# and only the remainder is reported as trucks we do not need.
# Contractor walls hold: RIM works BLB/TF, SMA works KR/TF, never swapped.
_LD_ADJ_KM15 = {"TF": {"RIM": 105.4}, "KR": {"SMA": 160.4}}
_LD_ADJ_KM0 = {"TF": {"RIM": 78.4}, "KR": {"SMA": 102.6}}
_LD_ADJ_P12 = {"TF": {"RIM": 113.4}, "KR": {"SMA": 201.9}}
_LD_ADJ_P12K = {"TF": {"RIM": 63.1}, "KR": {"SMA": 120.1}}
_LD_ADJ_PIT = {"RIM": "TF", "SMA": "KR"}


def _ld_year_adjustment(cards):
    """Return the year-level LD adjustment, or None when LD is under its line."""
    Y = _year_alloc_totals(cards)
    if not Y:
        return None
    ld = (Y.get("materials") or {}).get("ld") or {}
    pred = ld.get("pred_after")
    line = ld.get("sales_target") or ld.get("target")
    if not pred or not line:
        return None
    surplus = pred - line
    months = []
    for c in cards:
        m = ((c.get("alloc") or {}).get("materials") or {}).get("ld") or {}
        p = m.get("pred_after_month") or 0
        t = m.get("target_month") or 0
        if not t:
            continue
        months.append({"card": c, "name": c.get("name"), "month": c.get("month"),
                       "pred": p, "tgt": t, "cov": p / t, "head": max(0.0, p - t),
                       "nd": len(_days_in(c.get("month") or ""))or 30})
    out = {"pred": pred, "line": line, "surplus": surplus,
           "cov_before": pred / line, "plan": [], "freed": 0,
           "absorbed": 0, "park": 0, "removed": 0.0}
    if surplus <= 0 or not months:
        out["cov_after"] = out["cov_before"]
        return out
    # FLAT FLEET (owner, 2026-08-29): park to a single RUNNING LEVEL, not a
    # different number each month. Uneven parking (3.1.2 ran 1,216 in Nov then
    # 1,119 in Dec) forces them to buy for the peak and leaves those trucks
    # standing by a month later. So we find the HIGHEST level every month can
    # hold while the YEAR's LIM-LD still lands on its line: lower level removes
    # more tonnage, so the largest feasible level is the answer. Months whose
    # pool is already under the level are untouched.
    def _removable(mm, ndt):
        """Tonnes/month removed by parking ndt trucks off this month's LD."""
        rws = _plan_rows_by_material(mm["card"])
        rem, t = ndt, 0.0
        for x in sorted([z for z in rws if z["mat"] == "LIM-LD"],
                        key=lambda z: z["rate"]):
            if rem <= 0:
                break
            k = min(x["dt"], rem)
            t += k * x["rate"]
            rem -= k
        return t * mm["nd"], rem

    _pools = [int(mm["card"].get("dt") or 0) for mm in months]
    _flat = None
    if _pools:
        for F in range(max(_pools), 0, -1):
            tot, ok = 0.0, True
            for mm in months:
                pool = int(mm["card"].get("dt") or 0)
                t, short = _removable(mm, max(0, pool - F))
                if short > 0:
                    ok = False
                    break
                tot += t
            if ok and tot >= surplus:
                _flat = F
                break
    out["flat_level"] = _flat
    need = surplus
    for m in sorted(months, key=lambda x: -x["cov"]):
        if need <= 0:
            continue
        pool = int(m["card"].get("dt") or 0)
        if _flat is not None:
            take_dt = max(0, pool - _flat)
            if take_dt <= 0:
                continue
            take = _removable(m, take_dt)[0]
        else:
            if m["head"] <= 0:
                continue
            take = min(need, m["head"])
        rows = _plan_rows_by_material(m["card"])
        if not rows:
            continue
        # Take the trucks off the LOWEST-yield LIM-LD rows first: those are the
        # ones doing least where they stand, so parking them costs the least
        # tonnage per truck. No contractor preference is needed any more —
        # nothing is re-routed, so absorption capacity does not matter.
        cuts = []
        if _flat is not None:
            rem_dt = max(0, int(m["card"].get("dt") or 0) - _flat)
            for x in sorted([r for r in rows if r["mat"] == "LIM-LD"],
                            key=lambda z: z["rate"]):
                if rem_dt <= 0:
                    break
                k = min(x["dt"], rem_dt)
                if k <= 0:
                    continue
                cuts.append({"key": x["key"], "con": x["con"], "n": k,
                             "rate": x["rate"], "dt_before": x["dt"]})
                rem_dt -= k
        else:
            rem = take / m["nd"]
            for x in sorted([r for r in rows if r["mat"] == "LIM-LD"],
                            key=lambda z: z["rate"]):
                if rem <= 0:
                    break
                k = min(x["dt"], int(rem // x["rate"]) if x["rate"] else 0)
                if k <= 0:
                    continue
                cuts.append({"key": x["key"], "con": x["con"], "n": k,
                             "rate": x["rate"], "dt_before": x["dt"]})
                rem -= k * x["rate"]
        got = sum(c["n"] * c["rate"] for c in cuts) * m["nd"]
        free = {"RIM": 0, "SMA": 0}
        for c in cuts:
            free[c["con"]] += c["n"]
        # OPTION 1 ONLY (owner, 2026-08-28): the trucks that come off LIM-LD
        # are PARKED. They are not re-routed onto FeNi KM0 or anywhere else —
        # that was option 2 and it is dropped. So every freed truck is parked
        # and no other row of the plan moves.
        moves = []
        parked = dict(free)
        out["plan"].append({
            "name": m["name"], "cov_before": m["cov"],
            "cov_after": (m["pred"] - got) / m["tgt"], "take": got,
            "dt": sum(c["n"] for c in cuts), "cuts": cuts, "moves": moves,
            "absorbed": sum(x["use"] for x in moves),
            "park": sum(parked.values())})
        need -= got
    out["freed"] = sum(p["dt"] for p in out["plan"])
    out["removed"] = sum(p["take"] for p in out["plan"])
    out["absorbed"] = sum(p["absorbed"] for p in out["plan"])
    out["park"] = sum(p["park"] for p in out["plan"])
    out["cov_after"] = (pred - out["removed"]) / line
    return out


def _plan_rows_by_material(card):
    """Allocated own rows of one month as {key, con, mat, dt, t_day, rate}."""
    rows = []
    src = (card.get("alloc") or {}).get("rows") or []
    if not src:
        d = card.get("alloc_source_date")
        if d:
            try:
                with open(os.path.join(_SAVED_DIR, d + ".json"), encoding="utf-8") as fh:
                    src = ((json.load(fh).get("allocation") or {}).get("rows")) or []
            except Exception:  # noqa: BLE001
                src = []
    for r in src:
        if r.get("foreign") or r.get("_tenant"):
            continue
        dt = r.get("dt_after") or 0
        if dt <= 0:
            continue
        mat = str(r.get("material") or "").upper()
        ot = str(r.get("otype") or "").upper()
        mk = "SAP" if mat == "SAP" else ("LIM-LD" if ot == "LD" else "LIM-TOS")
        pd_ = r.get("pred_after") or 0
        rows.append({"key": r.get("key"), "con": r.get("contractor"), "mat": mk,
                     "dt": dt, "t_day": pd_, "rate": pd_ / dt})
    return rows


def _ld_headline_cov(cards, mats):
    """LIM-LD coverage AFTER the year-level adjustment, so the board reads
    100% like SAP and LIM-TOS. Falls back to the raw figure when nothing is
    adjusted (owner, 2026-08-28)."""
    try:
        adj = _ld_year_adjustment(cards)
        if adj and adj.get("freed"):
            return round(100 * adj["cov_after"], 1)
    except Exception:  # noqa: BLE001
        pass
    return (mats.get("ld") or {}).get("cov_pred")


def _ld_year_adjust_payload(cards):
    """JSON-safe view of the year-level LIM-LD adjustment for the dashboard."""
    try:
        adj = _ld_year_adjustment(cards)
    except Exception:  # noqa: BLE001
        return None
    if not adj:
        return None
    return {
        "pred": round(adj["pred"]),
        "line": round(adj["line"]),
        "surplus": round(adj["surplus"]),
        "cov_before": round(100 * adj["cov_before"], 1),
        "cov_after": round(100 * adj["cov_after"], 1),
        "freed": adj["freed"],
        "absorbed": adj["absorbed"],
        "park": adj["park"],
        "removed": round(adj["removed"]),
        "flat_level": adj.get("flat_level"),
        "months": [{"name": p["name"], "dt": p["dt"], "absorbed": p["absorbed"],
                    "park": p["park"], "take": round(p["take"]),
                    "cov_before": round(100 * p["cov_before"], 1),
                    "cov_after": round(100 * p["cov_after"], 1),
                    "moves": [{"con": m["con"], "from": m["from"], "to": m["to"],
                               "use": m["use"], "t": round(m["t"]),
                               "dt_off": round(m["dt_off"]), "dt_on": round(m["dt_on"])}
                              for m in p["moves"]]}
                   for p in adj["plan"] if p["dt"]],
    }


def _xlsx_ld_park_box(ws, start_col, cards, top_row=25):
    """Right-side PARKING box (owner, 2026-08-28: "make a box on the side,
    show which month and the number of DT you parked, so we see clearly how
    we reach our target and this is the DT we have to park").

    Sits under the constraints panel in the same column. It answers one
    question per month: to bring the YEAR's LIM-LD to 100%, how many trucks
    come off LD, how many keep working on a longer haul, and how many we
    simply do not need."""
    from openpyxl.styles import PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    adj = _ld_year_adjustment(cards)
    if not adj:
        return
    box = _xlsx_sides()[0]
    mid = _xlsx_mid()
    col = start_col
    r = top_row
    head = ws.cell(row=r, column=col, value="DT TO PARK — to land the year on 100%")
    head.font = _xlsx_font(True, 12, "FFFFFF")
    head.alignment = mid
    ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 3)
    for cc in range(col, col + 4):
        ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=_XLSX_NAVY)
        ws.cell(row=r, column=cc).border = box
    r += 1
    if adj["surplus"] <= 0:
        c = ws.cell(row=r, column=col,
                    value="LIM-LD year is %.1f%% — under its line. Nothing to park."
                          % (100 * adj["cov_before"]))
        c.font = _xlsx_font(True, 10, "1B7A41")
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=r, start_column=col, end_row=r + 1, end_column=col + 3)
        for cc in range(col, col + 4):
            ws.cell(row=r, column=cc).border = box
        return
    _fl = adj.get("flat_level")
    sub = ws.cell(row=r, column=col, value=(
        ("The fleet runs at ONE LEVEL — %d DT every month — instead of a different "
         "number each month. That is the peak they need to own: no truck is bought "
         "for a single month and then left standing a month later. LIM-LD was %.1f%% "
         "of its year line; parking to this level takes %s t off it and lands the "
         "year on %.1f%%. Nothing is re-routed, so SAP and LIM-TOS are untouched."
         % (_fl, 100 * adj["cov_before"], format(int(round(adj["removed"])), ","),
            100 * adj["cov_after"])) if _fl else
        ("The LIM-LD headline above is ALREADY this adjustment: it was %.1f%% of its "
         "year line, and parking these trucks takes %s t off LIM-LD and lands it on "
         "%.1f%%. Nothing is re-routed, so SAP and LIM-TOS are untouched."
         % (100 * adj["cov_before"], format(int(round(adj["removed"])), ","),
            100 * adj["cov_after"]))))
    sub.font = _xlsx_font(False, 9, _XLSX_MUTED)
    sub.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=col, end_row=r + 2, end_column=col + 3)
    for cc in range(col, col + 4):
        ws.cell(row=r, column=cc).border = box
    r += 3
    for i, h in enumerate(("Month", "PARKED", "RUNNING", "")):
        c = ws.cell(row=r, column=col + i, value=h)
        c.font = _xlsx_font(True, 10, "FFFFFF")
        c.alignment = mid
        c.fill = PatternFill("solid", fgColor=_XLSX_NAVY)
        c.border = box
    r += 1
    by_month = {p["name"]: p for p in adj["plan"]}
    for c0 in cards:
        nm = c0.get("name")
        p = by_month.get(nm)
        _xlsx_text(ws.cell(row=r, column=col), nm, True, center=True)
        if not p or not p["dt"]:
            _pool = int(c0.get("dt") or 0)
            cc = ws.cell(row=r, column=col + 1, value=0)
            cc.font = _xlsx_font(True, 10, "1B7A41")
            cc.alignment = mid
            cc = ws.cell(row=r, column=col + 2, value=_pool or None)
            cc.number_format = "#,##0"
            cc.font = _xlsx_font(True, 10, _XLSX_INK)
            cc.alignment = mid
        else:
            _pool = int(c0.get("dt") or 0)
            _xlsx_num(ws.cell(row=r, column=col + 1), p["park"], True, center=True)
            ws.cell(row=r, column=col + 1).font = _xlsx_font(True, 11, "A52929")
            ws.cell(row=r, column=col + 1).fill = PatternFill("solid", fgColor="FBE9E9")
            _xlsx_num(ws.cell(row=r, column=col + 2), _pool - p["park"], True, center=True)
            ws.cell(row=r, column=col + 2).font = _xlsx_font(True, 11, "1B7A41")
            ws.cell(row=r, column=col + 2).fill = PatternFill("solid", fgColor="D9F2E2")
            _xlsx_text(ws.cell(row=r, column=col + 3), "", center=True)
        for cc in range(col, col + 4):
            ws.cell(row=r, column=cc).border = box
        r += 1
    _xlsx_text(ws.cell(row=r, column=col), "TOTAL", True, _XLSX_NAVY, center=True)
    _xlsx_num(ws.cell(row=r, column=col + 1), adj["park"], True, center=True)
    ws.cell(row=r, column=col + 1).font = _xlsx_font(True, 12, "A52929")
    ws.cell(row=r, column=col + 1).fill = PatternFill("solid", fgColor="FBE9E9")
    if _fl:
        _xlsx_num(ws.cell(row=r, column=col + 2), _fl, True, center=True)
        ws.cell(row=r, column=col + 2).font = _xlsx_font(True, 12, "1B7A41")
        ws.cell(row=r, column=col + 2).fill = PatternFill("solid", fgColor="D9F2E2")
    _xlsx_text(ws.cell(row=r, column=col + 3), "", center=True)
    for cc in range(col, col + 4):
        ws.cell(row=r, column=cc).border = box
        _xlsx_total_border(ws.cell(row=r, column=cc))
    r += 1
    note = ws.cell(row=r, column=col, value=(
        ("FLEET TO OWN: %d DT, held every month. Park %d DT in total and the "
         "year's LIM-LD lands on %.1f%%."
         % (_fl, adj["park"], 100 * adj["cov_after"])) if _fl else
        ("PARK %d DT and the year's LIM-LD lands on %.1f%%."
         % (adj["park"], 100 * adj["cov_after"]))))
    note.font = _xlsx_font(True, 10, "1B7A41" if _fl else "A52929")
    note.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 3)
    for cc in range(col, col + 4):
        ws.cell(row=r, column=cc).border = box


def _xlsx_ld_year_adjustment_block(ws, r, cards):
    """Print the year-level LD adjustment on the Year sheet."""
    from openpyxl.styles import PatternFill
    adj = _ld_year_adjustment(cards)
    if not adj:
        return r
    box = _xlsx_sides()[0]
    mid = _xlsx_mid()
    r = _xlsx_section(
        ws, r, "LIM-LD · bringing the YEAR to 100%",
        "The YEAR is the judged line, not each month. Only the year's LD surplus is removed, "
        "taken from the most-over month first, and no month is pushed below its own line. "
        "Freed trucks are kept working by moving the SAME SAP tonnage onto a longer haul "
        "(FeNi KM15 to KM0, POS 12 to KM0); only what no haul can take is capacity we do not need.")
    if adj["surplus"] <= 0:
        c = ws.cell(row=r, column=1,
                    value="Year LIM-LD is %.1f%% of its line — already at or under it. No adjustment needed."
                          % (100 * adj["cov_before"]))
        c.font = _xlsx_font(True, 11, "1B7A41")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        return r + 2
    _xlsx_headers(ws, r, ["", "LIM-LD year", "Year line", "Year %",
                          "DT freed", "Kept working", "NOT NEEDED", "Year % after"],
                  center=True)
    r += 1
    _xlsx_text(ws.cell(row=r, column=1), "Year total", True, center=True)
    _xlsx_num(ws.cell(row=r, column=2), round(adj["pred"]), True, center=True)
    _xlsx_num(ws.cell(row=r, column=3), round(adj["line"]), center=True)
    _xlsx_pct_cell(ws.cell(row=r, column=4), 100 * adj["cov_before"])
    ws.cell(row=r, column=4).font = _xlsx_font(True, 11, "A52929")
    ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor="FBE9E9")
    _xlsx_num(ws.cell(row=r, column=5), adj["freed"], True, center=True)
    _xlsx_num(ws.cell(row=r, column=6), adj["absorbed"], True, center=True)
    ws.cell(row=r, column=6).font = _xlsx_font(True, 11, "1B7A41")
    _xlsx_num(ws.cell(row=r, column=7), adj["park"], True, center=True)
    ws.cell(row=r, column=7).font = _xlsx_font(
        True, 11, "A52929" if adj["park"] else "1B7A41")
    ws.cell(row=r, column=7).fill = PatternFill(
        "solid", fgColor="FBE9E9" if adj["park"] else "D9F2E2")
    _xlsx_paint_cov(ws.cell(row=r, column=8), 100 * adj["cov_after"])
    for col in range(1, 9):
        ws.cell(row=r, column=col).border = box
    r += 2
    _xlsx_headers(ws, r, ["Month", "LIM-LD before", "LIM-LD after", "Tonnes removed",
                          "DT freed", "Kept working", "Not needed", "Where those trucks come from"],
                  center=True)
    r += 1
    for p in adj["plan"]:
        if not p["dt"]:
            continue
        _xlsx_text(ws.cell(row=r, column=1), p["name"], True, center=True)
        _xlsx_pct_cell(ws.cell(row=r, column=2), 100 * p["cov_before"])
        ws.cell(row=r, column=2).font = _xlsx_font(True, 11, "8A6100")
        _xlsx_paint_cov(ws.cell(row=r, column=3), 100 * p["cov_after"])
        _xlsx_num(ws.cell(row=r, column=4), round(p["take"]), center=True)
        _xlsx_num(ws.cell(row=r, column=5), p["dt"], True, center=True)
        _xlsx_num(ws.cell(row=r, column=6), p["absorbed"], True, center=True)
        _xlsx_num(ws.cell(row=r, column=7), p["park"], True, center=True)
        mix = {}
        for c in p["cuts"]:
            mix[c["con"]] = mix.get(c["con"], 0) + c["n"]
        _xlsx_text(ws.cell(row=r, column=8),
                   " + ".join("%s %d DT" % (k, v) for k, v in sorted(mix.items()))
                   + " off LIM-LD", size=9, color=_XLSX_MUTED)
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = box
        r += 1
    moves = [(p, m) for p in adj["plan"] for m in p["moves"]]
    if moves:
        r += 1
        c = ws.cell(row=r, column=1,
                    value="Where the kept trucks work — the SAME SAP tonnage, hauled further")
        c.font = _xlsx_font(True, 11, _XLSX_NAVY)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        _xlsx_headers(ws, r, ["Month", "Contractor", "Moved from", "Moved to",
                              "SAP t/day", "DT off", "DT on", "Extra DT used"], center=True)
        r += 1
        for p, m in moves:
            _xlsx_text(ws.cell(row=r, column=1), p["name"], center=True)
            _xlsx_text(ws.cell(row=r, column=2), m["con"], True, center=True)
            _xlsx_text(ws.cell(row=r, column=3), m["from"], color="A52929", center=True)
            _xlsx_text(ws.cell(row=r, column=4), m["to"], True, "1B7A41", center=True)
            _xlsx_num(ws.cell(row=r, column=5), round(m["t"]), center=True)
            _xlsx_num(ws.cell(row=r, column=6), round(m["dt_off"]), center=True)
            _xlsx_num(ws.cell(row=r, column=7), round(m["dt_on"]), True, center=True)
            _xlsx_num(ws.cell(row=r, column=8), m["use"], True, center=True)
            ws.cell(row=r, column=8).fill = PatternFill("solid", fgColor="D9F2E2")
            for col in range(1, 9):
                ws.cell(row=r, column=col).border = box
            r += 1
    return r + 1


def _plant_consumption(cards):
    """Tonnes delivered to each PLANT per month (owner, 2026-08-31: "a graph
    of the evolution of the total and daily consumption of FeNi KM0, KM15 and
    Huafei for each month").

    Read straight off the allocated rows by destination, so it is the plan's
    own delivery, not a re-derivation. Tenant rows are excluded (not our
    tonnage). BSE is kept as its own key for when Huafei/BSE separate."""
    PLANTS = ("FENI KM0", "FENI KM15", "HUAFEI", "BSE")
    out = []
    for c in cards:
        rows = ((c.get("alloc") or {}).get("rows") or [])
        if not rows:
            d = c.get("alloc_source_date")
            if d:
                try:
                    with open(os.path.join(_SAVED_DIR, d + ".json"), encoding="utf-8") as fh:
                        rows = ((json.load(fh).get("allocation") or {}).get("rows")) or []
                except Exception:  # noqa: BLE001
                    rows = []
        nd = c.get("n_days") or len(_days_in(c.get("month") or "")) or 30
        per = {p: 0.0 for p in PLANTS}
        for r in rows:
            if r.get("_tenant"):
                continue
            key = str(r.get("key") or "")
            if ">" not in key:
                continue
            dest = key.split(">", 1)[1].strip().upper()
            if dest in per:
                per[dest] += (r.get("pred_after") or 0) * nd
        out.append({"name": c.get("name"), "month": c.get("month"), "days": nd,
                    "plants": per, "total": sum(per.values())})
    return out


def _xlsx_plant_consumption_block(ws, r, cards):
    """Plant consumption table + charts: total per month and per day."""
    data = _plant_consumption(cards)
    if not data or not any(d["total"] for d in data):
        return r
    PLANTS = ("FENI KM0", "FENI KM15", "HUAFEI", "BSE")
    live = [p for p in PLANTS if any(d["plants"].get(p) for d in data)]
    box = _xlsx_sides()[0]
    r = _xlsx_section(
        ws, r, "Plant consumption — what each plant receives",
        "Tonnes delivered by this plan, per plant per month, and the same figure "
        "per day. Straight off the allocated rows by destination; tenant fleets "
        "excluded. BSE appears as its own column once Huafei/BSE separate.")
    head_row = r
    _xlsx_headers(ws, r, ["Month"] + live + ["TOTAL"], center=True)
    r += 1
    first = r
    for d in data:
        _xlsx_text(ws.cell(row=r, column=1), d["name"], True, center=True)
        for i, p in enumerate(live):
            _xlsx_num(ws.cell(row=r, column=2 + i), round(d["plants"].get(p) or 0), center=True)
        _xlsx_num(ws.cell(row=r, column=2 + len(live)), round(d["total"]), True, center=True)
        for cc in range(1, 3 + len(live)):
            ws.cell(row=r, column=cc).border = box
        r += 1
    last = r - 1
    _xlsx_text(ws.cell(row=r, column=1), "TOTAL", True, _XLSX_NAVY, center=True)
    for i, p in enumerate(live):
        _xlsx_num(ws.cell(row=r, column=2 + i),
                  round(sum(d["plants"].get(p) or 0 for d in data)), True, center=True)
    _xlsx_num(ws.cell(row=r, column=2 + len(live)),
              round(sum(d["total"] for d in data)), True, center=True)
    for cc in range(1, 3 + len(live)):
        _xlsx_total_border(ws.cell(row=r, column=cc))
    r += 2
    # per-day table (same plants, tonnes / day)
    day_head = r
    _xlsx_headers(ws, r, ["Month · t/day"] + live + ["TOTAL"], center=True)
    r += 1
    day_first = r
    for d in data:
        _xlsx_text(ws.cell(row=r, column=1), d["name"], True, center=True)
        for i, p in enumerate(live):
            _xlsx_num(ws.cell(row=r, column=2 + i),
                      round((d["plants"].get(p) or 0) / d["days"]), center=True)
        _xlsx_num(ws.cell(row=r, column=2 + len(live)),
                  round(d["total"] / d["days"]), True, center=True)
        for cc in range(1, 3 + len(live)):
            ws.cell(row=r, column=cc).border = box
        r += 1
    day_last = r - 1
    try:
        _xlsx_line_chart(ws, "Plant consumption · tonnes per month", "t / month",
                         2, 1 + len(live), head_row, last,
                         "A%d" % (r + 1), height=8, width=17)
        _xlsx_line_chart(ws, "Plant consumption · tonnes per day", "t / day",
                         2, 1 + len(live), day_head, day_last,
                         "K%d" % (r + 1), height=8, width=17)
    except Exception:  # noqa: BLE001
        pass
    return r + 18


def _xlsx_fill_year_dashboard(ws, year, cards, title_prefix="", achv=False):
    """Year dashboard sheet: KPIs, five-clock charts, coverage table.
    achv=True adds old / optimized achievable (Plan simulate) next to predicted."""
    Y = _year_alloc_totals(cards)
    n_alloc = sum(1 for c in cards if c.get("has_alloc") or c.get("alloc"))
    head = "Year dashboard · %s" % year
    if title_prefix:
        head = "%s · %s" % (title_prefix, head)
    r = _xlsx_board_header(
        ws, head,
        ("%s month%s with Allocate snapshots. Target vs predicted plan."
         % (n_alloc, "" if n_alloc == 1 else "s"))
        + (" Achievable = /api/simulate after Allocate."
           if achv else ""),
        start=1)
    r = _xlsx_plan_source_block(ws, r, cards)
    if Y:
        mats = Y.get("materials") or {}
        # The board is judged against the constant 17,003,193 haulage sales
        # total (owner 2026-08-27) — the same in both scenario families, so
        # the two books are comparable on one line. The fleet-credited sum
        # stays as "Monthly plan targets" beside it. Audit V2 (2026-08-27):
        # this number was previously nowhere in the workbook.
        # Owner 2026-08-27 final: 17,003,193 everywhere. The internal
        # monthly-sum clock is no longer a requirement and is not shown.
        _sales_tot = Y.get("sales_target") or Y.get("target")
        _sales_cov = (Y.get("cov_sales") if Y.get("sales_target")
                      else Y.get("cov_new_pred"))
        year_kpis = [
            ("Sales target", _sales_tot, _XLSX_TGT, "Year tonnes"),
            ("Predicted plan", Y.get("new_pred"), _XLSX_PRED, "Year tonnes"),
        ]
        if achv:
            year_kpis.append(("Achievable",
                              Y.get("new_achv_raw") or Y.get("new_achv"),
                              _XLSX_ACHV, "Year tonnes"))
        if _sales_cov is not None:
            year_kpis.append(("Sales target coverage", _sales_cov,
                              "059669" if _sales_cov >= 100 else "D97706", "pct"))
        r = _xlsx_kpi_strip(ws, r, year_kpis, start=1)
        pct_kpis = [
            ("Together · % of target", Y.get("cov_new_pred"),
             _xlsx_cov_tone(Y.get("cov_new_pred"))[0] or _XLSX_MUTED, "pct"),
            ("SAP · % of target", (mats.get("sap") or {}).get("cov_pred"),
             _xlsx_cov_tone((mats.get("sap") or {}).get("cov_pred"))[0] or _XLSX_MUTED, "pct"),
            ("LIM-TOS · % of target", (mats.get("tos") or {}).get("cov_pred"),
             _xlsx_cov_tone((mats.get("tos") or {}).get("cov_pred"))[0] or _XLSX_MUTED, "pct"),
            # LIM-LD shows the SAME clock as the rest of this strip: the plan
            # BEFORE the adjustment. Owner, 2026-08-31: "do not mix the
            # scenarios - the table says 101.7% (before removing DT) but LD
            # 99.9% (after removing DT); it should be 103%." The adjusted
            # figure lives in the DT-TO-PARK box, which is labelled as such.
            ("LIM-LD · % of target", (mats.get("ld") or {}).get("cov_pred"),
             _xlsx_cov_tone((mats.get("ld") or {}).get("cov_pred"))[0] or _XLSX_MUTED, "pct"),
        ]
        if achv:
            pct_kpis.append(
                ("Together · achievable %", Y.get("cov_new_achv_raw") or Y.get("cov_new_achv"),
                 _xlsx_cov_tone(Y.get("cov_new_achv_raw") or Y.get("cov_new_achv"))[0]
                 or _XLSX_MUTED, "pct"))
        r = _xlsx_kpi_strip(ws, r, pct_kpis, start=1)

        def pts(getter):
            out = []
            for c in cards:
                a = c.get("alloc") or {}
                out.append({"name": c.get("name"), **getter(a, c)})
            return out

        _mats_sales = Y.get("materials") or {}
        r = _xlsx_five_clock_block(
            ws, r, "Together · year",
            "Month on X · tonnes on Y. Target vs predicted plan. TOTAL target "
            "is the sales line"
            + (" · old / optimized achievable." if achv else "."),
            pts(lambda a, c: {
                "target": a.get("target_month") if a else c.get("target_month"),
                "old_pred": a.get("old_pred_month") if a else c.get("pred_month"),
                "new_pred": a.get("new_pred_month") if a else None,
                "old_achv": _pick_achv(a, True, "month") if a else None,
                "new_achv": _pick_achv(a, False, "month") if a else None,
            }),
            start=1, chart_col="I", achv=achv,
            total_target=Y.get("sales_target"))
        for key_m, title in (("sap", "SAP · year"), ("tos", "LIM-TOS · year"),
                             ("ld", "LIM-LD · year")):
            r = _xlsx_five_clock_block(
                ws, r, title,
                "Same clocks for this material only. TOTAL target is the "
                "sales line"
                + (" — predicted and achievable." if achv else "."),
                pts(lambda a, c, k=key_m: {
                    "target": ((a.get("materials") or {}).get(k) or {}).get("target_month"),
                    "old_pred": ((a.get("materials") or {}).get(k) or {}).get("pred_before_month"),
                    "new_pred": ((a.get("materials") or {}).get(k) or {}).get("pred_after_month"),
                    "old_achv": _pick_mat_achv((a.get("materials") or {}).get(k) or {}, True, "month"),
                    "new_achv": _pick_mat_achv((a.get("materials") or {}).get(k) or {}, False, "month"),
                }),
                start=1, chart_col="I", achv=achv,
                total_target=(_mats_sales.get(key_m) or {}).get("sales_target"))
        if achv:
            cov_heads = ["Month", "Target", "Predicted plan",
                         "Achievable"]
        else:
            cov_heads = ["Month", "Target", "Predicted plan"]
        cov_heads += ["Predicted %", "SAP %", "LIM-TOS %", "LIM-LD %", "Delta"]
        if achv:
            cov_heads.append("Achievable %")
        r = _xlsx_section(
            ws, r, "Coverage table",
            "Predicted plan ÷ target"
            + (" · achievable is /api/simulate (Your plan vs after Allocate), not predicted."
               if achv else "."))
        _xlsx_headers(ws, r, cov_heads, start=1)
        for c in cards:
            r += 1
            a = c.get("alloc") or {}
            _xlsx_text(ws.cell(row=r, column=1), c.get("name"))
            if a:
                _xlsx_num(ws.cell(row=r, column=2), a.get("target_month"))
                if achv:
                    _xlsx_num(ws.cell(row=r, column=3), a.get("new_pred_month"), True)
                    na = _pick_achv(a, False, "month")
                    ac = ws.cell(row=r, column=4)
                    _xlsx_num(ac, na, True)
                    col = 5
                else:
                    _xlsx_num(ws.cell(row=r, column=3), a.get("new_pred_month"), True)
                    col = 4
                _xlsx_paint_cov(ws.cell(row=r, column=col), a.get("cov_new_pred"))
                _xlsx_paint_cov(ws.cell(row=r, column=col + 1),
                                ((a.get("materials") or {}).get("sap") or {}).get("cov_pred"))
                _xlsx_paint_cov(ws.cell(row=r, column=col + 2),
                                ((a.get("materials") or {}).get("tos") or {}).get("cov_pred"))
                _xlsx_paint_cov(ws.cell(row=r, column=col + 3),
                                ((a.get("materials") or {}).get("ld") or {}).get("cov_pred"))
                # Always derive from the CURRENT target/predicted on the card,
                # so the sales rescale can never leave a stale or blank delta.
                _tg = a.get("target_month")
                _pv = a.get("new_pred_month")
                _dv = ((_pv - _tg) if (_tg is not None and _pv is not None)
                       else a.get("delta_new_pred_month"))
                if _dv is None and a.get("left_new_pred_month") is not None:
                    _dv = -a["left_new_pred_month"]
                _dc = ws.cell(row=r, column=col + 4)
                _xlsx_num(_dc, _dv)
                if isinstance(_dv, (int, float)) and _dv:
                    _dc.number_format = "+#,##0;-#,##0"
                    _dc.font = _xlsx_font(True, 11, "1B7A41" if _dv > 0 else "A52929")
                if achv:
                    _xlsx_paint_cov(
                        ws.cell(row=r, column=col + 5),
                        _cov_pct(_pick_achv(a, False, "month"), a.get("target_month")))
            else:
                _xlsx_num(ws.cell(row=r, column=2), c.get("target_month"))
                if achv:
                    _xlsx_num(ws.cell(row=r, column=4), c.get("achv_month"))
                else:
                    _xlsx_num(ws.cell(row=r, column=3), c.get("pred_month"))
        r += 1
        tot_lab = ws.cell(row=r, column=1, value="TOTAL · %s months" % Y.get("n"))
        tot_lab.font = _xlsx_font(True, 11, _XLSX_NAVY)
        _xlsx_total_border(tot_lab)
        if achv:
            tot_row = [
                (2, Y.get("target"), None),
                (3, Y.get("new_pred"), None),
                (4, Y.get("new_achv_raw") or Y.get("new_achv"), None),
            ]
            col = 5
        else:
            # The old-pred column left this table on 2026-08-26; its TOTAL
            # row kept writing it, shifting every % one column right and
            # stranding an unlabelled figure past the last header (found by
            # the 2026-08-27 audit swarm at Year!C127).
            tot_row = [
                (2, Y.get("target"), None),
                (3, Y.get("new_pred"), None),
            ]
            col = 4
        tot_row += [
            (col, None, Y.get("cov_new_pred")),
            (col + 1, None, (mats.get("sap") or {}).get("cov_pred")),
            (col + 2, None, (mats.get("tos") or {}).get("cov_pred")),
            (col + 3, None, (mats.get("ld") or {}).get("cov_pred")),
            (col + 4, Y.get("left_new_pred"), None),
        ]
        if achv:
            tot_row.append(
                (col + 5, None,
                 _cov_pct(Y.get("new_achv_raw") or Y.get("new_achv"), Y.get("target"))))
        for col, val, pctv in tot_row:
            cell = ws.cell(row=r, column=col)
            if pctv is not None:
                _xlsx_paint_cov(cell, pctv)
            else:
                _xlsx_num(cell, val, True)
            _xlsx_total_border(cell)
        r += 2
        note = ws.cell(
            row=r, column=1,
            value="Path table for every month (origin, destination, NB Days, Predicted / month) → Paths sheet.")
        note.font = _xlsx_font(False, 10, _XLSX_MUTED)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    else:
        tot_p = sum(c.get("pred_month") or 0 for c in cards)
        tot_t = sum(c.get("target_month") or 0 for c in cards)
        r = _xlsx_kpi_strip(ws, r, [
            ("Target", tot_t or None, _XLSX_TGT, "Year tonnes"),
            ("Predicted plan", tot_p or None, _XLSX_PRED, "Year tonnes"),
        ], start=1)
        if achv:
            tot_a = sum(c.get("achv_month") or 0 for c in cards) or None
            r = _xlsx_kpi_strip(ws, r, [
                ("Achievable", tot_a, _XLSX_ACHV, "Year tonnes"),
            ], start=1)
        r = _xlsx_five_clock_block(
            ws, r, "Year · monthly tonnes",
            "No Allocate snapshots yet — matrix predicted plan / target"
            + (" · achievable from /api/simulate." if achv else "."),
            # pred_month is the matrix path-model figure — with the old-plan
            # clock gone it IS the "Predicted plan" column here.
            [{"name": c.get("name"), "target": c.get("target_month"),
              "new_pred": c.get("pred_month"),
              "new_achv": c.get("achv_month") if achv else None} for c in cards],
            start=1, chart_col="I", achv=achv)

    r = _xlsx_ld_year_adjustment_block(ws, r + 2, cards)
    r = _xlsx_plant_consumption_block(ws, r + 2, cards)
    r = _xlsx_road_corridor_block(ws, r + 2, cards)
    r = _xlsx_fill_saturation_block(
        ws, r + 1, _xlsx_plan_sat_payload(),
        table_col=16, chart_col=1, point_to_last_sheet=True)

    # Right-side constraints panel (owner, 2026-08-26: "started from
    # column G"). The dashboard's tables live in A-F above the coverage
    # block; the panel occupies G4 downward, clear of the coverage table's
    # G/H tail which begins ~100 rows lower.
    _xlsx_scenario_constraints_block(ws, 7, scenario_label=title_prefix)
    # Parking box under the constraints panel (owner, 2026-08-28).
    _xlsx_ld_park_box(ws, 7, cards, top_row=26)

    _xlsx_widths(ws, [16, 14, 14, 14, 12, 14, 14, 12, 14, 12, 10, 12, 12, 12])
    _xlsx_widths(ws, [18, 14, 22, 14, 18], start=16)
    ws.freeze_panes = "A4"


def _xlsx_append_month_sheets(wb, year, cards, used, prefix="", achv=False):
    """One old-vs-new sheet per month card. Cards may carry alloc_raw for scenarios."""
    # One year-level LIM-LD adjustment, split per month (owner, 2026-08-28).
    try:
        _adj = _ld_year_adjustment(cards) or {}
        _yr = {"cov_before": round(100 * _adj["cov_before"], 1),
               "cov_after": round(100 * _adj["cov_after"], 1),
               "flat_level": _adj.get("flat_level")} if _adj else {}
        # The YEAR result after this parking, per material — the number the
        # owner actually signs off (2026-08-29: "your total yearly target
        # should be reached 100%"). Same arithmetic the Year sheet uses.
        if _adj:
            _by = {p2["name"]: p2 for p2 in (_adj.get("plan") or [])}
            _acc = {"SAP": [0.0, 0.0], "LIM-TOS": [0.0, 0.0], "LIM-LD": [0.0, 0.0]}
            for _c2 in cards:
                _mm = ((_c2.get("alloc") or {}).get("materials") or {})
                _p2 = _by.get(_c2.get("name"))
                for _k, _api in (("SAP", "sap"), ("LIM-TOS", "tos"), ("LIM-LD", "ld")):
                    _t = (_mm.get(_api) or {}).get("target_month") or 0
                    _pr = (_mm.get(_api) or {}).get("pred_after_month") or 0
                    if _k == "LIM-LD" and _p2:
                        _pr -= _p2["take"]
                    _acc[_k][0] += _pr
                    _acc[_k][1] += _t
            _yr["materials"] = {k: {"pred": v[0], "tgt": v[1],
                                    "cov": (100 * v[0] / v[1]) if v[1] else None}
                                for k, v in _acc.items()}
            _tp = sum(v[0] for v in _acc.values())
            _tt = sum(v[1] for v in _acc.values())
            _yr["together"] = {"pred": _tp, "tgt": _tt,
                               "cov": (100 * _tp / _tt) if _tt else None}
        _park_by_month = {}
        for _p in (_adj.get("plan") or []):
            _p = dict(_p)
            _p["year"] = _yr
            _park_by_month[_p["name"]] = _p
        for _c in cards:
            _park_by_month.setdefault(_c.get("name"),
                                      {"dt": 0, "absorbed": 0, "park": 0, "year": _yr})
    except Exception:  # noqa: BLE001
        _park_by_month = {}
    for c in cards:
        month = c["month"]
        st = _load_state(month) or {"month": month}
        st["month"] = month
        if c.get("alloc_raw"):
            raw, src = c["alloc_raw"], c.get("alloc_source")
        else:
            raw, src = _resolve_allocation(month, st, day=c.get("_alloc_day"))
        n = c.get("n_days") or len(_days_in(month))
        alloc = _alloc_view(raw, n, src, include_detail=True)
        name = _xlsx_unique_sheet_name(prefix + (c.get("name") or month), used)
        ws = wb.create_sheet(name)
        label = "%s %s" % (c.get("name") or "", year)
        if alloc:
            _xlsx_fill_month_alloc(
                ws, month,
                "%s — %s" % (label.strip(),
                             "plan · predicted · achievable" if achv else "old vs new"),
                alloc, st, achv=achv,
                park=(_park_by_month or {}).get(c.get("name")))
        elif st.get("prediction") or st.get("manual"):
            _xlsx_fill_month(
                ws, st, "%s — production, capacity & SAP" % label.strip(), achv=achv)
        else:
            ws["A1"] = label
            ws["A1"].font = _xlsx_font(True, 16, _XLSX_NAVY)
            ws["A2"] = "No saved Allocate snapshot and no month file yet."
            ws["A2"].font = _xlsx_font(False, 10, _XLSX_MUTED)


def append_year_book_sheets(wb, year, cards, used=None, prefix="", achv=False):
    """Add a year dashboard + month sheets to an existing workbook (multi-scenario export)."""
    if used is None:
        used = set(wb.sheetnames)
    yr = _xlsx_unique_sheet_name(prefix + "Year", used)
    ws = wb.create_sheet(yr)
    _xlsx_fill_year_dashboard(ws, year, cards, title_prefix=prefix.rstrip(" · "), achv=achv)
    crowd = _xlsx_append_crowding_sheet(wb, cards, used, prefix=prefix, after_sheet=yr)
    _xlsx_append_paths_sheet(wb, year, cards, used, prefix=prefix, achv=achv,
                             after_sheet=crowd or yr)
    _xlsx_append_month_sheets(wb, year, cards, used, prefix=prefix, achv=achv)
    _xlsx_append_saturation_sheet(wb, used, prefix=prefix)
    return used


def _xlsx_year_book(year, cards, achv=False, scenario_label=""):
    """Key = year dashboard, all-months path table, then one sheet per month."""
    from openpyxl import Workbook
    # One clock in the workbook too: month targets are the sales line
    # distributed in the plan's shape (idempotent — the five-clock block's
    # own distribution then scales by 1.0).
    _rescale_cards_to_sales(cards, _year_alloc_totals(cards))
    wb = Workbook()
    key = wb.active
    key.title = "Year"
    used = {"Year"}
    _xlsx_fill_year_dashboard(key, year, cards, title_prefix=scenario_label,
                              achv=achv)
    crowd = _xlsx_append_crowding_sheet(wb, cards, used, prefix="", after_sheet="Year")
    _xlsx_append_paths_sheet(wb, year, cards, used, prefix="", achv=achv,
                             after_sheet=crowd or "Year")
    _xlsx_append_month_sheets(wb, year, cards, used, prefix="", achv=achv)
    _xlsx_append_saturation_sheet(wb, used, prefix="")
    return wb


@bp.route("/api/monthly/export")
def api_monthly_export():
    """One xlsx: Key page + month sheet (production, capacity, SAP, daily chart)."""
    month = (request.args.get("month") or "").strip()
    st = _load_state(month)
    if not st:
        return jsonify({"ok": False, "error": "nothing stored for %s" % month}), 404
    if not st.get("month"):
        st["month"] = month
    return _xlsx_send(_xlsx_month_book(month, st),
                      "monthly_plan_%s.xlsx" % month)


@bp.route("/api/monthly/export-year")
def api_monthly_export_year():
    """Year workbook: Key + year chart + one production/capacity/SAP sheet per month.

    ONE WORKBOOK = ONE SCENARIO. `day` omitted resolves to
    DEFAULT_SCENARIO_DAY (S1 = day 01), never "latest save wins"; the resolved
    scenario, its day and the exact source file per month are printed on the
    Year sheet and echoed in the X-Plan-Scenario response headers.
    """
    year = (request.args.get("year") or str(date.today().year)).strip()
    if not re.fullmatch(r"\d{4}", year):
        return jsonify({"ok": False, "error": "year=YYYY"}), 400
    day = (request.args.get("day") or "").strip()
    day = int(day) if re.fullmatch(r"[0-9]{1,2}", day) and 1 <= int(day) <= 28 else None
    achv = (request.args.get("achv") or "").strip() in ("1", "true", "yes")
    resolved_day = DEFAULT_SCENARIO_DAY if day is None else day
    _yearly, cards = _year_cards(year, day=day)
    if not cards:
        return jsonify({
            "ok": False,
            "error": "no day-%02d (%s) plans stored for %s — load a matrix and "
                     "build the year first, or ask for another day"
                     % (resolved_day, _scenario_label_for_day(resolved_day), year),
        }), 404
    name = "monthly_plan_%s%s%s.xlsx" % (
        year, "" if not day else "_day%02d" % day, "_achievable" if achv else "")
    rv = _xlsx_send(_xlsx_year_book(
        year, cards, achv=achv,
        scenario_label=_scenario_label_for_day(resolved_day)), name)
    rv.headers["X-Plan-Scenario"] = _scenario_label_for_day(resolved_day)
    rv.headers["X-Plan-Scenario-Day"] = "%02d" % resolved_day
    rv.headers["X-Plan-Sources"] = ",".join(
        "%s=%s" % (c["month"], c.get("alloc_source_date") or "-") for c in cards)
    return rv



# ── Yearly plan matrix (owner 2026-08-13) ────────────────────────────────────
# "i will not paste for whole month, but paste for one day plan, this is my
#  whole plan for this year ... its one day for each month and we have to
#  calculate for the whole month."
#
# The matrix is the mine-planning export: rows = ORIGIN/MATERIAL/ACTIVITY/
# ORIGIN_TYPE/DESTINATION/CONTRACTOR with merged-cell carry-down, a Values
# column alternating WMT/day and NB_DT_, and one column per month. Subtotal
# rows ("LIM WMT/day", "BLB WMT/day", "Total WMT/day") are skipped because
# their Values cell is not exactly WMT/day / NB_DT_. Parser verified against
# the owner's own totals: Aug 83,988/581 vs stated 83,989/579 (their subtotal
# rounding), Dec 208,408/1,281 vs 208,407/1,280.
#
# One matrix load fills BOTH sides of every month it covers:
#   manual side     = the matrix's own WMT/day × days in month (verbatim)
#   prediction side = the SAME fleet (NB_DT_) run through plan_simulator's
#                     measured engine (achievable t/shift × 2 shifts × days)

_MONTH_NAMES = {"january": 1, "february": 2, "march": 3, "april": 4, "may": 5,
                "june": 6, "july": 7, "august": 8, "september": 9,
                "october": 10, "november": 11, "december": 12}

_YEARLY_PATH = os.path.join(_MONTH_DIR, "yearly_matrix.json")

# Matrix labels -> the engine's route vocabulary (route_lookup.csv).
_ORIGIN_MAP = {"TOFU": "TF"}


def _canon_dest(d):
    d = re.sub(r"\s+", " ", (d or "").strip().upper())
    d = d.replace("KM 0", "KM0").replace("KM 15", "KM15").replace("KM 10", "KM10")
    return d


def _num(s):
    s = re.sub(r"[^\d.]", "", str(s or ""))
    return float(s) if s else None


def _parse_yearly_matrix(rows):
    hdr = None
    for i, r in enumerate(rows):
        low = [str(c or "").strip().lower() for c in r]
        if "values" in low and any(c in _MONTH_NAMES for c in low):
            hdr = i
            break
    if hdr is None:
        return None, "no header row found (need a 'Values' column plus month columns like August, September...)"
    h = rows[hdr]
    low = [str(c or "").strip().lower() for c in h]
    vcol = low.index("values")
    mcols = {j: _MONTH_NAMES[c] for j, c in enumerate(low) if c in _MONTH_NAMES}
    if not mcols:
        return None, "no month columns found"
    carry = {"origin": "", "material": "", "activity": "", "otype": "", "dest": "", "contractor": ""}
    entries, pending = [], None
    for r in rows[hdr + 1:]:
        cells = [str(c or "").strip() for c in r] + [""] * 12
        v = cells[vcol]
        if v == "WMT/day":
            for k, idx in (("origin", 0), ("material", 1), ("activity", 2),
                           ("otype", 3), ("dest", 4), ("contractor", 5)):
                if cells[idx]:
                    carry[k] = cells[idx]
            pending = {"origin": carry["origin"], "material": carry["material"],
                       "activity": carry["activity"], "otype": carry["otype"],
                       "dest": carry["dest"], "contractor": carry["contractor"],
                       "wmt": {str(m): _num(cells[j]) for j, m in mcols.items()},
                       "dt": {}}
            entries.append(pending)
        elif v == "NB_DT_" and pending is not None:
            pending["dt"] = {str(m): _num(cells[j]) for j, m in mcols.items()}
            pending = None
    entries = [e for e in entries if any(e["wmt"].values()) or any(e["dt"].values())]
    if not entries:
        return None, "found the header but no WMT/day / NB_DT_ row pairs under it"
    return {"entries": entries, "months": sorted({int(m) for e in entries for m, v in e["wmt"].items() if v})}, None


@bp.route("/api/monthly/yearly", methods=["GET", "POST"])
def api_monthly_yearly():
    if request.method == "GET":
        if not os.path.isfile(_YEARLY_PATH):
            return jsonify({"ok": True, "exists": False})
        with open(_YEARLY_PATH, encoding="utf-8") as fh:
            return jsonify({"ok": True, "exists": True, "yearly": json.load(fh)})
    rows, src_name = [], "pasted"
    f = request.files.get("file")
    if f and f.filename:
        src_name = f.filename
        data = f.read()
        if not data:
            return jsonify({"ok": False, "error": "the file is empty"}), 400
        if f.filename.lower().endswith((".xlsx", ".xlsm")):
            from openpyxl import load_workbook
            try:
                wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            except Exception:  # noqa: BLE001
                return jsonify({"ok": False, "error": "could not read that file as .xlsx"}), 400
            ws = wb.active
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
        else:
            text = data.decode("utf-8", "replace")
            rows = [l.split("\t" if "\t" in text else ",") for l in text.splitlines()]
    else:
        body = request.get_json(silent=True) or {}
        text = body.get("pasted") or ""
        if not text.strip():
            return jsonify({"ok": False, "error": "paste the matrix or upload the file"}), 400
        rows = [l.split("\t" if "\t" in text else ",") for l in text.splitlines()]
    parsed, err = _parse_yearly_matrix(rows)
    if err:
        return jsonify({"ok": False, "error": err}), 400
    parsed["source"] = src_name
    parsed["loaded_at"] = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    os.makedirs(_MONTH_DIR, exist_ok=True)
    with open(_YEARLY_PATH, "w", encoding="utf-8") as fh:
        json.dump(parsed, fh, indent=1)
    return jsonify({"ok": True, "months": parsed["months"], "routes": len(parsed["entries"])})


@bp.route("/api/monthly/targets")
def api_monthly_targets():
    """Year-matrix P1/P2/P3 rows for one month — Plan page stamps chips from this."""
    month = (request.args.get("month") or "").strip()
    date = (request.args.get("date") or "").strip()
    if not month and len(date) >= 7:
        month = date[:7]
    if not _month_path(month):
        return jsonify({"ok": False, "error": "supply month=YYYY-MM"}), 400
    yearly = _load_yearly()
    if not yearly:
        return jsonify({"ok": True, "month": month, "rows": [], "exists": False})
    mnum = str(int(month[5:7]))
    rows = []
    for e in yearly.get("entries") or []:
        wmt = (e.get("wmt") or {}).get(mnum) or 0
        dt = (e.get("dt") or {}).get(mnum) or 0
        if wmt <= 0 and dt <= 0:
            continue
        src = _ORIGIN_MAP.get((e.get("origin") or "").upper(), (e.get("origin") or "").upper())
        dst = _canon_dest(e.get("dest"))
        mat = (e.get("material") or "").upper()
        otype = (e.get("otype") or "").upper()
        prio = 1 if mat == "SAP" else (2 if mat == "LIM" and otype == "TOS" else 3)
        rows.append({
            "src": src, "dst": dst, "contractor": (e.get("contractor") or "").upper(),
            "mat": mat, "otype": otype, "prio": prio,
            "target": round(wmt) if wmt else 0, "dt": round(dt) if dt else 0,
        })
    return jsonify({"ok": True, "month": month, "exists": True, "rows": rows})


_MONTH_LABELS = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
                 7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}


def _load_yearly():
    if not os.path.isfile(_YEARLY_PATH):
        return None
    with open(_YEARLY_PATH, encoding="utf-8") as fh:
        return json.load(fh)


def _routes_for_month(yearly, mnum):
    """Matrix rows for one calendar month, merged to engine routes."""
    mnum = str(int(mnum))
    routes = {}
    for e in yearly.get("entries") or []:
        if not ((e.get("wmt") or {}).get(mnum) or 0) and not ((e.get("dt") or {}).get(mnum) or 0):
            continue
        src = _ORIGIN_MAP.get((e.get("origin") or "").upper(), (e.get("origin") or "").upper())
        dst = _canon_dest(e.get("dest"))
        key = (src, dst, (e.get("contractor") or "").upper())
        rec = routes.setdefault(key, {"src": src, "dst": dst,
                                      "contractor": (e.get("contractor") or "").upper(),
                                      "dt": 0.0, "wmt_day": 0.0, "materials": set()})
        rec["dt"] += (e.get("dt") or {}).get(mnum) or 0
        rec["wmt_day"] += (e.get("wmt") or {}).get(mnum) or 0
        if e.get("material"):
            rec["materials"].add(e["material"])
    return [r for r in routes.values() if r["dt"] > 0]


def _fetch_sim_json(path):
    """Same payloads the Plan tab fetches. Prefer live app, then local server, then fixture."""
    from flask import current_app, has_app_context
    if has_app_context():
        try:
            rv = current_app.test_client().get(path)
            if rv.status_code == 200:
                data = rv.get_json()
                if data:
                    return data
        except Exception:
            pass
    try:
        import urllib.request
        with urllib.request.urlopen("http://127.0.0.1:5055" + path, timeout=30) as resp:
            return json.load(resp)
    except Exception:
        pass
    fx = os.path.join(_ROOT, "fixtures", path.rsplit("/", 1)[-1] + ".json")
    if os.path.isfile(fx):
        with open(fx, encoding="utf-8") as fh:
            return json.load(fh)
    return {}


def _path_model_context():
    pr = _fetch_sim_json("/api/simulator/path-response") or {}
    cap = _fetch_sim_json("/api/simulator/capability") or {}
    paths = pr.get("paths") or {}
    fleet = (cap.get("kpi") or {}).get("tripsPerDT")
    contr_by = {}
    for c in cap.get("contractorProd") or []:
        name = (c.get("contractor") or "").upper()
        if name:
            contr_by[name] = c
    return paths, fleet, contr_by


_CONG_TRIPS_CACHE = {}


def _congestion_trips_per_dt(route, contractor, n_comb):
    """Trips/DT/day from the ONE contractor-pricing owner: the calibrated
    congestion model, asked with the contractor.

    Until 2026-08-23 this module priced trips/DT as `dayRate × cf`, where cf
    was the contractor's fleet-wide Trips/DT ÷ the site-wide Trips/DT, clamped
    [0.5,1.5]. plan.js deleted exactly that factor the same day, because it is
    fleet-size-confounded (RIM ran ~140 trucks/day against SMA's ~58) and
    INVERTS on the TF corridor against matched-day history. Leaving it here
    would have made monthly_api the third owner of one concept — the shape
    that produced the 0.85 availability override (J55) and the capacity card
    (J71).

    congestion.predictor.predict() applies the contractor's own
    overhead_per_trip_min where calibration found a matched-day baseline, and
    answers POOLED where it did not — the exact transform
    trips_c = 1440 / (1440/trips_pooled − ovh_pooled + ovh_c) that plan.js and
    /api/congestion_model both use. Returns None for an uncalibrated route so
    the caller falls back to the measured ticket history, pooled.

    Basis note: priced on this route's OWN combined fleet, exactly like
    /api/congestion_model without `others`. Cross-route segment traffic is a
    known, disclosed gap here (the Plan tab supplies it via `others`); it is
    not silently approximated.
    """
    try:
        from congestion.config import route_params
        from congestion.predictor import predict
    except ImportError:
        return None
    n = max(1.0, float(n_comb or 0))
    key = (route, (contractor or "").upper(), round(n, 1))
    if key in _CONG_TRIPS_CACHE:
        return _CONG_TRIPS_CACHE[key]
    out = None
    try:
        if route_params(route).get("calibrated"):
            rec = predict(route, n, None, contractor=(contractor or None) and
                          str(contractor).strip().upper())
            v = rec.get("trips_per_DT_per_day")
            if v and float(v) > 0:
                out = float(v)
    except (ValueError, ArithmeticError, KeyError, TypeError, OSError):
        out = None
    if len(_CONG_TRIPS_CACHE) > 20000:
        _CONG_TRIPS_CACHE.clear()
    _CONG_TRIPS_CACHE[key] = out
    return out


def _path_row_wmt(src, dst, contractor, dt, n_comb, paths, fleet, contr_by):
    """One contractor-path through the Plan Step 1 path model (day grain, rain=0)."""
    k = "%s>%s" % (src, dst)
    m = paths.get(k) or {}
    dt = float(dt or 0)
    n_comb = float(n_comb or dt) or dt
    c = contr_by.get((contractor or "").upper()) or {}
    pay = c.get("tf") or m.get("tf")
    day_rate = m.get("dayRate") or m.get("avgTr")
    cap_trips = m.get("dayTripsCap")
    # Contractor pricing has exactly ONE owner (see _congestion_trips_per_dt).
    cong_tr = _congestion_trips_per_dt(k, contractor, n_comb)
    if not pay or dt <= 0 or (not day_rate and not cong_tr):
        return {"wmt": None, "trips": None,
                "flag": "no measured day history — using cycle × payload",
                "pay": pay, "cap_trips": cap_trips, "avg_dt": m.get("avgDt"),
                "contractor_basis": None}
    flag = None
    if cong_tr:
        # Calibrated route: the congestion model already carries saturation on
        # the combined fleet, so the legacy demonstrated-ceiling divide and the
        # 30% floor are NOT applied on top — same guard plan.js uses
        # (`if(!_hybridInfo && capEff>0 …)`). Two saturation models on one
        # number is the defect this repo has paid for three times.
        tr = cong_tr
        basis = "congestion model (contractor=%s)" % ((contractor or "-").upper())
        cap_trips = None
    else:
        tr = float(day_rate)  # POOLED — no contractor factor exists for this route
        basis = "measured day history, pooled (route not calibrated)"
        linear = tr * n_comb
        if cap_trips and cap_trips > 0 and linear > cap_trips:
            tr *= cap_trips / linear
            flag = "at demonstrated ceiling (%d trips/day)" % round(cap_trips)
        floor = 0.3 * float(day_rate)
        if tr < floor:
            tr = floor
    dt_max = m.get("dtMaxDayAll") or m.get("dtMaxAll") or m.get("dtMax")
    if dt_max and n_comb > dt_max:
        flag = "beyond measured fleet (max ever %d DT)" % round(dt_max)
    trips = dt * tr
    return {"wmt": trips * float(pay), "trips": trips, "flag": flag,
            "pay": pay, "cap_trips": cap_trips, "avg_dt": m.get("avgDt"),
            "trips_per_dt": tr, "contractor_basis": basis}


def _plan_predict_for_routes(route_list):
    """Plan tab Step 1 WMT for one typical day (rain=0, no WB / other-traffic knobs).

    Same ingredients the Plan tab uses: trips/DT from the calibrated congestion
    model priced for THIS contractor (the model's own saturation, no legacy
    ceiling divide on top), or the measured pooled day rate with the
    demonstrated ceiling where the route is not calibrated; tonnes = trips ×
    contractor t/trip.
    """
    paths, fleet, contr_by = _path_model_context()
    combined = {}
    for r in route_list:
        k = "%s>%s" % (r["src"], r["dst"])
        combined[k] = combined.get(k, 0.0) + float(r["dt"] or 0)
    rows = []
    pred_day = 0.0
    for r in route_list:
        k = "%s>%s" % (r["src"], r["dst"])
        dt = float(r["dt"] or 0)
        pr = _path_row_wmt(r["src"], r["dst"], r.get("contractor"), dt,
                           combined[k] or dt, paths, fleet, contr_by)
        if pr.get("wmt") is not None:
            pred_day += pr["wmt"]
        rows.append(pr)
    return pred_day, rows


def _required_dt_day(src, dst, contractor, target_day, others_dt, paths, fleet, contr_by):
    """Invert the path model for a day target. Same solver as planDtForWmt (rain=0).

    `cap_trips` (and therefore the "target above path ceiling" pre-check) only
    exists on routes the congestion model has no calibration for; on calibrated
    routes the model's own saturation is the ceiling and the solver walks it."""
    if not target_day or target_day <= 0:
        return None, None
    seed = _path_row_wmt(src, dst, contractor, 1, others_dt + 1, paths, fleet, contr_by)
    pay = seed.get("pay")
    cap = seed.get("cap_trips")
    if pay:
        max_wmt = (float(cap) * float(pay)) if cap else None
        if max_wmt is not None and target_day > max_wmt * 0.999:
            return None, "target above path ceiling (%.0f t/day max)" % max_wmt
    dt = max(1.0, float(seed.get("avg_dt") or 30))
    last_wmt = None
    for _ in range(60):
        row = _path_row_wmt(src, dst, contractor, dt, others_dt + dt,
                            paths, fleet, contr_by)
        wmt = row.get("wmt")
        if not wmt or wmt <= 0:
            return None, row.get("flag") or "could not size"
        last_wmt = wmt
        nxt = dt * target_day / wmt
        if not math.isfinite(nxt) or nxt > 1e6:
            return None, "could not size"
        if abs(nxt - dt) < 0.01:
            dt = nxt
            break
        dt = dt + 0.6 * (nxt - dt)
    if last_wmt is not None and last_wmt < target_day * 0.995:
        return None, "target above path ceiling"
    return max(1, int(math.ceil(dt))), None


def _sap_entries_for_month(yearly, mnum):
    """Un-merged P1 SAP rows (kept for callers that still ask SAP-only)."""
    return [r for r in _priority_entries_for_month(yearly, mnum) if r.get("prio") == 1]


def _priority_entries_for_month(yearly, mnum):
    """Un-merged P1 SAP and P2 LIM-TOS rows. TOS vs LD is not collapsed."""
    mnum = str(int(mnum))
    rows = {}
    for e in (yearly or {}).get("entries") or []:
        mat = (e.get("material") or "").strip().upper()
        otype = (e.get("otype") or "").strip().upper()
        if mat == "SAP":
            prio = 1
        elif mat == "LIM" and otype == "TOS":
            prio = 2
        else:
            continue
        if not ((e.get("wmt") or {}).get(mnum) or 0) and not ((e.get("dt") or {}).get(mnum) or 0):
            continue
        src = _ORIGIN_MAP.get((e.get("origin") or "").upper(), (e.get("origin") or "").upper())
        dst = _canon_dest(e.get("dest"))
        contr = (e.get("contractor") or "").upper()
        key = (src, dst, contr, mat, otype)
        rec = rows.setdefault(key, {"src": src, "dst": dst, "contractor": contr,
                                    "mat": mat, "otype": otype, "prio": prio,
                                    "dt": 0.0, "wmt_day": 0.0})
        rec["dt"] += (e.get("dt") or {}).get(mnum) or 0
        rec["wmt_day"] += (e.get("wmt") or {}).get(mnum) or 0
    return [r for r in rows.values() if r["wmt_day"] > 0 or r["dt"] > 0]


def _sap_table_rows(st, paths, sim_rows):
    """Priority board for Excel: P1 SAP + P2 LIM-TOS. Achievable is this row's share of the route."""
    month = (st or {}).get("month")
    yearly = _load_yearly()
    sap_src = []
    if yearly and month:
        sap_src = _priority_entries_for_month(yearly, month[5:7])
    alloc_by = {}
    for row in ((st or {}).get("allocation") or {}).get("rows") or []:
        key = (row.get("route"), (row.get("contractor") or "").upper(),
               (row.get("mat") or "").upper(), (row.get("otype") or "").upper())
        alloc_by[key] = row
    if not sap_src:
        for p in paths or []:
            mat = (p.get("material") or "").upper()
            if "SAP" not in mat and "LIM" not in mat:
                continue
            key = p.get("key") or ""
            if ">" not in key:
                continue
            src, dst = key.split(">", 1)
            sap_src.append({"src": src, "dst": dst,
                            "contractor": (p.get("contractor") or "").upper(),
                            "mat": "SAP" if "SAP" in mat else "LIM",
                            "otype": "TOS", "prio": 1 if "SAP" in mat else 2,
                            "dt": float(p.get("dt") or 0),
                            "wmt_day": float(p.get("manual_wmt_day") or 0)})
    if not sap_src:
        return []
    ctx = _path_model_context()
    path_models, fleet, contr_by = ctx
    path_dt = {}
    route_achv = {}
    for i, p in enumerate(paths or []):
        k = p.get("key")
        path_dt[k] = path_dt.get(k, 0.0) + float(p.get("dt") or 0)
        if i < len(sim_rows or []) and sim_rows[i].get("achievable_production_t") is not None:
            route_achv[k] = route_achv.get(k, 0.0) + float(sim_rows[i]["achievable_production_t"]) * 2
        elif p.get("achv_wmt_day") is not None and k not in route_achv:
            route_achv[k] = float(p["achv_wmt_day"])
    out = []
    for s in sap_src:
        route = "%s>%s" % (s["src"], s["dst"])
        arow = alloc_by.get((route, s["contractor"], s.get("mat", ""), s.get("otype", "")))
        alloc = float((arow or {}).get("alloc_dt") or s["dt"] or 0)
        target = float(s["wmt_day"] or 0)
        n_comb = path_dt.get(route) or alloc
        others = max(0.0, n_comb - alloc)
        pred_row = _path_row_wmt(s["src"], s["dst"], s["contractor"], alloc,
                                 n_comb or alloc, path_models, fleet, contr_by)
        pred = pred_row.get("wmt")
        req, why = _required_dt_day(s["src"], s["dst"], s["contractor"], target,
                                    others, path_models, fleet, contr_by)
        route_a = route_achv.get(route)
        tot = path_dt.get(route) or alloc
        achv = (route_a * (alloc / tot)) if (route_a is not None and tot > 0) else None
        if req is None:
            status = why or "target above path ceiling"
        elif pred is not None and pred >= target * 0.995:
            status = "on target"
        else:
            extra = max(0, req - int(round(alloc)))
            status = "add %s DT" % format(extra, ",")
        out.append({
            "path": route.replace(">", " → "),
            "contractor": s["contractor"],
            "mat": s.get("mat") or "",
            "otype": s.get("otype") or "",
            "prio": s.get("prio") or 1,
            "target": round(target) if target else None,
            "pred": round(pred) if pred is not None else None,
            "achv": round(achv) if achv is not None else None,
            "alloc_dt": int(round(alloc)),
            "req_dt": req,
            "status": status,
        })
    return out


def _three_for_month(month, yearly):
    """Target (matrix) · Prediction (path model) · Achievable (simulate).

    Same fleet every day of the month; day = 2 × 12 h shifts. The three
    numbers are never averaged.
      • target      = matrix WMT/day (verbatim)
      • prediction  = measured trips/DT × DT, hard-capped at demonstrated
                      day trips (Plan tab path model — the bet on tomorrow)
      • achievable  = /api/simulate achievable_production_t × 2
                      (effective cycle, then loader / dump point capacity)
    """
    mnum = str(int(month[5:7]))
    route_list = _routes_for_month(yearly, mnum)
    if not route_list:
        return None, "the matrix has no rows for month %s" % month
    plans = [{"route": "%s>%s" % (r["src"], r["dst"]), "source": r["src"],
              "destination": r["dst"], "n_trucks": int(round(r["dt"])),
              "contractor": r["contractor"]}
             for r in route_list]
    import plan_simulator
    sim = plan_simulator.simulate({"plans": plans})
    if sim.get("error"):
        return None, sim["error"]
    sim_rows = sim.get("results") or []
    pred_sum, pred_rows = _plan_predict_for_routes(route_list)
    target_day = sum(r["wmt_day"] for r in route_list)
    pred_day = achv_day = 0.0
    paths, extrapolated, conditions = [], [], []
    for i, r in enumerate(route_list):
        rt = "%s>%s" % (r["src"], r["dst"])
        sr = sim_rows[i] if i < len(sim_rows) else {}
        planned_shift = float(sr.get("planned_production_t") or 0)
        achv_shift = float(sr.get("achievable_production_t") or 0)
        pr = pred_rows[i] if i < len(pred_rows) else {}
        if pr.get("wmt") is not None:
            row_pred = pr["wmt"]
            flag = pr.get("flag")
        else:
            row_pred = planned_shift * 2
            flag = pr.get("flag") or "no measured day history — using cycle × payload"
        row_achv = achv_shift * 2
        pred_day += row_pred
        achv_day += row_achv
        if flag:
            extrapolated.append("%s (%s %d DT): %s" % (rt, r["contractor"], round(r["dt"]), flag))
        if flag and flag.startswith("at demonstrated ceiling"):
            locked = max(0.0, (planned_shift * 2) - row_pred)
            if locked > 1:
                conditions.append({
                    "route": rt, "contractor": r["contractor"],
                    "locked_wmt_day": round(locked),
                    "condition": flag,
                })
        paths.append({"key": rt, "contractor": r["contractor"],
                      "dt": int(round(r["dt"])),
                      "material": "+".join(sorted(r["materials"])) or None,
                      "manual_wmt_day": round(r["wmt_day"]),
                      "pred_wmt_day": round(row_pred),
                      "achv_wmt_day": round(row_achv),
                      "envelope_flag": flag,
                      "cycle_basis": (sr.get("assumptions") or {}).get("cycle_time") or sr.get("cycle_source")})
    days = _days_in(month)
    src = "yearly matrix (%s)" % (yearly.get("source") or "pasted")
    st = _load_state(month) or {"month": month}
    st["prediction"] = {
        "source_date": src,
        "per_shift_wmt": round(pred_day / 2),
        "per_day_wmt": round(pred_day),
        "per_day_achv_wmt": round(achv_day),
        "dt": int(round(sum(r["dt"] for r in route_list))),
        "rain_mm": None,
        "paths": paths,
        "days": [{"date": d, "wmt": round(pred_day),
                  "wmt_achv": round(achv_day), "wmt_upside": round(achv_day)}
                 for d in days],
        "built_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "note": ("Same matrix fleet every day. Prediction is the Plan tab path "
                 "model (measured trips/DT × contractor factor × fleet, capped "
                 "at demonstrated day trips). Target is the matrix WMT/day. "
                 "Day = 2 × 12 h shifts."),
        "extrapolated": extrapolated,
        "per_day_upside_wmt": round(achv_day),
        "upside_conditions": sorted(conditions, key=lambda c: -c["locked_wmt_day"]),
    }
    st["manual"] = {
        "source": src,
        "days": [{"date": d, "wmt": round(target_day)} for d in days],
        "loaded_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
    }
    _save_state(month, st)
    n = len(days)
    return st, {
        "month": month, "n_days": n, "routes": len(plans),
        "target_day": round(target_day), "pred_day": round(pred_day),
        "achv_day": round(achv_day),
        "target_month": round(target_day * n),
        "pred_month": round(pred_day * n),
        "achv_month": round(achv_day * n),
        "dt": int(round(sum(r["dt"] for r in route_list))),
        "extrapolated": extrapolated,
        "warnings": (sim.get("summary") or {}).get("capacity_warnings") or [],
    }


@bp.route("/api/monthly/build-from-yearly", methods=["POST"])
def api_monthly_build_from_yearly():
    """One month: target + prediction + achievable from the loaded yearly matrix."""
    body = request.get_json(silent=True) or {}
    month = (body.get("month") or "").strip()
    if not _month_path(month):
        return jsonify({"ok": False, "error": "supply month=YYYY-MM"}), 400
    yearly = _load_yearly()
    if not yearly:
        return jsonify({"ok": False, "error": "no yearly matrix loaded yet"}), 404
    st, meta = _three_for_month(month, yearly)
    if st is None:
        return jsonify({"ok": False, "error": meta}), 400
    return jsonify({"ok": True, "month": month, "state": st,
                    "manual_day": meta["target_day"], "pred_day": meta["pred_day"],
                    "achv_day": meta["achv_day"], "routes": meta["routes"],
                    "extrapolated": meta["extrapolated"], "warnings": meta["warnings"]})


@bp.route("/api/monthly/build-year", methods=["POST"])
def api_monthly_build_year():
    """Paste once → every month in the matrix, same daily plan × days."""
    body = request.get_json(silent=True) or {}
    year = str(body.get("year") or date.today().year).strip()
    if not re.fullmatch(r"\d{4}", year):
        return jsonify({"ok": False, "error": "supply year=YYYY"}), 400
    yearly = _load_yearly()
    if not yearly:
        return jsonify({"ok": False, "error": "no yearly matrix loaded yet"}), 404
    months = yearly.get("months") or []
    if not months:
        return jsonify({"ok": False, "error": "the matrix has no month columns"}), 400
    cards, errors = [], []
    for mnum in months:
        month = "%s-%02d" % (year, int(mnum))
        st, meta = _three_for_month(month, yearly)
        if st is None:
            errors.append({"month": month, "error": meta})
            continue
        cards.append(meta)
    return jsonify({"ok": True, "year": year, "cards": cards, "errors": errors})


_ALLOC_MATS = (("sap", "SAP"), ("tos", "LIM-TOS"), ("ld", "LIM-LD"))


def _finite(x):
    try:
        v = float(x)
    except (TypeError, ValueError):
        return None
    return v if math.isfinite(v) else None


def _hz_to_day(alloc):
    """Saved allocation numbers are already per selected horizon.

    Plan tab day grain multiplies by 2 (two 12 h shifts), so those figures
    are already t/day. Shift grain is one shift — scale ×2 to a day.
    """
    hz = (alloc or {}).get("horizon") or "day"
    if hz in ("shift", "12h", "night"):
        return 2
    return 1


def _find_saved_allocation(month, day=None):
    """Frozen Plan-tab allocation in data/saved_plans for YYYY-MM.

    day=None keeps the old rule (latest date wins). day=N restricts to that
    day of month - the scenario convention (owner, 2026-08-19): the 1st holds
    S1, the 2nd S2, the 3rd S3, so one month stores one plan per scenario."""
    if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", month or ""):
        return None, None
    if not os.path.isdir(_SAVED_DIR):
        return None, None
    best, best_date = None, None
    prefix = month + "-" if day is None else "%s-%02d" % (month, int(day))
    try:
        names = os.listdir(_SAVED_DIR)
    except OSError:
        return None, None
    for fn in names:
        if not (fn.startswith(prefix) and fn.endswith(".json")):
            continue
        path = os.path.join(_SAVED_DIR, fn)
        try:
            with open(path, encoding="utf-8") as f:
                plan = json.load(f)
        except (OSError, ValueError):
            continue
        alloc = plan.get("allocation")
        if not isinstance(alloc, dict) or not alloc.get("frozen"):
            continue
        d = fn[:-5]
        if best_date is None or d > best_date:
            best_date, best = d, alloc
    return best, best_date


def _resolve_allocation(month, st=None, day=None):
    """Prefer the month file's copy; else the latest saved daily plan.

    With day=N the month-state copy is SKIPPED: that copy was frozen from
    whichever save last built the month, which may be another scenario's."""
    if st is None:
        st = _load_state(month)
    if day is not None:
        return _find_saved_allocation(month, day)
    pred = (st or {}).get("prediction") or {}
    for blob, src in (
        ((st or {}).get("saved_day_allocation"), pred.get("source_date")),
        (pred.get("allocation"), pred.get("source_date")),
    ):
        if isinstance(blob, dict) and blob.get("frozen"):
            return blob, src
    return _find_saved_allocation(month, day)


def _cov_pct(num, den):
    if num is None or not den:
        return None
    return round(100.0 * float(num) / float(den), 1)


def _plan_achv(pred, tgt):
    """Target-credited prediction for management presentation.

    This is deliberately not called achievable: achievable is reserved for
    raw /api/simulate output throughout the API and exports.
    """
    if pred is None:
        return None
    if tgt:
        return min(float(pred), float(tgt))
    return float(pred)


def _cap_at(val, cap):
    if val is None:
        return None
    if cap is None:
        return val
    return min(float(val), float(cap))


def _path_rates(row):
    """Trips/DT and WMT/DT from a saved allocation row. Old trips inferred from payload if missing."""
    dt_b, dt_a = _finite(row.get("dt_before")), _finite(row.get("dt_after"))
    pr_b, pr_a = _finite(row.get("pred_before")), _finite(row.get("pred_after"))
    tr_a = _finite(row.get("trips"))
    tr_b = _finite(row.get("trips_before"))
    if tr_b is None and tr_a and pr_a:
        pay = pr_a / tr_a
        if pay:
            tr_b = (pr_b / pay) if pr_b is not None else None
    achv_a = _finite(row.get("achv_after"))
    achv_b = _finite(row.get("achv_before"))
    sim_a = _finite(row.get("achv_sim"))
    if sim_a is None:
        sim_a = achv_a
    tgt = _finite(row.get("target"))

    def tpd(tr, dt):
        if tr is None or not dt:
            return None
        return round(tr / dt, 2)

    def wpd(pr, dt):
        if pr is None or not dt:
            return None
        return round(pr / dt, 1)

    def wpt(pr, tr):
        if pr is None or not tr:
            return None
        return round(pr / tr, 2)

    def rnd(v):
        return None if v is None else int(round(v))

    return {
        "trips_before": rnd(tr_b),
        "trips_after": rnd(tr_a),
        "trips_per_dt_before": tpd(tr_b, dt_b),
        "trips_per_dt_after": tpd(tr_a, dt_a),
        "wmt_per_dt_before": wpd(pr_b, dt_b),
        "wmt_per_dt_after": wpd(pr_a, dt_a),
        "wmt_per_trip_before": wpt(pr_b, tr_b),
        "wmt_per_trip_after": wpt(pr_a, tr_a),
        "achv_before_capped": rnd(_plan_achv(pr_b, tgt)),
        "achv_after_capped": rnd(_plan_achv(pr_a, tgt)),
        "achv_after_raw": rnd(sim_a),
        "over_cap": rnd(max(0, sim_a - (_plan_achv(pr_a, tgt) or 0))) if sim_a is not None else None,
    }


def _alloc_view(alloc, n_days, source_date=None, include_detail=False):
    """Day clocks × n_days → month totals. Never averages pred and achv."""
    if not isinstance(alloc, dict) or not alloc.get("frozen"):
        return None
    fac = _hz_to_day(alloc)
    n = int(n_days or 0)

    def day(v):
        x = _finite(v)
        return None if x is None else x * fac

    def rnd(v):
        return None if v is None else int(round(v))

    def month(v):
        d = day(v)
        return None if d is None else rnd(d * n)

    old = alloc.get("old") or {}
    new = alloc.get("new") or {}
    goals = alloc.get("goals") or {}
    buckets = alloc.get("buckets") or {}
    tgt_raw = new.get("target")
    if tgt_raw is None:
        tgt_raw = goals.get("total")
    materials = {}
    for key, label in _ALLOC_MATS:
        b = buckets.get(key) or {}
        tgt = day(b.get("target") if b.get("target") is not None else goals.get(key))
        pb, pa = day(b.get("pred_before")), day(b.get("pred_after"))
        ab_raw, aa_stored = day(b.get("achv_before")), day(b.get("achv_after"))
        aa_sim = day(b.get("achv_sim"))
        if aa_sim is None:
            aa_sim = aa_stored
        ab, aa = _plan_achv(pb, tgt), _plan_achv(pa, tgt)
        materials[key] = {
            "label": label,
            "n": int(b.get("n") or 0),
            "dt_before": b.get("dt_before"),
            "dt_after": b.get("dt_after"),
            "target_day": rnd(tgt),
            "pred_before_day": rnd(pb),
            "pred_after_day": rnd(pa),
            "credited_pred_before_day": rnd(ab),
            "credited_pred_after_day": rnd(aa),
            "achv_before_day": rnd(ab_raw),
            "achv_after_day": rnd(aa_sim),
            "achv_before_raw_day": rnd(ab_raw),
            "achv_after_raw_day": rnd(aa_sim),
            "over_achv_day": rnd(max(0, aa_sim - aa)) if aa_sim is not None and aa is not None else None,
            "target_month": rnd(tgt * n) if tgt is not None else None,
            "pred_before_month": rnd(pb * n) if pb is not None else None,
            "pred_after_month": rnd(pa * n) if pa is not None else None,
            "credited_pred_before_month": rnd(ab * n) if ab is not None else None,
            "credited_pred_after_month": rnd(aa * n) if aa is not None else None,
            "achv_before_month": rnd(ab_raw * n) if ab_raw is not None else None,
            "achv_after_month": rnd(aa_sim * n) if aa_sim is not None else None,
            "achv_before_raw_month": rnd(ab_raw * n) if ab_raw is not None else None,
            "achv_after_raw_month": rnd(aa_sim * n) if aa_sim is not None else None,
            "cov_pred": _cov_pct(pa, tgt),
            "cov_credited_pred": _cov_pct(aa, tgt),
            "cov_achv": _cov_pct(aa_sim, tgt),
            "cov_achv_raw": _cov_pct(aa_sim, tgt),
            "left_pred_month": rnd(max(0, tgt - pa) * n) if tgt is not None and pa is not None else None,
            # DELTA is signed (owner, 2026-08-31: "replace the column Leaving
            # by Delta to also consider the part over the plan"). Negative =
            # short of target, positive = over it. `left_*` stays for callers
            # that want the one-sided shortfall.
            "delta_pred_month": rnd((pa - tgt) * n) if tgt is not None and pa is not None else None,
            "over_pred_month": rnd(max(0, pa - tgt) * n) if tgt is not None and pa is not None else None,
        }
    op, np_ = day(old.get("pred")), day(new.get("pred"))
    oa_raw, na_stored = day(old.get("achv")), day(new.get("achv"))
    na_sim = day(new.get("achv_sim"))
    if na_sim is None:
        na_sim = na_stored
    tgt = day(tgt_raw)
    rows = alloc.get("rows") or []
    prio_map = {1: "sap", 2: "tos", 3: "ld"}
    if rows:
        tot_aa = tot_ab = 0.0
        sums = {k: {"aa": 0.0, "ab": 0.0} for k, _ in _ALLOC_MATS}
        for row in rows:
            rt = day(row.get("target"))
            pa = day(row.get("pred_after"))
            pb = day(row.get("pred_before"))
            aa = _plan_achv(pa, rt)
            ab = _plan_achv(pb, rt)
            k = prio_map.get(row.get("prio"))
            if aa is not None:
                tot_aa += aa
                if k in sums:
                    sums[k]["aa"] += aa
            if ab is not None:
                tot_ab += ab
                if k in sums:
                    sums[k]["ab"] += ab
        credited_oa, credited_na = tot_ab, tot_aa
        for k, s in sums.items():
            mat = materials.get(k)
            if not mat:
                continue
            mat["credited_pred_after_day"] = rnd(s["aa"])
            mat["credited_pred_before_day"] = rnd(s["ab"])
            mat["credited_pred_after_month"] = rnd(s["aa"] * n) if n else None
            mat["credited_pred_before_month"] = rnd(s["ab"] * n) if n else None
            mat["cov_credited_pred"] = _cov_pct(s["aa"], mat.get("target_day"))
            if mat.get("achv_after_day") is not None and mat.get("credited_pred_after_day") is not None:
                mat["over_achv_day"] = rnd(max(0, mat["achv_after_day"] - mat["credited_pred_after_day"]))
    else:
        credited_oa, credited_na = _plan_achv(op, tgt), _plan_achv(np_, tgt)
    oa, na = oa_raw, na_sim
    fleet = alloc.get("fleet") or {}
    out = {
        "has": True,
        "source_date": source_date,
        "horizon": alloc.get("horizon") or "day",
        "dt_before": fleet.get("before") if fleet.get("before") is not None else old.get("dt"),
        "dt_after": fleet.get("after") if fleet.get("after") is not None else new.get("dt"),
        "old_pred_day": rnd(op), "new_pred_day": rnd(np_),
        "old_achv_day": rnd(oa), "new_achv_day": rnd(na),
        "old_credited_pred_day": rnd(credited_oa),
        "new_credited_pred_day": rnd(credited_na),
        "old_achv_raw_day": rnd(oa_raw), "new_achv_raw_day": rnd(na_sim),
        "over_new_achv_day": rnd(max(0, na_sim - na)) if na_sim is not None and na is not None else None,
        "target_day": rnd(tgt),
        "old_pred_month": rnd(op * n) if op is not None else None,
        "new_pred_month": rnd(np_ * n) if np_ is not None else None,
        "old_achv_month": rnd(oa * n) if oa is not None else None,
        "new_achv_month": rnd(na * n) if na is not None else None,
        "old_credited_pred_month": rnd(credited_oa * n) if credited_oa is not None else None,
        "new_credited_pred_month": rnd(credited_na * n) if credited_na is not None else None,
        "old_achv_raw_month": rnd(oa_raw * n) if oa_raw is not None else None,
        "new_achv_raw_month": rnd(na_sim * n) if na_sim is not None else None,
        "target_month": rnd(tgt * n) if tgt is not None else None,
        "cov_old_pred": _cov_pct(op, tgt),
        "cov_new_pred": _cov_pct(np_, tgt),
        "cov_old_achv": _cov_pct(oa, tgt),
        "cov_new_achv": _cov_pct(na, tgt),
        "cov_old_credited_pred": _cov_pct(credited_oa, tgt),
        "cov_new_credited_pred": _cov_pct(credited_na, tgt),
        "cov_new_achv_raw": _cov_pct(na_sim, tgt),
        "left_new_pred_month": rnd(max(0, tgt - np_) * n) if tgt is not None and np_ is not None else None,
        "delta_new_pred_month": rnd((np_ - tgt) * n) if tgt is not None and np_ is not None else None,
        "over_new_pred_month": rnd(max(0, np_ - tgt) * n) if tgt is not None and np_ is not None else None,
        "over_new_achv_month": rnd(max(0, na_sim - na) * n) if na_sim is not None and na is not None else None,
        "moved_total": alloc.get("moved_total"),
        "materials": materials,
        # LIM-LD capacity (never clipped) vs credited production, and whether
        # the month's targets fit the pool at all. Both travel with the view so
        # the workbook and the API cannot report one without the other.
        "capacity": alloc.get("capacity"),
        "feasible": alloc.get("feasible", True),
    }
    if include_detail:
        out["rows"] = alloc.get("rows") or []
        out["moves"] = alloc.get("moves") or []
        out["notes"] = alloc.get("notes")
        out["fleet"] = fleet
        out["goals"] = goals
    return out


def _year_alloc_totals(cards):
    """Sum allocation months only. Two clocks stay separate."""
    keys = ("old_pred", "new_pred", "old_achv", "new_achv", "target")
    tot = {k: 0 for k in keys}
    tot["n"] = 0
    mats = {mk: {"target": 0, "pred_before": 0, "pred_after": 0,
                 "achv_before": 0, "achv_after": 0}
            for mk, _ in _ALLOC_MATS}
    for c in cards:
        a = c.get("alloc")
        if not a:
            continue
        tot["n"] += 1
        for k in keys:
            v = a.get(k + "_month")
            if v is not None:
                tot[k] += v
        for mk, mv in (a.get("materials") or {}).items():
            b = mats.setdefault(mk, {"target": 0, "pred_before": 0, "pred_after": 0,
                                     "achv_before": 0, "achv_after": 0})
            for fk, sk in (("target", "target_month"),
                           ("pred_before", "pred_before_month"),
                           ("pred_after", "pred_after_month"),
                           ("achv_before", "achv_before_month"),
                           ("achv_after", "achv_after_month")):
                v = mv.get(sk)
                if v is not None:
                    b[fk] += v
    if not tot["n"]:
        return None
    tot["cov_old_pred"] = _cov_pct(tot["old_pred"], tot["target"])
    tot["cov_new_pred"] = _cov_pct(tot["new_pred"], tot["target"])
    tot["cov_old_achv"] = _cov_pct(tot["old_achv"], tot["target"])
    tot["cov_new_achv"] = _cov_pct(tot["new_achv"], tot["target"])
    tot["left_new_pred"] = max(0, tot["target"] - tot["new_pred"])
    tot["over_new_pred"] = max(0, tot["new_pred"] - tot["target"])
    tot["old_achv_raw"] = 0
    tot["new_achv_raw"] = 0
    tot["over_new_achv"] = 0
    for c in cards:
        a = c.get("alloc")
        if not a:
            continue
        if a.get("old_achv_raw_month") is not None:
            tot["old_achv_raw"] += a["old_achv_raw_month"]
        if a.get("new_achv_raw_month") is not None:
            tot["new_achv_raw"] += a["new_achv_raw_month"]
        if a.get("over_new_achv_month") is not None:
            tot["over_new_achv"] += a["over_new_achv_month"]
    tot["cov_old_achv_raw"] = _cov_pct(tot["old_achv_raw"], tot["target"])
    tot["cov_new_achv_raw"] = _cov_pct(tot["new_achv_raw"], tot["target"])
    labels = dict(_ALLOC_MATS)
    for mk, b in mats.items():
        b["label"] = labels.get(mk, mk)
        b["cov_pred"] = _cov_pct(b["pred_after"], b["target"])
        b["cov_achv"] = _cov_pct(b["achv_after"], b["target"])
        b["left"] = max(0, b["target"] - b["pred_after"])
        b["over"] = max(0, b["pred_after"] - b["target"])
    # The LD year card's headline judgement is against the SALES target
    # (planning team 2026-08-26: 6,644,306 t), not the sum of the saves'
    # internal monthly LD targets. Those monthly targets are the waterfall's
    # fleet-credited amounts — 5,290,224 across the day-03 saves — and the
    # owner read that as "the monthly page shows a different LD target"
    # (2026-08-26, sales-table check). Both clocks stay visible: `target`
    # keeps the monthly sum (labelled by the UI), `sales_target` carries the
    # quarter's sales line and the coverage the owner judges by.
    if "ld" in mats:
        import scenario_api as _sa
        b = mats["ld"]
        _b_keep = b
        # Scenario-aware LD line (owner 2026-08-27): the total haulage
        # target is constant at 17.0 Mt, so 3.0 (days 03/04) carries the
        # ~1 Mt its LIM-TOS did not take. The day is read from the cards'
        # source dates.
        _day = None
        for c in cards:
            src = c.get("alloc_source_date") or (c.get("alloc") or {}).get("source_date")
            if src and len(str(src)) == 10:
                _day = int(str(src)[8:10])
                break
        _sid = {3: "S3", 4: "S4", 5: "S5", 6: "S6", 7: "S7", 8: "S8", 9: "S9"}.get(_day, "")
        b["sales_target"] = _sa.ld_target_for_scenario(_sid)
        b["cov_sales"] = _cov_pct(b["pred_after"], b["sales_target"])
        b["left_sales"] = max(0, b["sales_target"] - b["pred_after"])
        b["over_sales"] = max(0, b["pred_after"] - b["sales_target"])
        # The TOGETHER line gets the same treatment (owner 2026-08-27: "our
        # total should remain 17 million"): SAP + LIM-TOS + LIM-LD sales
        # lines sum to the SAME constant in both scenario families, because
        # 3.0's smaller TOS is exactly offset by its larger LD.
        tot["sales_target"] = _sa.total_target_for_scenario(_sid)
        tot["cov_sales"] = _cov_pct(tot.get("new_pred"), tot["sales_target"])
        tot["left_sales"] = max(0, tot["sales_target"] - (tot.get("new_pred") or 0))
        tot["over_sales"] = max(0, (tot.get("new_pred") or 0) - tot["sales_target"])
        # Owner 2026-08-27 (final ruling): "17 million everywhere — the
        # monthly plan target no longer exists as a requirement." EVERY
        # material's headline number and coverage is its SALES line now:
        # SAP 5,718,686 · LIM-TOS by scenario · LIM-LD by scenario · the
        # internal fleet-credited sums stay in the payload as *_plan keys
        # for anyone who needs the old clock, but nothing headlines them.
        if "sap" in mats:
            bs = mats["sap"]
            bs["sales_target"] = _sa.sap_target_for_scenario(_sid)
            bs["cov_sales"] = _cov_pct(bs.get("pred_after"), bs["sales_target"])
            bs["left_sales"] = max(0, bs["sales_target"] - (bs.get("pred_after") or 0))
            bs["over_sales"] = max(0, (bs.get("pred_after") or 0) - bs["sales_target"])
        if "tos" in mats:
            bt = mats["tos"]
            bt["sales_target"] = _sa.tos_target_for_scenario(_sid)
            bt["cov_sales"] = _cov_pct(bt.get("pred_after"), bt["sales_target"])
            bt["left_sales"] = max(0, bt["sales_target"] - (bt.get("pred_after") or 0))
            bt["over_sales"] = max(0, (bt.get("pred_after") or 0) - bt["sales_target"])
        for _k in ("sap", "tos", "ld"):
            _m = mats.get(_k)
            if _m and _m.get("sales_target"):
                _m["plan_target"] = _m.get("target")
                _m["target"] = _m["sales_target"]
                _m["cov_pred"] = _m["cov_sales"]
                _m["left"] = _m["left_sales"]
                _m["over"] = _m["over_sales"]
        tot["plan_target"] = tot.get("target")
        tot["target"] = tot["sales_target"]
        tot["cov_new_pred"] = tot["cov_sales"]
        tot["left_new_pred"] = tot["left_sales"]
        tot["over_new_pred"] = tot["over_sales"]
    tot["materials"] = mats
    return tot


# The day-of-month saved-plan convention (owner, 2026-08-19): day 01 holds
# Scenario 1, day 03 Scenario 3, day 04 Scenario 4. Day 02 is reserved (S2 was
# deleted from the app on 2026-08-21).
#
# ONE WORKBOOK = ONE SCENARIO. `day=None` used to mean "latest file wins" per
# month, and that silently mixed scenarios inside a single workbook: on
# 2026-08-23 the year book named monthly_plan_2026.xlsx resolved Aug from a
# legacy 2026-08-13 daily and Sep/Oct/Nov/Dec from the day-04 (S4) saves, so
# its Sep target read 2,410,410 t where the real S1 day-01 plan gives
# 2,840,190 t (-15%) and the Year TOTAL summed ACROSS scenarios with no
# disclosure. The default is now EXPLICIT: S1 = day 01.
DEFAULT_SCENARIO_DAY = 1
# Planning-team names 2026-08-26: 03=3.0.1, 04=3.0.2, 05=3.1.1, 06=3.1.2.
_SCENARIO_FOR_DAY = {1: "S1", 3: "S3 (3.0.1)", 4: "S4 (3.0.2)",
                     5: "S5 (3.1.1)", 6: "S6 (3.1.2)", 7: "S7 (4.1)",
                     8: "S8 (4.2)", 9: "S9 (4.2.1)"}


def _scenario_label_for_day(day):
    """'S1' / 'S3' / 'S4' for the day-of-month convention, else 'day-NN plans'."""
    return _SCENARIO_FOR_DAY.get(int(day), "day-%02d plans" % int(day))


def _year_cards(year, day=None):
    """Month cards for the year board and the year Excel download.

    day=N picks that day-of-month's saved plan (the scenario convention).
    day=None resolves to DEFAULT_SCENARIO_DAY (=1, Scenario 1) rather than
    letting the latest file in each month win, so one workbook can only ever
    contain one scenario.

    A month with NO save on the requested day is dropped entirely - no target,
    no old predicted plan, and never another day's plan substituted (owner,
    2026-08-19: August belongs to Scenario 1 only). The dropped months are
    reported on every card as `_missing_months` so the Year sheet and the API
    can say so instead of leaving a silent hole. day=3 (S3) and day=4 (S4)
    therefore never include August."""
    resolved_day = DEFAULT_SCENARIO_DAY if day is None else int(day)
    yearly = _load_yearly()
    mnums = set(int(m) for m in (yearly or {}).get("months") or [])
    if os.path.isdir(_MONTH_DIR):
        for f in os.listdir(_MONTH_DIR):
            if re.fullmatch(r"%s-(0[1-9]|1[0-2])\.json" % year, f):
                mnums.add(int(f[5:7]))
    if os.path.isdir(_SAVED_DIR):
        try:
            for f in os.listdir(_SAVED_DIR):
                if re.fullmatch(r"%s-(0[1-9]|1[0-2])-\d{2}\.json" % year, f):
                    mnums.add(int(f[5:7]))
        except OSError:
            pass
    with_day = set()
    if os.path.isdir(_SAVED_DIR):
        try:
            for f in os.listdir(_SAVED_DIR):
                m = re.fullmatch(r"%s-(0[1-9]|1[0-2])-%02d\.json" % (year, resolved_day), f)
                if m:
                    with_day.add(int(m.group(1)))
        except OSError:
            pass
    asked = set(mnums)
    mnums &= with_day
    # August belongs to Scenario 1 only (owner 2026-08-19). Day 3 = S3,
    # day 4 = S4: start at September. A leftover 08-04 save is a legacy
    # daily plan, not S4 — do not treat it as a scenario month.
    if resolved_day != 1:
        mnums.discard(8)
        asked.discard(8)
    missing = sorted(asked - mnums)
    cards = []
    for mnum in sorted(mnums):
        month = "%s-%02d" % (year, mnum)
        st = _load_state(month)
        n = len(_days_in(month))
        p = (st or {}).get("prediction") or {}
        man = (st or {}).get("manual") or {}
        pred_day = p.get("per_day_wmt")
        achv_day = _achv_day_from_pred(p)
        if achv_day is None and p.get("paths"):
            achv_day = _fill_achv_from_plan(p)
            if achv_day is not None:
                _save_state(month, st)
        tgt_days = man.get("days") or []
        tgt_day = tgt_days[0].get("wmt") if tgt_days else None
        alloc, src = _resolve_allocation(month, st, day=resolved_day)
        view = _alloc_view(alloc, n, src, include_detail=False)
        card = {
            "month": month, "name": _MONTH_LABELS[mnum], "n_days": n,
            "dt": p.get("dt"),
            "pred_day": pred_day, "achv_day": achv_day, "target_day": tgt_day,
            "pred_month": round(pred_day * n) if pred_day is not None else None,
            "achv_month": round(achv_day * n) if achv_day is not None else None,
            "target_month": round(tgt_day * n) if tgt_day is not None else None,
            "built": bool(p and man),
            "has_alloc": bool(view),
            # Disclosure travels WITH the card, so the Year sheet, the month
            # sheets and the API cannot disagree about which plan was read.
            "_alloc_day": resolved_day,
            "_scenario": _scenario_label_for_day(resolved_day),
            "_missing_months": [_MONTH_LABELS.get(m, str(m)) for m in missing],
            "_source_note": (
                "%s — day-%02d saved plans only. Every month on this sheet comes "
                "from data/saved_plans/%s-MM-%02d.json; no other scenario's plan "
                "appears anywhere in this workbook."
                % (_scenario_label_for_day(resolved_day), resolved_day,
                   year, resolved_day)),
            "alloc_source_date": src,
        }
        if view:
            card["alloc"] = view
            if card.get("dt") is None:
                card["dt"] = view.get("dt_after")
        cards.append(card)
    return yearly, cards


def _rescale_cards_to_sales(cards, Y):
    """One clock on every surface (owner 2026-08-27): each month's TARGET is
    its share of the SALES total (plan shape × sales/plan-sum), per material
    and together, with coverages recomputed. A month can then never show
    green while the year shows short — the same months, the same predicted,
    one judgement line. Mutates the cards in place; the original plan
    amounts stay in *_plan keys."""
    if not Y or not Y.get("sales_target"):
        return cards
    mats_y = Y.get("materials") or {}
    plan_sum = Y.get("plan_target") or 0
    fam = {"all": (plan_sum, Y["sales_target"])}
    for k in ("sap", "tos", "ld"):
        m = mats_y.get(k) or {}
        if m.get("sales_target") and m.get("plan_target"):
            fam[k] = (m["plan_target"], m["sales_target"])
    for c in cards:
        a = c.get("alloc")
        if not a:
            continue
        ps, ss = fam["all"]
        if ps and a.get("target_month") is not None:
            a["target_plan_month"] = a["target_month"]
            a["target_month"] = int(round(a["target_month"] * ss / ps))
            a["cov_new_pred"] = _cov_pct(a.get("new_pred_month"), a["target_month"])
            a["left_new_pred_month"] = max(0, a["target_month"] - (a.get("new_pred_month") or 0))
            a["delta_new_pred_month"] = (a.get("new_pred_month") or 0) - a["target_month"]
            a["over_new_pred_month"] = max(0, (a.get("new_pred_month") or 0) - a["target_month"])
            if a.get("target_day") is not None:
                a["target_day"] = int(round(a["target_day"] * ss / ps))
        for k in ("sap", "tos", "ld"):
            m = (a.get("materials") or {}).get(k)
            if not m or k not in fam:
                continue
            ps, ss = fam[k]
            if ps and m.get("target_month") is not None:
                m["target_plan_month"] = m["target_month"]
                m["target_month"] = int(round(m["target_month"] * ss / ps))
                m["cov_pred"] = _cov_pct(m.get("pred_after_month"), m["target_month"])
                if m.get("target_day") is not None:
                    m["target_day"] = int(round(m["target_day"] * ss / ps))
        if c.get("target_month") is not None and fam["all"][0]:
            c["target_plan_month"] = c["target_month"]
            c["target_month"] = int(round(c["target_month"] * fam["all"][1] / fam["all"][0]))
    return cards


@bp.route("/api/monthly/year-board")
def api_monthly_year_board():
    """Cards + year totals for the loaded matrix (and any stored months)."""
    year = (request.args.get("year") or str(date.today().year)).strip()
    if not re.fullmatch(r"\d{4}", year):
        return jsonify({"ok": False, "error": "year=YYYY"}), 400
    day = (request.args.get("day") or "").strip()
    day = int(day) if re.fullmatch(r"[0-9]{1,2}", day) and 1 <= int(day) <= 28 else None
    yearly, cards = _year_cards(year, day=day)
    resolved_day = DEFAULT_SCENARIO_DAY if day is None else day
    _Y = _year_alloc_totals(cards)
    _rescale_cards_to_sales(cards, _Y)
    return jsonify({
        "ok": True, "year": year, "day": day,
        # Which plan this board is actually showing. `day` echoes what was
        # asked; `resolved_day`/`scenario` say what was read, and `sources`
        # names the exact saved file per month — one scenario, disclosed.
        "resolved_day": resolved_day,
        "scenario": _scenario_label_for_day(resolved_day),
        "default_day_applied": day is None,
        "sources": {c["month"]: c.get("alloc_source_date") for c in cards},
        "missing_months": (cards[0].get("_missing_months") if cards else []),
        "has_matrix": yearly is not None,
        "source": (yearly or {}).get("source"),
        "routes": len((yearly or {}).get("entries") or []),
        "matrix_months": (yearly or {}).get("months") or [],
        "cards": cards,
        "alloc_year": _Y,
        # Year-level LIM-LD adjustment: how many DT come off LD to land the
        # YEAR on 100%, how many keep working on longer SAP hauls, and how
        # many must park (owner, 2026-08-28). Same maths the Excel books use.
        "ld_adjust": _ld_year_adjust_payload(cards),
    })


@bp.route("/api/monthly/record-attempt", methods=["POST"])
def api_monthly_record_attempt():
    """First-principles 'what if we really run it': no history ceiling."""
    body = request.get_json(silent=True) or {}
    month = (body.get("month") or "").strip()
    if not _month_path(month):
        return jsonify({"ok": False, "error": "supply month=YYYY-MM"}), 400
    if not os.path.isfile(_YEARLY_PATH):
        return jsonify({"ok": False, "error": "no yearly matrix loaded yet"}), 404
    with open(_YEARLY_PATH, encoding="utf-8") as fh:
        yearly = json.load(fh)
    mnum = str(int(month[5:7]))
    active = [e for e in yearly["entries"]
              if (e["wmt"].get(mnum) or 0) > 0 or (e["dt"].get(mnum) or 0) > 0]
    if not active:
        return jsonify({"ok": False, "error": "the matrix has no rows for month %s" % month}), 400
    routes = {}
    for e in active:
        src = _ORIGIN_MAP.get(e["origin"].upper(), e["origin"].upper())
        dst = _canon_dest(e["dest"])
        key = (src, dst, e["contractor"].upper())
        rec = routes.setdefault(key, {"src": src, "dst": dst,
                                      "contractor": e["contractor"].upper(),
                                      "dt": 0.0, "wmt_day": 0.0})
        rec["dt"] += e["dt"].get(mnum) or 0
        rec["wmt_day"] += e["wmt"].get(mnum) or 0
    route_list = [r for r in routes.values() if r["dt"] > 0]
    # Cycle + payload from measured route history.
    import csv as _csv
    cyc, pay = {}, {}
    try:
        with open(os.path.join(_ROOT, "data", "route_lookup.csv"), encoding="utf-8") as fh:
            for r in _csv.DictReader(fh):
                cyc[r["route"]] = float(r["mean_cycle_min"] or 0) or None
                pay[r["route"]] = float(r["median_payload_t"] or 0) or None
    except OSError:
        pass
    for r in route_list:
        r["payload"] = pay.get("%s>%s" % (r["src"], r["dst"]))
    out = _record_physics(route_list, {k: v for k, v in cyc.items() if v})
    if out is None:
        return jsonify({"ok": False, "error": "point_capacity.csv missing"}), 500
    man_day = sum(r["wmt_day"] for r in route_list)
    out["manual_day"] = round(man_day)
    out["month"] = month
    return jsonify({"ok": True, **out})


# ── "Record attempt" physics (owner 2026-08-13) ──────────────────────────────
# "no matter what was the history and if we want to run this plan, i want to
#  know what will happen ... what if we want to make new record, dont limit
#  your thinking ... think we really put this plan in work what will happen?"
#
# So: NO historical trip ceiling. Every truck is assumed to run the route's
# free-flow cycle all day (demand = DT x 1440 / cycle). The only limits left
# are physical infrastructure, each taken at its RECORD rate - the best hour
# ever measured at that point, sustained for 24 hours straight, which no
# operation has ever done. This is the most generous physically-grounded
# ceiling that exists. What cannot be served queues, and the queue is
# reported as trucks standing idle.

def _record_physics(route_list, cyc_lookup):
    import csv as _csv
    cap = {}
    try:
        with open(os.path.join(_ROOT, "data", "point_capacity.csv"), encoding="utf-8") as fh:
            for r in _csv.DictReader(fh):
                cap[(r["point"].upper(), r["kind"])] = {
                    "cap_hr": float(r["capacity_trips_hr"] or 0),
                    "peak_hr": float(r["peak_trips_hr"] or 0),
                }
    except OSError:
        return None
    # Free-flow demand per route.
    demands = []
    from collections import defaultdict as _dd
    load_d, dump_d = _dd(float), _dd(float)
    for r in route_list:
        cycle = cyc_lookup.get("%s>%s" % (r["src"], r["dst"]))
        est = cycle is None
        if est:
            # unmeasured route: median of routes out of the same source
            same = [v for k, v in cyc_lookup.items() if k.startswith(r["src"] + ">")]
            cycle = sorted(same)[len(same) // 2] if same else 120.0
        dem = r["dt"] * 1440.0 / cycle
        demands.append({"r": r, "cycle": cycle, "demand": dem, "cycle_estimated": est})
        load_d[r["src"]] += dem
        dump_d[r["dst"]] += dem
    # Served fraction at each point, at RECORD pace (best hour ever x 24).
    def frac(point, kind, demand):
        c = cap.get((point.upper(), kind))
        if not c or not c["peak_hr"]:
            return 1.0, None                      # no measured device: assume open
        day_cap = c["peak_hr"] * 24.0
        return (min(1.0, day_cap / demand) if demand > 0 else 1.0), day_cap
    bottlenecks = []
    lfrac, dfrac = {}, {}
    for p, dem in load_d.items():
        f, day_cap = frac(p, "loading", dem)
        lfrac[p] = f
        if f < 0.999 and day_cap:
            bottlenecks.append({"point": p, "kind": "loading", "demand": round(dem),
                                "record_day_cap": round(day_cap), "served_pct": round(f * 100),
                                "queued_trucks_per_hour": round((dem - day_cap) / 24.0, 1)})
    for p, dem in dump_d.items():
        f, day_cap = frac(p, "dumping", dem)
        dfrac[p] = f
        if f < 0.999 and day_cap:
            bottlenecks.append({"point": p, "kind": "dumping", "demand": round(dem),
                                "record_day_cap": round(day_cap), "served_pct": round(f * 100),
                                "queued_trucks_per_hour": round((dem - day_cap) / 24.0, 1)})
    paths, total_wmt, total_served, total_demand = [], 0.0, 0.0, 0.0
    for d in demands:
        r = d["r"]
        f = min(lfrac.get(r["src"], 1.0), dfrac.get(r["dst"], 1.0))
        served = d["demand"] * f
        payload = r.get("payload") or 50.0
        wmt = served * payload
        total_wmt += wmt
        total_served += served
        total_demand += d["demand"]
        assumptions = []
        if d.get("cycle_estimated"):
            assumptions.append("cycle estimated from other %s routes (this one unmeasured)" % r["src"])
        if (r["dst"].upper(), "dumping") not in cap:
            assumptions.append("no capacity data for %s dump - assumed open" % r["dst"])
        if not r.get("payload"):
            assumptions.append("payload assumed 50 t (route unmeasured)")
        paths.append({"key": "%s>%s" % (r["src"], r["dst"]), "contractor": r["contractor"],
                      "dt": int(round(r["dt"])), "cycle_min": round(d["cycle"]),
                      "demand_trips_day": round(d["demand"]), "served_trips_day": round(served),
                      "served_pct": round(f * 100), "wmt_day": round(wmt),
                      "assumptions": assumptions})
    return {"per_day_wmt": round(total_wmt), "demand_trips": round(total_demand),
            "served_trips": round(total_served), "paths": paths,
            "bottlenecks": sorted(bottlenecks, key=lambda b: b["served_pct"]),
            "note": ("NO history ceiling. Demand = every truck running the free-flow "
                     "cycle all day. Limits = each loading/dumping point at its RECORD "
                     "rate (best hour ever measured, held for 24 h straight). Trucks "
                     "beyond a point's record rate queue - they do not vanish.")}


# ── SAP-fixed rebalance advisor (owner 2026-08-13) ───────────────────────────
# "we can't reduce the tonnages of our SAP material ... come up with a plan
#  which will keep our SAP material quantity same. And if you want to add more
#  trucks in this, you can add there. And give me suggestions from where we can
#  reduce the trucks from Lemonite ... we can't switch the contractors."
#
# HARD CONSTRAINTS honoured:
#   1. SAP rows: tonnage NEVER goes down. Trucks may be ADDED to SAP rows that
#      still have measured headroom (below the fleet where the road saturates).
#   2. Contractor stays on its own route - no row changes contractor or route.
#   3. Only LIM rows may give up trucks, and only trucks that are past the
#      route's saturation point (they add ~nothing where they are).
# The advisor therefore only DOCUMENTS waste and headroom - it moves LIM's
# stranded trucks to SAP rows of the SAME contractor where the road still pays.

def _day_stats_from_snapshot():
    """Per-route day model: trimmed-mean rate, demonstrated cap, fleet max."""
    stats = {}
    try:
        import simulator_api as _sa
        rows_snap, _rain = _sa._path_snapshot()
        from collections import defaultdict as _dd
        day_agg = _dd(lambda: [0.0, 0.0])
        env = _dd(float)
        for r in rows_snap:
            k = (r["o"], r["dd"])
            day_agg[(k, r.get("d"))][0] += r.get("dt") or 0
            day_agg[(k, r.get("d"))][1] += r.get("trips") or 0
            env[k] = max(env[k], r.get("dt") or 0)
        by_path = _dd(list)
        for (k, _dte), (dtv, tr) in day_agg.items():
            if dtv > 0 and tr > 0:
                by_path[k].append((dtv, tr))
        for k, pts in by_path.items():
            if len(pts) < 5:
                continue
            rates = sorted(t / d for d, t in pts)
            core = rates[int(len(rates) * .2):max(int(len(rates) * .2) + 1, int(len(rates) * .8))]
            stats["%s>%s" % k] = {
                "rate": sum(core) / len(core),
                "cap": max(t for _d, t in pts),
                "dt_max": max(max(d for d, _t in pts), env[k]),
            }
    except Exception:  # noqa: BLE001
        pass
    return stats


@bp.route("/api/monthly/rebalance", methods=["POST"])
def api_monthly_rebalance():
    """SAP-fixed truck reallocation suggestions for one matrix month."""
    body = request.get_json(silent=True) or {}
    month = (body.get("month") or "").strip()
    if not _month_path(month):
        return jsonify({"ok": False, "error": "supply month=YYYY-MM"}), 400
    if not os.path.isfile(_YEARLY_PATH):
        return jsonify({"ok": False, "error": "no yearly matrix loaded yet"}), 404
    with open(_YEARLY_PATH, encoding="utf-8") as fh:
        yearly = json.load(fh)
    mnum = str(int(month[5:7]))
    active = [e for e in yearly["entries"]
              if (e["wmt"].get(mnum) or 0) > 0 or (e["dt"].get(mnum) or 0) > 0]
    if not active:
        return jsonify({"ok": False, "error": "no matrix rows for %s" % month}), 400
    # Row per origin+dest+contractor+MATERIAL (material drives the constraint).
    rows = {}
    for e in active:
        src = _ORIGIN_MAP.get(e["origin"].upper(), e["origin"].upper())
        dst = _canon_dest(e["dest"])
        mat = (e.get("material") or "").upper()
        key = (src, dst, e["contractor"].upper(), mat)
        rec = rows.setdefault(key, {"src": src, "dst": dst, "route": "%s>%s" % (src, dst),
                                    "contractor": e["contractor"].upper(), "material": mat,
                                    "dt": 0.0, "wmt_day": 0.0})
        rec["dt"] += e["dt"].get(mnum) or 0
        rec["wmt_day"] += e["wmt"].get(mnum) or 0
    rows = [r for r in rows.values() if r["dt"] > 0]
    stats = _day_stats_from_snapshot()
    if not stats:
        return jsonify({"ok": False, "error": "no measured day history available (DB down?)"}), 503
    # Combined fleet per route (saturation is a road property).
    from collections import defaultdict as _dd
    comb = _dd(float)
    for r in rows:
        comb[r["route"]] += r["dt"]
    # Per-route saturation fleet N* = cap/rate.
    def n_star(route):
        s = stats.get(route)
        return (s["cap"] / s["rate"]) if s and s["rate"] > 0 else None
    # 1) LIM donors: trucks beyond saturation on their route (stranded).
    donors = []
    for r in rows:
        if r["material"] != "LIM":
            continue
        ns = n_star(r["route"])
        if ns is None:
            continue
        over = comb[r["route"]] - ns
        if over <= 0:
            continue
        # This row's share of the stranded trucks (pro-rata by its fleet).
        share = r["dt"] / comb[r["route"]] if comb[r["route"]] else 0
        stranded = min(r["dt"], over * share)
        if stranded >= 1:
            donors.append({"row": r, "stranded_dt": stranded, "n_star": ns})
    # 2) SAP receivers: same-contractor rows with measured headroom.
    receivers = []
    for r in rows:
        if r["material"] != "SAP":
            continue
        s = stats.get(r["route"])
        ns = n_star(r["route"])
        if s is None or ns is None:
            continue
        head = ns - comb[r["route"]]
        if head >= 1:
            receivers.append({"row": r, "headroom_dt": head, "rate": s["rate"],
                              "dt_max": s["dt_max"]})
    receivers.sort(key=lambda x: -x["rate"])            # best-paying road first
    # 3) Match: same contractor only; SAP tonnage only goes UP.
    moves = []
    pay_default = 49.9
    for d in donors:
        rem = d["stranded_dt"]
        for rc in receivers:
            if rem < 1:
                break
            if rc["row"]["contractor"] != d["row"]["contractor"]:
                continue
            take = min(rem, rc["headroom_dt"])
            if take < 1:
                continue
            gain = take * rc["rate"] * pay_default
            beyond = comb[rc["row"]["route"]] + take > rc["dt_max"]
            moves.append({
                "contractor": d["row"]["contractor"],
                "from_route": d["row"]["route"], "from_material": "LIM",
                "to_route": rc["row"]["route"], "to_material": "SAP",
                "trucks": int(take),
                "lim_wmt_lost_day": 0,
                "sap_wmt_gain_day": round(gain),
                "note": ("these %d trucks sit past %s's saturation (N*≈%d) and add "
                         "almost nothing there; on %s they still earn the full rate "
                         "(%.1f trips/DT·day)%s"
                         % (int(take), d["row"]["route"], round(d["n_star"]),
                            rc["row"]["route"], rc["rate"],
                            " — takes the road past its biggest measured fleet, label as trial" if beyond else "")),
            })
            rc["headroom_dt"] -= take
            rem -= take
    total_gain = sum(m["sap_wmt_gain_day"] for m in moves)
    return jsonify({
        "ok": True, "month": month,
        "constraints": ["SAP tonnage never reduced (only increased)",
                        "contractors stay on their own routes",
                        "only LIM trucks past their road's saturation point move"],
        "donors": [{"route": d["row"]["route"], "contractor": d["row"]["contractor"],
                    "dt": round(d["row"]["dt"]), "stranded_dt": round(d["stranded_dt"]),
                    "n_star": round(d["n_star"])} for d in donors],
        "moves": moves,
        "sap_gain_day": round(total_gain),
        "lim_loss_day": 0,
        "note": ("LIM tonnage is UNCHANGED by these moves: the donated trucks are "
                 "the ones already past the road's demonstrated saturation - the "
                 "road cannot serve them, so removing them does not remove trips. "
                 "SAP gains are earned at the receiving road's measured day rate."),
    })


# ── Priority allocation (owner 2026-08-13) ───────────────────────────────────
# "we have fixed amount of SAP material and TOS LIM material ... First
#  priority is our SAP ... Second priority is the TOS LIM ... and the third
#  is the rest of the Lemonite left [LD]. We don't get more trucks, so we
#  have to allocate the trucks in a way that we will fulfill our first two
#  priorities ... if we are moving trucks, we can move in between same
#  contractor, first try to move from same plans [same origin], if still we
#  are less ... look for other plans with same contractor."
#
# Fleet is FIXED at the matrix's month total per contractor. Allocation:
#   P1  SAP           — must-move. Filled first from the contractor fleet.
#   P2  LIM from TOS  — filled only with trucks left after SAP. May run short.
#   P3  LIM from LD   — first donor; leftover trucks stay here.
# If LD is not enough for SAP, LIM-TOS trucks move to SAP. Same contractor,
# same-origin donors first. Persisted beside the original month for Excel.

@bp.route("/api/monthly/allocate", methods=["POST"])
def api_monthly_allocate():
    body = request.get_json(silent=True) or {}
    month = (body.get("month") or "").strip()
    if not _month_path(month):
        return jsonify({"ok": False, "error": "supply month=YYYY-MM"}), 400
    yearly = _load_yearly()
    if not yearly:
        return jsonify({"ok": False, "error": "no yearly matrix loaded yet"}), 404
    mnum = str(int(month[5:7]))
    path_models, fleet_kpi, contr_by = _path_model_context()
    if not path_models:
        return jsonify({"ok": False, "error": "no measured day history (path-response empty)"}), 503

    rows = []
    for e in yearly["entries"]:
        wmt = e["wmt"].get(mnum) or 0
        dt = e["dt"].get(mnum) or 0
        if wmt <= 0 and dt <= 0:
            continue
        src = _ORIGIN_MAP.get(e["origin"].upper(), e["origin"].upper())
        dst = _canon_dest(e["dest"])
        route = "%s>%s" % (src, dst)
        mat = (e.get("material") or "").upper()
        otype = (e.get("otype") or "").upper()
        prio = 1 if mat == "SAP" else (
            2 if (mat == "LIM" and otype == "TOS") else (3 if mat == "LIM" else 9))
        rows.append({"origin": e["origin"].upper(), "route": route, "src": src,
                     "dst": dst, "mat": mat, "otype": otype,
                     "contractor": e["contractor"].upper(),
                     "prio": prio, "target": wmt, "matrix_dt": dt})
    if not rows:
        return jsonify({"ok": False, "error": "no matrix rows for %s" % month}), 400

    from collections import defaultdict as _dd
    comb_matrix = _dd(float)
    for r in rows:
        comb_matrix[r["route"]] += r["matrix_dt"]

    for r in rows:
        others = max(0.0, comb_matrix[r["route"]] - r["matrix_dt"])
        if r["prio"] in (1, 2, 3) and r["target"] > 0:
            req, why = _required_dt_day(r["src"], r["dst"], r["contractor"],
                                        r["target"], others, path_models,
                                        fleet_kpi, contr_by)
            if req is None:
                seed = _path_row_wmt(r["src"], r["dst"], r["contractor"], 1,
                                     others + 1, path_models, fleet_kpi, contr_by)
                cap = seed.get("cap_trips")
                pay = seed.get("pay")
                max_wmt = (float(cap) * float(pay)) if cap and pay else None
                if max_wmt and max_wmt > 0:
                    req, why2 = _required_dt_day(
                        r["src"], r["dst"], r["contractor"], max_wmt * 0.995,
                        others, path_models, fleet_kpi, contr_by)
                    why = why or why2 or "target above path ceiling"
                r["capped"] = True
                r["req_dt"] = req if req is not None else r["matrix_dt"]
                r["cap_why"] = why
            else:
                r["req_dt"] = req
                r["capped"] = False
        else:
            r["req_dt"] = r["matrix_dt"]
            r["capped"] = False

    fleet = _dd(float)
    for r in rows:
        fleet[r["contractor"]] += r["matrix_dt"]
    moves, shortfalls = [], []
    unused_fleet = {}
    for cont in sorted(fleet):
        crows = [r for r in rows if r["contractor"] == cont]
        remaining = fleet[cont]
        # Supplied targets consume the fixed fleet strictly P1 -> P2 -> P3.
        # Unclassified work can use only what remains after all three targets.
        for prio in (1, 2, 3):
            for r in sorted([x for x in crows if x["prio"] == prio],
                            key=lambda x: -x["target"]):
                need = float(r["req_dt"] or 0)
                give = min(remaining, int(math.ceil(need))) if need else 0
                r["alloc_dt"] = round(give)
                remaining -= give
                if give + 0.5 < need:
                    if prio == 1:
                        shortfalls.append("%s SAP still short %d DT — %s fleet exhausted (LD + LIM-TOS used)"
                                          % (r["route"], round(need - give), cont))
                    else:
                        label = "LIM-TOS" if prio == 2 else "LIM-LD"
                        shortfalls.append("%s %s short %d DT after higher priorities used the fleet"
                                          % (r["route"], label, round(need - give)))
        for r in sorted([x for x in crows if x["prio"] == 9],
                        key=lambda x: -x["matrix_dt"]):
            give = min(remaining, round(r["matrix_dt"]))
            r["alloc_dt"] = max(0, give)
            remaining -= give
        unused_fleet[cont] = max(0.0, remaining)
        donors = [x for x in crows if x.get("alloc_dt", 0) < x["matrix_dt"]]
        receivers = [x for x in crows if x.get("alloc_dt", 0) > x["matrix_dt"]]
        for rec in sorted(receivers, key=lambda x: x["prio"]):
            need = rec["alloc_dt"] - rec["matrix_dt"]
            for prio_don in (3, 2, 9, 1):
                for same_origin in (True, False):
                    if need <= 0:
                        break
                    for don in donors:
                        if need <= 0:
                            break
                        if don["prio"] != prio_don:
                            continue
                        if (don["origin"] == rec["origin"]) != same_origin:
                            continue
                        avail = don["matrix_dt"] - don.get("alloc_dt", 0) - don.get("_given", 0)
                        if avail <= 0:
                            continue
                        take = min(need, avail)
                        don["_given"] = don.get("_given", 0) + take
                        need -= take
                        why = ("LD buffer" if don["prio"] == 3
                               else ("LIM-TOS → SAP" if rec["prio"] == 1 and don["prio"] == 2
                                     else "rebalance"))
                        moves.append({"contractor": cont,
                                      "from": "%s (%s %s)" % (don["route"], don["mat"], don["otype"]),
                                      "to": "%s (%s %s)" % (rec["route"], rec["mat"], rec["otype"]),
                                      "trucks": round(take),
                                      "same_origin": same_origin,
                                      "reason": why})

    comb = _dd(float)
    for r in rows:
        comb[r["route"]] += r.get("alloc_dt", 0)
    for r in rows:
        a = r.get("alloc_dt", 0)
        pr = _path_row_wmt(r["src"], r["dst"], r["contractor"], a,
                           comb[r["route"]] or a, path_models, fleet_kpi, contr_by)
        r["pred_wmt"] = round(pr["wmt"]) if pr.get("wmt") is not None else None
        r["met"] = r["pred_wmt"] is not None and r["target"] > 0 \
            and r["pred_wmt"] >= r["target"] * 0.995

    merged = {}
    for r in rows:
        key = (r["src"], r["dst"], r["contractor"])
        rec = merged.setdefault(key, {"src": r["src"], "dst": r["dst"],
                                      "contractor": r["contractor"],
                                      "dt": 0.0, "wmt_day": 0.0, "materials": set()})
        rec["dt"] += r.get("alloc_dt", 0)
        rec["wmt_day"] += r["target"]
        if r.get("mat"):
            rec["materials"].add(r["mat"])
    route_list = [v for v in merged.values() if v["dt"] > 0]
    plans = [{"route": "%s>%s" % (r["src"], r["dst"]), "source": r["src"],
              "destination": r["dst"], "n_trucks": int(round(r["dt"])),
              "contractor": r["contractor"]}
             for r in route_list]
    import plan_simulator
    sim = plan_simulator.simulate({"plans": plans})
    sim_rows = (sim or {}).get("results") or []
    pred_day, pred_rows = _plan_predict_for_routes(route_list)
    achv_day = 0.0
    for i, r in enumerate(route_list):
        sr = sim_rows[i] if i < len(sim_rows) else {}
        achv_shift = float(sr.get("achievable_production_t") or 0)
        achv_day += achv_shift * 2
        if i < len(pred_rows) and pred_rows[i].get("wmt") is not None:
            pass
    target_day = sum(r["target"] for r in rows)

    st = _load_state(month) or {"month": month}
    old_pred = (st.get("prediction") or {}).get("per_day_wmt")
    old_achv = (st.get("prediction") or {}).get("per_day_achv_wmt")
    if old_pred is None:
        orig_list = _routes_for_month(yearly, mnum)
        if orig_list:
            old_pred, _ = _plan_predict_for_routes(orig_list)
            old_pred = round(old_pred)
            orig_plans = [{"route": "%s>%s" % (r["src"], r["dst"]), "source": r["src"],
                           "destination": r["dst"], "n_trucks": int(round(r["dt"])),
                           "contractor": r["contractor"]} for r in orig_list]
            orig_sim = plan_simulator.simulate({"plans": orig_plans})
            old_achv = round(sum(float(x.get("achievable_production_t") or 0) * 2
                                 for x in (orig_sim.get("results") or [])))
    prio_sum = {str(p): {"target": round(sum(r["target"] for r in rows if r["prio"] == p)),
                         "pred": round(sum(r["pred_wmt"] or 0 for r in rows if r["prio"] == p))}
                for p in (1, 2, 3)}
    payload = {
        "applied_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "old": {"pred": old_pred, "achv": old_achv, "target": round(target_day),
                "dt": int(round(sum(r["matrix_dt"] for r in rows)))},
        "new": {"pred": round(pred_day), "achv": round(achv_day),
                "target": round(target_day),
                "dt": int(round(sum(r.get("alloc_dt", 0) for r in rows)))},
        "rows": [{k: r.get(k) for k in
                  ("origin", "route", "mat", "otype", "contractor", "prio", "target",
                   "matrix_dt", "alloc_dt", "pred_wmt", "met", "capped")}
                 for r in sorted(rows, key=lambda x: (x["prio"], x["contractor"]))],
        "moves": moves,
        "shortfalls": shortfalls,
        "prio_summary": prio_sum,
        "fleet": {c: round(v) for c, v in fleet.items()},
        "unused_fleet": {c: round(v, 1) for c, v in unused_fleet.items()},
    }
    st["allocation"] = payload
    _save_state(month, st)
    return jsonify({
        "ok": True, "month": month, "saved": True,
        "priorities": ["P1 SAP (fixed supply)", "P2 LIM from TOS", "P3 LIM from LD (target)"],
        "fleet": payload["fleet"],
        "unused_fleet": payload["unused_fleet"],
        "old": payload["old"], "new": payload["new"],
        "rows": payload["rows"], "moves": moves, "shortfalls": shortfalls,
        "prio_summary": prio_sum,
        "note": ("Supplied targets consume the fixed fleet in strict order: "
                 "SAP, then LIM-TOS, then LIM-LD. Any fleet beyond all targets "
                 "is reported as unused rather than credited as production. "
                 "Required DT inverts the Plan path model (hard ceiling). Original "
                 "month prediction is kept; this allocation is stored beside it."),
    })
