"""Scenario planning: P1 SAP -> P2 LIM-TOS -> P3 LIM-LD waterfall.

A *scenario* is a set of pit x material monthly ROM targets (t/day). The fleet
is NOT part of a scenario: every scenario runs on the same fixed fleet - the
per-contractor DT pools of the loaded yearly matrix (Scenario 1). What changes
between scenarios is where those trucks go.

The waterfall, in the owner's words (2026-08-18):
  1. allocate DTs until every SAP target is met,
  2. then until every LIM-TOS target is met,
  3. every truck still free hauls LIM-LD (Tofu limonite dump -> Huafei).

THE LIM-LD SALES-TARGET RULE, as two labelled numbers (owner ruling,
2026-08-23; target value 6,644,306 t since the planning team's 2026-08-26
sales table — it was 8 Mt before that).
Both owner statements are true at once and the repo's standing pattern is
two clocks, never merged:

  * CAPACITY IS NEVER CLIPPED. What the free fleet could physically move on
    LD is computed and reported in full, above the target when the trucks are
    there (`ld_t_day_capacity`, `ld_t_month_capacity`, `ld_t_capacity`,
    `dt_p3_capacity`, `free`). Destroying that information is what the
    2026-08-19 rule forbids: "LIM-LD is the only place extra trucks go - it
    has no kind of cap."
  * CREDITED PRODUCTION IS BOUNDED BY THE SUPPLIED TARGET. Tonnage credited
    against a target never exceeds it (`ld_t_month_planned`, `ld_t_planned`,
    `p3`, `dt_p3`); the remainder is reported as explicit unused/excess
    capacity (`unused`, `ld_t_month_excess`) and never folded into headline
    production.

So: no clip on capacity, no credit beyond target, both present in the
payload and in the workbook with names that say which is which. The target
(LIM_LD_TARGET_T below) is the sales line the plan is judged against, not a
limit on what the fleet can do.

Hard rules:
  * BLB pit accepts RIM trucks only - never SMA or another contractor.
  * Free DTs are pooled per contractor: a spare SMA truck cannot cover a
    BLB (RIM) shortfall, but it can cover TOFU/KRENE work and LIM-LD.
  * DT counts per month never change; only the allocation does.
  * Credited tonnage never exceeds what the allocated DT can physically
    move: an impossible target starves P3, reports a deficit and marks the
    month `feasible: false` - it is never credited as delivered production.

Storage: data/scenarios/{id}.json - one file per scenario, S1 is derived
live from the yearly matrix so it can never drift from the real plan.
"""

import io
import json
import os
import re
from collections import defaultdict
from datetime import datetime

from flask import Blueprint, jsonify, request

bp = Blueprint("scenario_api", __name__)

_ROOT = os.path.dirname(os.path.abspath(__file__))
_SCEN_DIR = os.path.join(_ROOT, "data", "scenarios")
_YEARLY_PATH = os.path.join(_ROOT, "data", "monthly_plans", "yearly_matrix.json")

# Matrix origin labels -> scenario pit labels (the mine-plan workbook vocabulary).
_PIT = {"BLB": "BLB", "KR": "KRENE", "KRENE": "KRENE", "TOFU": "TOFU", "TF": "TOFU"}
_PITS = ("BLB", "KRENE", "TOFU")
_MONTHS = (8, 9, 10, 11, 12)
_DAYS = {8: 31, 9: 30, 10: 31, 11: 30, 12: 31}
_MN = {"aug": 8, "august": 8, "sept": 9, "sep": 9, "september": 9,
       "oct": 10, "october": 10, "nov": 11, "november": 11,
       "dec": 12, "december": 12}

# H2 LIM-LD sales TARGET: the line credited production is measured against,
# never a limit on the fleet. Capacity above it is computed and reported in
# full (see the two-labelled-numbers rule in the module docstring); only the
# CREDITED tonnage stops here. Planning team sales table 2026-08-26:
# Limonite LD 6,644,306 wmt declared (was 8 Mt).
LIM_LD_TARGET_T = 6_644_306
# Owner 2026-08-27: the TOTAL haulage target is constant at 17,003,193
# (sales table 20,003,193 minus the out-of-scope LGS POS 3 Mt). Scenario
# 3.0 does not add the ~1 Mt to LIM-TOS, so that tonnage TRANSFERS to
# LIM-LD: 3.0's LD line is 6,644,306 + 990,000. One total, two splits.
LIM_LD_TARGET_30_T = LIM_LD_TARGET_T + 990_000   # 7,634,306
HAULAGE_TOTAL_T = 17_003_193
SAP_SALES_T = 5_718_686
LIM_TOS_SALES_31_T = 4_640_201
LIM_TOS_SALES_30_T = 3_650_201

# ── Scenario 4.1 (manager's ROM table, 2026-08-31) ──────────────────────
# From the 20260828 mine plan "Scen 3 (Without IPPKH6) LimRec80%" ROM sheet:
#   SAP 6,383,425  = SAP ORE 5,756,186 (ROM 6,541,121 x 88% ore-sales
#                    recovery; the TOFU grizzly keeps ~10% of ROM as reject
#                    in TOFU, hence 6.5 Mt mined -> 6.38 Mt hauled)
#                    + 627,239 topped up from HGS stock
#   LIM-TOS 4,581,137 = the LIM ORE line, fresh limonite from the pit (exact)
#   LIM-LD 6,035,439  = drawn from the LIM stockpile (7,243,968 recovered
#                    available: TOFU 8,009,107 x 88% + BLB/CBB/CSW x 90%)
#   TOTAL 17,000,001 — one clock, same as ever.
# Hauling rule (Killian, 2026-08-31): RIM preferentially works the FeNi
# routes; other contractors' fleets take the HUAFEI legs where walls allow.
SAP_SALES_41_T = 6_383_425
LIM_TOS_SALES_41_T = 4_581_137
LIM_LD_SALES_41_T = 6_035_439
HAULAGE_TOTAL_41_T = SAP_SALES_41_T + LIM_TOS_SALES_41_T + LIM_LD_SALES_41_T


def sap_target_for_scenario(sid):
    """SAP sales line by scenario: 4.1 carries the ROM-table SAP."""
    s = str(sid or "").upper()
    if s == "S7" or "4.1" in s:
        return SAP_SALES_41_T
    return SAP_SALES_T


def total_target_for_scenario(sid):
    s = str(sid or "").upper()
    if s == "S7" or "4.1" in s:
        return HAULAGE_TOTAL_41_T
    return HAULAGE_TOTAL_T


def tos_target_for_scenario(sid):
    """LIM-TOS sales line by scenario family."""
    s = str(sid or "").upper()
    if s == "S7" or "4.1" in s:
        return LIM_TOS_SALES_41_T
    if s in ("S3", "S4") or "3.0" in s:
        return LIM_TOS_SALES_30_T
    return LIM_TOS_SALES_31_T


def ld_target_for_scenario(sid):
    """LD sales line by scenario family: 3.0 carries the transferred ~1 Mt."""
    s = str(sid or "").upper()
    if s == "S7" or "4.1" in s:
        return LIM_LD_SALES_41_T
    if s in ("S3", "S4") or "3.0" in s:
        return LIM_LD_TARGET_30_T
    return LIM_LD_TARGET_T
LIM_LD_CAP_T = LIM_LD_TARGET_T  # back-compat alias (old name, same number)
# P3 LIM-LD always runs Tofu limonite dump -> Huafei. Named once so the
# draft-plan sizing and the LD rows cannot disagree about which road the
# leftover fleet joins.
LD_ROUTE_KEY = "TF>HUAFEI"
RIM_ONLY_PITS = ("BLB",)

# SAP routing for imported scenarios (owner, 2026-08-26 — INVERTING the
# 2026-08-25 rule, which the owner called a blunder the next morning):
#   ~2 kt/day of each pit's SAP goes to POS as the BUFFER (±2 kt landing
#   band); the REST goes DIRECT to FeNi, to the destination the pit's own
#   matrix rows name. _split_sap_conditions distributes the rest PRO-RATA
#   over the matching matrix rows, so a pit whose plan ships to BOTH FeNi
#   plants splits between them automatically ("if from one pit it's going to
#   two FeNis, add the two FeNis accordingly, as per our plans").
#   KRENE's matrix carries no FeNi SAP row; its rest goes to its corridor's
#   most-used direct haul, FENI KM15 (dispatch history: 375 direct rows to
#   KM15 vs 214 to KM0). S1 keeps the matrix's own split untouched.
SAP_ROUTING = {
    "BLB":  {"fixed": [("POS 14", 2000.0)], "rest": "FENI KM0"},
    "TOFU": {"fixed": [("POS 12", 2000.0)], "rest": "FENI KM15"},
    "KRENE": {"fixed": [("POS 12", 2000.0)], "rest": "FENI KM15"},
}


def _norm_sap_dest(d):
    s = re.sub(r"\s+", " ", str(d or "").upper().strip())
    return s.replace("KM 0", "KM0").replace("KM 15", "KM15").replace("KM 10", "KM10")


def _split_sap_conditions(pit, T, grp, m):
    """[(row, wmt_day)] under SAP_ROUTING. grp = the pit's SAP matrix routes."""
    rule = SAP_ROUTING.get(pit)
    out, left = [], float(T)
    if not rule:
        return None
    for dest, amt in rule["fixed"]:
        want = _norm_sap_dest(dest)
        rows = [r for r in grp if _norm_sap_dest(r["dest"]) == want]
        if left <= 0:
            continue
        if not rows:
            # The 2,000 t/day POS buffer is the owner's standing instruction
            # (2026-08-26); the matrix rows predate it. A matrix without the
            # POS SAP row still buffers — clone one, as the rest leg does.
            probe = dict(grp[0])
            probe["dest"] = dest
            rows = [probe]
        w = min(left, amt)
        out.append((rows[0], w))
        left -= w
    rest_want = _norm_sap_dest(rule["rest"])
    # The plan's own destinations are the authority (owner, 2026-08-26): the
    # rest spreads pro-rata over EVERY FeNi SAP row the pit's matrix carries,
    # so a pit shipping to both plants splits between them. The named rest is
    # only the answer when the matrix itself is silent — it must not override
    # a two-FeNi plan down to one plant.
    rest_rows = [r for r in grp if "FENI" in _norm_sap_dest(r["dest"])]
    if left > 0 and not rest_rows:
        # Matrix has no FeNi SAP haul for this pit (KRENE today). Clone onto
        # the named rest dest so the leftover is not silently dumped on POS.
        probe = dict(grp[0])
        probe["dest"] = rule["rest"]
        rest_rows = [probe]
    if left > 0:
        base = sum(r["wmt"].get(m, 0) for r in rest_rows)
        for r in rest_rows:
            share = (r["wmt"].get(m, 0) / base) if base else 1.0 / len(rest_rows)
            out.append((r, left * share))
    return out


# ---------------------------------------------------------------- storage

def _safe_id(sid):
    sid = re.sub(r"[^A-Za-z0-9_-]", "", str(sid or ""))
    return sid[:32] or None


def _scen_path(sid):
    sid = _safe_id(sid)
    return os.path.join(_SCEN_DIR, sid + ".json") if sid else None


def _load_yearly():
    if not os.path.isfile(_YEARLY_PATH):
        return None
    with open(_YEARLY_PATH, encoding="utf-8") as fh:
        return json.load(fh)


def _yearly_rows(yearly):
    """Yearly-matrix entries in the allocator's vocabulary."""
    rows = []
    for e in (yearly or {}).get("entries") or []:
        pit = _PIT.get((e.get("origin") or "").upper())
        if not pit:
            continue
        rows.append({
            "pit": pit,
            "mat": (e.get("material") or "").upper(),
            "otype": (e.get("otype") or "").upper(),
            "dest": (e.get("dest") or "").upper(),
            "contractor": (e.get("contractor") or "").upper(),
            "wmt": {int(k): v for k, v in (e.get("wmt") or {}).items() if v},
            "dt": {int(k): v for k, v in (e.get("dt") or {}).items() if v},
        })
    return rows


def _route_rate(row, m):
    """Demonstrated t/DT/day for one matrix row, month m (fallback: its mean)."""
    w, d = row["wmt"].get(m), row["dt"].get(m)
    if w and d:
        return w / d
    vals = [row["wmt"][k] / row["dt"][k]
            for k in row["dt"] if row["dt"].get(k) and row["wmt"].get(k)]
    return (sum(vals) / len(vals)) if vals else None


def _s1_targets(rows):
    """Scenario 1 targets = the matrix's own SAP / LIM-TOS wmt by pit."""
    tgt = defaultdict(float)
    for r in rows:
        if r["mat"] == "LIM" and r["otype"] == "LD":
            continue  # LD is the P3 sink, not a target
        for m, w in r["wmt"].items():
            tgt[(r["pit"], r["mat"], m)] += w
    return dict(tgt)


def _scenario_ids():
    out = []
    if os.path.isdir(_SCEN_DIR):
        for f in sorted(os.listdir(_SCEN_DIR)):
            if f.endswith(".json"):
                out.append(f[:-5])
    return out


def _load_scenario(sid):
    """S1 is always derived live from the yearly matrix; others from disk."""
    if sid == "S1":
        yearly = _load_yearly()
        if not yearly:
            return None
        rows = _yearly_rows(yearly)
        tgt = _s1_targets(rows)
        return {
            "id": "S1", "label": "Scenario 1 - current plan",
            "source": "yearly matrix (%s)" % (yearly.get("source") or "pasted"),
            "derived": True,
            "targets": [{"pit": p, "mat": mt, "month": m, "wmt_day": round(v, 1)}
                        for (p, mt, m), v in sorted(tgt.items())],
        }
    p = _scen_path(sid)
    if not p or not os.path.isfile(p):
        return None
    with open(p, encoding="utf-8") as fh:
        return json.load(fh)


def _save_scenario(sc):
    os.makedirs(_SCEN_DIR, exist_ok=True)
    p = _scen_path(sc["id"])
    # The J59 lesson: identical targets keep their stamp. A re-import of the
    # same workbook must not churn source/loaded_at, so git status stays a
    # usable signal.
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as fh:
                old = json.load(fh)
            if old.get("targets") == sc.get("targets"):
                return
        except (OSError, ValueError):
            pass
    tmp = p + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(sc, fh, indent=1)
        fh.write("\n")
    os.replace(tmp, p)


# ---------------------------------------------------------------- importer

def _parse_mine_plan_db(ws_rows, src_name):
    """Long-format 'Mine Plan DB' sheet: Scenario | Month | Nb Days | Mining Pit |
    Material | Type Ore | wmt ROM | ... One scenario dict per distinct label."""
    hdr = None
    for i, r in enumerate(ws_rows[:10]):
        low = [str(c or "").strip().lower() for c in r]
        if "scenario" in low and "month" in low:
            hdr = i
            break
    if hdr is None:
        return None, "no header row (need Scenario / Month / Mining Pit / Material / wmt ROM columns)"
    low = [str(c or "").strip().lower() for c in ws_rows[hdr]]

    def col(*names):
        for n in names:
            if n in low:
                return low.index(n)
        return None

    c_sc, c_m = col("scenario"), col("month")
    c_days, c_pit = col("nb days"), col("mining pit", "pit")
    c_mat, c_wmt = col("material"), col("wmt rom", "wmt")
    c_otype = col("type ore", "ore type", "otype")
    if None in (c_sc, c_m, c_pit, c_mat, c_wmt):
        return None, "missing one of: Scenario, Month, Mining Pit, Material, wmt ROM"
    acc = defaultdict(float)
    for r in ws_rows[hdr + 1:]:
        sc = str(r[c_sc] or "").strip()
        mon = _MN.get(str(r[c_m] or "").strip().lower())
        pit = _PIT.get(str(r[c_pit] or "").strip().upper())
        mat = str(r[c_mat] or "").strip().upper()
        raw_otype = str(r[c_otype] or "").strip().upper() if c_otype is not None else ""
        otype = "LD" if (mat == "LIM" and "LD" in raw_otype) else "TOS"
        try:
            wmt = float(r[c_wmt] or 0)
        except (TypeError, ValueError):
            wmt = 0
        if not sc or not mon or not pit or mat not in ("SAP", "LIM") or wmt <= 0:
            continue
        days = None
        if c_days is not None:
            try:
                days = float(r[c_days] or 0) or None
            except (TypeError, ValueError):
                days = None
        acc[(sc, pit, mat, otype, mon)] += wmt / (days or _DAYS[mon])
    if not acc:
        return None, "found the header but no scenario rows under it"
    scens = {}
    for (sc, pit, mat, otype, mon), v in acc.items():
        sid = "S" + re.sub(r"\D", "", sc) if re.search(r"\d", sc) else _safe_id(sc)
        rec = scens.setdefault(sid, {"id": sid, "label": sc, "source": src_name,
                                     "targets": []})
        rec["targets"].append({"pit": pit, "mat": mat, "otype": otype, "month": mon,
                               "wmt_day": round(v, 1)})
    for rec in scens.values():
        rec["targets"].sort(key=lambda t: (t["month"], t["pit"], t["mat"], t.get("otype", "")))
        rec["loaded_at"] = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    return list(scens.values()), None


# ---------------------------------------------------------------- allocator

def waterfall(sc, yearly=None, ld_cap=LIM_LD_TARGET_T):
    """Run the P1 -> P2 -> P3 waterfall for one scenario.

    Fleet = yearly-matrix DT pools per contractor per month (fixed).
    Targets = scenario pit x material t/day; split across that pit's matrix
    routes by S1 wmt share, each at its demonstrated t/DT/day.
    Months the scenario does not cover fall back to S1 targets (August).
    """
    yearly = yearly or _load_yearly()
    if not yearly:
        return None, "no yearly matrix loaded yet"
    rows = _yearly_rows(yearly)
    if not rows:
        return None, "the yearly matrix has no usable rows"
    s1 = _s1_targets(rows)
    tgt = dict(s1) if sc["id"] == "S1" else {}
    ld_targets = defaultdict(float)
    filled_from_s1 = []
    if sc["id"] != "S1":
        for t in sc.get("targets") or []:
            mat = str(t.get("mat") or "").upper()
            otype = str(t.get("otype") or "").upper()
            month = int(t["month"])
            value = float(t["wmt_day"] or 0)
            if mat in ("LIM-LD", "LD") or (mat == "LIM" and otype == "LD"):
                ld_targets[month] += value
            else:
                tgt[(t["pit"], mat, month)] = value
        scen_months = {int(t["month"]) for t in sc.get("targets") or []}
        for (pit, mat, m), v in s1.items():
            if m not in scen_months and (pit, mat, m) not in tgt:
                tgt[(pit, mat, m)] = v
                if m not in filled_from_s1:
                    filled_from_s1.append(m)
    work = [r for r in rows if not (r["mat"] == "LIM" and r["otype"] == "LD")]
    ld_rows = [r for r in rows if r["mat"] == "LIM" and r["otype"] == "LD"]
    ld_rate = {}
    for r in ld_rows:
        rt = _route_rate(r, 11)
        if rt:
            ld_rate[r["contractor"]] = rt
    ld_rate.setdefault("RIM", 120.0)
    ld_rate.setdefault("SMA", 100.0)

    months_out, violations = [], []
    ld_cum = 0.0
    ld_capacity_total = 0.0
    for m in _MONTHS:
        pool = defaultdict(float)
        for r in rows:
            pool[r["contractor"]] += r["dt"].get(m, 0)
        if not sum(pool.values()):
            continue
        used = defaultdict(float)
        alloc_rows = []
        deficit = []
        for prio, mat in ((1, "SAP"), (2, "LIM")):
            for pit in _PITS:
                T = tgt.get((pit, mat, m)) or 0
                if T <= 0:
                    continue
                grp = [r for r in work if r["pit"] == pit and r["mat"] == mat]
                base = sum(r["wmt"].get(m, 0) for r in grp)
                if not grp:
                    deficit.append({"pit": pit, "mat": mat, "wmt_day": round(T),
                                    "why": "no matrix route hauls this"})
                    continue
                # Imported scenarios route SAP by the owner's conditions
                # (fixed FENI tonnages, rest to POS). S1 keeps the matrix split.
                pieces = None
                if mat == "SAP" and sc["id"] != "S1":
                    pieces = _split_sap_conditions(pit, T, grp, m)
                if pieces is None:
                    pieces = [(r, T * ((r["wmt"].get(m, 0) / base) if base
                                       else 1.0 / len(grp))) for r in grp]
                for r, w in pieces:
                    if w <= 0:
                        continue
                    rate = _route_rate(r, m)
                    if not rate:
                        deficit.append({"pit": pit, "mat": mat, "wmt_day": round(w),
                                        "why": "no demonstrated rate for %s>%s" % (pit, r["dest"])})
                        continue
                    if pit in RIM_ONLY_PITS and r["contractor"] != "RIM":
                        violations.append("%s row with %s contractor in month %d"
                                          % (pit, r["contractor"], m))
                        continue
                    d = w / rate
                    used[r["contractor"]] += d
                    alloc_rows.append({
                        "prio": prio, "pit": pit, "mat": mat, "otype": r["otype"],
                        "dest": r["dest"], "contractor": r["contractor"],
                        # wmt_day is the TARGET asked of this route. dt is the
                        # fleet actually fielded and wmt_day_credited is what
                        # that fleet moves - the feasibility pass below may cut
                        # dt when the pool cannot supply the target.
                        "wmt_day": round(w), "dt": round(d, 1),
                        "dt_target": round(d, 1),
                        "wmt_day_credited": round(w),
                        "feasible": True,
                        "rate_t_dt_day": round(rate, 1),
                    })
        # Lending: overflow in one contractor covered by the other's spare,
        # never into a RIM-only pit.
        lends = []
        for c in list(used):
            over = used[c] - pool[c]
            if over <= 0.01:
                continue
            for c2 in pool:
                if c2 == c:
                    continue
                spare = pool[c2] - used[c2]
                take = min(over, max(0.0, spare))
                if take <= 0:
                    continue
                movable = [a for a in alloc_rows
                           if a["contractor"] == c and a["pit"] not in RIM_ONLY_PITS]
                if not movable:
                    continue
                used[c] -= take
                used[c2] += take
                lends.append({"from": c2, "to_work_of": c, "dt": round(take, 1),
                              "note": "spare %s trucks cover %s work outside %s"
                                      % (c2, c, "/".join(RIM_ONLY_PITS))})
                over -= take
            if over > 0.01:
                deficit.append({"pit": "*", "mat": "*", "wmt_day": None,
                                "why": "%s short %.0f DT even after lending" % (c, over)})
        # ---- feasibility: credited tonnage <= what the fielded DT can move --
        # A target the pool cannot supply used to be credited in full: SAP x100
        # on S3 allocated 29,187 DT out of a 1,281-DT December pool and reported
        # 576,659,545 t against an honest 6,927,160 t (83x). The waterfall was
        # already right to starve P3 and emit deficits - that stays untouched -
        # but sap_t_day/limtos_t_day read the ASK, not the delivery. Trucks are
        # never invented here: rows are cut back in strict P1 -> P2 order until
        # the contractor's rows fit its pool (plus whatever was lent to them).
        month_feasible = True
        infeasible_dt = 0.0
        for c in list(used):
            over = used[c] - pool[c]
            if over <= 0.01:
                continue
            crows = [a for a in alloc_rows if a["contractor"] == c]
            room = max(0.0, sum(a["dt"] for a in crows) - over)
            for a in sorted(crows, key=lambda x: (x["prio"], -x["wmt_day"])):
                give = min(a["dt"], room)
                room -= give
                if give + 0.05 < a["dt"]:
                    month_feasible = False
                    infeasible_dt += a["dt"] - give
                    a["feasible"] = False
                    deficit.append({
                        "pit": a["pit"], "mat": a["mat"],
                        "wmt_day": round(a["wmt_day"] - give * a["rate_t_dt_day"]),
                        "why": "%s pool supplies %.0f of the %.0f DT this target needs"
                               % (c, give, a["dt"]),
                    })
                a["dt"] = round(give, 1)
                a["wmt_day_credited"] = round(give * a["rate_t_dt_day"])
            used[c] = pool[c]
        free = {c: max(0.0, pool[c] - used[c]) for c in pool}
        # CAPACITY (never clipped): what every free truck could move on LD.
        ld_day_capacity = sum(free[c] * ld_rate.get(c, 100.0) for c in free)
        ld_month_capacity = ld_day_capacity * _DAYS[m]
        ld_capacity_total += ld_month_capacity
        # CREDITED (bounded by the supplied target): a month-specific imported
        # LD target wins; otherwise the horizon target (LIM_LD_TARGET_T) is
        # filled in chronological order. Capacity above it is reported as
        # unused/excess, never folded into headline production.
        explicit_month_target = ld_targets.get(m)
        if explicit_month_target is not None:
            target_month = max(0.0, explicit_month_target * _DAYS[m])
        elif ld_cap is not None:
            target_month = max(0.0, float(ld_cap) - ld_cum)
        else:
            target_month = ld_month_capacity
        ld_month = min(ld_month_capacity, target_month)
        use_scale = (ld_month / ld_month_capacity) if ld_month_capacity > 0 else 0.0
        p3 = {c: free[c] * use_scale for c in free}
        unused = {c: max(0.0, free[c] - p3[c]) for c in free}
        ld_cum += ld_month
        sap_credited = sum(a["wmt_day_credited"] for a in alloc_rows if a["prio"] == 1)
        tos_credited = sum(a["wmt_day_credited"] for a in alloc_rows if a["prio"] == 2)
        months_out.append({
            "month": m, "days": _DAYS[m],
            "pool": {c: round(v) for c, v in pool.items()},
            "dt_p1": round(sum(a["dt"] for a in alloc_rows if a["prio"] == 1), 1),
            "dt_p2": round(sum(a["dt"] for a in alloc_rows if a["prio"] == 2), 1),
            # free = physical P3 capacity. p3 = the part credited against the
            # LD target. unused = the labelled remainder. Never merged.
            "free": {c: round(v, 1) for c, v in free.items()},
            "p3": {c: round(v, 1) for c, v in p3.items()},
            "unused": {c: round(v, 1) for c, v in unused.items()},
            "dt_p3": round(sum(p3.values()), 1),
            "dt_p3_capacity": round(sum(free.values()), 1),
            "dt_p3_unused": round(sum(unused.values()), 1),
            "rows": alloc_rows,
            "lends": lends,
            "deficit": deficit,
            "feasible": month_feasible,
            "infeasible_dt": round(infeasible_dt, 1),
            "sap_t_day": round(sap_credited),
            "limtos_t_day": round(tos_credited),
            "sap_t_day_target": round(sum(a["wmt_day"] for a in alloc_rows if a["prio"] == 1)),
            "limtos_t_day_target": round(sum(a["wmt_day"] for a in alloc_rows if a["prio"] == 2)),
            "ld_t_day_capacity": round(ld_day_capacity),
            "ld_t_day_planned": round(ld_month / _DAYS[m]) if _DAYS[m] else 0,
            "ld_t_month_capacity": round(ld_month_capacity),
            "ld_t_month_planned": round(ld_month),
            "ld_t_month_excess": round(max(0.0, ld_month_capacity - ld_month)),
            "ld_target_month": round(target_month),
            "ld_capped": bool(ld_month + 0.5 < ld_month_capacity),
        })
    total = {
        "sap_t": round(sum(mo["sap_t_day"] * mo["days"] for mo in months_out)),
        "limtos_t": round(sum(mo["limtos_t_day"] * mo["days"] for mo in months_out)),
        "sap_t_target": round(sum(mo["sap_t_day_target"] * mo["days"] for mo in months_out)),
        "limtos_t_target": round(sum(mo["limtos_t_day_target"] * mo["days"] for mo in months_out)),
        "ld_t_capacity": round(ld_capacity_total),
        "ld_t_planned": round(ld_cum),
        "ld_t_excess_capacity": round(max(0.0, ld_capacity_total - ld_cum)),
        "ld_cap": ld_cap,
        "ld_target": ld_cap,
        "ld_cap_reached": bool(ld_cap and ld_cum >= ld_cap - 0.5),
        "ld_shortfall_t": round(max(0.0, (ld_cap or 0) - ld_cum)),
        "ld_over_target_t": round(max(0.0, ld_cum - (ld_cap or 0))),
        "feasible": all(mo["feasible"] for mo in months_out),
        "infeasible_months": [mo["month"] for mo in months_out if not mo["feasible"]],
    }
    return {
        "id": sc["id"], "label": sc.get("label") or sc["id"],
        "months": months_out, "total": total,
        "months_filled_from_s1": sorted(filled_from_s1),
        "violations": violations,
        "feasible": total["feasible"],
        "priority_note": "Targets run P1 SAP -> P2 LIM-TOS -> P3 LIM-LD (Tofu dump -> Huafei). "
                         "Fixed fleet: the yearly matrix's DT per contractor per month. "
                         "LIM-LD capacity is never clipped and is reported in full; the "
                         "tonnage CREDITED against the target stops at it and the rest is "
                         "reported as unused/excess. Credited tonnage never exceeds what "
                         "the fielded DT can move (feasible=false says a target outran "
                         "the pool). BLB accepts RIM only.",
    }, None


# ---------------------------------------------------------------- full Excel (all scenarios)

_PIT_SRC = {"BLB": "BLB", "KRENE": "KR", "TOFU": "TF"}


def _wf_route_key(a):
    import monthly_api as ma
    src = _PIT_SRC.get(a["pit"], a["pit"])
    src = ma._ORIGIN_MAP.get(src, src)
    return (src, ma._canon_dest(a["dest"]), a["contractor"], a["mat"], a.get("otype") or "TOS")


def _matrix_route_key(e, mstr):
    import monthly_api as ma
    src = ma._ORIGIN_MAP.get((e.get("origin") or "").upper(), (e.get("origin") or "").upper())
    return (src, ma._canon_dest(e.get("dest")), (e.get("contractor") or "").upper(),
            (e.get("material") or "").upper(), (e.get("otype") or "").upper())


def _dt_moves(rows):
    """Net DT transfers per contractor, same shape as Plan-tab Allocate moves."""
    from collections import defaultdict
    by_c = defaultdict(list)
    for r in rows:
        by_c[r.get("contractor") or ""].append(r)
    moves = []
    for c, rs in by_c.items():
        donors, recvs = [], []
        for r in rs:
            delta = (r.get("dt_before") or 0) - (r.get("dt_after") or 0)
            if delta > 0.5:
                donors.append([r, delta])
            elif delta < -0.5:
                recvs.append([r, -delta])
        i = j = 0
        while i < len(donors) and j < len(recvs):
            take = min(donors[i][1], recvs[j][1])
            if take >= 0.5:
                drow, rrow = donors[i][0], recvs[j][0]
                fm = "%s%s" % (drow.get("material") or "",
                               ("-" + drow["otype"]) if drow.get("otype") else "")
                tm = "%s%s" % (rrow.get("material") or "",
                               ("-" + rrow["otype"]) if rrow.get("otype") else "")
                origin_d = (drow.get("key") or "").split(">")[0]
                origin_r = (rrow.get("key") or "").split(">")[0]
                moves.append({
                    "contractor": c,
                    "from": drow.get("key"),
                    "to": rrow.get("key"),
                    "trucks": int(round(take)),
                    "tag": "%s → %s" % (fm, tm),
                    "reason": "%s → %s" % (fm, tm),
                    "from_mat": drow.get("material"),
                    "to_mat": rrow.get("material"),
                    "same_origin": origin_d == origin_r,
                })
            donors[i][1] -= take
            recvs[j][1] -= take
            if donors[i][1] < 0.5:
                i += 1
            if recvs[j][1] < 0.5:
                j += 1
    return moves


def _allocation_target_day(raw):
    """THE target owner for a scenario workbook: t/day the frozen allocation
    itself asks for, summed over its own rows.

    One function, one number. The month card used to compute
    sap_t_day + limtos_t_day + ld_month_planned/n while the sheets rendered
    _build_synthetic_allocation's row targets, and the two diverged wherever
    the scenario zeroed a matrix route (Sep 93,169 vs 80,563 t/day, +15.6%;
    Dec 188,441 vs 183,441, +2.7%). Card, month sheet and Year total now all
    read this, so they cannot drift again.
    """
    if not isinstance(raw, dict):
        return 0.0
    return sum(float(r.get("target") or 0) for r in (raw.get("rows") or []))


def _build_synthetic_allocation(mo, yearly, mnum, sc_id):
    """Frozen Plan-tab allocation shape from a waterfall month (for Excel only)."""
    import monthly_api as ma
    mstr = str(int(mnum))
    wf_map = {_wf_route_key(a): a for a in (mo.get("rows") or [])}
    detail = []
    ld_pending = []
    for e in (yearly or {}).get("entries") or []:
        dt_b = (e.get("dt") or {}).get(mstr) or (e.get("dt") or {}).get(int(mnum)) or 0
        mat = (e.get("material") or "").upper()
        otype = (e.get("otype") or "").upper()
        if dt_b <= 0 and mat == "LIM" and otype == "LD":
            continue
        rk = _matrix_route_key(e, mstr)
        src, dst, contractor, _, _ = rk
        key = "%s>%s" % (src, dst)
        wf = wf_map.get(rk)
        if mat == "LIM" and otype == "LD":
            if dt_b > 0:
                ld_pending.append({
                    "key": key, "src": src, "dst": dst, "contractor": contractor,
                    "material": mat, "otype": otype, "prio": 3,
                    "dt_before": float(dt_b), "dt_after": 0.0, "target": 0,
                })
            continue
        if wf:
            detail.append({
                "key": key, "src": src, "dst": dst, "contractor": contractor,
                "material": mat, "otype": otype, "prio": wf["prio"],
                "dt_before": float(dt_b), "dt_after": float(wf["dt"]),
                "target": float(wf["wmt_day"] or 0),
            })
        elif dt_b > 0:
            # A matrix route the SCENARIO does not run. Its trucks are shown
            # leaving (dt_before > 0, dt_after = 0) but it carries NO target:
            # the matrix wmt here belongs to S1's routing, and this scenario's
            # own target for the same pit x material is already carried by the
            # routes the waterfall did allocate. Counting it again is the
            # 533,180 t phantom (Sep BLB>POS 14 12,606 t/d, Dec BLB>POS 16
            # 5,000 t/d) that inflated the headline row and the Year total.
            detail.append({
                "key": key, "src": src, "dst": dst, "contractor": contractor,
                "material": mat, "otype": otype, "prio": 1 if mat == "SAP" else 2,
                "dt_before": float(dt_b), "dt_after": 0.0,
                "target": 0.0,
                "matrix_target": float((e.get("wmt") or {}).get(mstr)
                                       or (e.get("wmt") or {}).get(int(mnum)) or 0),
                "not_in_scenario": True,
            })
    # CREDITED P3 fleet: only the trucks needed to reach the LD target become
    # planned production. `free` (physical post-P1/P2 capacity) may be larger;
    # the difference is reported as unused/excess capacity below, never as
    # production and never deleted.
    free = mo.get("p3") or mo.get("free") or {}
    capacity_free = mo.get("free") or free
    days = float(mo.get("days") or _DAYS.get(int(mnum)) or 30)
    ld_planned_day = (float(mo.get("ld_t_month_planned") or 0) / days) if days else 0.0
    free_total = sum(float(v or 0) for v in free.values())
    for contractor, fdt in free.items():
        fdt = float(fdt or 0)
        if fdt <= 0.01:
            continue
        crs = [r for r in ld_pending if r["contractor"] == contractor]
        if not crs:
            crs = [{
                "key": "TF>HUAFEI", "src": "TF", "dst": "HUAFEI",
                "contractor": contractor, "material": "LIM", "otype": "LD",
                "prio": 3, "dt_before": 0.0, "dt_after": 0.0, "target": 0,
            }]
            ld_pending.extend(crs)
        tot_b = sum(r["dt_before"] for r in crs)
        share_c = (fdt / free_total) if free_total else 1.0 / max(1, len(free))
        for r in crs:
            if tot_b > 0:
                share = r["dt_before"] / tot_b
            else:
                share = 1.0 / len(crs)
            r["dt_after"] = fdt * share
            r["target"] = round(ld_planned_day * share_c * share) if ld_planned_day else 0
    for r in ld_pending:
        if r["dt_before"] > 0 or r["dt_after"] > 0:
            detail.append(r)
    detail = [d for d in detail if d["dt_before"] > 0 or d["dt_after"] > 0]
    if not detail:
        return None

    def _route_list(which):
        return [{"src": d["src"], "dst": d["dst"], "contractor": d["contractor"],
                 "dt": d[which], "key": d["key"]} for d in detail]

    before_list = _route_list("dt_before")
    after_list = _route_list("dt_after")
    pred_b, rows_b = ma._plan_predict_for_routes(before_list)
    pred_a, rows_a = ma._plan_predict_for_routes(after_list)
    sim_b, _ = ma._simulate_for_paths(before_list)
    sim_a, _ = ma._simulate_for_paths(after_list)
    sim_rows_b = (sim_b or {}).get("results") or []
    sim_rows_a = (sim_a or {}).get("results") or []

    def _achv(sim_rows, idx, d):
        if idx >= len(sim_rows):
            return None
        v = sim_rows[idx].get("achievable_production_t")
        return float(v) * 2 if v is not None else None

    achv_b = achv_a = 0.0
    alloc_rows = []
    for i, d in enumerate(detail):
        pr_b = rows_b[i] if i < len(rows_b) else {}
        pr_a = rows_a[i] if i < len(rows_a) else {}
        pb, pa = pr_b.get("wmt"), pr_a.get("wmt")
        tr_b, tr_a = pr_b.get("trips"), pr_a.get("trips")
        ab = _achv(sim_rows_b, i, d)
        aa = _achv(sim_rows_a, i, d)
        if pb is not None:
            achv_b += ab or 0
        if pa is not None:
            achv_a += aa or 0
        rid = "%s|%s>%s" % (d["contractor"], d["src"], d["dst"])
        alloc_rows.append({
            "id": rid, "key": d["key"], "contractor": d["contractor"],
            "material": d["material"], "otype": d["otype"], "prio": d["prio"],
            "target": round(d["target"]), "foreign": False,
            "dt_before": round(d["dt_before"]), "dt_after": round(d["dt_after"]),
            "pred_before": round(pb) if pb is not None else None,
            "pred_after": round(pa) if pa is not None else None,
            "achv_before": round(ab) if ab is not None else None,
            "achv_after": round(aa) if aa is not None else None,
            "achv_sim": round(aa) if aa is not None else None,
            "trips": round(tr_a) if tr_a is not None else None,
            "trips_before": round(tr_b) if tr_b is not None else None,
        })

    def _bucket(prio):
        rs = [r for r in alloc_rows if r.get("prio") == prio]
        if not rs:
            return {"n": 0, "target": 0, "dt_before": 0, "dt_after": 0,
                    "pred_before": 0, "pred_after": 0, "achv_before": 0, "achv_after": 0,
                    "achv_sim": 0}
        return {
            "n": len(rs),
            "target": sum(r.get("target") or 0 for r in rs),
            "dt_before": sum(r.get("dt_before") or 0 for r in rs),
            "dt_after": sum(r.get("dt_after") or 0 for r in rs),
            "pred_before": sum(r.get("pred_before") or 0 for r in rs),
            "pred_after": sum(r.get("pred_after") or 0 for r in rs),
            "achv_before": sum(r.get("achv_before") or 0 for r in rs),
            "achv_after": sum(r.get("achv_after") or 0 for r in rs),
            "achv_sim": sum(r.get("achv_sim") or 0 for r in rs),
        }

    fleet_b = sum(d["dt_before"] for d in detail)
    fleet_a = sum(d["dt_after"] for d in detail)
    buckets = {"sap": _bucket(1), "tos": _bucket(2), "ld": _bucket(3)}
    goals = {
        "sap": buckets["sap"]["target"],
        "tos": buckets["tos"]["target"],
        "ld": buckets["ld"]["target"],
        "total": sum(buckets[k]["target"] for k in buckets),
    }
    moves = _dt_moves(alloc_rows)
    # Two labelled numbers, never merged: what the free fleet COULD move on LD
    # and what is CREDITED against the target. `unused_dt` is the difference,
    # reported as excess capacity - it is not production and is not deleted.
    unused_dt = {c: round(max(0.0, float(capacity_free.get(c) or 0)
                              - float(free.get(c) or 0)), 1)
                 for c in capacity_free}
    capacity = {
        "ld_dt_capacity": round(sum(float(v or 0) for v in capacity_free.values()), 1),
        "ld_dt_credited": round(free_total, 1),
        "ld_dt_unused": round(sum(unused_dt.values()), 1),
        "unused_by_contractor": unused_dt,
        "ld_t_day_capacity": mo.get("ld_t_day_capacity"),
        "ld_t_day_credited": round(ld_planned_day),
        "ld_t_day_excess": round(max(0.0, float(mo.get("ld_t_day_capacity") or 0)
                                     - ld_planned_day)),
        "ld_t_month_capacity": mo.get("ld_t_month_capacity"),
        "ld_t_month_credited": mo.get("ld_t_month_planned"),
        "ld_t_month_excess": mo.get("ld_t_month_excess"),
        "ld_target_month": mo.get("ld_target_month"),
        "feasible": mo.get("feasible", True),
        "infeasible_dt": mo.get("infeasible_dt", 0),
        "note": "LIM-LD capacity is never clipped; the tonnage credited against "
                "the target stops at it and the rest is unused/excess capacity.",
    }
    return {
        "frozen": True,
        "horizon": "day",
        "old": {"pred": round(pred_b), "achv": round(achv_b), "dt": round(fleet_b)},
        "new": {"pred": round(pred_a), "achv": round(achv_a), "dt": round(fleet_a),
                "target": goals["total"]},
        "fleet": {"before": round(fleet_b), "after": round(fleet_a)},
        "goals": goals,
        "buckets": buckets,
        "rows": alloc_rows,
        "moves": moves,
        "capacity": capacity,
        "feasible": mo.get("feasible", True),
        "moved_total": sum(m.get("trucks") or 0 for m in moves),
        "notes": "Scenario %s — same layout as Monthly year Excel. "
                 "Old = yearly-matrix fleet. Optimized = P1 SAP → P2 LIM-TOS → P3 LIM-LD. "
                 "Routes this scenario does not run keep their trucks-out row but "
                 "carry no target."
                 % sc_id,
    }


def _scenario_month_card(sc, mo, year, mnum, yearly):
    import monthly_api as ma
    month = "%s-%02d" % (year, int(mnum))
    n = len(ma._days_in(month))
    raw = _build_synthetic_allocation(mo, yearly, mnum, sc["id"])
    if not raw:
        return None
    view = ma._alloc_view(raw, n, "scenario:%s" % sc["id"], include_detail=False)
    # ONE target owner (see _allocation_target_day): the card, the month sheet
    # and the Year total all read the frozen allocation's own row targets.
    tgt_day = _allocation_target_day(raw)
    return {
        "month": month,
        "name": ma._MONTH_LABELS.get(int(mnum), month),
        "n_days": n,
        "target_day": round(tgt_day),
        "target_month": round(tgt_day * n),
        "has_alloc": bool(view),
        "alloc": view,
        "alloc_raw": raw,
        "alloc_source": "scenario:%s" % sc["id"],
        "scenario": sc["id"],
        "feasible": mo.get("feasible", True),
        "synthetic": True,
    }


def _scenario_year_cards(sc, year):
    """Year board cards for S2+ from the waterfall (S1 uses saved Plan allocations)."""
    import monthly_api as ma
    yearly = _load_yearly()
    if not yearly:
        return None, "no yearly matrix loaded yet"
    res, err = waterfall(sc, yearly, ld_cap=ld_target_for_scenario(sc.get("id")))
    if err:
        return None, err
    mo_by = {mo["month"]: mo for mo in res.get("months") or []}
    mnums = set(int(m) for m in (yearly.get("months") or []))
    mnums.update(mo_by.keys())
    mnums.update(_MONTHS)
    cards, missing = [], []
    for mnum in sorted(mnums):
        if mnum < 1 or mnum > 12:
            continue
        # August is S1 only. S3/S4 (and any later imported scenario) start
        # at September — do not fill Aug from S1 into the workbook.
        if sc["id"] != "S1" and mnum == 8:
            continue
        mo = mo_by.get(mnum)
        if not mo:
            missing.append(ma._MONTH_LABELS.get(mnum, str(mnum)))
            continue
        card = _scenario_month_card(sc, mo, year, mnum, yearly)
        if card:
            cards.append(card)
        else:
            missing.append(ma._MONTH_LABELS.get(mnum, str(mnum)))
    if not cards:
        return None, "no months to export for %s" % sc.get("id")
    # ONE workbook = ONE scenario, and it says which one on its Year sheet.
    for c in cards:
        c["_source_note"] = (
            "Scenario %s — every month on this sheet is the %s waterfall on the "
            "yearly-matrix fleet. No other scenario's plan appears anywhere in "
            "this workbook." % (sc["id"], sc["id"]))
        c["_missing_months"] = list(missing)
        c["_scenario"] = sc["id"]
    return cards, None


def _scenario_results_for_export(ids=None):
    """(results, skipped). `ids` filters; None means every waterfall scenario.

    Returns SKIPPED ids as well, because this sheet can only ever show
    waterfall runs. A day-convention scenario like S4 has no targets file to
    run a waterfall on, so asking for it here used to be dropped in silence
    and `?ids=S3,S4` returned the same workbook as no arguments at all. Naming
    the skip — and where the scenario CAN be had — is the difference between a
    gap and a lie.
    """
    known = ["S1"] + [s for s in _scenario_ids() if s != "S1"]
    want = list(ids) if ids else known
    results, skipped = [], []
    for sid in want:
        sc = _load_scenario(sid)
        if not sc:
            day = _day_for_scenario_id(sid)
            skipped.append((sid, (
                "day-%02d saved-plan scenario: no waterfall targets file, so it "
                "cannot appear on a Compare sheet. Export it from "
                "/api/scenarios/export-full?id=%s or "
                "/api/monthly/export-year?day=%d, or read the saved_day_plans "
                "block of /api/scenarios/compare?ids=%s" % (day, sid, day, sid))
                if day else "no such scenario"))
            continue
        res, err = waterfall(sc, ld_cap=ld_target_for_scenario(sc.get("id")))
        if err:
            skipped.append((sid, err))
            continue
        results.append(res)
    return results, skipped


def _write_compare_sheet(ws, results):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(bold=True, color="FFFFFF")
    num = "#,##0"
    ws.append(["Mine-plan scenarios - same fleet, different allocation"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Waterfall: P1 SAP -> P2 LIM-TOS -> P3 LIM-LD (Tofu dump -> Huafei). "
               "LIM-LD capacity is never clipped; %s t is the sales TARGET the "
               "credited tonnage stops at, and the rest is excess capacity. "
               "BLB accepts RIM only. Fleet = yearly matrix DT per contractor."
               % format(LIM_LD_TARGET_T, ",")])
    ws.append([])
    hdr = ["", ""]
    for r in results:
        hdr += [r["label"], "", "", "", ""]
    ws.append(hdr)
    sub = ["Month", "Fleet (RIM+SMA)"]
    for _ in results:
        sub += ["P1 SAP DT", "P2 LIM-TOS DT", "P3 DT credited",
                "LIM-LD credited t", "LIM-LD capacity t"]
    ws.append(sub)
    for c in range(1, len(sub) + 1):
        ws.cell(row=5, column=c).fill = head_fill
        ws.cell(row=5, column=c).font = head_font
        top = ws.cell(row=4, column=c)
        if top.value:
            top.font = bold
    if not results:
        return
    for i, mo0 in enumerate(results[0]["months"]):
        m = mo0["month"]
        pool = mo0["pool"]
        row = [_MONN.get(m, m), "%d (%d+%d)" % (sum(pool.values()),
                                                pool.get("RIM", 0), pool.get("SMA", 0))]
        for r in results:
            mo = r["months"][i] if i < len(r["months"]) else {}
            row += [round(mo.get("dt_p1", 0)), round(mo.get("dt_p2", 0)),
                    round(mo.get("dt_p3", 0)),
                    mo.get("ld_t_month_planned", 0),
                    mo.get("ld_t_month_capacity", 0)]
        ws.append(row)
    tot = ["Total", ""]
    for r in results:
        t = r["total"]
        tot += ["", "", "", t["ld_t_planned"], ""]
    ws.append(tot)
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([])
    for label, key in (("SAP credited (t)", "sap_t"),
                       ("SAP target asked (t)", "sap_t_target"),
                       ("LIM-TOS credited (t)", "limtos_t"),
                       ("LIM-TOS target asked (t)", "limtos_t_target"),
                       ("LIM-LD credited (t)", "ld_t_planned"),
                       ("LIM-LD capacity, never clipped (t)", "ld_t_capacity"),
                       ("LIM-LD excess capacity above target (t)", "ld_t_excess_capacity"),
                       ("LD sales target met", "ld_cap_reached"),
                       ("Short of target by (t)", "ld_shortfall_t"),
                       ("Targets fit the fleet", "feasible")):
        row = [label, ""]
        for r in results:
            v = r["total"].get(key)
            if key in ("ld_cap_reached", "feasible"):
                v = "YES" if v else "no"
            row += ["", "", "", v, ""]
        ws.append(row)
        ws.cell(row=ws.max_row, column=1).font = bold
    for row in ws.iter_rows(min_row=6):
        for cell in row:
            if isinstance(cell.value, (int, float)) and abs(cell.value) >= 1000:
                cell.number_format = num
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    for c in range(3, len(sub) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13


def _scenario_cards_for_excel(sc, year):
    """Cards for one scenario workbook. S1 prefers saved Plan Allocate snapshots
    (identical to Monthly ⬇ Download Excel). S2+ always use the waterfall."""
    import monthly_api as ma
    if sc["id"] == "S1":
        _, cards = ma._year_cards(year)
        if any(c.get("has_alloc") for c in (cards or [])):
            return cards, None
    return _scenario_year_cards(sc, year)


def _scenario_year_book(sc, year, achv=False):
    """One workbook = Monthly year Excel: Year + one sheet per month.
    S1 is Aug–Dec. S3/S4 start at September (no August sheet).
    achv=True adds the engine's achievable next to every predicted figure."""
    import monthly_api as ma
    cards, err = _scenario_cards_for_excel(sc, year)
    if err:
        return None, err
    if not cards:
        return None, "no months to export for %s" % sc.get("id")
    lab = sc.get("id") or ""
    disp = SCENARIO_DISPLAY.get(lab)
    if disp:
        lab = "%s (%s)" % (lab, disp)
    return ma._xlsx_year_book(year, cards, achv=achv, scenario_label=lab), None


def _export_filename(year, sid):
    if sid == "S1":
        return "monthly_plan_%s.xlsx" % year
    return "monthly_plan_%s_%s.xlsx" % (year, sid)


def _exportable_scenario_ids():
    """Every scenario the app OFFERS, not every scenario that has a file.

    A scenario reaches the user two different ways and this function is the
    only place that has to know both:

      * waterfall imports  -> data/scenarios/<id>.json  (S1 derived, S3, ...)
      * day-of-month saves -> data/saved_plans/YYYY-MM-<day>.json

    S4 is the second kind and always will be (see _day_for_scenario_id): it is
    S3 with the leftover TF LD trucks split 50/50, expressed as day-04 Plan-tab
    saves. Listing only the files is why `export-full` shipped a zip with S1
    and S3 in it and `?id=S4` returned 404, while /monthly's own
    `export-year?day=4` built the S4 workbook perfectly well.
    """
    ids = ["S1"] + [s for s in _scenario_ids() if s != "S1"]
    for sid in _saved_day_scenario_ids():
        if sid not in ids:
            ids.append(sid)
    return ids


# The day-of-month scenario convention, spelled out. ONE owner for it.
#   01 = S1   03 = S3 (3.0.1)   04 = 3.0.2   05 = S5 (3.1.1)   06 = 3.1.2
#   (02 is reserved: S2 deleted 2026-08-21)
# It is a CLOSED list on purpose. Deriving it from whatever days happen to have
# saves turns the legacy August dailies (2026-08-05 / -07 / -13, which predate
# the convention and are plain daily plans) into phantom scenarios S5/S7/S13 —
# measured, all three appeared. Adding a scenario day must be a deliberate edit
# here, not an accident of someone saving a plan on the 9th.
_DAY_SCENARIOS = {1: "S1", 3: "S3", 4: "S4", 5: "S5", 6: "S6", 7: "S7"}
# Planning-team names (2026-08-26): mining plan 3.0 = as-is, 3.1 = +1 Mt BLB
# LIM Oct-Dec; hauling .1 = all leftovers HUAFEI/BSE, .2 = half to POS 6 from
# October. The day convention stays the machine key; these are the labels.
SCENARIO_DISPLAY = {"S3": "3.0.1", "S4": "3.0.2", "S5": "3.1.1", "S6": "3.1.2",
                    "S7": "4.1"}


def _saved_day_scenario_ids():
    """Convention scenario ids that actually have saves on their day."""
    import monthly_api as ma
    out = []
    d = getattr(ma, "_SAVED_DIR", None)
    if not d or not os.path.isdir(d):
        return out
    days = set()
    for name in sorted(os.listdir(d)):
        if not name.endswith(".json"):
            continue
        m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", name[:-5])
        if m:
            days.add(int(m.group(3)))
    for day in sorted(_DAY_SCENARIOS):
        if day in days:
            out.append(_DAY_SCENARIOS[day])
    return out


def _year_book_for_scenario(sid, year, achv=False):
    """The workbook for `sid`, whichever kind of scenario it is.

    Returns (workbook, error). A day-convention scenario is built from the
    SAME cards /api/monthly/export-year?day=N uses, so the two endpoints can
    never disagree about what S4 contains.
    """
    import monthly_api as ma
    sc = _load_scenario(sid)
    if sc:
        return _scenario_year_book(sc, year, achv=achv)
    day = _day_for_scenario_id(sid)
    if not day:
        return None, "no such scenario"
    _yearly, cards = ma._year_cards(year, day=day)
    if not cards:
        return None, ("no day-%02d plans stored for %s — %s is the day-%02d "
                      "saved-plan convention, so it needs saves on that day"
                      % (day, year, sid, day))
    lab = sid
    disp = SCENARIO_DISPLAY.get(sid)
    if disp:
        lab = "%s (%s)" % (sid, disp)
    return ma._xlsx_year_book(year, cards, achv=achv, scenario_label=lab), None


def _xlsx_all_scenarios_zip(year, achv=False, ids=None):
    """Zip of one monthly_plan workbook per scenario — same files as Year board Excel."""
    import zipfile
    ids = list(ids) if ids else _exportable_scenario_ids()
    buf = io.BytesIO()
    written, errors = [], []
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for sid in ids:
            wb, err = _year_book_for_scenario(sid, year, achv=achv)
            if err or wb is None:
                errors.append("%s: %s" % (sid, err or "no workbook"))
                continue
            inner = io.BytesIO()
            wb.save(inner)
            name = _export_filename(year, sid)
            zf.writestr(name, inner.getvalue())
            written.append(name)
        if written and errors:
            # Say what is NOT in the zip. A member silently missing is how the
            # S4 gap survived: the download succeeded, so nothing looked wrong.
            zf.writestr("_MISSING.txt",
                        "These scenarios were requested but could not be "
                        "built:\n" + "\n".join(errors) + "\n")
    if not written:
        return None, ("no scenario year sheets could be built — load the matrix "
                      "and import scenarios (%s)" % "; ".join(errors))
    buf.seek(0)
    return buf, None


# ---------------------------------------------------------------- routes

@bp.route("/api/scenarios")
def api_scenarios():
    ids = ["S1"] + [s for s in _scenario_ids() if s != "S1"]
    out = []
    for sid in ids:
        sc = _load_scenario(sid)
        if sc:
            out.append({"id": sc["id"], "label": sc.get("label") or sc["id"],
                        "source": sc.get("source"), "derived": bool(sc.get("derived")),
                        "loaded_at": sc.get("loaded_at")})
    return jsonify({"ok": True, "scenarios": out})


@bp.route("/api/scenarios/<sid>")
def api_scenario_one(sid):
    sc = _load_scenario(_safe_id(sid))
    if not sc:
        return jsonify({"ok": False, "error": "no such scenario"}), 404
    return jsonify({"ok": True, "scenario": sc})


@bp.route("/api/scenarios/<sid>", methods=["DELETE"])
def api_scenario_delete(sid):
    sid = _safe_id(sid)
    if sid == "S1":
        return jsonify({"ok": False, "error": "S1 is the live plan - it cannot be deleted"}), 400
    p = _scen_path(sid)
    if not p or not os.path.isfile(p):
        return jsonify({"ok": False, "error": "no such scenario"}), 404
    os.remove(p)
    return jsonify({"ok": True, "deleted": sid})


@bp.route("/api/scenarios/import", methods=["POST"])
def api_scenarios_import():
    """Upload the mine-plan workbook; every 'Scenario N' sheet-row group becomes a scenario."""
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "upload the mine-plan .xlsx"}), 400
    data = f.read()
    if not data:
        return jsonify({"ok": False, "error": "the file is empty"}), 400
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return jsonify({"ok": False, "error": "could not read that file as .xlsx"}), 400
    ws = None
    for name in wb.sheetnames:
        if "plan db" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    scens, err = _parse_mine_plan_db(rows, f.filename)
    if err:
        return jsonify({"ok": False, "error": err}), 400
    saved = []
    for sc in scens:
        if sc["id"] == "S1":
            continue  # never overwrite the live plan
        _save_scenario(sc)
        saved.append(sc["id"])
    return jsonify({"ok": True, "imported": saved,
                    "skipped_s1": any(s["id"] == "S1" for s in scens)})


@bp.route("/api/scenarios/<sid>/allocate")
def api_scenario_allocate(sid):
    sc = _load_scenario(_safe_id(sid))
    if not sc:
        return jsonify({"ok": False, "error": "no such scenario"}), 404
    res, err = waterfall(sc, ld_cap=ld_target_for_scenario(sc.get("id")))
    if err:
        return jsonify({"ok": False, "error": err}), 400
    return jsonify({"ok": True, "allocation": res})


def _day_for_scenario_id(sid):
    """S1 -> day 01, S3 -> 03, S4 -> 04, S5 -> 05, S6 -> 06 (saved-plan
    convention). Planning-team names: S3=3.0.1, S4=3.0.2, S5=3.1.1, S6=3.1.2.

    S4/S6 have no data/scenarios/*.json and never will: they are not mine-plan
    imports, they are the day-04/06 Plan-tab saves (identical to S3/S5 except
    the TF LD trucks split 50/50 HUAFEI/BSE vs POS 6 from October - owner
    2026-08-25, planning team 2026-08-26). Day 02 is reserved - S2 was
    deleted from the app on 2026-08-21 - so it maps to nothing.
    """
    m = re.fullmatch(r"S(\d{1,2})", str(sid or "").upper())
    if not m:
        return None
    day = int(m.group(1))
    if day == 2 or not 1 <= day <= 28:
        return None
    return day


def _saved_day_summary(sid, day, year):
    """The saved-plan side of a day-convention scenario, on its OWN basis.

    This is NOT a waterfall run and is never presented as one: it reads the
    frozen Plan-tab allocations at data/saved_plans/YYYY-MM-{day}.json, the
    same source the Year board and the year workbook use. Nothing is
    fabricated - a month with no save on that day simply is not here.
    """
    import monthly_api as ma
    _yearly, cards = ma._year_cards(year, day=day)
    months = []
    for c in cards or []:
        a = c.get("alloc") or {}
        if not a:
            continue
        months.append({
            "month": c["month"], "name": c.get("name"),
            "source_date": a.get("source_date"),
            "n_days": c.get("n_days"),
            "target_month": a.get("target_month"),
            "new_pred_month": a.get("new_pred_month"),
            "new_achv_month": a.get("new_achv_month"),
            "dt_after": a.get("dt_after"),
            "materials": {k: {"target_month": (v or {}).get("target_month"),
                              "pred_after_month": (v or {}).get("pred_after_month")}
                          for k, v in (a.get("materials") or {}).items()},
        })
    if not months:
        return None
    tot = ma._year_alloc_totals(cards)
    return {
        "id": sid, "day": day, "year": year,
        "label": "%s - saved Plan-tab allocations, day-%02d convention" % (sid, day),
        "basis": "saved_plans/%s-MM-%02d.json (frozen Allocate snapshots)" % (year, day),
        "months": months,
        "total": {"target": (tot or {}).get("target"),
                  "new_pred": (tot or {}).get("new_pred"),
                  "new_achv_raw": (tot or {}).get("new_achv_raw"),
                  "n_months": (tot or {}).get("n")},
        "note": "Target and predicted tonnage from frozen saves. This is a "
                "different clock from the waterfall `scenarios` block above - "
                "compare like with like.",
    }


@bp.route("/api/scenarios/compare")
def api_scenarios_compare():
    """Compare scenarios. `ids=S3,S4` works.

    Two kinds of "scenario" exist in this app and they are kept apart on
    purpose (the two-panels-one-concept lesson):
      * `scenarios`      - waterfall runs of data/scenarios/{id}.json (S1, S3).
      * `saved_day_plans`- the day-of-month saved-plan convention (01=S1,
        03=S3, 04=S4), read from frozen Plan-tab Allocate snapshots.
    S4 lives only in the second block; asking for it used to return
    ok:true + errors:[no such scenario], which read as success and left the
    owner's natural S3-vs-S4 comparison unreachable. Unknown ids now fail
    loudly (ok:false) and say where to look.
    """
    year = (request.args.get("year") or str(datetime.utcnow().year)).strip()
    if not re.fullmatch(r"\d{4}", year):
        return jsonify({"ok": False, "error": "year=YYYY"}), 400
    ids = request.args.get("ids")
    explicit = bool(ids)
    ids = [_safe_id(s) for s in ids.split(",")] if ids else (["S1"] + _scenario_ids())
    seen, out, errors, saved_days = set(), [], [], []
    for sid in ids:
        if not sid or sid in seen:
            continue
        seen.add(sid)
        sc = _load_scenario(sid)
        day = _day_for_scenario_id(sid)
        if sc:
            res, err = waterfall(sc, ld_cap=ld_target_for_scenario(sc.get("id")))
            if err:
                errors.append({"id": sid, "error": err})
            else:
                out.append(res)
        if explicit and day is not None:
            summary = _saved_day_summary(sid, day, year)
            if summary:
                saved_days.append(summary)
            elif not sc:
                errors.append({
                    "id": sid, "day": day,
                    "error": "no scenario file and no day-%02d saved plans for %s"
                             % (day, year),
                    "where": "a scenario lives in data/scenarios/%s.json; the "
                             "day-%02d convention lives in "
                             "data/saved_plans/%s-MM-%02d.json (Year board "
                             "/api/monthly/year-board?year=%s&day=%d)"
                             % (sid, day, year, day, year, day),
                })
        elif not sc:
            errors.append({
                "id": sid,
                "error": "no such scenario",
                "where": "scenarios are files in data/scenarios/*.json; S4 is not "
                         "one of them - it is the day-04 saved-plan convention "
                         "(data/saved_plans/YYYY-MM-04.json). Ask for it with "
                         "?ids=...,S4 and read the saved_day_plans block.",
            })
    return jsonify({"ok": not errors, "scenarios": out,
                    "saved_day_plans": saved_days, "errors": errors,
                    "year": year,
                    "ld_target": LIM_LD_TARGET_T,
                    "ld_cap": LIM_LD_CAP_T,
                    "note": "`scenarios` are waterfall runs; `saved_day_plans` are "
                            "frozen Plan-tab saves under the day-of-month "
                            "convention (01=S1, 03=S3, 04=S4). Two clocks, never "
                            "merged. LIM-LD capacity is reported in full; only the "
                            "credited tonnage stops at the %s t target."
                            % format(LIM_LD_TARGET_T, ",")})


_MONN = {8: "Aug", 9: "Sept", 10: "Oct", 11: "Nov", 12: "Dec"}


@bp.route("/api/scenarios/export")
def api_scenarios_export():
    """One workbook: Compare sheet + a full-allocation sheet per scenario."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from flask import send_file

    want = (request.args.get("ids") or "").strip()
    ids = [_safe_id(x) for x in want.split(",") if _safe_id(x)] if want else None
    results, skipped = _scenario_results_for_export(ids)
    if not results:
        return jsonify({
            "ok": False,
            "error": "no scenarios to export",
            "skipped": [{"id": s, "why": w} for s, w in skipped],
        }), 404

    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(bold=True, color="FFFFFF")
    num = "#,##0"

    wb = Workbook()
    ws = wb.active
    ws.title = "Compare"
    _write_compare_sheet(ws, results)
    if skipped:
        # In the workbook, not just the log. A requested scenario missing from
        # a file that opened fine reads as "there was nothing to show".
        sk = wb.create_sheet("Not in this file")
        sk.append(["Requested but NOT included"])
        sk["A1"].font = Font(bold=True, size=13)
        sk.append([])
        sk.append(["Scenario", "Why", ])
        for c in sk[3]:
            c.font = head_font
            c.fill = head_fill
        for sid, why in skipped:
            sk.append([sid, why])
        sk.column_dimensions["A"].width = 14
        sk.column_dimensions["B"].width = 120

    for r in results:
        d = wb.create_sheet(r["id"][:28])
        d.append([r["label"] + " - full DT allocation"])
        d["A1"].font = Font(bold=True, size=13)
        if r["months_filled_from_s1"]:
            d.append(["Months taken from S1 (not in this scenario's file): "
                      + ", ".join(_MONN.get(m, str(m)) for m in r["months_filled_from_s1"])])
        d.append([])
        d.append(["Credited = tonnage counted against the target. Capacity = what the "
                  "fleet on this row could move, never clipped. On P1/P2 they differ "
                  "only when a target outran the pool (Feasible = no)."])
        d.append([])
        # ONE P3 LIM-LD row per month (gate J72 counts them). The capacity /
        # credited pair lives in its own two columns rather than a second row,
        # so "t/day" means the same thing on every row of the sheet.
        cols = ["Month", "Priority", "Pit", "Material", "Destination", "Contractor",
                "t/day credited", "DT credited", "t/DT/day",
                "t/day capacity", "DT capacity", "Feasible"]
        d.append(cols)
        hr = d.max_row
        for c in range(1, len(cols) + 1):
            d.cell(row=hr, column=c).fill = head_fill
            d.cell(row=hr, column=c).font = head_font
        for mo in r["months"]:
            mn = _MONN.get(mo["month"], mo["month"])
            for a in mo["rows"]:
                d.append([mn, "P%d" % a["prio"], a["pit"],
                          a["mat"] + ("-" + a["otype"] if a["otype"] not in ("", "TOS") else ""),
                          a["dest"], a["contractor"],
                          a.get("wmt_day_credited", a["wmt_day"]), a["dt"],
                          a["rate_t_dt_day"],
                          a["wmt_day"], a.get("dt_target", a["dt"]),
                          "yes" if a.get("feasible", True) else "NO"])
            free_txt = " + ".join("%s %d" % (c, round(v)) for c, v in mo["free"].items())
            d.append([mn, "P3", "TOFU", "LIM-LD", "HUAFEI", free_txt,
                      mo.get("ld_t_day_planned", 0), round(mo["dt_p3"], 1), "",
                      mo["ld_t_day_capacity"], round(mo.get("dt_p3_capacity", 0), 1),
                      "yes"])
            d.cell(row=d.max_row, column=2).font = bold
            if mo.get("dt_p3_unused"):
                d.append([mn, "", "TOFU", "LIM-LD", "UNUSED CAPACITY",
                          " + ".join("%s %d" % (c, round(v))
                                     for c, v in (mo.get("unused") or {}).items() if v),
                          "", "", "",
                          round((mo.get("ld_t_month_excess") or 0) / (mo["days"] or 1)),
                          round(mo.get("dt_p3_unused", 0), 1),
                          "capacity above target — not credited"])
            for l in mo["lends"]:
                d.append([mn, "", "", "lend", "", "%s -> %s work" % (l["from"], l["to_work_of"]),
                          "", l["dt"], l["note"]])
            for df in mo["deficit"]:
                d.append([mn, "", df["pit"], df["mat"], "DEFICIT", "", df.get("wmt_day") or "",
                          "", df["why"]])
        for row in d.iter_rows(min_row=hr + 1):
            for cell in row:
                if isinstance(cell.value, (int, float)) and abs(cell.value) >= 1000:
                    cell.number_format = num
        widths = [8, 9, 8, 10, 16, 22, 13, 12, 10, 13, 12, 40]
        for c, w in enumerate(widths, start=1):
            d.column_dimensions[get_column_letter(c)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name = "scenario_compare_%s.xlsx" % datetime.utcnow().strftime("%Y%m%d")
    return send_file(buf, as_attachment=True, download_name=name,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/api/scenarios/export-full")
def api_scenarios_export_full():
    """One Monthly-style year workbook per scenario.
    S1 = Year + Aug–Dec; S3/S4 = Year + Sep–Dec (August is S1 only).

    ?id=S2 → a single xlsx named like monthly_plan_2026_S2.xlsx
    no id → a zip of every scenario, S1 named monthly_plan_2026.xlsx
    """
    from flask import send_file
    import monthly_api as ma
    year = (request.args.get("year") or str(datetime.utcnow().year)).strip()
    if not re.fullmatch(r"\d{4}", year):
        return jsonify({"ok": False, "error": "year=YYYY"}), 400
    achv = (request.args.get("achv") or "").strip() in ("1", "true", "yes")
    sid = _safe_id(request.args.get("id") or "")
    if sid:
        # _year_book_for_scenario handles BOTH kinds of scenario, so S4 (which
        # has no data/scenarios/S4.json and never will) exports here exactly as
        # it does from /api/monthly/export-year?day=4 instead of 404ing.
        wb, err = _year_book_for_scenario(sid, year, achv=achv)
        if err or wb is None:
            return jsonify({"ok": False, "error": err or "no such scenario",
                            "known": _exportable_scenario_ids()}), 404
        name = _export_filename(year, sid)
        if achv:
            name = name.replace(".xlsx", "_achievable.xlsx")
        return ma._xlsx_send(wb, name)
    want = (request.args.get("ids") or "").strip()
    ids = [_safe_id(x) for x in want.split(",") if _safe_id(x)] if want else None
    buf, err = _xlsx_all_scenarios_zip(year, achv=achv, ids=ids)
    if err:
        return jsonify({"ok": False, "error": err}), 404
    zname = "monthly_plan_%s_all_scenarios%s.zip" % (year, "_achievable" if achv else "")
    return send_file(buf, as_attachment=True,
                     download_name=zname,
                     mimetype="application/zip")


# ------------------------------------------------- Plan-tab draft plans

def _scenario_draft_paths(sc, mnum, yearly):
    """Plan-tab `paths` dict for one scenario month, sized by the REAL path
    model: P1 SAP, P2 LIM-TOS and the supplied P3 LIM-LD target are sized in
    strict order.

    The LD rows carry the CREDITED fleet; whatever is left over is returned as
    `unused` (and stored on the plan's meta as `unused_fleet`) so the capacity
    stays visible without being credited as production. BLB stays RIM-only
    because targets only land on matrix routes, which are RIM there."""
    import monthly_api as ma
    res, err = waterfall(sc, yearly, ld_cap=ld_target_for_scenario(sc.get("id")))
    if err:
        return None, err
    mo = next((m for m in res["months"] if m["month"] == int(mnum)), None)
    if not mo:
        return None, "the scenario has no month %s" % mnum
    path_models, fleet, contr_by = ma._path_model_context()
    pool = {c: float(v) for c, v in (mo.get("pool") or {}).items()}
    wf_rows = mo.get("rows") or []
    p3_plan = mo.get("p3") or mo.get("free") or {}
    ld_rate = {}
    for r in _yearly_rows(yearly):
        if r["mat"] == "LIM" and r["otype"] == "LD":
            rt = _route_rate(r, 11)
            if rt:
                ld_rate[r["contractor"]] = rt

    # ---- size against the FINAL shared-corridor fleet, not the seed ------
    # The P3 LIM-LD block lands on LD_ROUTE_KEY after the targeted rows are
    # sized, so sizing each target against the waterfall's DT alone priced it
    # on a road that later got hundreds more trucks: Sep S3 sized TF>HUAFEI
    # LIM-TOS at 53 SMA DT for a 5,533 t/day target, then 215 LD trucks joined
    # the same key and the row delivered 3,566 t (64%). Same trap as the
    # 2026-08-21 finalTrim pass in plan_sap_target.js. Iterate to a fixed
    # point: everyone is priced on the fleet everyone ends up sharing.
    sized = {i: float(a["dt"]) for i, a in enumerate(wf_rows)}
    ld_fleet = sum(float(v or 0) for v in p3_plan.values())
    order = sorted(range(len(wf_rows)),
                   key=lambda i: (wf_rows[i]["prio"], wf_rows[i]["pit"],
                                  wf_rows[i]["dest"]))
    used, warnings = {c: 0.0 for c in pool}, []
    for _pass in range(5):
        combined = {}
        for i, a in enumerate(wf_rows):
            rk = _wf_route_key(a)
            k = "%s>%s" % (rk[0], rk[1])
            combined[k] = combined.get(k, 0.0) + sized[i]
        combined[LD_ROUTE_KEY] = combined.get(LD_ROUTE_KEY, 0.0) + ld_fleet
        used = {c: 0.0 for c in pool}
        warnings, nxt = [], {}
        for i in order:
            a = wf_rows[i]
            src, dst, contractor, mat, otype = _wf_route_key(a)
            key = "%s>%s" % (src, dst)
            target = float(a["wmt_day"] or 0)
            others = max(0.0, (combined.get(key) or 0) - sized[i])
            req, why = ma._required_dt_day(src, dst, contractor, target, others,
                                           path_models, fleet, contr_by)
            dt = float(req) if req else float(a["dt"])
            if why:
                warnings.append("%s %s (%s): %s - kept waterfall %s DT"
                                % (key, mat, contractor, why, round(a["dt"])))
            avail = pool.get(contractor, 0.0) - used.get(contractor, 0.0)
            if dt > avail:
                warnings.append("%s %s (%s): needs %d DT, only %d left - clipped"
                                % (key, mat, contractor, round(dt), round(avail)))
                dt = max(0.0, avail)
            nxt[i] = dt
            used[contractor] = used.get(contractor, 0.0) + dt
        ld_next = sum(min(max(0.0, pool.get(c, 0.0) - used.get(c, 0.0)),
                          float(p3_plan.get(c) or 0.0)) for c in pool)
        settled = (all(abs(nxt[i] - sized[i]) <= 0.5 for i in nxt)
                   and abs(ld_next - ld_fleet) <= 0.5)
        sized, ld_fleet = nxt, ld_next
        if settled:
            break

    paths = {}
    for i, a in enumerate(wf_rows):
        src, dst, contractor, mat, otype = _wf_route_key(a)
        key = "%s>%s" % (src, dst)
        slot = "%s|%s" % (contractor, key)
        if mat == "LIM":
            slot += "|LIM|%s" % (otype or "TOS")
        paths[slot] = {
            "key": key, "dt": int(round(sized[i])), "contractor": contractor,
            "source": src, "dest": dst, "material": mat,
            "otype": (otype or "TOS") if mat == "LIM" else "",
            "targetWmt": int(round(float(a["wmt_day"] or 0))),
            "_targetManual": mat == "LIM",
        }
    unused, capacity = {}, {}
    for contractor, have in sorted(pool.items()):
        capacity_dt = max(0.0, have - used.get(contractor, 0.0))
        capacity[contractor] = capacity_dt
        free = min(capacity_dt, float(p3_plan.get(contractor) or 0.0))
        unused[contractor] = max(0.0, capacity_dt - free)
        if free < 0.5:
            continue
        rate = ld_rate.get(contractor, 120.0 if contractor == "RIM" else 100.0)
        paths["%s|%s|LIM|LD" % (contractor, LD_ROUTE_KEY)] = {
            "key": LD_ROUTE_KEY, "dt": int(round(free)), "contractor": contractor,
            "source": LD_ROUTE_KEY.split(">")[0], "dest": LD_ROUTE_KEY.split(">")[1],
            "material": "LIM", "otype": "LD",
            "targetWmt": int(round(free * rate)),
            "_targetManual": True,
        }
    if not paths:
        return None, "no routes with targets for month %s" % mnum
    return {"paths": paths, "warnings": warnings,
            "pool": {c: round(v) for c, v in pool.items()},
            "used": {c: round(v, 1) for c, v in used.items()},
            # P3 capacity (never clipped) vs the part credited to the LD rows.
            "p3_capacity": {c: round(v, 1) for c, v in capacity.items()},
            "unused": {c: round(v, 1) for c, v in unused.items()}}, None


@bp.route("/api/scenarios/<sid>/draft-plans", methods=["POST"])
def api_scenario_draft_plans(sid):
    """Write Plan-tab saved plans for a scenario, one per month on the given
    day. Body: {year: 2026, day: 2, months: [9,10,11,12], overwrite: false}.
    They appear in the Plan tab's saved list: open, Check capacity, Allocate
    DT, Save. An existing date is never overwritten unless overwrite=true."""
    sid = _safe_id(sid)
    sc = _load_scenario(sid)
    if not sc:
        return jsonify({"ok": False, "error": "no such scenario"}), 404
    body = request.get_json(silent=True) or {}
    year = str(body.get("year") or datetime.utcnow().year)
    if not re.fullmatch(r"\d{4}", year):
        return jsonify({"ok": False, "error": "year=YYYY"}), 400
    try:
        day = int(body.get("day") or 2)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "day must be 1-28"}), 400
    if not 1 <= day <= 28:
        return jsonify({"ok": False, "error": "day must be 1-28"}), 400
    months = body.get("months") or [9, 10, 11, 12]
    overwrite = bool(body.get("overwrite"))
    yearly = _load_yearly()
    if not yearly:
        return jsonify({"ok": False, "error": "no yearly matrix loaded yet"}), 404
    saved_dir = os.path.join(_ROOT, "data", "saved_plans")
    os.makedirs(saved_dir, exist_ok=True)
    out, errors = [], []
    for m in months:
        try:
            m = int(m)
        except (TypeError, ValueError):
            errors.append({"month": m, "error": "bad month"})
            continue
        date_s = "%s-%02d-%02d" % (year, m, day)
        fp = os.path.join(saved_dir, date_s + ".json")
        if os.path.isfile(fp) and not overwrite:
            errors.append({"month": m, "date": date_s,
                           "error": "a plan already exists on this date (set overwrite=true)"})
            continue
        draft, err = _scenario_draft_paths(sc, m, yearly)
        if err:
            errors.append({"month": m, "date": date_s, "error": err})
            continue
        plan = {
            "date": date_s,
            "paths": draft["paths"],
            "rain_mm": 0,
            "hours": 12,
            "wb": None,
            "meta": {"scenario": sid,
                     "p3_capacity_fleet": draft.get("p3_capacity") or {},
                     "unused_fleet": draft.get("unused") or {},
                     "note": "draft from %s waterfall - open on the Plan tab, "
                             "Check capacity, Allocate DT, Save. LD rows carry "
                             "the fleet credited against the target; "
                             "unused_fleet is capacity above it." % sid,
                     "generated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")},
            "saved_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        }
        tmp = fp + ".tmp"
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(plan, fh, indent=2)
            fh.write("\n")
        os.replace(tmp, fp)
        out.append({"month": m, "date": date_s,
                    "routes": len(draft["paths"]),
                    "dt_total": sum(p["dt"] for p in draft["paths"].values()),
                    "pool": draft["pool"],
                    "p3_capacity": draft.get("p3_capacity") or {},
                    "unused": draft.get("unused") or {},
                    "warnings": draft["warnings"]})
    return jsonify({"ok": bool(out), "scenario": sid, "written": out, "errors": errors})
