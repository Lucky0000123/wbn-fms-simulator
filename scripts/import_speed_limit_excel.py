"""Import FMS posted speed-limit zones from the Safety Excel workbook.

Writes data/speed_limit_zones_public.csv (chainage + limit only — no tonnages).

Stick rule: TOFU + KAORAHAI use the TF→FENI corridor chainage and paint on
the Tab 1 stick (onStick=1). BLB / BIRI BIRI use spur-local KM and must NOT be
drawn on the main stick by raw km (onStick=0, road=BLB|BB).

Default source (override with SPEED_LIMIT_XLSX=...):
  /Volumes/LUCKY_SSD/WORK_WBN/WORK/FMS/Safety/FMS_Speed_Limit_Geofencing_Professional.xlsx
"""
from __future__ import annotations

import csv
import os
import re
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data", "speed_limit_zones_public.csv")
DEFAULT_XLSX = (
    "/Volumes/LUCKY_SSD/WORK_WBN/WORK/FMS/Safety/"
    "FMS_Speed_Limit_Geofencing_Professional.xlsx"
)

# Operating Area → (region code for stick, road tag, onStick)
_AREA = {
    "TOFU": ("TF", "TOFU", 1),
    "KAORAHAI": ("KR", "KR", 1),
    "KAO RAHAI": ("KR", "KR", 1),
    "BUKET LIMBER": ("BLB", "BLB", 0),
    "BIRI BIRI": ("BB", "BB", 0),
}


def parse_chainage(text):
    """'KM65.0 - KM68.0' → (fromKm=68.0, toKm=65.0) — higher chainage first."""
    nums = [float(x) for x in re.findall(r"[\d.]+", str(text or "").replace(",", ""))]
    if len(nums) < 2:
        return None, None
    a, b = nums[0], nums[1]
    return (max(a, b), min(a, b))


def load_rows(path):
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl required: pip install openpyxl") from exc
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    out = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        segment = str(r[0]).strip()
        chainage = str(r[1] or "").strip()
        limit = r[2]
        area = str(r[6] or "").strip().upper()
        from_km, to_km = parse_chainage(chainage)
        if from_km is None or limit is None:
            continue
        region, road, on_stick = _AREA.get(area, ("?", area.replace(" ", "_")[:8], 0))
        out.append({
            "segment": segment,
            "region": region,
            "road": road,
            "fromKm": round(from_km, 3),
            "toKm": round(to_km, 3),
            "limit": float(limit),
            "operatingArea": area.title() if area != "KAORAHAI" else "KAORAHAI",
            "chainage": chainage,
            "onStick": on_stick,
        })
    return out


def main():
    src = os.environ.get("SPEED_LIMIT_XLSX") or (
        sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    )
    if not os.path.isfile(src):
        raise SystemExit("workbook not found: %s" % src)
    rows = load_rows(src)
    if not rows:
        raise SystemExit("no zones parsed from %s" % src)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    fields = ["segment", "region", "road", "fromKm", "toKm", "limit",
              "operatingArea", "chainage", "onStick"]
    with open(OUT, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    stick = sum(1 for r in rows if r["onStick"])
    print("wrote %s  (%d zones, %d on TF→FENI stick)" % (OUT, len(rows), stick))
    # Sanity: TOFU KM65-68 = 30
    hit = [r for r in rows if r["segment"] == "SL_TF_10_30"]
    if hit and abs(hit[0]["limit"] - 30) < 0.01 and abs(hit[0]["fromKm"] - 68) < 0.1:
        print("  OK SL_TF_10_30 → fromKm=68 toKm=65 limit=30")
    else:
        print("  WARN SL_TF_10_30 check:", hit)


if __name__ == "__main__":
    main()
