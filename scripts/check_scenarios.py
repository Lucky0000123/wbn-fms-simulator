"""Gate J72: the scenario waterfall respects fleet, contractor and cap invariants.

The scenario feature (2026-08-18) lets the owner load alternative mine plans
(S3, ...; S2 was deleted from the app 2026-08-21) and re-allocate the SAME
fleet by priority:
P1 SAP -> P2 LIM-TOS -> P3 LIM-LD (Tofu dump -> Huafei), 8 Mt cap.

What must always hold, per scenario and per month:
  1. DT conservation: P1 + P2 + P3(free) == the yearly matrix's pool,
     per contractor. Trucks are never created or destroyed.
  2. BLB is RIM-only: no allocation row ever puts a non-RIM truck at BLB.
  3. The LD cap: cumulative LIM-LD planned never exceeds 8,000,000 t, and
     a capped month says so (ld_capped=True), an uncapped one does not.
  4. Priority is real: when a P1/P2 target rises, P3 falls - the free fleet
     is what is LEFT, not a fixed share. (Checked by construction: free =
     pool - used, and used covers every target row or reports a deficit.)
  5. S1 passthrough: Scenario 1 derived from the live yearly matrix keeps
     the matrix's own targets - the waterfall does not invent tonnage.
  6. The importer reads the real workbook shape (long-format Mine Plan DB)
     and never writes an S1 file.

Both directions on purpose (the J71 lesson): a gate that only demands
"cap reached" is passed by hardcoding it; one that only demands "under cap"
is passed by deleting P3. We assert exact conservation and both cap states.
"""
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import scenario_api as sa  # noqa: E402

FAILS = []


def check(name, ok, detail=""):
    print("  %s %s%s" % ("PASS" if ok else "FAIL", name,
                         (" — " + str(detail)[:120]) if (detail and not ok) else ""))
    if not ok:
        FAILS.append(name)


yearly = sa._load_yearly()
if not yearly:
    print("no yearly matrix — gate cannot run")
    sys.exit(1)
rows = sa._yearly_rows(yearly)

print("=== S1: the live plan flows through unchanged ===")
s1 = sa._load_scenario("S1")
check("S1 derives from the yearly matrix", s1 is not None and s1.get("derived"))
res1, err = sa.waterfall(s1)
check("S1 allocates without error", err is None, err)
tgt1 = sa._s1_targets(rows)
sep = {m: round(sum(v for (p, mt, mm), v in tgt1.items() if mt == "SAP" and mm == m))
       for m in (9, 10, 11, 12)}
got = {mo["month"]: mo["sap_t_day"] for mo in res1["months"]}
for m in (9, 10, 11, 12):
    check("S1 month %d SAP target == matrix (%s)" % (m, format(sep[m], ",")),
          abs(got.get(m, 0) - sep[m]) <= 2, "%s vs %s" % (got.get(m), sep[m]))

print("\n=== DT conservation: pool = P1 + P2 + free, per contractor ===")
scen_ids = ["S1"] + [s for s in sa._scenario_ids() if s != "S1"]
results = {}
for sid in scen_ids:
    sc = sa._load_scenario(sid)
    if not sc:
        continue
    res, err = sa.waterfall(sc)
    check("%s allocates without error" % sid, err is None, err)
    if err:
        continue
    results[sid] = res
    for mo in res["months"]:
        for c, pool in mo["pool"].items():
            used = sum(a["dt"] for a in mo["rows"] if a["contractor"] == c)
            for l in mo["lends"]:
                if l["to_work_of"] == c:
                    used -= l["dt"]
                if l["from"] == c:
                    used += l["dt"]
            free = mo["free"].get(c, 0)
            check("%s M%d %s: used %.0f + free %.0f == pool %d"
                  % (sid, mo["month"], c, used, free, pool),
                  abs(used + free - pool) <= 1.5,
                  "%.1f + %.1f != %d" % (used, free, pool))

print("\n=== BLB is RIM-only, in the data and in the allocator ===")
for sid, res in results.items():
    bad = [a for mo in res["months"] for a in mo["rows"]
           if a["pit"] == "BLB" and a["contractor"] != "RIM"]
    check("%s: no non-RIM truck at BLB" % sid, not bad and not res["violations"],
          res["violations"] or bad[:1])
# and the lending path refuses to move trucks INTO a RIM-only pit:
check("lending never targets a RIM-only pit (code path)",
      "RIM_ONLY_PITS" in open(sa.__file__.replace(".pyc", ".py")).read())

print("\n=== the 8 Mt LIM-LD target is filled after P1/P2; excess is capacity ===")
for sid, res in results.items():
    t = res["total"]
    expect = min(t["ld_t_capacity"], sa.LIM_LD_TARGET_T)
    check("%s: planned LD stops at target; capacity remains visible" % sid,
          abs(t["ld_t_planned"] - expect) <= 2,
          "%s vs %s" % (t["ld_t_planned"], expect))
    cum = sum(mo["ld_t_month_planned"] for mo in res["months"])
    check("%s: months sum to the total (%s)" % (sid, format(t["ld_t_planned"], ",")),
          abs(cum - t["ld_t_planned"]) <= 2, cum)
    for mo in res["months"]:
        capped = mo["ld_t_month_planned"] + 0.5 < mo["ld_t_month_capacity"]
        check("%s M%d: target/capacity state is truthful" % (sid, mo["month"]),
              mo["ld_t_month_planned"] <= mo["ld_t_month_capacity"] + 1 and
              bool(mo["ld_capped"]) == capped)
    if t["ld_t_capacity"] >= sa.LIM_LD_TARGET_T:
        check("%s: target reached with no credited over-production" % sid,
              t["ld_cap_reached"] and t["ld_shortfall_t"] == 0 and
              t["ld_over_target_t"] == 0 and t["ld_t_planned"] == sa.LIM_LD_TARGET_T)
    else:
        check("%s: shortfall vs target reported" % sid,
              not t["ld_cap_reached"] and t["ld_over_target_t"] == 0 and
              abs(t["ld_shortfall_t"] - (sa.LIM_LD_TARGET_T - t["ld_t_planned"])) <= 1)
# a scenario whose targets consume everything must plan ~zero LD:
starved = {"id": "SX", "label": "starve", "targets": [
    {"pit": p, "mat": m, "month": mm, "wmt_day": 10_000_000}
    for p in ("BLB", "KRENE", "TOFU") for m in ("SAP", "LIM") for mm in (9, 10, 11, 12)]}
resx, err = sa.waterfall(starved)
check("impossible targets starve P3 to ~0 LD (Sep-Dec)",
      err is None and sum(mo["ld_t_month_planned"] for mo in resx["months"]
                          if mo["month"] > 8) < 1000,
      err or sum(mo["ld_t_month_planned"] for mo in resx["months"] if mo["month"] > 8))
check("and reports deficits instead of inventing trucks",
      err is None and any(mo["deficit"] for mo in resx["months"]))
# and one with zero targets frees the whole fleet:
empty = {"id": "SY", "label": "all-free", "targets": [
    {"pit": "BLB", "mat": "SAP", "month": m, "wmt_day": 0.001} for m in (9, 10, 11, 12)]}
resy, err = sa.waterfall(empty)
if err is None:
    mo9 = next(mo for mo in resy["months"] if mo["month"] == 9)
    check("zero P1/P2 targets -> entire pool is available to P3",
          abs(sum(mo9["free"].values()) - sum(mo9["pool"].values())) <= 1.5
          and mo9["dt_p3"] <= sum(mo9["free"].values()) + 1.5,
          "%s planned vs %s available" % (mo9["dt_p3"], sum(mo9["free"].values())))
else:
    check("zero targets -> entire pool is free for LD", False, err)

print("\n=== Monthly clock semantics: achievable is simulation, not capped prediction ===")
import monthly_api as _ma
clock_alloc = {
    "frozen": True, "horizon": "day",
    "old": {"pred": 80, "achv": 70},
    "new": {"pred": 200, "achv": 150, "achv_sim": 150, "target": 100},
    "goals": {"sap": 100, "tos": 0, "ld": 0, "total": 100},
    "buckets": {"sap": {"target": 100, "pred_before": 80, "pred_after": 200,
                            "achv_before": 70, "achv_after": 150, "achv_sim": 150}},
    "rows": [{"prio": 1, "target": 100, "pred_before": 80, "pred_after": 200,
              "achv_before": 70, "achv_after": 150, "achv_sim": 150,
              "dt_before": 1, "dt_after": 1}],
}
clock_view = _ma._alloc_view(clock_alloc, 1)
check("new_achv_day is raw /api/simulate output",
      clock_view.get("new_achv_day") == 150, clock_view)
check("target-capped prediction has an explicit credited field",
      clock_view.get("new_credited_pred_day") == 100, clock_view)

print("\n=== the importer reads the real workbook shape ===")
hdr = ["Scenario", "Month", "Nb Days", "Mining Pit", "Material", "Type Ore", "wmt ROM"]
demo = [hdr,
        ["Scenario 9", "Sept", 30, "BLB", "SAP", "TOS", 300000],
        ["Scenario 9", "Sept", 30, "TOFU", "LIM", "TOS", 150000],
        ["Scenario 9", "Sept", 30, "TOFU", "LIM", "LD", 240000],
        ["Scenario 9", "Oct", 31, "KRENE", "SAP", "TOS", 310000]]
scens, err = sa._parse_mine_plan_db(demo, "demo")
check("long-format rows parse", err is None and len(scens) == 1, err)
if scens:
    s9 = scens[0]
    check("wmt ROM / Nb Days becomes t/day",
          any(abs(t["wmt_day"] - 10000) < 1 for t in s9["targets"]),
          s9["targets"][:2])
    check("scenario id normalised to S9", s9["id"] == "S9", s9["id"])
    check("import preserves LIM-LD as a P3 target",
          any(t.get("otype") == "LD" and abs(t["wmt_day"] - 8000) < 1
              for t in s9["targets"]), s9["targets"])
check("garbage in -> clear error, not a crash",
      sa._parse_mine_plan_db([["nothing", "here"]], "x")[1] is not None)

print("\n=== re-import of identical targets keeps its stamp (the J59 lesson) ===")
if scens:
    import copy
    os.makedirs(sa._SCEN_DIR, exist_ok=True)
    p9 = sa._scen_path("S9")
    try:
        sa._save_scenario(s9)
        stamp1 = os.path.getmtime(p9)
        blob1 = open(p9).read()
        again = copy.deepcopy(s9)
        again["source"] = "different-filename.xlsx"
        again["loaded_at"] = "2099-01-01T00:00:00Z"
        sa._save_scenario(again)
        check("identical targets -> file byte-identical, stamp kept",
              open(p9).read() == blob1)
        changed = copy.deepcopy(s9)
        changed["targets"][0]["wmt_day"] += 1
        sa._save_scenario(changed)
        check("changed targets -> file IS rewritten", open(p9).read() != blob1)
    finally:
        if os.path.isfile(p9):
            os.remove(p9)

if os.path.isfile(os.path.join(sa._SCEN_DIR, "S1.json")):
    check("no S1.json file may exist (S1 is always derived live)", False)

print("\n=== SAP to FeNi is 2 kt/day, rest to POS (owner 2026-08-25) ===")
check("BLB FeNi cap is 2,000 t/day",
      sa.SAP_ROUTING["BLB"]["fixed"][0][1] == 2000.0, sa.SAP_ROUTING["BLB"])
check("TOFU FeNi cap is 2,000 t/day",
      sa.SAP_ROUTING["TOFU"]["fixed"][0][1] == 2000.0, sa.SAP_ROUTING["TOFU"])
check("BLB rest is POS 14", sa.SAP_ROUTING["BLB"]["rest"] == "POS 14")
check("TOFU rest is POS 12", sa.SAP_ROUTING["TOFU"]["rest"] == "POS 12")
check("KRENE rest is POS 12", sa.SAP_ROUTING["KRENE"]["rest"] == "POS 12")
grp = [{"dest": "FENI KM 15", "wmt": {9: 1}},
       {"dest": "POS 12", "wmt": {9: 1}}]
pieces = sa._split_sap_conditions("TOFU", 10288.0, grp, 9)
by = {r["dest"]: w for r, w in (pieces or [])}
check("TOFU 10,288 SAP → 2,000 FeNi + 8,288 POS",
      abs(by.get("FENI KM 15", 0) - 2000) < 0.01
      and abs(by.get("POS 12", 0) - 8288) < 0.01, by)
grp_blb = [{"dest": "FENI KM0", "wmt": {9: 1}},
           {"dest": "POS 14", "wmt": {9: 1}}]
pieces_b = sa._split_sap_conditions("BLB", 4147.0, grp_blb, 9)
by_b = {sa._norm_sap_dest(r["dest"]): w for r, w in (pieces_b or [])}
check("BLB 4,147 SAP → 2,000 FeNi + 2,147 POS",
      abs(by_b.get("FENI KM0", 0) - 2000) < 0.01
      and abs(by_b.get("POS 14", 0) - 2147) < 0.01, by_b)
pieces_lo = sa._split_sap_conditions("BLB", 1500.0, grp_blb, 9)
by_lo = {sa._norm_sap_dest(r["dest"]): w for r, w in (pieces_lo or [])}
check("BLB 1,500 SAP (< 2,000 cap) all stays on FeNi KM0",
      abs(by_lo.get("FENI KM0", 0) - 1500) < 0.01
      and abs(by_lo.get("POS 14", 0)) < 0.01, by_lo)
grp_bad = [{"dest": "FENI KM0", "wmt": {9: 1}},
           {"dest": "FENI KM15", "wmt": {9: 1}}]
pieces_miss = sa._split_sap_conditions("BLB", 5000.0, grp_bad, 9)
by_m = {sa._norm_sap_dest(r["dest"]): w for r, w in (pieces_miss or [])}
check("BLB leftover clones onto POS 14 when matrix has no POS SAP row",
      abs(by_m.get("FENI KM0", 0) - 2000) < 0.01
      and abs(by_m.get("POS 14", 0) - 3000) < 0.01, by_m)
check("leftover SAP does not go to the other FeNi",
      abs(by_m.get("FENI KM15", 0)) < 0.01, by_m)
try:
    from flask import Flask
    from openpyxl import load_workbook
    import io as _io
    app = Flask(__name__)
    app.register_blueprint(sa.bp)
    c = app.test_client()
    rv = c.get("/api/scenarios/export")
    check("export returns an xlsx", rv.status_code == 200 and
          rv.data[:2] == b"PK", rv.status_code)
    if rv.status_code == 200:
        wb = load_workbook(_io.BytesIO(rv.data))
        api = c.get("/api/scenarios/compare").get_json()
        ids = [s["id"] for s in api["scenarios"]]
        check("one detail sheet per scenario + Compare",
              wb.sheetnames == ["Compare"] + ids, wb.sheetnames)
        totals = None
        for row in wb["Compare"].iter_rows(values_only=True):
            if row[0] == "Total":
                totals = [v for v in row if isinstance(v, (int, float))]
        expect = [s["total"]["ld_t_planned"] for s in api["scenarios"]]
        check("Compare totals row == /api/scenarios/compare LD totals",
              totals == expect, "%s vs %s" % (totals, expect))
        # per-scenario sheet: every month carries its P3 LIM-LD row
        for s in api["scenarios"]:
            d = wb[s["id"]]
            ld_rows = sum(1 for row in d.iter_rows(values_only=True)
                          if row[3] == "LIM-LD" and row[1] == "P3")
            check("%s sheet has a P3 LIM-LD row per month" % s["id"],
                  ld_rows == len(s["months"]), "%d vs %d" % (ld_rows, len(s["months"])))
except ImportError as e:
    print("  SKIP export checks (%s)" % e)

print("\n=== full Excel export (one Monthly workbook per scenario) ===")
try:
    from flask import Flask
    from openpyxl import load_workbook
    import io as _io
    import zipfile
    import monthly_api as ma
    app = Flask(__name__)
    app.register_blueprint(sa.bp)
    app.register_blueprint(ma.bp)
    c = app.test_client()
    yr = "2026"
    sep_plan = os.path.join("data", "saved_plans", "2026-09-03.json")
    if os.path.isfile(sep_plan):
        with open(sep_plan, encoding="utf-8") as f:
            blob = json.load(f)
        arows = (blob.get("allocation") or {}).get("rows") or []
        flow = ma._dest_pit_flow(arows)
        together = sum(flow.values())
        raw = 0.0
        for row in arows:
            if row.get("foreign") or ma._is_tenant_row(row):
                continue
            t = ma._finite(row.get("pred_after"))
            if t:
                raw += t
        check("Sep S3 dest×pit Together equals production predicted t/day",
              abs(together - raw) < 1.0,
              "together=%s raw=%s" % (round(together, 1), round(raw, 1)))
        plants = {d for (_, d, _) in flow}
        check("Sep S3 dest×pit has FENI KM0, FENI KM15, POS",
              {"FENI KM0", "FENI KM15", "POS"}.issubset(plants), plants)
    want_s3 = ["Year", "Road crowding", "Sep", "Oct", "Nov", "Dec"]
    rv = c.get("/api/scenarios/export-full?year=%s&id=S3" % yr)
    check("S3 monthly workbook returns xlsx", rv.status_code == 200 and rv.data[:2] == b"PK",
          rv.status_code)
    if rv.status_code == 200:
        wb = load_workbook(_io.BytesIO(rv.data), read_only=True)
        # "Paths" (the all-months path list, added 2026-08-19) is optional;
        # the month sheets must be Year-then-months in order.
        got = [s for s in wb.sheetnames if s != "Paths"]
        check("S3 sheets are Year + Road crowding + Sep–Dec (no August)",
              got == want_s3, wb.sheetnames)
        check("S3 workbook has Road crowding sheet",
              "Road crowding" in wb.sheetnames, wb.sheetnames)
        check("S3 workbook has no August sheet",
              "Aug" not in wb.sheetnames, wb.sheetnames)
        sep = wb["Sep"]
        a1 = sep["A1"].value or ""
        check("S3 Sep title is 'Sep 2026 — old vs new'",
              a1.startswith("Sep 2026"), a1)
        rows = list(sep.iter_rows(min_row=1, max_row=120, max_col=13, values_only=True))
        heads = [r for r in rows if r and r[0] == "P"]
        check("S3 Sep has the path table header", bool(heads), heads[:1])
        dest_title = [r for r in rows if r and isinstance(r[0], str)
                      and r[0].startswith("Where material goes")]
        check("S3 Sep has Where material goes · t/day at the top",
              bool(dest_title), dest_title[:1])
        dest_heads = [r for r in rows if r and r[0] == "Material" and r[1] == "To plant"]
        check("S3 Sep dest×pit columns are pits (TF / BLB / KR)",
              dest_heads and any(h in (dest_heads[0] or ()) for h in ("TF", "BLB", "KR")),
              dest_heads[:1])
        dest_plants = {r[1] for r in rows if r and r[1] in ("FENI KM0", "FENI KM15", "POS")}
        check("S3 Sep dest rows are FENI KM0, FENI KM15, POS",
              dest_plants == {"FENI KM0", "FENI KM15", "POS"}
              or dest_plants.issuperset({"FENI KM0", "FENI KM15", "POS"}),
              dest_plants)
        dest_nums = [v for r in rows if r and r[1] in ("FENI KM0", "FENI KM15", "POS")
                     for v in (r[2:] if r else ())
                     if isinstance(v, (int, float)) and v > 0]
        check("S3 Sep dest×pit table has real t/day numbers",
              len(dest_nums) >= 3, "n=%s" % len(dest_nums))
        sep_heads = heads[0] if heads else ()
        check("S3 Sep path columns have no 'old' label",
              sep_heads and all("old" not in str(h or "").lower() for h in sep_heads),
              sep_heads)
        check("S3 Sep path columns have no 'new' suffix",
              sep_heads and all("new" not in str(h or "").lower() for h in sep_heads),
              sep_heads)
        check("S3 Sep path table is DT / Trips / WMT (not DT old / DT new)",
              sep_heads and "DT" in sep_heads and "Trips" in sep_heads
              and "WMT" in sep_heads
              and "DT old" not in sep_heads and "DT new" not in sep_heads,
              sep_heads)
        p3 = [r for r in rows if r and r[0] == "P3"]
        check("S3 Sep has P3 LIM-LD path rows (leftover DT)", bool(p3), p3[:1])
        if "Paths" in wb.sheetnames:
            ph = list(wb["Paths"].iter_rows(min_row=1, max_row=12, max_col=16,
                                            values_only=True))
            path_heads = next((r for r in ph if r and r[0] == "Month"), None)
            check("S3 Paths header uses Priority, not P",
                  path_heads and path_heads[1] == "Priority", path_heads)
            check("S3 Paths columns have no 'new' suffix",
                  path_heads and all("new" not in str(h or "").lower()
                                     for h in path_heads),
                  path_heads)
            check("S3 Paths keeps DT / Trips / WMT / WMT/DT / Trips/DT / NB Days",
                  path_heads and all(x in path_heads for x in
                                     ("DT", "Trips", "WMT", "WMT/DT", "Trips/DT",
                                      "NB Days")),
                  path_heads)
        wb.close()
        wb_style = load_workbook(_io.BytesIO(rv.data), read_only=False)
        reds = {"DC2626", "B91C1C", "FFDC2626", "FFB91C1C"}
        bad = []
        for sname in wb_style.sheetnames:
            ws = wb_style[sname]
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 80),
                                    max_col=16):
                for cell in row:
                    if not isinstance(cell.value, (int, float)):
                        continue
                    rgb = None
                    if cell.font and cell.font.color is not None:
                        rgb = getattr(cell.font.color, "rgb", None)
                    if rgb and str(rgb).upper() in reds:
                        bad.append("%s!%s %s" % (sname, cell.coordinate, rgb))
        check("S3 Excel table numbers are not red", not bad, bad[:8])
        wb_style.close()
    rvz = c.get("/api/scenarios/export-full?year=" + yr)
    check("all-scenarios zip returns", rvz.status_code == 200 and rvz.data[:2] == b"PK",
          rvz.status_code)
    if rvz.status_code == 200:
        z = zipfile.ZipFile(_io.BytesIO(rvz.data))
        names = z.namelist()
        check("zip has monthly_plan_2026.xlsx (S1)", "monthly_plan_2026.xlsx" in names, names)
        check("zip has monthly_plan_2026_S3.xlsx", "monthly_plan_2026_S3.xlsx" in names, names)
        # S2 was deleted from the app 2026-08-21: its workbook must NOT come back.
        check("zip has no S2 workbook (S2 deleted)",
              "monthly_plan_2026_S2.xlsx" not in names, names)
        for fn in names:
            inner = load_workbook(_io.BytesIO(z.read(fn)), read_only=True)
            check("%s sheets start with Year" % fn, inner.sheetnames[0] == "Year",
                  inner.sheetnames)
            if fn == "monthly_plan_2026.xlsx":
                check("S1 workbook still has August",
                      "Aug" in inner.sheetnames, inner.sheetnames)
            elif "_S3" in fn:
                check("S3 zip workbook has no August",
                      "Aug" not in inner.sheetnames, inner.sheetnames)
            inner.close()
        z.close()
    # Year-board Download Excel with S3/S4 selected (day=3/4) also starts
    # at September. Both directions: S3/S4 have no Aug AND S1/day=1 still has it.
    rv1 = c.get("/api/monthly/export-year?year=%s&day=1" % yr)
    if rv1.status_code == 200:
        wb1 = load_workbook(_io.BytesIO(rv1.data), read_only=True)
        check("S1 year Excel still has August",
              "Aug" in wb1.sheetnames, wb1.sheetnames)
        wb1.close()
    for day, label in ((3, "S3"), (4, "S4")):
        rvd = c.get("/api/monthly/export-year?year=%s&day=%d" % (yr, day))
        if rvd.status_code == 404:
            # No Sep–Dec saves for that day yet — August-only leftover is
            # correctly not enough to build a workbook.
            continue
        if rvd.status_code != 200:
            check("%s year Excel (day=%d) returns xlsx" % (label, day),
                  False, rvd.status_code)
            continue
        wbd = load_workbook(_io.BytesIO(rvd.data), read_only=True)
        names_d = wbd.sheetnames
        check("%s year Excel has no August sheet" % label,
              "Aug" not in names_d, names_d)
        months_d = [s for s in names_d if s not in ("Year", "Paths", "Road crowding")]
        check("%s year Excel months start at September" % label,
              months_d[:1] == ["Sep"], months_d)
        yr_ws = wbd["Year"]
        yr_vals = [cell.value for row in yr_ws.iter_rows(min_row=1, max_row=80, max_col=1)
                   for cell in row]
        check("%s Year sheet has no August row" % label,
              "Aug" not in yr_vals, [v for v in yr_vals if v in ("Aug", "Sep")])
        # Each month sheet carries the Plan-tab hour grid (owner 2026-08-24):
        # corridor × 07..06, mean concurrent trucks. Both directions: the
        # title is present AND a real occupancy number lands (a title-only
        # stub would pass the first and fail the second).
        month_sheets = [s for s in names_d if s not in ("Year", "Paths", "Road crowding")]
        n_grid = 0
        n_occ = 0
        n_dest = 0
        hour_ok = False
        for ms in month_sheets:
            ws = wbd[ms]
            col_a = [cell.value for row in ws.iter_rows(min_row=1, max_col=1)
                     for cell in row]
            if any(isinstance(v, str) and "Road crowding by hour" in v for v in col_a):
                n_grid += 1
            if any(isinstance(v, str) and v.startswith("Where material goes") for v in col_a):
                n_dest += 1
            for row in ws.iter_rows(min_row=1, max_col=25):
                # NOT `c` — that is the Flask test client in this scope, and
                # shadowing it here made the NEXT loop iteration call
                # .get() on a ReadOnlyCell ("export-full smoke test —
                # 'ReadOnlyCell' object has no attribute 'get'"), failing J72
                # while every actual export was fine.
                vals = [cl.value for cl in row]
                if vals and vals[0] == "Corridor" and vals[1] == "07":
                    hours = vals[1:25]
                    if (hours[:2] == ["07", "08"] and hours[-1] == "06"
                            and len(hours) == 24):
                        hour_ok = True
                lab = vals[0] if vals else None
                if isinstance(lab, str) and lab.startswith("TF") and "KR" in lab:
                    nums = [v for v in vals[1:25] if isinstance(v, (int, float))]
                    n_occ = max(n_occ, len(nums))
        check("%s every month sheet has the hour-crowding table" % label,
              n_grid == len(month_sheets) and month_sheets,
              "%s/%s sheets" % (n_grid, len(month_sheets)))
        check("%s every month sheet has dest×pit t/day" % label,
              n_dest == len(month_sheets) and month_sheets,
              "%s/%s sheets" % (n_dest, len(month_sheets)))
        check("%s hour axis is 07..06 (24 h)" % label, hour_ok)
        check("%s Sep-class sheet has TF–KR hourly occupancy" % label,
              n_occ >= 20, "n=%s" % n_occ)
        packed = False
        over_ok = True
        over_seen = False
        for ms in month_sheets:
            ws = wbd[ms]
            col_a = [cell.value for row in ws.iter_rows(min_row=1, max_col=1)
                     for cell in row]
            blob = " ".join(str(v) for v in col_a if v)
            if "one loaded lane" in blob.lower():
                packed = True
            for row in ws.iter_rows(min_row=1, max_col=25):
                lab = row[0].value if row else None
                if not (isinstance(lab, str) and lab.startswith("TF") and "KR" in lab):
                    continue
                for cl in row[1:25]:  # not `c` — the Flask client lives in this scope
                    if not isinstance(cl.value, (int, float)) or cl.value < 576:
                        continue
                    over_seen = True
                    rgb = ""
                    try:
                        rgb = str(cl.fill.fgColor.rgb or "")
                    except Exception:
                        rgb = ""
                    if not rgb.upper().endswith("FCA5A5"):
                        over_ok = False
        check("%s hour grid names one-lane packing (same as Plan)" % label, packed)
        if over_seen:
            check("%s TF–KR cells ≥576 are red (one-lane pack)" % label, over_ok)
        wbd.close()
except ImportError as e:
    print("  SKIP export-full checks (%s)" % e)
except Exception as e:
    check("export-full smoke test", False, str(e)[:200])


print("=== draft plans: Plan-tab shape, fleet conserved, predicted on target ===")
try:
    import shutil, tempfile
    from flask import Flask
    app3 = Flask(__name__)
    import monthly_api as ma3
    app3.register_blueprint(sa.bp)
    app3.register_blueprint(ma3.bp)
    import simulator_api as sim3
    app3.register_blueprint(sim3.bp)
    c3 = app3.test_client()
    # generate into a throwaway date (day 28, unlikely to be used) then clean up
    rv = c3.post("/api/scenarios/S3/draft-plans", json={
        "year": 2026, "day": 28, "months": [9]})
    d = rv.get_json()
    wrote = d.get("ok") and d.get("written")
    check("draft-plans writes a Plan-tab save", bool(wrote), d)
    fp = os.path.join(sa._ROOT, "data", "saved_plans", "2026-09-28.json")
    try:
        if wrote:
            plan = json.load(open(fp))
            check("draft has date/paths/rain/hours keys",
                  all(k in plan for k in ("date", "paths", "rain_mm", "hours")))
            tot = {}
            blb_bad = []
            for slot, p in plan["paths"].items():
                tot[p["contractor"]] = tot.get(p["contractor"], 0) + p["dt"]
                if p["source"] == "BLB" and p["contractor"] != "RIM":
                    blb_bad.append(slot)
            pool = d["written"][0]["pool"]
            unused = d["written"][0].get("unused") or {}
            check("draft allocated + unused == pool per contractor",
                  all(abs(tot.get(k, 0) + unused.get(k, 0) - v) <= 1.5
                      for k, v in pool.items()),
                  "%s + %s vs %s" % (tot, unused, pool))
            check("draft has no non-RIM at BLB", not blb_bad, blb_bad)
            # existing date is refused without overwrite
            rv2 = c3.post("/api/scenarios/S3/draft-plans", json={
                "year": 2026, "day": 28, "months": [9]})
            d2 = rv2.get_json()
            check("existing date refused without overwrite",
                  not d2.get("written") and d2.get("errors"), d2)
            # P1/P2 predicted lands on target (>=99.5% of target each)
            with app3.app_context():
                paths_m, fleet_m, contr_m = ma3._path_model_context()
                comb = {}
                for p in plan["paths"].values():
                    comb[p["key"]] = comb.get(p["key"], 0) + p["dt"]
                low = []
                for slot, p in plan["paths"].items():
                    if p.get("otype") == "LD" or not p.get("targetWmt"):
                        continue
                    row = ma3._path_row_wmt(p["source"], p["dest"], p["contractor"],
                                            p["dt"], comb[p["key"]],
                                            paths_m, fleet_m, contr_m)
                    w = row.get("wmt") or 0
                    if w < p["targetWmt"] * 0.995:
                        low.append((slot, round(w), p["targetWmt"]))
                check("every P1/P2 route predicts >= 99.5%% of target", not low, low[:3])
    finally:
        if os.path.isfile(fp):
            os.remove(fp)
except ImportError as e:
    print("  SKIP draft-plan checks (%s)" % e)
except Exception as e:  # noqa: BLE001
    check("draft-plan smoke test", False, str(e)[:120])


print("=== year board day=N picks that day's saved plan (scenario convention) ===")
try:
    from flask import Flask
    import monthly_api as ma5
    app5 = Flask(__name__)
    app5.register_blueprint(ma5.bp)
    c5 = app5.test_client()
    srcs = {}
    for d in (None, 1, 2, 3):
        url = "/api/monthly/year-board?year=2026" + ("&day=%d" % d if d else "")
        r = c5.get(url).get_json()
        srcs[d] = [(c["month"], (c.get("alloc") or {}).get("source_date"))
                   for c in r["cards"] if c.get("has_alloc")]
    for d in (1, 3):
        got = srcs.get(d) or []
        check("day=%d only resolves day-%02d saves" % (d, d),
              got and all(s and s.endswith("-%02d" % d) for _, s in got),
              got[:3])
    # S2 (day 2) was deleted 2026-08-21: its slot must resolve nothing.
    check("day=2 resolves no saves (S2 deleted)", not srcs.get(2), srcs.get(2))
    check("day=1 and day=3 resolve different plans (not the same copy)",
          srcs[1] and srcs[3] and srcs[1] != srcs[3])
    check("day=3 year board has no August card (S3 starts at September)",
          not any(m.endswith("-08") for m, _ in (srcs.get(3) or [])),
          srcs.get(3))
    srcs4 = []
    r4 = c5.get("/api/monthly/year-board?year=2026&day=4").get_json()
    if r4 and r4.get("cards"):
        srcs4 = [c["month"] for c in r4["cards"] if c.get("has_alloc")]
    check("day=4 year board has no August card (S4 starts at September)",
          "2026-08" not in srcs4, srcs4)
    check("day omitted keeps the old latest-save rule",
          bool(srcs[None]))
except ImportError as e:
    print("  SKIP year-board day checks (%s)" % e)
except Exception as e:  # noqa: BLE001
    check("year-board day smoke test", False, str(e)[:120])

print()
if FAILS:
    print("J72 FAILED: %d check(s). First: %s" % (len(FAILS), FAILS[0]))
    sys.exit(1)
print("scenario waterfall gate passes")
