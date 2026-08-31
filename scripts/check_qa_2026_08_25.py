#!/usr/bin/env python3
"""Gates for the 2026-08-25 three-agent QA audit fixes.

Each check asserts BOTH directions where a one-sided assertion could be passed
by deleting the feature (the J71 lesson), and each targets the REAL caller
rather than a payload this script invents (the J52 lesson).

  J81  the corridor flow readout divides by the OFFICIAL road geometry
       (/api/road_segments: S1-S3 600/hr, S4 400/hr at 50 m), not the ~54 tph
       Jul GPS "struggle extract" demonstrated peak, and its sections are the
       official S1-S4 split rather than the legacy POS 10 boundaries.
  J82  tenant rows never reach the flow readout's production math, and their
       road flow is still charged (excluded from trips AND present as flow).
  J83  a frozen saved plan opens with the New Allocation panel visible, no
       stale pre-allocation "add N DT" board, and predicted totals that match
       what was saved (pricing state rebuilt, not just the DT).
  J84  the priority board totals sum on ONE fleet basis and never print an
       unknown achievable as a confident zero.
  J85  Excel TOTAL DT is our fleet only, with IWIP counted beside it, while
       every other TOTAL column is unchanged.
  J86  every scenario the app offers is exportable, including the day-04
       convention scenario S4 which has no data/scenarios/S4.json.

Run standalone:  .venv/bin/python scripts/check_qa_2026_08_25.py
"""
import io
import json
import os
import re
import sys
import urllib.request

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.environ.get("SIM_BASE", "http://127.0.0.1:5055")
PLAN_DATE = os.environ.get("QA_PLAN_DATE", "2026-12-04")

FAILS = []
CHECKS = []


def ok(name, cond, detail=""):
    CHECKS.append((name, bool(cond), detail))
    if not cond:
        FAILS.append("%s — %s" % (name, detail))
    return bool(cond)


def _get(path, timeout=120):
    with urllib.request.urlopen(BASE + path, timeout=timeout) as r:
        return r.status, r.read(), dict(r.headers)


def _json(path, timeout=120):
    st, body, _ = _get(path, timeout)
    return st, json.loads(body.decode("utf-8"))


def src(rel):
    return io.open(os.path.join(ROOT, rel), encoding="utf-8").read()


# ---------------------------------------------------------------- J81
def j81_official_geometry():
    st, d = _json("/api/road_segments")
    ok("J81 road_segments serves", st == 200 and d.get("ok"), "status %s" % st)
    caps = {s["id"]: s["cap_hr"] for s in d["segments"]}
    ok("J81 official caps", caps == {"S1": 600, "S2": 600, "S3": 600, "S4": 400},
       "caps=%s" % caps)

    js = src("static/js/flow_sim.js")
    # Strip comments first. The fix DOCUMENTS what it removed, so a raw
    # substring search finds "measuredCapacity" in the very comment explaining
    # that it is no longer a divisor — a gate that fails on its own rationale
    # teaches people to delete the rationale.
    code = re.sub(r"/\*.*?\*/", "", js, flags=re.S)
    code = "\n".join(re.sub(r"//.*$", "", ln) for ln in code.split("\n"))
    # POSITIVE: the readout takes capacity from the served geometry.
    ok("J81 flow_sim fetches road_segments", "/api/road_segments" in js,
       "flow_sim.js never fetches the official geometry")
    ok("J81 sections come from geometry",
       "flowRoadSegments().map(" in js,
       "sectionDefs is not derived from flowRoadSegments()")
    # NEGATIVE: the demonstrated GPS peak may not come back as the divisor.
    ok("J81 no measuredCapacity divisor",
       "measuredCapacity" not in code,
       "flow_sim.js still reads _D.corridor.measuredCapacity as a capacity")
    ok("J81 no legacy POS 10 sections",
       "POS 12\u2013POS 10" not in code and "POS 10\u2013FENI" not in code,
       "legacy POS 10 section boundaries still present in code")
    # The fallback literals must equal the served geometry, or they rot.
    m = re.search(r"FLOW_SEG_FALLBACK\s*=\s*\[(.*?)\];", js, re.S)
    ok("J81 fallback literal present", bool(m), "no FLOW_SEG_FALLBACK")
    if m:
        fb = dict(re.findall(r"id:'(S\d)'.*?cap_hr:(\d+)", m.group(1)))
        ok("J81 fallback matches served caps",
           {k: int(v) for k, v in fb.items()} == caps,
           "fallback %s vs served %s" % (fb, caps))


# ---------------------------------------------------------------- J82
def j82_tenants_out_of_production():
    st, d = _json("/api/congestion_tenants")
    ok("J82 tenant register serves", st == 200 and d.get("ok"), "status %s" % st)
    flow = d.get("segment_flow_hr") or {}
    # POSITIVE: the tenants really are on the road (flow > 0 on shared segments).
    ok("J82 tenants carry road flow",
       sum(float(v) for v in flow.values()) > 0,
       "register reports no segment flow at all — nothing to charge")

    plan = src("static/js/plan_scenario.js")
    seed = plan[plan.index("function planDraftToFlowSeed"):]
    seed = seed[:seed.index("\nfunction ")]
    # NEGATIVE: the seed must recognise and drop tenant rows.
    ok("J82 seed drops tenant rows", "planIsTenantRow" in seed,
       "planDraftToFlowSeed has no tenant recogniser — tenant DT reach the "
       "readout and get priced at our trips/DT")
    ok("J82 seed marks foreign rows", "foreign" in seed,
       "planDraftToFlowSeed does not mark road-only rows, so IWIP DT are "
       "counted as our production")

    js = src("static/js/flow_sim.js")
    # POSITIVE: tenant flow is added back as FLOW, not as trucks at our tempo.
    ok("J82 tenant flow charged to sections",
       "flowTenantSegmentFlow" in js and "tenantHourly" in js,
       "flow_sim never adds the tenant register's own flow to a section")
    ok("J82 foreign excluded from production",
       "foreignDt" in js and "r.foreign&&_flowHost==='plan'" in js,
       "flow_sim still sums road-only rows into trips/tonnes")


# ---------------------------------------------------------------- J83
def j83_frozen_load_state():
    """Browser gate: a frozen saved plan opens correct AND priced correctly.

    Drives the REAL page through the REAL loader (planLoadSavedForDate), not a
    payload this script assembles — the J52 lesson. Waits on CONDITIONS, never
    on the clock: in fixture mode start-up fetches stall for seconds and a
    fixed sleep once failed 17 assertions with nothing actually wrong.
    """
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        CHECKS.append(("J83 skipped (no playwright)", True, ""))
        return
    saved_path = os.path.join(ROOT, "data", "saved_plans", "%s.json" % PLAN_DATE)
    if not os.path.exists(saved_path):
        CHECKS.append(("J83 skipped (no %s save)" % PLAN_DATE, True, ""))
        return
    saved = json.load(io.open(saved_path, encoding="utf-8"))
    want_wmt = ((saved.get("meta") or {}).get("predict") or {}).get("wmt")

    with sync_playwright() as pw:
        br = pw.chromium.launch()
        pg = br.new_page(viewport={"width": 1560, "height": 1000})
        errs = []
        pg.on("console", lambda m: errs.append(m.text) if m.type == "error" else None)
        pg.goto(BASE + "/simulator", wait_until="domcontentloaded")
        pg.wait_for_function(
            "typeof setSimTab==='function'&&typeof planLoadSavedForDate==='function'")
        pg.evaluate("setSimTab('plan');window.confirm=()=>true;")
        pg.wait_for_function("typeof planDraftEntries==='function'")
        pg.evaluate("document.getElementById('plan-date').value=%r" % PLAN_DATE)
        pg.evaluate("planLoadSavedForDate({quiet:true})")
        pg.wait_for_function("planDraftEntries().length>0", timeout=180_000)
        pg.wait_for_function("planAllocFrozen&&planAllocFrozen()", timeout=180_000)

        st = pg.evaluate("""()=>{
          const wrap=document.getElementById('plan-alloc-wrap');
          const est=document.getElementById('plan-scenario-estimate');
          const cap=document.getElementById('plan-sap-board-cap');
          return {wrap: wrap?getComputedStyle(wrap).display:'(absent)',
                  stale: !!(cap&&cap.innerHTML.trim().length),
                  addDt: (((est||{}).textContent)||'').match(/add [0-9,]+ DT/g)||[],
                  label: !!(est&&est.querySelector('.plan-cap-frozen-lab'))};}""")
        # POSITIVE: the New Allocation panel is on screen ...
        ok("J83 alloc panel visible on load", st["wrap"] not in ("none", "(absent)"),
           "plan-alloc-wrap display=%s on a frozen plan" % st["wrap"])
        ok("J83 original card is labelled frozen", st["label"],
           "no .plan-cap-frozen-lab — the saved card is not marked as original")
        # ... NEGATIVE: and no pre-allocation shortfall board survives.
        ok("J83 no stale required-DT board", not st["stale"],
           "plan-sap-board-cap still rendered on a frozen plan")
        ok("J83 no pre-alloc 'add N DT'", not st["addDt"],
           "frozen plan still tells the planner to %s" % st["addDt"])

        # Pricing state, not just the DT: predicted totals must come back to
        # what was saved. A DT-only round-trip passes while the tonnes drift.
        if want_wmt:
            # Bounded wait, then MEASURE. An unbounded wait turns "pricing
            # drifted" into a raw TimeoutError that fails the whole gate
            # function instead of this named check (mutation-proved).
            try:
                pg.wait_for_function(
                    """(w)=>{const t=planPredictTotals();
                            return t&&Math.abs(t.wmt-w)/w<0.005;}""",
                    arg=want_wmt, timeout=120_000)
            except Exception:
                pass
            got = pg.evaluate("planPredictTotals()")
            drift = abs(got["wmt"] - want_wmt) / want_wmt
            ok("J83 predicted totals round-trip", drift < 0.005,
               "saved %.1f vs reloaded %.1f (%+.2f%%)"
               % (want_wmt, got["wmt"], 100 * (got["wmt"] - want_wmt) / want_wmt))
            # NEGATIVE: a gate that only demands closeness is passed by a stub
            # that returns the saved number. Assert the fleet is real too.
            ok("J83 priced fleet is the real one", got.get("dt", 0) > 0,
               "planPredictTotals reports no DT at all")
        ok("J83 console clean", not errs, "console errors: %s" % errs[:3])
        br.close()


# ---------------------------------------------------------------- J84
def j84_board_one_basis():
    js = src("static/js/plan_sap_target.js")
    code = re.sub(r"/\*.*?\*/", "", js, flags=re.S)
    code = "\n".join(re.sub(r"//.*$", "", ln) for ln in code.split("\n"))
    # NEGATIVE first: the pre-allocation DT may not be divided by the
    # allocated one ANYWHERE. The positive substring below survives a
    # single-site regression (two call sites share it — mutation-proved),
    # so the absence check is the one that bites.
    ok("J84 board share uses working DT",
       "achievableShare(r.key,r.dt)" not in code,
       "a call site passes r.dt (pre-allocation) while routeDt() sums "
       "workingDt() — the share exceeds the route's own achievable")
    ok("J84 strip total uses working DT",
       code.count("achievableShare(r.key,workingDt(r))") >= 2,
       "fewer than 2 call sites price on workingDt — board and strip must both")
    # NEGATIVE: an unknown achievable may not be coerced to a confident zero.
    ok("J84 snapshot keeps unknown as null",
       "achv:Number.isFinite(c.achv)?c.achv:null" in js,
       "snapshotRow still does achv: c.achv||0, so 'OLD ACHIEVABLE 0' prints "
       "for a value that was never measured")
    ok("J84 strip renders null as dash",
       "(v==null?'—':fmt(v))" in js,
       "the KPI cell cannot express 'not measured'")
    # POSITIVE: the caption must state which priorities it actually summed.
    ok("J84 caption names real priorities",
       "rows with a target" in js and "P1+P2 rows only" not in js.split("//")[0],
       "the strip still claims 'P1+P2 rows only' while P3 LD rows carry "
       "targets and land in the same sum")


# ---------------------------------------------------------------- J85
def j85_excel_total_dt():
    import openpyxl
    html = open(os.path.join(ROOT, "templates", "monthly.html"), encoding="utf-8").read()
    ok("J85 download refuses a non-xlsx body",
       "u8[0]===0x50" in html and "u8[1]===0x4B" in html,
       "moExportYear must check zip magic so an HTML timeout is not saved as .xlsx")
    ok("J85 download names the file .xlsx",
       'if(!/\\.xlsx$/i.test(name)) name+=\'.xlsx\'' in html
       or 'name+=\'.xlsx\'' in html)
    st, body, hdrs = _get("/api/monthly/export-year?year=2026&day=4&achv=1", timeout=600)
    ok("J85 export-year day=4 serves", st == 200, "status %s" % st)
    ok("J85 export-year is an xlsx zip, not HTML",
       isinstance(body, (bytes, bytearray)) and body[:2] == b"PK",
       (hdrs, body[:60] if isinstance(body, (bytes, bytearray)) else type(body)))
    wb = openpyxl.load_workbook(io.BytesIO(body))

    import monthly_api as ma
    _y, cards = ma._year_cards("2026", day=4)
    # The SAME collector the sheet builds from (the J52 lesson: verify what the
    # real caller produces, not a fleet this script re-derives its own way).
    rows = ma._collect_year_path_rows(cards)
    pools, iwip = {}, {}
    for rr in rows:
        mkey = str(rr.get("month"))          # "Sep" / "Oct" / ... — sheet names
        if rr.get("_tenant"):
            continue
        bucket = iwip if rr.get("foreign") else pools
        bucket[mkey] = bucket.get(mkey, 0) + (rr.get("dt") or 0)
    ok("J85 fixture has IWIP to exclude", sum(iwip.values()) > 0,
       "no IWIP rows in the day-4 saves — this gate would pass vacuously")

    for sh in [s for s in wb.sheetnames if s in ("Sep", "Oct", "Nov", "Dec")]:
        ws = wb[sh]
        total_dt = None
        iwip_line = None
        for row in ws.iter_rows():
            v = row[0].value
            if v == "TOTAL" and total_dt is None and len(row) > 5:
                if isinstance(row[5].value, (int, float)):
                    total_dt = row[5].value
            if isinstance(v, str) and v.startswith("IWIP POS-transit"):
                iwip_line = v
        month = sh
        want = pools.get(month)
        # POSITIVE: TOTAL DT equals our fleet ...
        ok("J85 %s TOTAL DT == our fleet" % sh,
           want and total_dt and abs(total_dt - want) < 1,
           "sheet %s TOTAL DT %s vs our-fleet %s" % (sh, total_dt, want))
        # ... and NEGATIVE: it is not the pool + IWIP sum it used to be.
        ok("J85 %s TOTAL DT excludes IWIP" % sh,
           want and total_dt and abs(total_dt - (want + iwip.get(month, 0))) > 0.5,
           "sheet %s TOTAL DT still equals pool+IWIP (%s)" % (sh, total_dt))
        # ... and the IWIP fleet is still SHOWN, not just dropped.
        ok("J85 %s names the IWIP fleet" % sh, bool(iwip_line),
           "sheet %s dropped IWIP without saying so" % sh)


# ---------------------------------------------------------------- J86
def j86_every_scenario_exportable():
    sys.path.insert(0, ROOT)
    import scenario_api as sa
    ids = sa._exportable_scenario_ids()
    # POSITIVE: S4 is offered ...
    ok("J86 S4 is offered", "S4" in ids, "exportable ids = %s" % ids)
    # ... NEGATIVE: and legacy August dailies are NOT invented as scenarios.
    ok("J86 no phantom scenarios",
       # S5/S6 became deliberate scenarios on 2026-08-26 (3.1.1 / 3.1.2) and
       # S7 became deliberate on 2026-08-31 (4.1 — the manager's ROM-table
       # targets, day-07 saves). The remaining phantom sentinel is the legacy
       # August daily save that must NEVER surface: 13.
       not ({"S13"} & set(ids)),
       "legacy August daily saves surfaced as scenarios: %s" % ids)

    st, body, hdrs = _get(
        "/api/scenarios/export-full?year=2026&achv=1&id=S4", timeout=900)
    ok("J86 export-full?id=S4 is 200", st == 200, "status %s" % st)
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(body))
    ok("J86 S4 workbook has month sheets",
       {"Sep", "Oct", "Nov", "Dec"} <= set(wb.sheetnames),
       "sheets %s" % wb.sheetnames)
    ok("J86 S4 starts September (no Aug)", "Aug" not in wb.sheetnames,
       "S4 workbook carries an August sheet; S3/S4 start in September")

    import zipfile
    st, body, _ = _get("/api/scenarios/export-full?year=2026&achv=1", timeout=900)
    ok("J86 zip serves", st == 200, "status %s" % st)
    names = zipfile.ZipFile(io.BytesIO(body)).namelist()
    ok("J86 zip carries S4",
       any("S4" in n for n in names), "zip members %s" % names)
    ok("J86 zip carries every offered scenario",
       len([n for n in names if n.endswith(".xlsx")]) >= len(ids),
       "zip has %s workbooks for %s scenarios" % (len(names), len(ids)))


def main():
    sys.path.insert(0, ROOT)
    for fn in (j81_official_geometry, j82_tenants_out_of_production,
               j83_frozen_load_state, j84_board_one_basis, j85_excel_total_dt,
               j86_every_scenario_exportable):
        # ONE bounded retry on an environmental error (browser boot starvation,
        # a server mid-reload) — the D18b pattern. A retry that keeps failing
        # is a real failure; a gate that fails on a transient teaches people to
        # wave failures through. Assertion failures (ok(...) FAILS) never
        # retry: those are the gate's verdict, not the environment's.
        for attempt in (1, 2):
            n_fails, n_checks = len(FAILS), len(CHECKS)
            try:
                fn()
                break
            except Exception as exc:           # noqa: BLE001 - report, never hide
                if attempt == 1:
                    # drop the failed attempt's half-recorded checks, then retry
                    del FAILS[n_fails:]
                    del CHECKS[n_checks:]
                    import time as _t
                    _t.sleep(10)
                    continue
                FAILS.append("%s raised %s: %s" % (fn.__name__, type(exc).__name__, exc))
                CHECKS.append((fn.__name__, False, str(exc)))
    for name, good, detail in CHECKS:
        print("  %s  %s%s" % ("PASS" if good else "FAIL", name,
                              "" if good else "  <- " + detail))
    print("%d/%d checks passed" % (sum(1 for _, g, _ in CHECKS if g), len(CHECKS)))
    if FAILS:
        for f in FAILS:
            print("FAIL: %s" % f, file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
