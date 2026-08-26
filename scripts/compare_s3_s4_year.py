#!/usr/bin/env python
"""Build a Year-board comparison workbook: S3 predicted vs S4 predicted.

Reads the exported year sheets (plan_2026_S03.xlsx / plan_2026_S04.xlsx).
Old predicted = S3 optimized plan. New predicted = S4 optimized plan.
SAP and LIM-TOS tables/charts are omitted (they match). LIM-LD gets the charts.

    .venv/bin/python scripts/compare_s3_s4_year.py \\
        --s3 ~/Downloads/plan_2026_S03.xlsx \\
        --s4 ~/Downloads/plan_2026_S04.xlsx \\
        --out ~/Downloads/plan_2026_S3_vs_S4.xlsx
"""
from __future__ import annotations

import argparse
import os
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.chart import BarChart, Reference  # noqa: E402

from monthly_api import (  # noqa: E402
    _XLSX_CLOCKS, _XLSX_MUTED, _XLSX_NAVY, _XLSX_PRED, _XLSX_TGT,
    _cov_pct, _xlsx_board_header, _xlsx_font, _xlsx_headers, _xlsx_kpi_strip,
    _xlsx_line_chart, _xlsx_mid, _xlsx_num, _xlsx_paint_cov, _xlsx_section,
    _xlsx_text, _xlsx_total_border, _xlsx_widths,
)


def _parse_year_blocks(path):
    """Map section title → [{name, target, old, new}, ...] from a Year sheet."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Year"]
    rows = list(ws.iter_rows(values_only=True))
    blocks = {}
    title = "Together"
    i = 0
    while i < len(rows):
        row = rows[i]
        a = row[0] if row else None
        if isinstance(a, str) and a.endswith(" · year"):
            title = a.replace(" · year", "").strip()
        # Clock tables are Month/Target/Old/Optimized/%. The coverage table
        # reuses the same first four headers and would overwrite LIM-LD.
        if (a == "Month" and len(row) > 3 and row[1] == "Target"
                and (len(row) < 6 or row[5] in (None, ""))):
            months = []
            i += 1
            while i < len(rows):
                r = rows[i]
                name = r[0] if r else None
                if name in (None, "TOTAL") or (
                        isinstance(name, str) and name.startswith("TOTAL")):
                    break
                months.append({
                    "name": name,
                    "target": r[1],
                    "old": r[2],
                    "new": r[3],
                })
                i += 1
            blocks[title] = months
        i += 1
    return blocks


def _sum(points, key):
    return sum((p.get(key) or 0) for p in points)


def _write_clock_table(ws, row, title, sub, points, chart=False, extra_note=None):
    """Month × Target / S3 predicted / S4 predicted / S4 % / S3 % / S4−S3."""
    r = _xlsx_section(ws, row, title, sub)
    heads = ["Month", "Target", "S3 predicted", "S4 predicted",
             "S4 % of target", "S3 % of target", "S4 − S3"]
    _xlsx_headers(ws, r, heads, center=True)
    header_row = r
    rr = r
    for p in points:
        rr += 1
        _xlsx_text(ws.cell(row=rr, column=1), p["name"], center=True)
        _xlsx_num(ws.cell(row=rr, column=2), p["target"], center=True)
        _xlsx_num(ws.cell(row=rr, column=3), p["s3"], center=True)
        _xlsx_num(ws.cell(row=rr, column=4), p["s4"], True, center=True)
        _xlsx_paint_cov(ws.cell(row=rr, column=5), _cov_pct(p["s4"], p["target"]))
        _xlsx_paint_cov(ws.cell(row=rr, column=6), _cov_pct(p["s3"], p["target"]))
        delta = (p["s4"] or 0) - (p["s3"] or 0)
        _xlsx_num(ws.cell(row=rr, column=7), delta, True, center=True)
    data_last = rr
    rr += 1
    lab = ws.cell(row=rr, column=1, value="TOTAL")
    lab.font = _xlsx_font(True, 11, _XLSX_NAVY)
    lab.alignment = _xlsx_mid()
    _xlsx_total_border(lab)
    tot_t, tot_s3, tot_s4 = _sum(points, "target"), _sum(points, "s3"), _sum(points, "s4")
    for col, val, bold in ((2, tot_t, True), (3, tot_s3, True),
                           (4, tot_s4, True), (7, tot_s4 - tot_s3, True)):
        cell = ws.cell(row=rr, column=col)
        _xlsx_num(cell, val, bold, center=True)
        _xlsx_total_border(cell)
    c5 = ws.cell(row=rr, column=5)
    _xlsx_paint_cov(c5, _cov_pct(tot_s4, tot_t))
    _xlsx_total_border(c5)
    c6 = ws.cell(row=rr, column=6)
    _xlsx_paint_cov(c6, _cov_pct(tot_s3, tot_t))
    _xlsx_total_border(c6)
    if extra_note:
        note = ws.cell(row=rr + 1, column=1, value=extra_note)
        note.font = _xlsx_font(False, 9, _XLSX_MUTED)
        ws.merge_cells(start_row=rr + 1, start_column=1, end_row=rr + 1, end_column=7)
    if chart and data_last > header_row:
        chart_row = rr + 3
        anchor = "A%d" % chart_row
        _xlsx_line_chart(
            ws, title, "tonnes", 2, 4, header_row, data_last,
            anchor, height=9, width=18, cat_col=1, colors=_XLSX_CLOCKS)
        bc = BarChart()
        bc.type = "col"
        bc.grouping = "clustered"
        bc.title = title + " · columns"
        bc.y_axis.title = "tonnes"
        bc.y_axis.scaling.min = 0
        bc.y_axis.numFmt = "#,##0"
        bc.height, bc.width = 9, 18
        bc.legend.position = "t"
        bc.style = None
        data = Reference(ws, min_col=2, max_col=4,
                         min_row=header_row, max_row=data_last)
        cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=data_last)
        bc.add_data(data, titles_from_data=True)
        bc.set_categories(cats)
        palette = (_XLSX_TGT, _XLSX_MUTED, _XLSX_NAVY)
        for s, col in zip(bc.series, palette):
            s.graphicalProperties.solidFill = col
        ws.add_chart(bc, "A%d" % (chart_row + 18))
        return chart_row + 36
    return rr + 3


def build(s3_path, s4_path, out_path):
    s3 = _parse_year_blocks(s3_path)
    s4 = _parse_year_blocks(s4_path)
    together, ld = [], []
    for a, b in zip(s3["Together"], s4["Together"]):
        together.append({
            "name": a["name"], "target": b["target"],
            "s3": a["new"], "s4": b["new"],
        })
    for a, b in zip(s3["LIM-LD"], s4["LIM-LD"]):
        ld.append({
            "name": a["name"], "target": b["target"],
            "s3": a["new"], "s4": b["new"],
        })

    tot_t = _sum(together, "target")
    tot_s3 = _sum(together, "s3")
    tot_s4 = _sum(together, "s4")
    ld_t = _sum(ld, "target")
    ld_s3 = _sum(ld, "s3")
    ld_s4 = _sum(ld, "s4")

    wb = Workbook()
    ws = wb.active
    ws.title = "Year"
    r = _xlsx_board_header(
        ws, "S3 vs S4 · Year dashboard · 2026",
        "S3 predicted = Scenario 3 optimized plan. S4 predicted = Scenario 4 optimized "
        "plan (TF LIM-LD split 50/50 HUAFEI vs POS 12). SAP and LIM-TOS omitted — they match. "
        "Charts are LIM-LD only.")
    r = _xlsx_kpi_strip(ws, r, [
        ("Target", tot_t, _XLSX_TGT, "Year tonnes"),
        ("S3 predicted", tot_s3, _XLSX_MUTED, "Year tonnes"),
        ("S4 predicted", tot_s4, _XLSX_PRED, "Year tonnes"),
    ], start=1)
    r = _xlsx_kpi_strip(ws, r, [
        ("Together · S4 % of target", _cov_pct(tot_s4, tot_t), _XLSX_NAVY, "pct"),
        ("LIM-LD · S3 % of target", _cov_pct(ld_s3, ld_t), _XLSX_MUTED, "pct"),
        ("LIM-LD · S4 % of target", _cov_pct(ld_s4, ld_t), _XLSX_NAVY, "pct"),
        ("LIM-LD · S4 vs S3", _cov_pct(ld_s4, ld_s3), _XLSX_PRED, "pct"),
    ], start=1)

    r = _write_clock_table(
        ws, r, "Together · year",
        "Month on X · tonnes on Y. Target, S3 predicted, S4 predicted. "
        "No chart here — the gap is LIM-LD.",
        together, chart=False)

    r = _write_clock_table(
        ws, r, "LIM-LD · year",
        "The only material that moves between S3 and S4. S4 leftover TF LD trucks "
        "split 50/50 HUAFEI vs POS 12. Line + columns under the table.",
        ld, chart=True,
        extra_note="Sales line is the LIM-LD target total (%.0f t — planning team 2026-08-26)." % ld_t)

    r = _xlsx_section(
        ws, r, "Coverage table",
        "S4 predicted ÷ target. LIM-LD S3 % and S4 % sit next to each other so the lift is visible.")
    heads = ["Month", "Target", "S3 predicted", "S4 predicted",
             "S4 %", "LIM-LD S3 %", "LIM-LD S4 %", "LIM-LD S4 − S3"]
    _xlsx_headers(ws, r, heads, start=1)
    by_ld = {p["name"]: p for p in ld}
    for p in together:
        r += 1
        m = by_ld[p["name"]]
        _xlsx_text(ws.cell(row=r, column=1), p["name"])
        _xlsx_num(ws.cell(row=r, column=2), p["target"])
        _xlsx_num(ws.cell(row=r, column=3), p["s3"])
        _xlsx_num(ws.cell(row=r, column=4), p["s4"], True)
        _xlsx_paint_cov(ws.cell(row=r, column=5), _cov_pct(p["s4"], p["target"]))
        _xlsx_paint_cov(ws.cell(row=r, column=6), _cov_pct(m["s3"], m["target"]))
        _xlsx_paint_cov(ws.cell(row=r, column=7), _cov_pct(m["s4"], m["target"]))
        _xlsx_num(ws.cell(row=r, column=8), (m["s4"] or 0) - (m["s3"] or 0), True)
    r += 1
    tot_lab = ws.cell(row=r, column=1, value="TOTAL · %s months" % len(together))
    tot_lab.font = _xlsx_font(True, 11, _XLSX_NAVY)
    _xlsx_total_border(tot_lab)
    for col, val in ((2, tot_t), (3, tot_s3), (4, tot_s4), (8, ld_s4 - ld_s3)):
        cell = ws.cell(row=r, column=col)
        _xlsx_num(cell, val, True)
        _xlsx_total_border(cell)
    c5 = ws.cell(row=r, column=5)
    _xlsx_paint_cov(c5, _cov_pct(tot_s4, tot_t))
    _xlsx_total_border(c5)
    c6 = ws.cell(row=r, column=6)
    _xlsx_paint_cov(c6, _cov_pct(ld_s3, ld_t))
    _xlsx_total_border(c6)
    c7 = ws.cell(row=r, column=7)
    _xlsx_paint_cov(c7, _cov_pct(ld_s4, ld_t))
    _xlsx_total_border(c7)

    r += 2
    note = ws.cell(
        row=r, column=1,
        value="SAP and LIM-TOS year tables omitted — S3 and S4 land on the same targets. "
              "Source files: plan_2026_S03.xlsx and plan_2026_S04.xlsx.")
    note.font = _xlsx_font(False, 10, _XLSX_MUTED)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    _xlsx_widths(ws, [22, 22, 22, 22, 16, 16, 16, 18])
    ws.freeze_panes = "A4"
    ws.row_dimensions[1].height = 24
    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path, {
        "together_s3": tot_s3, "together_s4": tot_s4, "target": tot_t,
        "ld_s3": ld_s3, "ld_s4": ld_s4, "ld_target": ld_t,
    }


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--s3", default=os.path.expanduser("~/Downloads/plan_2026_S03.xlsx"))
    p.add_argument("--s4", default=os.path.expanduser("~/Downloads/plan_2026_S04.xlsx"))
    p.add_argument("--out", default=os.path.expanduser("~/Downloads/plan_2026_S3_vs_S4.xlsx"))
    args = p.parse_args()
    path, n = build(args.s3, args.s4, args.out)
    print("wrote", path)
    print("Together  S3 {:,.0f}  S4 {:,.0f}  target {:,.0f}".format(
        n["together_s3"], n["together_s4"], n["target"]))
    print("LIM-LD    S3 {:,.0f}  S4 {:,.0f}  target {:,.0f}  S4/S3 {:.1%}".format(
        n["ld_s3"], n["ld_s4"], n["ld_target"], n["ld_s4"] / n["ld_s3"]))


if __name__ == "__main__":
    main()
